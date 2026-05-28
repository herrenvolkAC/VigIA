"""
Base SQLite separada para Carga de Daily.

Esta base queda pensada para consumo directo desde Power BI, sin depender de
vigia.db ni de sus tablas operativas.
"""
from __future__ import annotations

import csv
import os
import re
from datetime import datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import aiosqlite


def _daily_db_path() -> Path:
    configured = os.getenv("DAILY_OPERATIVA_DB_PATH", "").strip()
    if configured:
        return Path(configured).expanduser()
    return Path(__file__).resolve().parent / "daily_operativa.db"


DAILY_DB_PATH = _daily_db_path()
DAILY_POWERBI_CSV_PATH = DAILY_DB_PATH.with_name("daily_powerbi.csv")
DAILY_CONSOLIDADO_CSV_PATH = DAILY_DB_PATH.with_name("daily_consolidado_powerbi.csv")
try:
    LOCAL_TZ = ZoneInfo("America/Argentina/Buenos_Aires")
except ZoneInfoNotFoundError:
    LOCAL_TZ = timezone(timedelta(hours=-3), name="America/Argentina/Buenos_Aires")
LOAD_CUTOFF = time(6, 30)

CONSOLIDADO_HEADERS = [
    "Semana",
    "Division",
    "UP",
    "Metrica",
    "Dia Daily",
    "Fecha Daily",
    "Semana",
    "REAL",
    "TARGET/PLAN",
    "%",
    "COLOR",
    "Filtro",
    "Fecha_hoy",
    "Filtro de semana",
    "Dia_hoy",
    "Filtro de avance",
]
CONSOLIDADO_SEED_DEFAULT = Path(r"C:\Users\207189\Downloads\DMS Daily Consolidado.xlsx")


CREATE_DAILY_CARGAS = """
CREATE TABLE IF NOT EXISTS daily_cargas (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    daily_key           TEXT NOT NULL,
    daily_label         TEXT NOT NULL,
    fecha_inicio        TEXT NOT NULL,
    fecha_fin           TEXT NOT NULL,
    fecha_carga         TEXT NOT NULL,
    timestamp_carga     TEXT NOT NULL,
    usuario_carga       TEXT NOT NULL,
    tipo_daily          TEXT NOT NULL,
    sector              TEXT,
    turno               TEXT,
    plan                TEXT,
    version             INTEGER DEFAULT 1,
    reemplazado         INTEGER DEFAULT 0,
    created_at          TEXT DEFAULT CURRENT_TIMESTAMP,
    updated_at          TEXT DEFAULT CURRENT_TIMESTAMP
);
"""

CREATE_DAILY_PARAMETROS = """
CREATE TABLE IF NOT EXISTS daily_parametros_cumplimiento (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    id_parametro        TEXT NOT NULL UNIQUE,
    tipo_daily          TEXT NOT NULL,
    sector_aplicable    TEXT,
    grupo               TEXT NOT NULL,
    proceso             TEXT,
    ventana_horaria     TEXT,
    fuente              TEXT,
    nombre              TEXT NOT NULL,
    descripcion         TEXT,
    unidad              TEXT,
    tipo_campo          TEXT NOT NULL,
    valor_esperado      REAL,
    regla_cumplimiento  TEXT DEFAULT 'informativo',
    orden               INTEGER DEFAULT 0,
    activo              INTEGER DEFAULT 1,
    created_at          TEXT DEFAULT CURRENT_TIMESTAMP,
    updated_at          TEXT DEFAULT CURRENT_TIMESTAMP
);
"""

CREATE_DAILY_DETALLE = """
CREATE TABLE IF NOT EXISTS daily_carga_detalle (
    id                          INTEGER PRIMARY KEY AUTOINCREMENT,
    carga_id                    INTEGER NOT NULL REFERENCES daily_cargas(id) ON DELETE CASCADE,
    id_parametro                TEXT NOT NULL,
    valor_real_texto            TEXT,
    valor_real_numero           REAL,
    valor_esperado_snapshot     REAL,
    desvio                      REAL,
    porcentaje_cumplimiento     REAL,
    porcentaje_desvio           REAL,
    estado_cumplimiento         TEXT,
    created_at                  TEXT DEFAULT CURRENT_TIMESTAMP
);
"""

CREATE_DAILY_CONSOLIDADO_HISTORICO = """
CREATE TABLE IF NOT EXISTS daily_consolidado_historico (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    source_file         TEXT NOT NULL,
    source_row          INTEGER NOT NULL,
    semana_label        TEXT,
    division            TEXT,
    up                  TEXT,
    metrica             TEXT,
    dia_daily           TEXT,
    fecha_daily         TEXT,
    semana_num          TEXT,
    real                REAL,
    target_plan         REAL,
    porcentaje          REAL,
    color               TEXT,
    filtro              TEXT,
    fecha_hoy           TEXT,
    filtro_semana       TEXT,
    dia_hoy             TEXT,
    filtro_avance       TEXT,
    imported_at         TEXT DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(source_file, source_row)
);
"""

CREATE_POWERBI_VIEW = """
CREATE VIEW IF NOT EXISTS vw_daily_powerbi AS
SELECT
    c.id AS carga_id,
    c.daily_key,
    c.daily_label,
    c.fecha_inicio,
    c.fecha_fin,
    c.fecha_carga,
    c.timestamp_carga,
    c.usuario_carga,
    c.tipo_daily,
    c.sector,
    c.turno,
    c.plan,
    c.version,
    c.reemplazado,
    p.id_parametro,
    p.grupo,
    p.proceso,
    p.ventana_horaria,
    p.fuente,
    p.nombre AS nombre_parametro,
    p.descripcion AS descripcion_parametro,
    p.unidad,
    p.tipo_campo,
    d.valor_esperado_snapshot AS valor_esperado,
    d.valor_real_texto,
    d.valor_real_numero,
    d.desvio,
    d.porcentaje_cumplimiento,
    d.porcentaje_desvio,
    d.estado_cumplimiento,
    p.regla_cumplimiento,
    p.orden
FROM daily_cargas c
JOIN daily_carga_detalle d ON d.carga_id = c.id
JOIN daily_parametros_cumplimiento p ON p.id_parametro = d.id_parametro
WHERE c.reemplazado = 0;
"""

CREATE_POWERBI_RESUMEN_VIEW = """
CREATE VIEW IF NOT EXISTS vw_daily_resumen_powerbi AS
SELECT
    c.id AS carga_id,
    c.daily_key,
    c.daily_label,
    c.fecha_inicio,
    c.fecha_fin,
    c.fecha_carga,
    c.timestamp_carga,
    c.usuario_carga,
    c.tipo_daily,
    c.sector,
    c.turno,
    c.plan,
    c.version,
    c.reemplazado,
    COUNT(d.id) AS cantidad_parametros,
    SUM(CASE WHEN d.estado_cumplimiento = 'Rojo' THEN 1 ELSE 0 END) AS cantidad_rojos,
    SUM(CASE WHEN d.estado_cumplimiento = 'Amarillo' THEN 1 ELSE 0 END) AS cantidad_amarillos,
    SUM(CASE WHEN d.estado_cumplimiento = 'Verde claro' THEN 1 ELSE 0 END) AS cantidad_verde_claro,
    SUM(CASE WHEN d.estado_cumplimiento = 'Verde oscuro' THEN 1 ELSE 0 END) AS cantidad_verde_oscuro,
    AVG(d.porcentaje_cumplimiento) AS porcentaje_promedio_cumplimiento
FROM daily_cargas c
LEFT JOIN daily_carga_detalle d ON d.carga_id = c.id
WHERE c.reemplazado = 0
GROUP BY c.id;
"""

CREATE_INDEXES = [
    "CREATE INDEX IF NOT EXISTS idx_daily_cargas_lookup ON daily_cargas(daily_key, tipo_daily, sector, reemplazado)",
    "CREATE INDEX IF NOT EXISTS idx_daily_cargas_timestamp ON daily_cargas(timestamp_carga DESC)",
    "CREATE INDEX IF NOT EXISTS idx_daily_detalle_carga ON daily_carga_detalle(carga_id)",
    "CREATE INDEX IF NOT EXISTS idx_daily_parametros_tipo ON daily_parametros_cumplimiento(tipo_daily, activo, orden)",
    "CREATE INDEX IF NOT EXISTS idx_daily_consolidado_hist_key ON daily_consolidado_historico(fecha_daily, division, up, metrica)",
]


PARAMETROS_SEED: list[dict[str, Any]] = [
    # Operacion - Productividad
    {"id_parametro": "OP_PROD_RECEPCION_SECOS_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Secos", "grupo": "PRODUCTIVIDAD", "proceso": "RECEPCION", "ventana_horaria": "6 a 6", "fuente": "Plataforma PI", "nombre": "Pallets recepcionados por legajo y turno", "descripcion": "Cuantos pallets se recepcionaron, por legajo y turno?", "unidad": "Pallet / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 160, "regla_cumplimiento": "productividad", "orden": 10},
    {"id_parametro": "OP_PROD_PICKING_SECOS_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Secos", "grupo": "PRODUCTIVIDAD", "proceso": "PICKING", "ventana_horaria": "6 a 6", "fuente": "Reporte WF", "nombre": "Bultos armados por legajo y turno", "descripcion": "Cuantos bultos se armaron, por legajo y turno? Filtrar Logisu + Mayores 45.", "unidad": "Bultos / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 1050, "regla_cumplimiento": "productividad", "orden": 20},
    {"id_parametro": "OP_PROD_CLARK_SECOS_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Secos", "grupo": "PRODUCTIVIDAD", "proceso": "CLARK", "ventana_horaria": "6 a 6", "fuente": "Reporte WF", "nombre": "Pallets movilizados por legajo y turno", "descripcion": "Cuantos pallets se movilizaron, por legajo y turno? Filtrar Logisu + Mayores 45.", "unidad": "Pallet / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 120, "regla_cumplimiento": "productividad", "orden": 30},
    {"id_parametro": "OP_PROD_DESPACHO_SECOS_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Secos", "grupo": "PRODUCTIVIDAD", "proceso": "DESPACHO", "ventana_horaria": "6 a 6", "fuente": "Reporte WF", "nombre": "Viajes despachados por legajo y turno", "descripcion": "Cuantos viajes se despacharon, por legajo y turno? Filtrar Logisu + Mayores 45.", "unidad": "Viajes / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 6, "regla_cumplimiento": "productividad", "orden": 40},
    {"id_parametro": "OP_PROD_RECEPCION_NOA_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Noa", "grupo": "PRODUCTIVIDAD", "proceso": "RECEPCION", "ventana_horaria": "6 a 6", "fuente": "Plataforma PI", "nombre": "Pallets recepcionados por legajo y turno", "descripcion": "Cuantos pallets se recepcionaron, por legajo y turno?", "unidad": "Pallet / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 125, "regla_cumplimiento": "productividad", "orden": 10},
    {"id_parametro": "OP_PROD_PICKING_NOA_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Noa", "grupo": "PRODUCTIVIDAD", "proceso": "PICKING", "ventana_horaria": "6 a 6", "fuente": "Reporte WF", "nombre": "Bultos armados por legajo y turno", "descripcion": "Cuantos bultos se armaron, por legajo y turno? Filtrar Logisu + Mayores 45.", "unidad": "Bultos / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 850, "regla_cumplimiento": "productividad", "orden": 20},
    {"id_parametro": "OP_PROD_CLARK_NOA_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Noa", "grupo": "PRODUCTIVIDAD", "proceso": "CLARK", "ventana_horaria": "6 a 6", "fuente": "Reporte WF", "nombre": "Pallets movilizados por legajo y turno", "descripcion": "Cuantos pallets se movilizaron, por legajo y turno? Filtrar Logisu + Mayores 45.", "unidad": "Pallet / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 120, "regla_cumplimiento": "productividad", "orden": 30},
    {"id_parametro": "OP_PROD_RECEPCION_REFRI_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Refrigerados", "grupo": "PRODUCTIVIDAD", "proceso": "RECEPCION", "ventana_horaria": "6 a 6", "fuente": "Plataforma PI", "nombre": "Pallets recepcionados por legajo y turno", "descripcion": "Cuantos pallets se recepcionaron, por legajo y turno?", "unidad": "Pallet / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 80, "regla_cumplimiento": "productividad", "orden": 10},
    {"id_parametro": "OP_PROD_PICKING_REFRI_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Refrigerados", "grupo": "PRODUCTIVIDAD", "proceso": "PICKING", "ventana_horaria": "6 a 6", "fuente": "Reporte WF", "nombre": "Bultos armados por legajo y turno", "descripcion": "Cuantos bultos se armaron, por legajo y turno? Filtrar Logisu + Mayores 45.", "unidad": "Bultos / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 910, "regla_cumplimiento": "productividad", "orden": 20},
    {"id_parametro": "OP_PROD_CLARK_REFRI_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Refrigerados", "grupo": "PRODUCTIVIDAD", "proceso": "CLARK", "ventana_horaria": "6 a 6", "fuente": "Reporte WF", "nombre": "Pallets movilizados por legajo y turno", "descripcion": "Cuantos pallets se movilizaron, por legajo y turno? Filtrar Logisu + Mayores 45.", "unidad": "Pallet / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 130, "regla_cumplimiento": "productividad", "orden": 30},
    {"id_parametro": "OP_PROD_DESPACHO_REFRI_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Refrigerados", "grupo": "PRODUCTIVIDAD", "proceso": "DESPACHO", "ventana_horaria": "6 a 6", "fuente": "Reporte WF", "nombre": "Viajes despachados por legajo y turno", "descripcion": "Cuantos viajes se despacharon, por legajo y turno? Filtrar Logisu + Mayores 45.", "unidad": "Viajes / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 6, "regla_cumplimiento": "productividad", "orden": 40},
    # Operacion - Cumplimiento
    {"id_parametro": "OP_CUMP_RECEPCION_PLAN_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "CUMPLIMIENTO", "proceso": "RECEPCION", "ventana_horaria": "6 a 6", "fuente": "Excel Planeamiento IN", "nombre": "Pallets planificados para recibir", "descripcion": "Cuantos pallets se planificaron, desde C.Proveedores, para recibir?", "unidad": "Pallets", "tipo_campo": "numerico", "valor_esperado": 100, "regla_cumplimiento": "informativo", "orden": 50},
    {"id_parametro": "OP_CUMP_RECEPCION_REAL_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "CUMPLIMIENTO", "proceso": "RECEPCION", "ventana_horaria": "6 a 6", "fuente": "Plataforma PI", "nombre": "Pallets reales recepcionados", "descripcion": "Cuantos pallets reales se recepcionaron mediante tareas de Recepcion por Plataforma PI?", "unidad": "Pallets", "tipo_campo": "numerico", "valor_esperado": 100, "regla_cumplimiento": "cumplimiento", "orden": 60},
    {"id_parametro": "OP_CUMP_PICKING_PLAN_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "CUMPLIMIENTO", "proceso": "PICKING", "ventana_horaria": "6 a 6", "fuente": "Excel Planeamiento OUT", "nombre": "Bultos planificados para armar", "descripcion": "Cuantos bultos se planificaron para armar mediante tarea de Picking por TMS?", "unidad": "Bultos", "tipo_campo": "numerico", "valor_esperado": 100, "regla_cumplimiento": "informativo", "orden": 70},
    {"id_parametro": "OP_CUMP_PICKING_REAL_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "CUMPLIMIENTO", "proceso": "PICKING", "ventana_horaria": "6 a 6", "fuente": "Reportes WF", "nombre": "Bultos reales armados", "descripcion": "Cuantos bultos reales se armaron mediante tarea de Picking por WMS?", "unidad": "Bultos", "tipo_campo": "numerico", "valor_esperado": 100, "regla_cumplimiento": "cumplimiento", "orden": 80},
    {"id_parametro": "OP_CUMP_SPC_PLAN_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "CUMPLIMIENTO", "proceso": "SPC", "ventana_horaria": "6 a 6", "fuente": "Excel Planeamiento OUT", "nombre": "Pallets planificados para SPC", "descripcion": "Cuantos pallets se planificaron para armar mediante tarea de SPC por TMS?", "unidad": "Pallets", "tipo_campo": "numerico", "valor_esperado": 100, "regla_cumplimiento": "informativo", "orden": 90},
    {"id_parametro": "OP_CUMP_SPC_REAL_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "CUMPLIMIENTO", "proceso": "SPC", "ventana_horaria": "6 a 6", "fuente": "Reportes WF", "nombre": "Pallets reales movilizados por SPC", "descripcion": "Cuantos pallets se movilizaron mediante tarea de SPC por WMS?", "unidad": "Pallets", "tipo_campo": "numerico", "valor_esperado": 100, "regla_cumplimiento": "cumplimiento", "orden": 100},
    {"id_parametro": "OP_CUMP_DESPACHO_PLAN_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "CUMPLIMIENTO", "proceso": "DESPACHO", "ventana_horaria": "6 a 6", "fuente": "Excel Planeamiento OUT", "nombre": "Viajes planificados para despachar", "descripcion": "Cuantos viajes se planificaron para despachar por TMS?", "unidad": "Camiones", "tipo_campo": "numerico", "valor_esperado": 100, "regla_cumplimiento": "informativo", "orden": 110},
    {"id_parametro": "OP_CUMP_DESPACHO_REAL_6A6", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "CUMPLIMIENTO", "proceso": "DESPACHO", "ventana_horaria": "6 a 6", "fuente": "Reportes WF", "nombre": "Viajes reales despachados", "descripcion": "Cuantos viajes reales se despacharon por WMS?", "unidad": "Camiones", "tipo_campo": "numerico", "valor_esperado": 100, "regla_cumplimiento": "cumplimiento", "orden": 120},
    # Operacion - Avance
    {"id_parametro": "OP_AVANCE_RECEPCION_PLAN_6A8", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "AVANCE", "proceso": "RECEPCION", "ventana_horaria": "6 a 8", "fuente": "Excel Planeamiento IN", "nombre": "Pallets planificados para recepcionar hasta 8", "descripcion": "Cuantos pallets se planificaron para recepcionar mediante tarea de Recepcion por C.Proveedores?", "unidad": "Pallets", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 130},
    {"id_parametro": "OP_AVANCE_RECEPCION_REAL_6A8", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "AVANCE", "proceso": "RECEPCION", "ventana_horaria": "6 a 8", "fuente": "Plataforma PI", "nombre": "Pallets recepcionados hasta el momento", "descripcion": "Cuantos pallets se recepcionaron hasta el momento por Plataforma PI?", "unidad": "Pallets", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 140},
    {"id_parametro": "OP_AVANCE_PICKING_PLAN_6A8", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "AVANCE", "proceso": "PICKING", "ventana_horaria": "6 a 8", "fuente": "Excel Planeamiento OUT", "nombre": "Bultos planificados para armar hasta 8", "descripcion": "Cuantos bultos se planificaron para armar mediante tarea de Picking por TMS?", "unidad": "Bultos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 150},
    {"id_parametro": "OP_AVANCE_PICKING_REAL_6A8", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "AVANCE", "proceso": "PICKING", "ventana_horaria": "6 a 8", "fuente": "Reportes WF", "nombre": "Bultos armados hasta el momento", "descripcion": "Cuantos bultos se armaron hasta el momento por WMS?", "unidad": "Bultos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 160},
    {"id_parametro": "OP_AVANCE_SPC_PLAN_6A8", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "AVANCE", "proceso": "SPC", "ventana_horaria": "6 a 8", "fuente": "Excel Planeamiento OUT", "nombre": "Pallets planificados para SPC hasta 8", "descripcion": "Cuantos pallets se planificaron para armar mediante tarea de SPC por TMS?", "unidad": "Pallets", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 170},
    {"id_parametro": "OP_AVANCE_SPC_REAL_6A8", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "AVANCE", "proceso": "SPC", "ventana_horaria": "6 a 8", "fuente": "Reportes WF", "nombre": "Pallets surtidos hasta el momento", "descripcion": "Cuantos pallets se surtieron hasta el momento por WMS?", "unidad": "Pallets", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 180},
    {"id_parametro": "OP_AVANCE_DESPACHO_PLAN_6A8", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "AVANCE", "proceso": "DESPACHO", "ventana_horaria": "6 a 8", "fuente": "Excel Planeamiento OUT", "nombre": "Viajes planificados para despachar hasta 8", "descripcion": "Cuantos viajes se planificaron para despachar por TMS?", "unidad": "Camiones", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 190},
    {"id_parametro": "OP_AVANCE_DESPACHO_REAL_6A8", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "AVANCE", "proceso": "DESPACHO", "ventana_horaria": "6 a 8", "fuente": "Reportes WF", "nombre": "Viajes despachados hasta el momento", "descripcion": "Cuantos viajes se despacharon hasta el momento por WMS?", "unidad": "Camiones", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 200},
    # Operacion - Desvios y dotacion
    {"id_parametro": "OP_DESVIO_IDENTIFICADO", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "DESVIOS", "proceso": "DESVIO", "ventana_horaria": "", "fuente": "Carga manual", "nombre": "Desvio identificado", "descripcion": "Cual es el desvio identificado? Colocar un numero ID adelante de cada desvio para trazabilidad.", "unidad": "", "tipo_campo": "texto", "valor_esperado": None, "regla_cumplimiento": "texto", "orden": 210},
    {"id_parametro": "OP_DESVIO_CAUSA_RAIZ", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "DESVIOS", "proceso": "DESVIO", "ventana_horaria": "", "fuente": "Carga manual", "nombre": "Causa raiz", "descripcion": "Cual es su principal causa raiz?", "unidad": "", "tipo_campo": "texto", "valor_esperado": None, "regla_cumplimiento": "texto", "orden": 220},
    {"id_parametro": "OP_DESVIO_PLAN_ACCION", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "DESVIOS", "proceso": "DESVIO", "ventana_horaria": "", "fuente": "Carga manual", "nombre": "Plan de accion asociado", "descripcion": "Tiene plan de accion asociado? Si se deja vacio se identifica como No.", "unidad": "", "tipo_campo": "texto", "valor_esperado": None, "regla_cumplimiento": "texto", "orden": 230},
    {"id_parametro": "OP_DESVIO_RESPONSABLE_ACCION", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "DESVIOS", "proceso": "DESVIO", "ventana_horaria": "", "fuente": "Carga manual", "nombre": "Responsable del plan de accion", "descripcion": "Quien es su responsable? Si se deja vacio se identifica como No.", "unidad": "", "tipo_campo": "texto", "valor_esperado": None, "regla_cumplimiento": "texto", "orden": 240},
    {"id_parametro": "OP_DESVIO_FECHA_AVANCE", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "DESVIOS", "proceso": "DESVIO", "ventana_horaria": "", "fuente": "Carga manual", "nombre": "Fecha planificada de avances", "descripcion": "Cual es la fecha planificada de avances concretos? Si se deja vacio se identifica como No.", "unidad": "", "tipo_campo": "fecha", "valor_esperado": None, "regla_cumplimiento": "texto", "orden": 250},
    {"id_parametro": "OP_DOT_PICKING_LEGAJOS_6A8", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "DOTACION", "proceso": "PICKING", "ventana_horaria": "6 a 8", "fuente": "Reportes WF", "nombre": "Legajos activos armando en turno manana", "descripcion": "Cuantos legajos activos tenemos armando en el Turno Manana Corriente por WMS?", "unidad": "Legajos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 260},
    {"id_parametro": "OP_DOT_SPC_LEGAJOS_6A8", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "DOTACION", "proceso": "SPC", "ventana_horaria": "6 a 8", "fuente": "Reportes WF", "nombre": "Legajos clarkistas activos en turno manana", "descripcion": "Cuantos legajos activos tenemos de Clarkistas en el Turno Manana Corriente por WMS?", "unidad": "Legajos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 270},
    {"id_parametro": "OP_DOT_DESPACHO_LEGAJOS_6A8", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "DOTACION", "proceso": "DESPACHO", "ventana_horaria": "6 a 8", "fuente": "Reportes WF", "nombre": "Legajos activos cargando en turno manana", "descripcion": "Cuantos legajos activos tenemos cargando en el Turno Manana Corriente por WMS?", "unidad": "Legajos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 280},
    {"id_parametro": "OP_DOT_RECEPCION_LEGAJOS_6A8", "tipo_daily": "Operacion", "sector_aplicable": "Noa|Secos|Refrigerados", "grupo": "DOTACION", "proceso": "RECEPCION", "ventana_horaria": "6 a 8", "fuente": "Plataforma PI", "nombre": "Legajos activos recepcionando en turno manana", "descripcion": "Cuantos legajos activos tenemos recepcionando en el Turno Manana Corriente por WMS?", "unidad": "Legajos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 290},
    # Logistica inversa
    {"id_parametro": "LI_PROD_CARGA_DESCARGA_MOVILES_6A6", "tipo_daily": "Logistica Inversa", "sector_aplicable": "Logistica Inversa", "grupo": "PRODUCTIVIDAD", "proceso": "CARGA Y DESCARGA MOVILES", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Moviles recepcionados o cargados por legajo y turno", "descripcion": "Cuantos moviles se recepcionaron y/o cargaron, por legajo y turno?", "unidad": "Mov / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 8.4, "regla_cumplimiento": "productividad", "orden": 300},
    {"id_parametro": "LI_PROD_RECEPCION_ENVASES_6A6", "tipo_daily": "Logistica Inversa", "sector_aplicable": "Logistica Inversa", "grupo": "PRODUCTIVIDAD", "proceso": "RECEPCION ENVASES", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Documentos recepcionados por legajo y turno", "descripcion": "Cuantos documentos se recepcionaron, por legajo y turno?", "unidad": "Doc / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 10, "regla_cumplimiento": "productividad", "orden": 310},
    {"id_parametro": "LI_PROD_RECEPCION_DEVOLUCION_6A6", "tipo_daily": "Logistica Inversa", "sector_aplicable": "Logistica Inversa", "grupo": "PRODUCTIVIDAD", "proceso": "RECEPCION DEVOLUCION", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Documentos recepcionados por legajo y turno", "descripcion": "Cuantos documentos se recepcionaron, por legajo y turno?", "unidad": "Doc / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 45, "regla_cumplimiento": "productividad", "orden": 320},
    {"id_parametro": "LI_PROD_ENFARDADORA_6A6", "tipo_daily": "Logistica Inversa", "sector_aplicable": "Logistica Inversa", "grupo": "PRODUCTIVIDAD", "proceso": "ENFARDADORA", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Fardos producidos por legajo y turno", "descripcion": "Cuantos fardos se produjeron, por legajo y turno?", "unidad": "Fardos / Turno x Operario", "tipo_campo": "numerico", "valor_esperado": 30, "regla_cumplimiento": "productividad", "orden": 330},
    {"id_parametro": "LI_CUMP_RECEPCION_ENVASES_PROX_VENCER_6A6", "tipo_daily": "Logistica Inversa", "sector_aplicable": "Logistica Inversa", "grupo": "CUMPLIMIENTO", "proceso": "RECEPCION ENVASES", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Documentos proximos a vencer", "descripcion": "Cuantos documentos se encuentran proximos a vencer?", "unidad": "Documentos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "cumplimiento", "orden": 340},
    {"id_parametro": "LI_CUMP_RECEPCION_ENVASES_VENCIDOS_6A6", "tipo_daily": "Logistica Inversa", "sector_aplicable": "Logistica Inversa", "grupo": "CUMPLIMIENTO", "proceso": "RECEPCION ENVASES", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Documentos vencidos", "descripcion": "Cuantos documentos se encuentran vencidos?", "unidad": "Documentos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "cumplimiento", "orden": 350},
    {"id_parametro": "LI_CUMP_RECEPCION_DEVOLUCION_PROX_VENCER_6A6", "tipo_daily": "Logistica Inversa", "sector_aplicable": "Logistica Inversa", "grupo": "CUMPLIMIENTO", "proceso": "RECEPCION DEVOLUCION", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Documentos proximos a vencer", "descripcion": "Cuantos documentos se encuentran proximos a vencer?", "unidad": "Documentos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "cumplimiento", "orden": 360},
    {"id_parametro": "LI_CUMP_RECEPCION_DEVOLUCION_VENCIDOS_6A6", "tipo_daily": "Logistica Inversa", "sector_aplicable": "Logistica Inversa", "grupo": "CUMPLIMIENTO", "proceso": "RECEPCION DEVOLUCION", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Documentos vencidos", "descripcion": "Cuantos documentos se encuentran vencidos?", "unidad": "Documentos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "cumplimiento", "orden": 370},
    {"id_parametro": "LI_CUMP_RECEPCION_ROTURAS_PROX_VENCER_6A6", "tipo_daily": "Logistica Inversa", "sector_aplicable": "Logistica Inversa", "grupo": "CUMPLIMIENTO", "proceso": "RECEPCION ROTURAS", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Documentos proximos a vencer", "descripcion": "Cuantos documentos se encuentran proximos a vencer?", "unidad": "Documentos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "cumplimiento", "orden": 380},
    {"id_parametro": "LI_CUMP_RECEPCION_ROTURAS_VENCIDOS_6A6", "tipo_daily": "Logistica Inversa", "sector_aplicable": "Logistica Inversa", "grupo": "CUMPLIMIENTO", "proceso": "RECEPCION ROTURAS", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Documentos vencidos", "descripcion": "Cuantos documentos se encuentran vencidos?", "unidad": "Documentos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "cumplimiento", "orden": 390},
    {"id_parametro": "LI_CUMP_ENFARDADORA_PLAN_6A6", "tipo_daily": "Logistica Inversa", "sector_aplicable": "Logistica Inversa", "grupo": "CUMPLIMIENTO", "proceso": "ENFARDADORA", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Fardos planificados para producir", "descripcion": "Cuantos fardos se planificaron para producir?", "unidad": "Fardos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 400},
    # Planeamiento
    {"id_parametro": "PL_PROD_HUECOS_6A6", "tipo_daily": "Planeamiento", "sector_aplicable": "Planeamiento", "grupo": "PRODUCTIVIDAD", "proceso": "HUECOS", "ventana_horaria": "6 a 6", "fuente": "MicroStrategy", "nombre": "Huecos disponibles", "descripcion": "Cuantos huecos se tienen?", "unidad": "Huecos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 410},
    {"id_parametro": "PL_PROD_PLAN_CORRIENTE_HORA_ENVIO", "tipo_daily": "Planeamiento", "sector_aplicable": "Planeamiento", "grupo": "PRODUCTIVIDAD", "proceso": "Programa", "ventana_horaria": "Plan corriente", "fuente": "Carga manual", "nombre": "Hora de envio del programa diario", "descripcion": "A que hora se envio el programa diario?", "unidad": "HH:MM", "tipo_campo": "hora", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 420},
    {"id_parametro": "PL_PROD_RECURSO_FTE_6A6", "tipo_daily": "Planeamiento", "sector_aplicable": "Planeamiento", "grupo": "PRODUCTIVIDAD", "proceso": "Recurso FTE", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Relacion entre FTE real y dotacion", "descripcion": "Cual es la relacion entre FTE Real y Dotacion?", "unidad": "FTE / Dot", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "productividad", "orden": 430},
    {"id_parametro": "PL_PROD_COORD_PROVEEDORES_6A6", "tipo_daily": "Planeamiento", "sector_aplicable": "Planeamiento", "grupo": "PRODUCTIVIDAD", "proceso": "Coordinacion de Proveedores", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Ocupacion del CD", "descripcion": "En que porcentaje de ocupacion se encuentra nuestro CD?", "unidad": "%", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "productividad", "orden": 440},
    {"id_parametro": "PL_CUMP_PROGRAMA_BULTOS_PROGRAMADOS_0A0", "tipo_daily": "Planeamiento", "sector_aplicable": "Planeamiento", "grupo": "CUMPLIMIENTO", "proceso": "Programa", "ventana_horaria": "0 a 0", "fuente": "Carga manual", "nombre": "Bultos programados para despachar", "descripcion": "Cuantos bultos fueron PROGRAMADOS para despachar en el total de los sectores?", "unidad": "Bultos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 450},
    {"id_parametro": "PL_CUMP_PROGRAMA_BULTOS_PLANIFICADOS_0A0", "tipo_daily": "Planeamiento", "sector_aplicable": "Planeamiento", "grupo": "CUMPLIMIENTO", "proceso": "Programa", "ventana_horaria": "0 a 0", "fuente": "Carga manual", "nombre": "Bultos planificados para despachar", "descripcion": "Cuantos bultos fueron PLANIFICADOS para despachar en el total de los sectores?", "unidad": "Bultos", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "cumplimiento", "orden": 460},
    {"id_parametro": "PL_CUMP_PROGRAMA_FTE_UTILIZADO_6A6", "tipo_daily": "Planeamiento", "sector_aplicable": "Planeamiento", "grupo": "CUMPLIMIENTO", "proceso": "Programa", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "FTE utilizado en la operacion", "descripcion": "Cuantos FTE fueron UTILIZADOS en la operacion?", "unidad": "FTE", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "informativo", "orden": 470},
    {"id_parametro": "PL_CUMP_PROGRAMA_FTE_PLANIFICADO_6A6", "tipo_daily": "Planeamiento", "sector_aplicable": "Planeamiento", "grupo": "CUMPLIMIENTO", "proceso": "Programa", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "FTE planificado en la operacion", "descripcion": "Cuantos FTE fueron PLANIFICADOS en la operacion?", "unidad": "FTE", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "cumplimiento", "orden": 480},
    {"id_parametro": "PL_CUMP_COORD_PROVEEDORES_MAILS_6A6", "tipo_daily": "Planeamiento", "sector_aplicable": "Planeamiento", "grupo": "CUMPLIMIENTO", "proceso": "Coordinacion de Proveedores", "ventana_horaria": "6 a 6", "fuente": "Carga manual", "nombre": "Mails vencidos", "descripcion": "Cuantos mails vencidos tenemos?", "unidad": "Mails", "tipo_campo": "numerico", "valor_esperado": None, "regla_cumplimiento": "cumplimiento", "orden": 490},
]


async def init_daily_db() -> None:
    DAILY_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    async with aiosqlite.connect(DAILY_DB_PATH) as db:
        await db.execute("PRAGMA journal_mode = WAL")
        await db.execute("PRAGMA busy_timeout = 10000")
        await db.execute(CREATE_DAILY_CARGAS)
        await db.execute(CREATE_DAILY_PARAMETROS)
        await db.execute(CREATE_DAILY_DETALLE)
        await db.execute(CREATE_DAILY_CONSOLIDADO_HISTORICO)
        for statement in CREATE_INDEXES:
            await db.execute(statement)
        await db.execute("DROP VIEW IF EXISTS vw_daily_powerbi")
        await db.execute("DROP VIEW IF EXISTS vw_daily_resumen_powerbi")
        await db.execute(CREATE_POWERBI_VIEW)
        await db.execute(CREATE_POWERBI_RESUMEN_VIEW)
        await _seed_parametros(db)
        await _apply_parametros_migrations(db)
        await _seed_consolidado_historico_if_available(db)
        await db.commit()


async def _seed_parametros(db: aiosqlite.Connection) -> None:
    await db.executemany(
        """
        INSERT OR IGNORE INTO daily_parametros_cumplimiento (
            id_parametro, tipo_daily, sector_aplicable, grupo, proceso,
            ventana_horaria, fuente, nombre, descripcion, unidad, tipo_campo,
            valor_esperado, regla_cumplimiento, orden, activo
        ) VALUES (
            :id_parametro, :tipo_daily, :sector_aplicable, :grupo, :proceso,
            :ventana_horaria, :fuente, :nombre, :descripcion, :unidad,
            :tipo_campo, :valor_esperado, :regla_cumplimiento, :orden, 1
        )
        """,
        PARAMETROS_SEED,
    )


async def _apply_parametros_migrations(db: aiosqlite.Connection) -> None:
    now = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
    old_operacion_ids = [
        "OP_PROD_RECEPCION_6A6",
        "OP_PROD_PICKING_6A6",
        "OP_PROD_CLARK_6A6",
        "OP_PROD_DESPACHO_6A6",
    ]
    placeholders = ",".join("?" for _ in old_operacion_ids)
    await db.execute(
        f"""
        UPDATE daily_parametros_cumplimiento
        SET activo = 0, updated_at = ?
        WHERE id_parametro IN ({placeholders})
        """,
        (now, *old_operacion_ids),
    )

    force_ids = {
        "OP_PROD_RECEPCION_SECOS_6A6",
        "OP_PROD_PICKING_SECOS_6A6",
        "OP_PROD_CLARK_SECOS_6A6",
        "OP_PROD_DESPACHO_SECOS_6A6",
        "OP_PROD_RECEPCION_NOA_6A6",
        "OP_PROD_PICKING_NOA_6A6",
        "OP_PROD_CLARK_NOA_6A6",
        "OP_PROD_RECEPCION_REFRI_6A6",
        "OP_PROD_PICKING_REFRI_6A6",
        "OP_PROD_CLARK_REFRI_6A6",
        "OP_PROD_DESPACHO_REFRI_6A6",
        "LI_PROD_CARGA_DESCARGA_MOVILES_6A6",
        "LI_PROD_RECEPCION_ENVASES_6A6",
        "LI_PROD_RECEPCION_DEVOLUCION_6A6",
        "LI_PROD_ENFARDADORA_6A6",
    }
    seed_by_id = {item["id_parametro"]: item for item in PARAMETROS_SEED}
    for id_parametro in sorted(force_ids):
        item = seed_by_id.get(id_parametro)
        if not item:
            continue
        await db.execute(
            """
            UPDATE daily_parametros_cumplimiento
            SET sector_aplicable = ?,
                unidad = ?,
                valor_esperado = ?,
                regla_cumplimiento = ?,
                orden = ?,
                activo = 1,
                updated_at = ?
            WHERE id_parametro = ?
            """,
            (
                item.get("sector_aplicable") or "",
                item.get("unidad") or "",
                item.get("valor_esperado"),
                item.get("regla_cumplimiento") or "informativo",
                item.get("orden") or 0,
                now,
                id_parametro,
            ),
        )


async def _seed_consolidado_historico_if_available(db: aiosqlite.Connection) -> None:
    async with db.execute("SELECT COUNT(*) FROM daily_consolidado_historico") as cur:
        count = int((await cur.fetchone())[0] or 0)
    if count:
        return

    configured = os.getenv("DAILY_CONSOLIDADO_SEED_XLSX", "").strip()
    source = Path(configured).expanduser() if configured else CONSOLIDADO_SEED_DEFAULT
    if not source.exists():
        return

    try:
        from openpyxl import load_workbook
    except ImportError:
        return

    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet_name = "CONSOLIDADO" if "CONSOLIDADO" in workbook.sheetnames else "Consolidado"
    if sheet_name not in workbook.sheetnames:
        return
    ws = workbook[sheet_name]
    imported_at = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
    rows_to_insert = []
    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_col=len(CONSOLIDADO_HEADERS), values_only=True), 2):
        if not any(value not in (None, "") for value in row):
            continue
        if not row[1] or not row[2] or not row[3]:
            continue
        rows_to_insert.append(
            (
                str(source),
                row_index,
                _cell_text(row[0]),
                _cell_text(row[1]),
                _cell_text(row[2]),
                _cell_text(row[3]),
                _cell_text(row[4]),
                _excel_date_text(row[5]),
                _cell_text(row[6]),
                _as_float(row[7]),
                _as_float(row[8]),
                _as_float(row[9]),
                _cell_text(row[10]),
                _cell_text(row[11]),
                _cell_text(row[12]),
                _cell_text(row[13]),
                _cell_text(row[14]),
                _cell_text(row[15]),
                imported_at,
            )
        )
    if not rows_to_insert:
        return
    await db.executemany(
        """
        INSERT OR IGNORE INTO daily_consolidado_historico (
            source_file, source_row, semana_label, division, up, metrica,
            dia_daily, fecha_daily, semana_num, real, target_plan, porcentaje,
            color, filtro, fecha_hoy, filtro_semana, dia_hoy, filtro_avance,
            imported_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        rows_to_insert,
    )


def _strip_accents_for_key(text: str) -> str:
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ú": "U",
        "ñ": "n", "Ñ": "N",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text


def normalize_tipo_daily(value: str) -> str:
    text = _strip_accents_for_key(str(value or "")).strip().lower()
    if text in {"operacion", "op"}:
        return "Operacion"
    if text in {"logistica inversa", "li"}:
        return "Logistica Inversa"
    if text in {"planeamiento", "planificacion"}:
        return "Planeamiento"
    return str(value or "").strip()


def normalize_sector(value: str) -> str:
    text = _strip_accents_for_key(str(value or "")).strip().lower()
    mapping = {
        "noa": "Noa",
        "secos": "Secos",
        "refrigerados": "Refrigerados",
        "logistica inversa": "Logistica Inversa",
        "planeamiento": "Planeamiento",
    }
    if text.startswith("logistica inversa"):
        return "Logistica Inversa"
    return mapping.get(text, str(value or "").strip())


def calculate_daily_window(now: datetime | None = None) -> dict[str, Any]:
    current = now.astimezone(LOCAL_TZ) if now else datetime.now(LOCAL_TZ)
    if current.time() < LOAD_CUTOFF:
        return {
            "can_load": False,
            "reason": "La Daily se habilita desde las 06:30.",
            "now": current.isoformat(timespec="seconds"),
        }

    weekday = current.weekday()  # lunes=0
    if weekday == 6:
        return {
            "can_load": False,
            "reason": "La Daily de fin de semana se carga el lunes desde las 06:30.",
            "now": current.isoformat(timespec="seconds"),
        }

    if weekday == 0:
        start_date = current.date() - timedelta(days=2)
    else:
        start_date = current.date() - timedelta(days=1)
    end_date = current.date()

    start = datetime.combine(start_date, time(6, 0), tzinfo=LOCAL_TZ)
    end = datetime.combine(end_date, time(6, 0), tzinfo=LOCAL_TZ)
    label = f"{_day_name(start)} 06:00 / {_day_name(end)} 06:00"
    daily_key = f"{start:%Y%m%d0600}_{end:%Y%m%d0600}"

    return {
        "can_load": True,
        "daily_key": daily_key,
        "daily_label": label,
        "fecha_inicio": start.isoformat(timespec="seconds"),
        "fecha_fin": end.isoformat(timespec="seconds"),
        "fecha_carga": current.date().isoformat(),
        "timestamp_carga_preview": current.isoformat(timespec="seconds"),
        "now": current.isoformat(timespec="seconds"),
    }


def _day_name(value: datetime) -> str:
    names = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]
    return names[value.weekday()]


def calculate_metrics(valor_real: float | None, valor_esperado: float | None, regla: str) -> dict[str, Any]:
    if valor_real is None or valor_esperado in (None, 0) or regla in {"informativo", "texto"}:
        return {
            "desvio": None,
            "porcentaje_cumplimiento": None,
            "porcentaje_desvio": None,
            "estado_cumplimiento": "Informativo",
        }

    desvio = valor_real - valor_esperado
    porcentaje_cumplimiento = (valor_real / valor_esperado) * 100
    porcentaje_desvio = (desvio / valor_esperado) * 100
    if regla == "cumplimiento":
        estado = _cumplimiento_status(porcentaje_cumplimiento)
    elif regla == "productividad":
        estado = _productividad_status(porcentaje_desvio)
    else:
        estado = "Informativo"
    return {
        "desvio": desvio,
        "porcentaje_cumplimiento": porcentaje_cumplimiento,
        "porcentaje_desvio": porcentaje_desvio,
        "estado_cumplimiento": estado,
    }


def _cumplimiento_status(porcentaje: float) -> str:
    if porcentaje < 85:
        return "Rojo"
    if porcentaje < 90:
        return "Amarillo"
    if porcentaje <= 95:
        return "Verde claro"
    return "Verde oscuro"


def _productividad_status(porcentaje_desvio: float) -> str:
    if porcentaje_desvio < -5:
        return "Rojo"
    if porcentaje_desvio < 0:
        return "Amarillo"
    if porcentaje_desvio <= 5:
        return "Verde claro"
    return "Verde oscuro"


def parse_number(value: Any) -> float | None:
    text = str(value if value is not None else "").strip()
    if not text:
        return None
    normalized = re.sub(r"[^0-9,.\-]", "", text)
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return None


async def get_parametros(tipo_daily: str, sector: str) -> list[dict[str, Any]]:
    tipo = normalize_tipo_daily(tipo_daily)
    sector_norm = normalize_sector(sector)
    async with aiosqlite.connect(DAILY_DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """
            SELECT *
            FROM daily_parametros_cumplimiento
            WHERE tipo_daily = ? AND activo = 1
            ORDER BY orden
            """,
            (tipo,),
        ) as cur:
            rows = [dict(row) for row in await cur.fetchall()]
    if tipo != "Operacion":
        return rows
    return [
        row for row in rows
        if not row.get("sector_aplicable")
        or sector_norm in str(row.get("sector_aplicable") or "").split("|")
    ]


async def get_all_parametros() -> list[dict[str, Any]]:
    async with aiosqlite.connect(DAILY_DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """
            SELECT *
            FROM daily_parametros_cumplimiento
            ORDER BY tipo_daily, orden
            """
        ) as cur:
            return [dict(row) for row in await cur.fetchall()]


async def update_parametros(rows: list[dict[str, Any]]) -> int:
    updated = 0
    now = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
    async with aiosqlite.connect(DAILY_DB_PATH) as db:
        await db.execute("PRAGMA busy_timeout = 10000")
        for row in rows:
            id_parametro = str(row.get("id_parametro") or "").strip()
            if not id_parametro:
                continue
            valor_esperado = parse_number(row.get("valor_esperado"))
            activo = 1 if row.get("activo") in (True, 1, "1", "true", "True", "SI", "Si", "si") else 0
            cursor = await db.execute(
                """
                UPDATE daily_parametros_cumplimiento
                SET valor_esperado = ?,
                    activo = ?,
                    updated_at = ?
                WHERE id_parametro = ?
                """,
                (valor_esperado, activo, now, id_parametro),
            )
            updated += cursor.rowcount
        await db.commit()
    return updated


async def export_powerbi_csv() -> Path:
    """
    Regenera el CSV completo de la vista plana para Power BI.

    Se escribe primero un archivo temporal y luego se reemplaza el destino para
    evitar que Power BI lea un archivo a medio escribir.
    """
    DAILY_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    target = DAILY_POWERBI_CSV_PATH
    temp = target.with_suffix(target.suffix + ".tmp")
    async with aiosqlite.connect(DAILY_DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM vw_daily_powerbi ORDER BY carga_id, orden") as cur:
            rows = await cur.fetchall()
            columns = [item[0] for item in (cur.description or [])]

    with temp.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerow(columns)
        for row in rows:
            writer.writerow([row[column] for column in columns])
    temp.replace(target)
    return target


async def export_consolidado_powerbi_csv() -> Path:
    DAILY_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    target = DAILY_CONSOLIDADO_CSV_PATH
    temp = target.with_suffix(target.suffix + ".tmp")
    async with aiosqlite.connect(DAILY_DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """
            SELECT
                semana_label AS "Semana",
                division AS "Division",
                up AS "UP",
                metrica AS "Metrica",
                dia_daily AS "Dia Daily",
                fecha_daily AS "Fecha Daily",
                semana_label AS "Semana",
                real AS "REAL",
                target_plan AS "TARGET/PLAN",
                porcentaje AS "%",
                color AS "COLOR",
                filtro AS "Filtro",
                fecha_hoy AS "Fecha_hoy",
                filtro_semana AS "Filtro de semana",
                dia_hoy AS "Dia_hoy",
                filtro_avance AS "Filtro de avance"
            FROM daily_consolidado_historico
            ORDER BY fecha_daily, division, metrica, up
            """
        ) as cur:
            historico_rows = [dict(row) for row in await cur.fetchall()]
        async with db.execute(
            """
            SELECT *
            FROM vw_daily_powerbi
            WHERE tipo_campo = 'numerico'
              AND grupo IN ('PRODUCTIVIDAD', 'CUMPLIMIENTO', 'AVANCE')
            ORDER BY fecha_inicio, tipo_daily, sector, grupo, orden
            """
        ) as cur:
            rows = [dict(row) for row in await cur.fetchall()]

    with temp.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerow(CONSOLIDADO_HEADERS)
        for row in rows:
            item = _consolidado_row(row)
            if item:
                historico_rows = _upsert_consolidado_item(historico_rows, item)
        for item in historico_rows:
            writer.writerow([item.get(column, "") for column in CONSOLIDADO_HEADERS])
    temp.replace(target)
    return target


def _upsert_consolidado_item(rows: list[dict[str, Any]], item: dict[str, Any]) -> list[dict[str, Any]]:
    key = _consolidado_key(item)
    for index, row in enumerate(rows):
        if _consolidado_key(row) == key:
            rows[index] = item
            return rows
    rows.append(item)
    return rows


def _consolidado_key(row: dict[str, Any]) -> tuple[str, str, str, str]:
    return (
        str(row.get("Fecha Daily") or ""),
        str(row.get("Division") or ""),
        str(row.get("UP") or ""),
        str(row.get("Metrica") or ""),
    )


def _consolidado_row(row: dict[str, Any]) -> dict[str, Any] | None:
    division = _consolidado_division(row)
    up = _consolidado_up(row)
    metrica = _consolidado_metrica(row.get("grupo"))
    if not division or not up or not metrica:
        return None
    fecha_daily = _parse_iso_date(row.get("fecha_fin")) or _parse_iso_date(row.get("fecha_carga"))
    real = row.get("valor_real_numero")
    target = row.get("valor_esperado")
    percent = _consolidado_percent(real, target, metrica)
    return {
        "Semana": _week_label(fecha_daily),
        "Division": division,
        "UP": up,
        "Metrica": metrica,
        "Dia Daily": _daily_day_label(fecha_daily),
        "Fecha Daily": f"{fecha_daily.isoformat()} 00:00:00" if fecha_daily else "",
        "REAL": real if real is not None else "",
        "TARGET/PLAN": target if target is not None else "",
        "%": percent if percent is not None else "",
        "COLOR": _consolidado_color(percent, metrica, real, target),
        "Filtro": "Si",
        "Fecha_hoy": datetime.now(LOCAL_TZ).day,
        "Filtro de semana": "No",
        "Dia_hoy": _short_day_label(datetime.now(LOCAL_TZ)),
        "Filtro de avance": "Si" if metrica == "Avance" else "No",
    }


def _consolidado_division(row: dict[str, Any]) -> str:
    tipo = str(row.get("tipo_daily") or "")
    sector = str(row.get("sector") or "")
    if tipo == "Operacion":
        return "NoA" if sector == "Noa" else sector
    if tipo == "Logistica Inversa":
        return "Logística Inversa - Suc 126"
    return tipo


def _consolidado_metrica(group: Any) -> str:
    mapping = {
        "PRODUCTIVIDAD": "Productividad",
        "CUMPLIMIENTO": "Cumplimiento",
        "AVANCE": "Avance",
    }
    return mapping.get(str(group or "").upper(), "")


def _consolidado_up(row: dict[str, Any]) -> str:
    tipo = str(row.get("tipo_daily") or "")
    proceso = str(row.get("proceso") or "").upper()
    if tipo == "Logistica Inversa":
        li_map = {
            "RECEPCION ENVASES": "1-Recep Env",
            "RECEPCION DEVOLUCION": "2-Recep Dev",
            "CARGA Y DESCARGA MOVILES": "3-Car Des / Recep Rot",
            "RECEPCION ROTURAS": "3-Car Des / Recep Rot",
            "ENFARDADORA": "4-Enf",
        }
        return li_map.get(proceso, "")
    if tipo == "Planeamiento":
        pl_map = {
            "HUECOS": "1-Hue",
            "PROGRAMA": "2-Pro",
            "RECURSO FTE": "3-Fte",
            "COORDINACION DE PROVEEDORES": "4-Cpr",
        }
        return pl_map.get(proceso, "")
    op_map = {
        "RECEPCION": "1-Recep",
        "PICKING": "2-Pick",
        "CLARK": "3-Clark",
        "SPC": "3-Clark",
        "DESPACHO": "4-Desp",
    }
    return op_map.get(proceso, "")


def _consolidado_percent(real: Any, target: Any, metrica: str) -> float | None:
    real_num = _as_float(real)
    target_num = _as_float(target)
    if real_num is None or target_num in (None, 0):
        return None
    if metrica == "Productividad":
        return real_num / target_num - 1
    return real_num / target_num


def _consolidado_color(percent: float | None, metrica: str, real: Any, target: Any) -> str:
    if percent is None:
        return "GRIS"
    if metrica == "Productividad":
        if percent < -0.05:
            return "ROJO"
        if percent < 0:
            return "AMARILLO"
        if percent <= 0.05:
            return "VERDE C"
        return "VERDE O"
    if percent < 0.85:
        return "ROJO"
    if percent < 0.90:
        return "AMARILLO"
    if percent <= 0.95:
        return "VERDE C"
    return "VERDE O"


def _parse_iso_date(value: Any):
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "")).date()
    except ValueError:
        try:
            return datetime.strptime(text[:10], "%Y-%m-%d").date()
        except ValueError:
            return None


def _week_label(value) -> str:
    if not value:
        return ""
    start = value - timedelta(days=value.weekday())
    end = start + timedelta(days=4)
    return f"{start.day:02d} al {end.day:02d}-{end.month:02d}"


def _daily_day_label(value) -> str:
    if not value:
        return ""
    labels = ["1-Lunes", "2-Martes", "3-Miercoles", "4-Jueves", "5-Viernes", "6-Sabado", "7-Domingo"]
    return labels[value.weekday()]


def _short_day_label(value: datetime) -> str:
    labels = ["1-L", "2-M", "3-Mi", "4-J", "5-V", "6-S", "7-D"]
    return labels[value.weekday()]


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return str(value)


def _excel_date_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return f"{value.date().isoformat()} 00:00:00"
    text = str(value).strip()
    if not text:
        return ""
    parsed = _parse_iso_date(text)
    return f"{parsed.isoformat()} 00:00:00" if parsed else text


async def get_existing_cargas(daily_key: str, tipo_daily: str, sector: str) -> list[dict[str, Any]]:
    async with aiosqlite.connect(DAILY_DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """
            SELECT *
            FROM daily_cargas
            WHERE daily_key = ?
              AND tipo_daily = ?
              AND COALESCE(sector, '') = COALESCE(?, '')
              AND reemplazado = 0
            ORDER BY version DESC, timestamp_carga DESC
            """,
            (daily_key, normalize_tipo_daily(tipo_daily), normalize_sector(sector)),
        ) as cur:
            return [dict(row) for row in await cur.fetchall()]


async def save_daily_carga(
    *,
    usuario_carga: str,
    tipo_daily: str,
    sector: str,
    turno: str,
    plan: str,
    respuestas: dict[str, Any],
    action: str,
) -> dict[str, Any]:
    daily = calculate_daily_window()
    if not daily.get("can_load"):
        raise ValueError(str(daily.get("reason") or "La Daily no esta habilitada."))

    tipo = normalize_tipo_daily(tipo_daily)
    sector_norm = normalize_sector(sector)
    parametros = await get_parametros(tipo, sector_norm)
    if not parametros:
        raise ValueError("No hay parametros activos para la seleccion.")

    existing = await get_existing_cargas(daily["daily_key"], tipo, sector_norm)
    if existing and action == "cancel":
        raise ValueError("Existe una carga previa y se selecciono cancelar.")

    next_version = (max(row["version"] for row in existing) + 1) if existing else 1
    now = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
    async with aiosqlite.connect(DAILY_DB_PATH) as db:
        await db.execute("PRAGMA busy_timeout = 10000")
        if existing and action == "replace":
            await db.execute(
                """
                UPDATE daily_cargas
                SET reemplazado = 1, updated_at = ?
                WHERE daily_key = ?
                  AND tipo_daily = ?
                  AND COALESCE(sector, '') = COALESCE(?, '')
                  AND reemplazado = 0
                """,
                (now, daily["daily_key"], tipo, sector_norm),
            )

        cursor = await db.execute(
            """
            INSERT INTO daily_cargas (
                daily_key, daily_label, fecha_inicio, fecha_fin, fecha_carga,
                timestamp_carga, usuario_carga, tipo_daily, sector, turno,
                plan, version, reemplazado, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
            """,
            (
                daily["daily_key"], daily["daily_label"], daily["fecha_inicio"], daily["fecha_fin"],
                daily["fecha_carga"], now, usuario_carga, tipo, sector_norm, str(turno or "").strip(),
                str(plan or "").strip(), next_version, now, now,
            ),
        )
        carga_id = int(cursor.lastrowid)
        for param in parametros:
            raw_value = respuestas.get(param["id_parametro"], "")
            text_value = "" if raw_value is None else str(raw_value).strip()
            numeric_value = parse_number(text_value) if param["tipo_campo"] == "numerico" else None
            metrics = calculate_metrics(numeric_value, param.get("valor_esperado"), param.get("regla_cumplimiento") or "informativo")
            await db.execute(
                """
                INSERT INTO daily_carga_detalle (
                    carga_id, id_parametro, valor_real_texto, valor_real_numero,
                    valor_esperado_snapshot, desvio, porcentaje_cumplimiento,
                    porcentaje_desvio, estado_cumplimiento, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    carga_id, param["id_parametro"], text_value, numeric_value,
                    param.get("valor_esperado"), metrics["desvio"],
                    metrics["porcentaje_cumplimiento"], metrics["porcentaje_desvio"],
                    metrics["estado_cumplimiento"], now,
                ),
            )
        await db.commit()

    csv_path = await export_powerbi_csv()
    consolidado_csv_path = await export_consolidado_powerbi_csv()
    return {
        "carga_id": carga_id,
        "daily": daily,
        "version": next_version,
        "parametros_guardados": len(parametros),
        "db_path": str(DAILY_DB_PATH),
        "csv_path": str(csv_path),
        "consolidado_csv_path": str(consolidado_csv_path),
    }
