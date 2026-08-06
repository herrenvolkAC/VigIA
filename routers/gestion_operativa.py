"""
VigIA · Gestion operativa.

Endpoints para herramientas de apoyo directo a la operacion.
"""
import asyncio
import hashlib
import json
import logging
import os
import time
from contextlib import suppress
from collections import defaultdict
from datetime import datetime, time as dt_time, timedelta
from pathlib import Path
from typing import Any

import aiosqlite
from fastapi import APIRouter, HTTPException, Query, Request
from pydantic import BaseModel

from db.daily_operativa import (
    DAILY_DB_PATH,
    LOCAL_TZ,
    calculate_daily_window,
    export_consolidado_powerbi_csv,
    export_powerbi_csv,
    get_all_parametros,
    get_daily_carga_auto_manual_comparacion,
    get_existing_cargas,
    get_parametros,
    normalize_sector,
    normalize_tipo_daily,
    save_daily_carga,
    update_parametros,
)
from db.daily_auto import (
    DAILY_AUTO_DB_PATH,
    get_daily_auto_manual_comparacion,
    get_cached_results,
    get_despacho_raw_cache_rows,
    get_latest_run,
    get_productividad_raw_cache_rows,
    save_despacho_summary_cache,
    save_despacho_raw_cache,
    init_daily_auto_db,
    mark_run_error,
    rebuild_daily_auto_manual_comparacion,
    replace_daily_manual_comparacion_rows,
    save_avance_summary_cache,
    save_clark_raw_summary_cache,
    save_picking_summary_cache,
    save_productividad_raw_cache,
    save_planificacion_summary_cache,
    save_recepcion_summary_cache,
    scheduled_daily_window,
)
from db.schema import DB_PATH
from routers.ai import _call_ai, _extract_json
from routers.auth_local import current_auth
from routers.productividad_analisis import (
    _build_picking_idle_analysis,
    _turn_label,
    _turn_range_for_date,
    QUERY_DAILY_CLARK_REAL,
    QUERY_DAILY_DESPACHO_REAL,
    QUERY_DAILY_DESPACHO_RAW,
    QUERY_DAILY_PLANIFICACION,
    QUERY_DAILY_PICKING_REAL,
    QUERY_DAILY_PRODUCTIVIDAD_RAW,
    QUERY_DAILY_RECEPCION_REAL,
    query_productive_db_daily_productividad_raw,
    query_productive_db_daily_planificacion,
    query_productive_db_daily_despacho_raw,
    query_productive_db_gestion_productividad_picking,
    query_productive_db_picking_tiempos_muertos,
)
from routers.rendimiento_online import (
    build_rendimiento_online_operations,
    query_rendimiento_online_rows,
)

logger = logging.getLogger("vigia.gestion_operativa")
router = APIRouter(prefix="/api/gestion-operativa", tags=["gestion-operativa"])
GESTION_PRODUCTIVIDAD_IA_PROMPT_VERSION = "gestion_productividad_picking_v1"
DAILY_AUTO_SCHEDULE_TIME = dt_time(7, 35)
DAILY_AUTO_SCHEDULE_GRACE_MINUTES = 10
DAILY_AUTO_FORCE_USERS = {"acucci"}
DAILY_AUTO_CLARK_QUERY_VERSION = "clark_raw_pallet_cuenta_v4"
DAILY_AUTO_PICKING_QUERY_VERSION = "picking_raw_antiguedad_legajo_v6"
DAILY_AUTO_RECEPCION_QUERY_VERSION = "recepcion_raw_cache_v1"
DAILY_AUTO_DESPACHO_RAW_QUERY_VERSION = "despacho_raw_f922traf_hojaruta_funcion_distinct_v5"
DAILY_AUTO_DESPACHO_QUERY_VERSION = "despacho_raw_cache_funcion_distinct_v4"
DAILY_AUTO_PLANIFICACION_QUERY_VERSION = "planificacion_unificada_v1"
DAILY_AUTO_PRODUCTIVIDAD_RAW_QUERY_VERSION = "productividad_raw_f132hist_fec_ing_v8"
DAILY_AUTO_AVANCE_QUERY_VERSION = "avance_6a730_cache_v1"
_daily_auto_scheduler_task: asyncio.Task | None = None
_daily_auto_scheduler_stop: asyncio.Event | None = None

DESPACHO_CARGADOR_FUNCIONES_VALIDAS = {
    "ARMADOR REFRI",
    "ARMADOR DE REFRIGERADOS",
    "CARGADOR (REFRIG)",
    "ARMADOR DE SECOS",
    "ARMADOR T06",
    "ARMADOR DE NO ALIMENTOS",
    "ARMADOR",
    "CARGADOR DE SECOS",
}


def _despacho_cargador_elegible(row: dict[str, Any]) -> bool:
    funcion = str(row.get("DESC_FUNCION") or row.get("desc_funcion") or "").strip().upper()
    return funcion in DESPACHO_CARGADOR_FUNCIONES_VALIDAS


def _daily_auto_schedule_label() -> str:
    return DAILY_AUTO_SCHEDULE_TIME.strftime("%H:%M")


def _daily_window_for_fecha_carga(fecha_daily: datetime) -> dict[str, Any]:
    fecha = fecha_daily.date() if isinstance(fecha_daily, datetime) else fecha_daily
    start_date = fecha - timedelta(days=2) if fecha.weekday() == 0 else fecha - timedelta(days=1)
    start = datetime.combine(start_date, dt_time(6, 0), tzinfo=LOCAL_TZ)
    end = datetime.combine(fecha, dt_time(6, 0), tzinfo=LOCAL_TZ)
    labels = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]
    return {
        "can_run": True,
        "can_load": True,
        "daily_key": f"{start:%Y%m%d0600}_{end:%Y%m%d0600}",
        "daily_label": f"{labels[start.weekday()]} 06:00 / {labels[end.weekday()]} 06:00",
        "fecha_inicio": start.isoformat(timespec="seconds"),
        "fecha_fin": end.isoformat(timespec="seconds"),
        "fecha_carga": fecha.isoformat(),
        "now": datetime.now(LOCAL_TZ).isoformat(timespec="seconds"),
    }


class ProductividadPickingIARequest(BaseModel):
    analisis_base: dict[str, Any]
    provider: str | None = None


class DailyCargaRequest(BaseModel):
    tipo_daily: str
    sector: str = ""
    turno: str = ""
    plan: str = ""
    action: str = "replace"
    respuestas: dict[str, Any] = {}


class DailyParametrosUpdateRequest(BaseModel):
    clave: str
    parametros: list[dict[str, Any]]


class DailyAutoRetryRequest(BaseModel):
    force: bool = False


class DailyManualComparacionImportRequest(BaseModel):
    source_path: str = r"C:\Users\207189\Documents\DailyManual.xlsx"
    days: int = 15
    force_auto: bool = False


DAILY_PICKING_REAL_PARAM_IDS_BY_SECTOR = {
    "Noa": "OP_PROD_PICKING_NOA_6A6",
    "Secos": "OP_PROD_PICKING_SECOS_6A6",
    "Refrigerados": "OP_PROD_PICKING_REFRI_6A6",
    "OC": "OP_PROD_PICKING_OC_6A6",
    "Congelados": "OP_PROD_PICKING_CONGELADOS_6A6",
}
DAILY_DESPACHO_REAL_PARAM_IDS_BY_SECTOR = {
    "Secos": "OP_PROD_DESPACHO_SECOS_6A6",
    "Refrigerados": "OP_PROD_DESPACHO_REFRI_6A6",
}
DAILY_CLARK_REAL_PARAM_IDS_BY_SECTOR = {
    "Noa": "OP_PROD_CLARK_NOA_6A6",
    "Secos": "OP_PROD_CLARK_SECOS_6A6",
    "Refrigerados": "OP_PROD_CLARK_REFRI_6A6",
    "OC": "OP_PROD_CLARK_OC_6A6",
    "Congelados": "OP_PROD_CLARK_CONGELADOS_6A6",
}
DAILY_PICKING_PLAN_PARAM_IDS_BY_SECTOR = {
    "Noa": "OP_CUMP_PICKING_PLAN_6A6",
    "Secos": "OP_CUMP_PICKING_PLAN_6A6",
    "Refrigerados": "OP_CUMP_PICKING_PLAN_6A6",
}
DAILY_DESPACHO_PLAN_PARAM_IDS_BY_SECTOR = {
    "Noa": "OP_CUMP_DESPACHO_PLAN_6A6",
    "Secos": "OP_CUMP_DESPACHO_PLAN_6A6",
    "Refrigerados": "OP_CUMP_DESPACHO_PLAN_6A6",
}
DAILY_SPC_PLAN_PARAM_IDS_BY_SECTOR = {
    "Noa": "OP_CUMP_SPC_PLAN_6A6",
    "Secos": "OP_CUMP_SPC_PLAN_6A6",
    "Refrigerados": "OP_CUMP_SPC_PLAN_6A6",
}


SHIFT_LABELS = {"manana": "Mañana", "tarde": "Tarde", "noche": "Noche"}
EXPECTED_SECTORS_BY_ALMACEN = {
    "VARIOS NO ALIMENTOS": {"CD-NOA"},
    "SECTOR SECOS": {"CD-SECOS - ZONA 1", "CD-SECOS - ZONA 2"},
    "SECTOR DE F&Q - CONGELADOS": {"CD-REFRIGERADOS"},
    "SECTOR FRUTA-VERDURA": {"CD-REFRIGERADOS"},
}

DAILY_RAW_SECTOR_TO_ALMACEN = {
    "Noa": "NOA",
    "Secos": "SECOS",
    "Refrigerados": "REFRIGERADOS",
    "OC": "OC",
    "Congelados": "CONGELADOS",
}

PRODUCTIVIDAD_JORNADA_BY_DIVISION = {
    "Congelados": 6.0,
}

DAILY_MANUAL_XLSX_DEFAULT = Path(r"C:\Users\207189\Documents\DailyManual.xlsx")

DAILY_MANUAL_PARAM_MAP = {
    ("Productividad", "1-Recep", "REAL"): {
        "Noa": "OP_PROD_RECEPCION_NOA_6A6",
        "Secos": "OP_PROD_RECEPCION_SECOS_6A6",
        "Refrigerados": "OP_PROD_RECEPCION_REFRI_6A6",
    },
    ("Productividad", "2-Pick", "REAL"): {
        "Noa": "OP_PROD_PICKING_NOA_6A6",
        "Secos": "OP_PROD_PICKING_SECOS_6A6",
        "Refrigerados": "OP_PROD_PICKING_REFRI_6A6",
    },
    ("Productividad", "3-Clark", "REAL"): {
        "Noa": "OP_PROD_CLARK_NOA_6A6",
        "Secos": "OP_PROD_CLARK_SECOS_6A6",
        "Refrigerados": "OP_PROD_CLARK_REFRI_6A6",
    },
    ("Productividad", "4-Desp", "REAL"): {
        "Secos": "OP_PROD_DESPACHO_SECOS_6A6",
        "Refrigerados": "OP_PROD_DESPACHO_REFRI_6A6",
    },
    ("Cumplimiento", "1-Recep", "REAL"): {
        "Noa": "OP_CUMP_RECEPCION_REAL_6A6",
        "Secos": "OP_CUMP_RECEPCION_REAL_6A6",
        "Refrigerados": "OP_CUMP_RECEPCION_REAL_6A6",
    },
    ("Cumplimiento", "2-Pick", "REAL"): {
        "Noa": "OP_CUMP_PICKING_REAL_6A6",
        "Secos": "OP_CUMP_PICKING_REAL_6A6",
        "Refrigerados": "OP_CUMP_PICKING_REAL_6A6",
    },
    ("Cumplimiento", "2-Pick", "PLAN"): {
        "Noa": "OP_CUMP_PICKING_PLAN_6A6",
        "Secos": "OP_CUMP_PICKING_PLAN_6A6",
        "Refrigerados": "OP_CUMP_PICKING_PLAN_6A6",
    },
    ("Cumplimiento", "3-Clark", "REAL"): {
        "Noa": "OP_CUMP_SPC_REAL_6A6",
        "Secos": "OP_CUMP_SPC_REAL_6A6",
        "Refrigerados": "OP_CUMP_SPC_REAL_6A6",
    },
    ("Cumplimiento", "3-Clark", "PLAN"): {
        "Noa": "OP_CUMP_SPC_PLAN_6A6",
        "Secos": "OP_CUMP_SPC_PLAN_6A6",
        "Refrigerados": "OP_CUMP_SPC_PLAN_6A6",
    },
    ("Cumplimiento", "4-Desp", "REAL"): {
        "Noa": "OP_CUMP_DESPACHO_REAL_6A6",
        "Secos": "OP_CUMP_DESPACHO_REAL_6A6",
        "Refrigerados": "OP_CUMP_DESPACHO_REAL_6A6",
    },
    ("Cumplimiento", "4-Desp", "PLAN"): {
        "Noa": "OP_CUMP_DESPACHO_PLAN_6A6",
        "Secos": "OP_CUMP_DESPACHO_PLAN_6A6",
        "Refrigerados": "OP_CUMP_DESPACHO_PLAN_6A6",
    },
    ("Avance", "1-Recep", "REAL"): {
        "Noa": "OP_AVANCE_RECEPCION_REAL_6A8",
        "Secos": "OP_AVANCE_RECEPCION_REAL_6A8",
        "Refrigerados": "OP_AVANCE_RECEPCION_REAL_6A8",
    },
    ("Avance", "2-Pick", "REAL"): {
        "Noa": "OP_AVANCE_PICKING_REAL_6A8",
        "Secos": "OP_AVANCE_PICKING_REAL_6A8",
        "Refrigerados": "OP_AVANCE_PICKING_REAL_6A8",
    },
    ("Avance", "2-Pick", "PLAN"): {
        "Noa": "OP_AVANCE_PICKING_PLAN_6A8",
        "Secos": "OP_AVANCE_PICKING_PLAN_6A8",
        "Refrigerados": "OP_AVANCE_PICKING_PLAN_6A8",
    },
    ("Avance", "3-Clark", "REAL"): {
        "Noa": "OP_AVANCE_SPC_REAL_6A8",
        "Secos": "OP_AVANCE_SPC_REAL_6A8",
        "Refrigerados": "OP_AVANCE_SPC_REAL_6A8",
    },
    ("Avance", "3-Clark", "PLAN"): {
        "Noa": "OP_AVANCE_SPC_PLAN_6A8",
        "Secos": "OP_AVANCE_SPC_PLAN_6A8",
        "Refrigerados": "OP_AVANCE_SPC_PLAN_6A8",
    },
    ("Avance", "4-Desp", "REAL"): {
        "Noa": "OP_AVANCE_DESPACHO_REAL_6A8",
        "Secos": "OP_AVANCE_DESPACHO_REAL_6A8",
        "Refrigerados": "OP_AVANCE_DESPACHO_REAL_6A8",
    },
    ("Avance", "4-Desp", "PLAN"): {
        "Noa": "OP_AVANCE_DESPACHO_PLAN_6A8",
        "Secos": "OP_AVANCE_DESPACHO_PLAN_6A8",
        "Refrigerados": "OP_AVANCE_DESPACHO_PLAN_6A8",
    },
}


def _parse_dt(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    text = str(value or "").strip()
    if not text:
        return None
    text = text.replace("T", " ")
    if text.endswith("Z"):
        text = text[:-1]
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text).replace(tzinfo=None)
    except ValueError:
        return None


def _fmt_dt(value: datetime | None) -> str | None:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else None


def _fmt_daily_oracle_dt(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text).strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return text.replace("T", " ")[:19]


def _to_float(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _norm_legajo(value: Any) -> str:
    text = str(value or "").strip()
    return text.lstrip("0") or text


def _manual_sector(value: Any) -> str:
    text = str(value or "").strip()
    if text.lower() == "noa":
        return "Noa"
    return normalize_sector(text)


def _manual_operacion(up: str) -> str:
    if up == "1-Recep":
        return "RECEPCION"
    if up == "2-Pick":
        return "PICKING"
    if up == "3-Clark":
        return "SPC"
    if up == "4-Desp":
        return "DESPACHO"
    return str(up or "").strip().upper()


def _manual_excel_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[:10], fmt)
        except ValueError:
            continue
    return None


def _manual_compare_dates_from_excel(path: Path, days: int) -> list[datetime]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl no esta disponible para leer DailyManual.xlsx.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "CONSOLIDADO" not in workbook.sheetnames:
        raise RuntimeError("DailyManual.xlsx no contiene la hoja CONSOLIDADO.")
    today = datetime.now(LOCAL_TZ).date()
    dates = set()
    for row in workbook["CONSOLIDADO"].iter_rows(min_row=2, max_col=6, values_only=True):
        dt = _manual_excel_date(row[5])
        if dt and dt.date() <= today:
            dates.add(dt.date())
    return [datetime.combine(item, dt_time(0, 0)) for item in sorted(dates)[-max(days, 1):]]


def _manual_rows_from_excel(path: Path, dates: list[datetime]) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl no esta disponible para leer DailyManual.xlsx.") from exc
    target_dates = {item.date().isoformat() for item in dates}
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "CONSOLIDADO" not in workbook.sheetnames:
        raise RuntimeError("DailyManual.xlsx no contiene la hoja CONSOLIDADO.")
    rows: list[dict[str, Any]] = []
    ws = workbook["CONSOLIDADO"]
    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_col=16, values_only=True), 2):
        fecha_dt = _manual_excel_date(row[5])
        if not fecha_dt or fecha_dt.date().isoformat() not in target_dates:
            continue
        sector = _manual_sector(row[1])
        if sector not in {"Noa", "Secos", "Refrigerados"}:
            continue
        up = str(row[2] or "").strip()
        metrica = str(row[3] or "").strip()
        if not up or metrica not in {"Productividad", "Cumplimiento", "Avance"}:
            continue
        daily = _daily_window_for_fecha_carga(fecha_dt)
        for concepto, value_idx in (("REAL", 7), ("PLAN", 8)):
            id_param = DAILY_MANUAL_PARAM_MAP.get((metrica, up, concepto), {}).get(sector)
            if not id_param:
                continue
            value = _to_float(row[value_idx])
            rows.append(
                {
                    "source_row": row_index,
                    "concepto": concepto,
                    "fecha_daily": fecha_dt.date().isoformat(),
                    "daily_key": daily["daily_key"],
                    "sector": sector,
                    "up": up,
                    "metrica": metrica,
                    "operacion": _manual_operacion(up),
                    "id_parametro": id_param,
                    "valor_manual": value,
                }
            )
    return rows


def _expected_sectors_for_almacen(almacen: Any) -> set[str]:
    key = str(almacen or "").strip().upper()
    return EXPECTED_SECTORS_BY_ALMACEN.get(key, set())


async def _load_latest_legajero_profiles() -> dict[str, dict[str, str]]:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("PRAGMA busy_timeout = 10000")
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """
            SELECT batch_id
            FROM rrhh_import_batches
            WHERE status = 'complete'
            ORDER BY imported_at DESC, batch_id DESC
            LIMIT 1
            """
        ) as cur:
            batch = await cur.fetchone()
        if not batch:
            return {}
        async with db.execute(
            """
            SELECT legajo, desc_sector_generico, desc_funcion, desc_posicion
            FROM rrhh_legajero
            WHERE batch_id = ?
            """,
            (batch["batch_id"],),
        ) as cur:
            rows = await cur.fetchall()
    return {
        _norm_legajo(row["legajo"]): {
            "sector": str(row["desc_sector_generico"] or "").strip(),
            "funcion": str(row["desc_funcion"] or "").strip(),
            "posicion": str(row["desc_posicion"] or "").strip(),
        }
        for row in rows
        if _norm_legajo(row["legajo"])
    }


def _date_only(value: Any) -> str:
    dt = _parse_dt(value)
    if dt:
        return dt.strftime("%Y-%m-%d")
    text = str(value or "").strip()
    if len(text) >= 8 and text[:8].isdigit():
        return f"{text[:4]}-{text[4:6]}-{text[6:8]}"
    return text[:10]


def _pct_change(current: float, previous: float) -> float | None:
    if previous == 0:
        return None
    return round(((current - previous) / previous) * 100, 2)


def _productividad_hora(produccion: float, segundos: float) -> float:
    return produccion / (segundos / 3600) if segundos > 0 else 0


def _legajos_chart(
    metric: str,
    label: str,
    unit: str,
    description: str,
    by_date_sector: dict[tuple[str, str], dict[str, Any]],
    totals_by_date: dict[str, dict[str, Any]],
    dates: list[str],
    sectors: list[str],
) -> dict[str, Any]:
    return {
        "metric": metric,
        "label": label,
        "unit": unit,
        "description": description,
        "total": [{"fecha": fecha, "valor": round(float(totals_by_date.get(fecha, {}).get(metric) or 0), 2)} for fecha in dates],
        "series_type": "sector",
        "series": [
            {
                "sector": sector,
                "points": [
                    {"fecha": fecha, "valor": round(float(by_date_sector.get((fecha, sector), {}).get(metric) or 0), 2)}
                    for fecha in dates
                ],
            }
            for sector in sectors
        ],
    }


def _legajos_operator_chart(
    metric: str,
    label: str,
    unit: str,
    description: str,
    dates: list[str],
    total_points: list[dict[str, Any]],
    operators: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "metric": metric,
        "label": label,
        "unit": unit,
        "description": description,
        "total": total_points,
        "series_type": "legajo",
        "series": [
            {
                "legajo": item["legajo"],
                "sector": item["label"],
                "almacen": item["label"],
                "points": [{"fecha": fecha, "valor": round(_to_float(item["daily"].get(fecha)), 2)} for fecha in dates],
            }
            for item in operators
        ],
    }


def _build_legajos_operator_trends(rows: list[dict[str, Any]], dates: list[str], totals_by_date: dict[str, dict[str, Any]]) -> dict[str, Any]:
    by_legajo: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if _to_float(row.get("produccion")) <= 0:
            continue
        legajo = _norm_legajo(row.get("legajo"))
        if legajo:
            by_legajo[legajo].append(row)

    trends = []
    for legajo, items in by_legajo.items():
        daily: dict[str, dict[str, float]] = defaultdict(lambda: {"produccion": 0.0, "tiempo_neto": 0.0})
        funciones: set[str] = set()
        tipos: set[str] = set()
        for item in items:
            day = str(item.get("fecha") or "")
            daily[day]["produccion"] += _to_float(item.get("produccion"))
            daily[day]["tiempo_neto"] += _to_float(item.get("tiempo_neto_segundos"))
            if item.get("funcion"):
                funciones.add(str(item["funcion"]))
            if item.get("tipo"):
                tipos.add(str(item["tipo"]))
        ordered_dates = sorted(day for day, value in daily.items() if value["produccion"] > 0 and value["tiempo_neto"] > 0)
        if len(ordered_dates) < 3:
            continue
        split = max(1, len(ordered_dates) // 2)
        first_dates = ordered_dates[:split]
        last_dates = ordered_dates[split:] or ordered_dates[-split:]
        def avg(days: list[str]) -> float:
            produccion = sum(daily[day]["produccion"] for day in days)
            tiempo_neto = sum(daily[day]["tiempo_neto"] for day in days)
            return _productividad_hora(produccion, tiempo_neto)

        first_avg = avg(first_dates)
        last_avg = avg(last_dates)
        delta = last_avg - first_avg
        delta_pct = _pct_change(last_avg, first_avg)
        sample = items[0]
        daily_productividad = {
            fecha: _productividad_hora(daily.get(fecha, {}).get("produccion", 0), daily.get(fecha, {}).get("tiempo_neto", 0))
            for fecha in dates
        }
        trends.append(
            {
                "legajo": legajo,
                "nombre": sample.get("nombre") or legajo,
                "sector": sample.get("sector") or "SIN SECTOR",
                "funciones": sorted(funciones)[:3],
                "tipos": sorted(tipos)[:3],
                "dias": len(ordered_dates),
                "primer_promedio": round(first_avg, 2),
                "ultimo_promedio": round(last_avg, 2),
                "delta": round(delta, 2),
                "delta_pct": delta_pct,
                "productividad_actual": round(daily_productividad.get(ordered_dates[-1], 0), 2),
                "daily": {fecha: round(daily_productividad.get(fecha, 0), 2) for fecha in dates},
            }
        )

    improving = sorted(
        [item for item in trends if item["delta"] > 0],
        key=lambda item: (item["delta_pct"] if item["delta_pct"] is not None else -999999, item["delta"]),
        reverse=True,
    )[:8]
    worsening = sorted(
        [item for item in trends if item["delta"] < 0],
        key=lambda item: (item["delta_pct"] if item["delta_pct"] is not None else 999999, item["delta"]),
    )[:8]
    total_points = [{"fecha": fecha, "valor": round(_to_float(totals_by_date.get(fecha, {}).get("productividad_promedio")), 2)} for fecha in dates]

    def with_label(item: dict[str, Any]) -> dict[str, Any]:
        label_name = str(item.get("nombre") or item.get("legajo") or "")
        if len(label_name) > 24:
            label_name = label_name[:24] + "..."
        return {**item, "label": f"{label_name} ({item.get('legajo')})"}

    return {
        "improving": [with_label(item) for item in improving],
        "worsening": [with_label(item) for item in worsening],
        "charts": [
            _legajos_operator_chart(
                "legajos_mejoran",
                "Legajos que mejoran",
                "",
                "Lineas por legajo con mayor mejora entre el inicio y el cierre del rango.",
                dates,
                total_points,
                [with_label(item) for item in improving[:6]],
            ),
            _legajos_operator_chart(
                "legajos_empeoran",
                "Legajos que empeoran",
                "",
                "Lineas por legajo con mayor caida entre el inicio y el cierre del rango.",
                dates,
                total_points,
                [with_label(item) for item in worsening[:6]],
            ),
        ],
        "count": len(trends),
    }


async def _load_legajos_sector_options() -> list[str]:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("PRAGMA busy_timeout = 10000")
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """
            SELECT DISTINCT COALESCE(NULLIF(TRIM(desc_sector_generico), ''), 'SIN SECTOR') AS sector
            FROM rrhh_personas
            WHERE active = 1
            ORDER BY sector
            """
        ) as cur:
            rows = await cur.fetchall()
    sectors = [str(row["sector"] or "SIN SECTOR") for row in rows]
    if "SIN SECTOR" not in sectors:
        sectors.append("SIN SECTOR")
    return sorted(sectors, key=lambda value: value.upper())


async def _load_legajos_filter_options() -> dict[str, Any]:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("PRAGMA busy_timeout = 10000")
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """
            SELECT DISTINCT COALESCE(NULLIF(TRIM(razon_social), ''), 'SIN PROVEEDOR') AS proveedor
            FROM rrhh_personas
            WHERE active = 1
              AND COALESCE(NULLIF(TRIM(desc_grupo_personal), ''), '') = 'Externos'
            ORDER BY proveedor
            """
        ) as cur:
            proveedores = [str(row["proveedor"] or "SIN PROVEEDOR") for row in await cur.fetchall()]
        async with db.execute(
            """
            SELECT DISTINCT COALESCE(NULLIF(TRIM(desc_funcion), ''), 'SIN FUNCION') AS funcion
            FROM rrhh_personas
            WHERE active = 1
            ORDER BY funcion
            """
        ) as cur:
            funciones = [str(row["funcion"] or "SIN FUNCION") for row in await cur.fetchall()]
    return {
        "sectores": await _load_legajos_sector_options(),
        "dotaciones": ["Todos", "Externos", "Propios"],
        "proveedores": proveedores,
        "funciones": funciones,
    }


def _years_between(start: Any, end_date: datetime) -> float | None:
    text = str(start or "").strip()
    if not text:
        return None
    try:
        start_dt = datetime.strptime(text[:10], "%Y-%m-%d")
    except ValueError:
        return None
    return round(max((end_date.date() - start_dt.date()).days, 0) / 365.25, 2)


def _matches_antiguedad(years: float | None, bucket: str) -> bool:
    if not bucket or bucket == "ALL":
        return True
    if years is None:
        return False
    try:
        low, high = [float(part) for part in bucket.split("-", 1)]
    except ValueError:
        return True
    return low <= years < high


async def _build_legajos_productividad_payload(
    fecha_desde: str,
    fecha_hasta: str,
    sectores: list[str],
    dotacion: str = "ALL",
    proveedor: str = "ALL",
    antiguedad: str = "ALL",
) -> dict[str, Any]:
    selected = [str(item).strip() for item in sectores if str(item).strip()]
    dotacion = str(dotacion or "ALL").strip()
    proveedor = str(proveedor or "ALL").strip()
    antiguedad = str(antiguedad or "ALL").strip()
    fecha_desde_dt = datetime.strptime(fecha_desde[:10], "%Y-%m-%d")
    fecha_hasta_dt = datetime.strptime(fecha_hasta[:10], "%Y-%m-%d")
    query_start = datetime.combine(fecha_desde_dt.date(), dt_time(6, 0))
    query_end = datetime.combine(fecha_hasta_dt.date() + timedelta(days=1), dt_time(6, 0))
    raw_rows = await asyncio.to_thread(
        query_rendimiento_online_rows,
        query_start.strftime("%Y-%m-%d %H:%M:%S"),
        query_end.strftime("%Y-%m-%d %H:%M:%S"),
    )
    logger.info(
        "[gestion-operativa:analisis-legajos] Rendimiento Online crudo rango=%s..%s movimientos=%s",
        fecha_desde,
        fecha_hasta,
        len(raw_rows),
    )

    oracle_now = next(
        (
            _parse_dt(row.get("ORACLE_NOW") or row.get("oracle_now"))
            for row in raw_rows
            if row.get("ORACLE_NOW") or row.get("oracle_now")
        ),
        None,
    ) or datetime.now()
    rows_by_window: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for raw in raw_rows:
        ts = _parse_dt(raw.get("FCREAREG") or raw.get("fcreareg"))
        if not ts or ts < query_start or ts >= query_end:
            continue
        logistic_day = (ts - timedelta(hours=6)).date().isoformat()
        if 6 <= ts.hour < 14:
            turno = "Manana"
        elif 14 <= ts.hour < 22:
            turno = "Tarde"
        else:
            turno = "Noche"
        rows_by_window[(logistic_day, turno)].append(raw)

    operations: list[dict[str, Any]] = []
    day_cursor = fecha_desde_dt.date()
    while day_cursor <= fecha_hasta_dt.date():
        windows = [
            ("Manana", datetime.combine(day_cursor, dt_time(6, 0)), datetime.combine(day_cursor, dt_time(14, 0))),
            ("Tarde", datetime.combine(day_cursor, dt_time(14, 0)), datetime.combine(day_cursor, dt_time(22, 0))),
            ("Noche", datetime.combine(day_cursor, dt_time(22, 0)), datetime.combine(day_cursor + timedelta(days=1), dt_time(6, 0))),
        ]
        for turno, window_start, window_end in windows:
            window_rows = rows_by_window.get((day_cursor.isoformat(), turno), [])
            if not window_rows:
                continue
            cutoff = min(oracle_now, window_end) if oracle_now >= window_start else window_start
            for operation in build_rendimiento_online_operations(window_rows, cutoff):
                operations.append({**operation, "fecha": day_cursor.isoformat(), "turno": turno})
        day_cursor += timedelta(days=1)

    grouped_operations: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for operation in operations:
        legajo = _norm_legajo(operation.get("legajo"))
        if not legajo:
            continue
        fecha = str(operation.get("fecha") or "")
        division = str(operation.get("division") or "SIN MAPEAR").strip() or "SIN MAPEAR"
        sector = str(operation.get("sector") or "SIN SECTOR").strip() or "SIN SECTOR"
        key = (fecha, division, sector, legajo)
        bucket = grouped_operations.setdefault(
            key,
            {
                "fecha": fecha,
                "division": division,
                "sector": sector,
                "legajo": legajo,
                "nombre_oracle": str(operation.get("nombre") or "").strip(),
                "produccion": 0.0,
                "tiempo_segundos": 0.0,
                "operaciones": 0,
                "operaciones_abiertas": 0,
                "pallets": 0,
                "pedidos": 0,
                "turnos": set(),
                "primer_movimiento": None,
                "ultimo_movimiento": None,
            },
        )
        bucket["produccion"] += _to_float(operation.get("bultos"))
        bucket["tiempo_segundos"] += _to_float(operation.get("segundos"))
        bucket["operaciones"] += 1
        bucket["operaciones_abiertas"] += int(str(operation.get("cierre") or "") == "ABIERTA")
        bucket["pallets"] += int(operation.get("pallets") or 0)
        bucket["pedidos"] += int(operation.get("pedidos") or 0)
        bucket["turnos"].add(str(operation.get("turno") or ""))
        start = operation.get("inicio")
        end = operation.get("fin")
        if start and (bucket["primer_movimiento"] is None or start < bucket["primer_movimiento"]):
            bucket["primer_movimiento"] = start
        if end and (bucket["ultimo_movimiento"] is None or end > bucket["ultimo_movimiento"]):
            bucket["ultimo_movimiento"] = end

    legajos = {key[3] for key in grouped_operations}
    legajos.discard("")
    people_by_legajo: dict[str, dict[str, Any]] = {}
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("PRAGMA busy_timeout = 10000")
        db.row_factory = aiosqlite.Row
        if legajos:
            placeholders = ",".join("?" for _ in legajos)
            async with db.execute(
                f"""
                SELECT legajo, nombre, desc_sector_generico, desc_funcion, desc_posicion
                     , desc_grupo_personal, proveedor, razon_social, fecha_ingreso
                FROM rrhh_personas
                WHERE LTRIM(legajo, '0') IN ({placeholders})
                """,
                tuple(legajos),
            ) as cur:
                people_by_legajo = {
                    _norm_legajo(row["legajo"]): dict(row)
                    for row in await cur.fetchall()
                }

    all_divisions = sorted({key[1] for key in grouped_operations}, key=lambda value: value.upper())
    all_sectors = sorted({key[2] for key in grouped_operations}, key=lambda value: value.upper())
    sector_filter_active = bool(selected) and set(selected) != set(all_sectors)
    rows: list[dict[str, Any]] = []
    for data in grouped_operations.values():
        legajo = data["legajo"]
        person = people_by_legajo.get(legajo, {})
        sector = data["sector"]
        grupo_personal = str(person.get("desc_grupo_personal") or "").strip() or "SIN GRUPO"
        razon_social = str(person.get("razon_social") or "").strip() or "SIN PROVEEDOR"
        fecha_ingreso = str(person.get("fecha_ingreso") or "").strip()
        funcion_legajero = str(person.get("desc_funcion") or "").strip() or "SIN FUNCION"
        sector_legajero = str(person.get("desc_sector_generico") or "").strip() or "SIN SECTOR LEGAJERO"
        produccion = _to_float(data.get("produccion"))
        tiempo_segundos = _to_float(data.get("tiempo_segundos"))
        productividad = _productividad_hora(produccion, tiempo_segundos)
        antiguedad_anios = _years_between(person.get("fecha_ingreso"), fecha_hasta_dt)
        if sector_filter_active and sector not in selected:
            continue
        if dotacion != "ALL" and grupo_personal != dotacion:
            continue
        if proveedor != "ALL" and razon_social != proveedor:
            continue
        if not _matches_antiguedad(antiguedad_anios, antiguedad):
            continue
        rows.append(
            {
                "fecha": data["fecha"],
                "division": data["division"],
                "legajo": legajo,
                "nombre": person.get("nombre") or data.get("nombre_oracle") or legajo,
                "sector": sector,
                "sector_legajero": sector_legajero,
                "funcion": funcion_legajero,
                "funcion_productiva": "PICKING",
                "posicion": str(person.get("desc_posicion") or "").strip(),
                "dotacion": grupo_personal,
                "proveedor": razon_social,
                "fecha_ingreso": fecha_ingreso[:10] if fecha_ingreso else "",
                "antiguedad_anios": antiguedad_anios,
                "tipo": "BULTOS",
                "produccion": round(produccion, 2),
                "tiempo_neto_segundos": round(tiempo_segundos, 2),
                "tiempo_total_segundos": round(tiempo_segundos, 2),
                "productividad": round(productividad, 2),
                "productividad_neta": round(productividad, 2),
                "productividad_bruta": round(productividad, 2),
                "operaciones": int(data.get("operaciones") or 0),
                "operaciones_abiertas": int(data.get("operaciones_abiertas") or 0),
                "pallets": int(data.get("pallets") or 0),
                "pedidos": int(data.get("pedidos") or 0),
                "turnos": ", ".join(sorted(item for item in data.get("turnos", set()) if item)),
                "primer_movimiento": data["primer_movimiento"].strftime("%Y-%m-%d %H:%M:%S") if data.get("primer_movimiento") else "",
                "ultimo_movimiento": data["ultimo_movimiento"].strftime("%Y-%m-%d %H:%M:%S") if data.get("ultimo_movimiento") else "",
            }
        )

    totals_rows_by_date: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        totals_rows_by_date[row["fecha"]].append(row)

    def aggregate(items: list[dict[str, Any]]) -> dict[str, Any]:
        produccion = sum(_to_float(row.get("produccion")) for row in items)
        tiempo_neto = sum(_to_float(row.get("tiempo_neto_segundos")) for row in items)
        tiempo_total = sum(_to_float(row.get("tiempo_total_segundos")) for row in items)
        operarios = len({_norm_legajo(row.get("legajo")) for row in items if _norm_legajo(row.get("legajo"))})
        funciones = len({str(row.get("funcion") or "").strip() for row in items if str(row.get("funcion") or "").strip()})
        tipos = len({str(row.get("tipo") or "").strip() for row in items if str(row.get("tipo") or "").strip()})
        productividad_neta = produccion / (tiempo_neto / 3600) if tiempo_neto > 0 else 0
        productividad_bruta = produccion / (tiempo_total / 3600) if tiempo_total > 0 else 0
        return {
            "operarios": operarios,
            "registros": len(items),
            "divisiones": len({str(row.get("division") or "") for row in items if row.get("division")}),
            "funciones": funciones,
            "tipos": tipos,
            "operaciones": sum(int(row.get("operaciones") or 0) for row in items),
            "produccion": round(produccion, 2),
            "tiempo_neto_segundos": round(tiempo_neto, 2),
            "tiempo_total_segundos": round(tiempo_total, 2),
            "productividad": round(productividad_neta, 2),
            "productividad_neta": round(productividad_neta, 2),
            "productividad_bruta": round(productividad_bruta, 2),
            "productividad_promedio": round(productividad_neta, 2),
        }

    dates = sorted(totals_rows_by_date)
    totals_by_date = {
        fecha: {"fecha": fecha, "sector": "TOTAL", **aggregate(items)}
        for fecha, items in totals_rows_by_date.items()
    }
    by_division_groups: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    by_sector_groups: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_division_groups[(row["fecha"], row["division"])].append(row)
        by_sector_groups[(row["fecha"], row["division"], row["sector"])].append(row)
    by_division_day = [
        {"fecha": fecha, "division": division, **aggregate(items)}
        for (fecha, division), items in sorted(by_division_groups.items())
    ]
    by_sector_day = [
        {"fecha": fecha, "division": division, "sector": sector, **aggregate(items)}
        for (fecha, division, sector), items in sorted(by_sector_groups.items())
    ]
    summary = aggregate(rows)
    first = totals_by_date.get(dates[0], {}) if dates else {}
    last = totals_by_date.get(dates[-1], {}) if dates else {}
    summary.update({
        "dias": len(dates),
        "divisiones": len({row["division"] for row in rows}),
        "sectores": len({(row["division"], row["sector"]) for row in rows}),
        "corridas": 0,
        "delta_productividad_pct": _pct_change(_to_float(last.get("productividad")), _to_float(first.get("productividad"))) if dates else None,
    })
    operator_trends = _build_legajos_operator_trends(rows, dates, totals_by_date)
    summary["legajos_con_tendencia"] = operator_trends["count"]
    rows.sort(key=lambda row: (row["fecha"], row["division"], row["sector"], row["legajo"]))
    return {
        "source": "rendimiento_online_raw",
        "source_label": "Mismo crudo de Rendimiento Online: F132HIST + F602ASEC + PV_LEGAJO + F002ARTI + VW_UBICACIONES_DIVISION; enriquecido con rrhh_personas.",
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "raw_rows_count": len(raw_rows),
        "operations_count": len(operations),
        "filtered_rows_count": len(rows),
        "sector_filter_active": sector_filter_active,
        "division_options": all_divisions,
        "sector_options": all_sectors,
        "selected_sectors": selected,
        "selected_filters": {"dotacion": dotacion, "proveedor": proveedor, "antiguedad": antiguedad},
        "legajero_matches": len(people_by_legajo),
        "legajero_missing": max(len(legajos) - len(people_by_legajo), 0),
        "summary": summary,
        "by_division_day": by_division_day,
        "by_sector_day": by_sector_day,
        "totals_by_day": list(totals_by_date.values()),
        "operator_trends": {
            "improving": operator_trends["improving"],
            "worsening": operator_trends["worsening"],
            "count": operator_trends["count"],
        },
        "charts": operator_trends["charts"],
        "grid": rows,
    }


def _cargo_grupo(funcion: str) -> str:
    value = str(funcion or "").strip().upper()
    if not value:
        return "SIN CARGO"
    return "ARMADOR" if "ARMADOR" in value else "NO ARMADOR"


def _net_minutes(row: dict[str, Any]) -> float:
    return max(
        0.0,
        round(
            float(row.get("minutos_productivos") or 0)
            - float(row.get("minutos_entrega_primer") or 0)
            - float(row.get("minutos_ultimo_devol") or 0),
            2,
        ),
    )


def _enrich_segments_with_legajero(segments: list[dict[str, Any]], profile_by_legajo: dict[str, dict[str, str]]) -> list[dict[str, Any]]:
    enriched = []
    for row in segments:
        profile = profile_by_legajo.get(_norm_legajo(row.get("copecrea")), {})
        sector = profile.get("sector", "")
        funcion = profile.get("funcion", "")
        posicion = profile.get("posicion", "")
        expected = sorted(_expected_sectors_for_almacen(row.get("almacen")))
        fuera_sector = int(bool(sector and expected and sector not in expected))
        enriched.append(
            {
                **row,
                "sector_legajero": sector or "SIN LEGAJO",
                "sector_esperado": " / ".join(expected),
                "dotacion_fuera_almacen": fuera_sector,
                "cargo_legajero": funcion or "SIN CARGO",
                "posicion_legajero": posicion or "SIN POSICION",
                "cargo_grupo": _cargo_grupo(funcion),
                "minutos_productivos_netos": _net_minutes(row),
            }
        )
    return enriched


def _event_uid(row: dict[str, Any]) -> str:
    parts = [
        str(row.get("COPECREA") or row.get("copecrea") or ""),
        str(row.get("FH_MOVIMIENTO") or row.get("fh_movimiento") or ""),
        str(row.get("OPERACION") or row.get("operacion") or ""),
        str(row.get("NRO_PALLET") or row.get("nro_pallet") or ""),
        str(row.get("PEDIDO") or row.get("pedido") or ""),
        str(row.get("CANTIDAD") or row.get("cantidad") or ""),
    ]
    return hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()


def _shift_for_dt(ts: datetime) -> tuple[str, str, datetime, datetime]:
    base = ts.date()
    if 6 <= ts.hour < 14:
        start = datetime.combine(base, datetime.min.time()) + timedelta(hours=6)
        return "manana", str(base), start, start + timedelta(hours=8)
    if 14 <= ts.hour < 22:
        start = datetime.combine(base, datetime.min.time()) + timedelta(hours=14)
        return "tarde", str(base), start, start + timedelta(hours=8)
    if ts.hour >= 22:
        start = datetime.combine(base, datetime.min.time()) + timedelta(hours=22)
        return "noche", str(base), start, start + timedelta(hours=8)
    start_base = base - timedelta(days=1)
    start = datetime.combine(start_base, datetime.min.time()) + timedelta(hours=22)
    return "noche", str(start_base), start, start + timedelta(hours=8)


def _shift_windows_between(start: datetime, end: datetime) -> list[tuple[str, str, datetime, datetime]]:
    windows: list[tuple[str, str, datetime, datetime]] = []
    cursor = start - timedelta(hours=8)
    cursor = datetime.combine(cursor.date(), datetime.min.time(), tzinfo=start.tzinfo)
    limit = end + timedelta(hours=8)
    while cursor <= limit:
        for key, hour in (("manana", 6), ("tarde", 14), ("noche", 22)):
            shift_start = cursor + timedelta(hours=hour)
            shift_end = shift_start + timedelta(hours=8)
            if shift_end > start and shift_start < end:
                windows.append((key, str(shift_start.date()), shift_start, shift_end))
        cursor += timedelta(days=1)
    return sorted(windows, key=lambda item: item[2])


def _normalize_productividad_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = []
    seen = set()
    for raw in rows:
        row = {str(k).upper(): v for k, v in raw.items()}
        ts = _parse_dt(row.get("FH_MOVIMIENTO"))
        copecrea = str(row.get("COPECREA") or "").strip()
        operacion = str(row.get("OPERACION") or "").strip().upper()
        if not ts or not copecrea or not operacion:
            continue
        uid = _event_uid(row)
        if uid in seen:
            continue
        seen.add(uid)
        turno_key, fecha_operativa, _, _ = _shift_for_dt(ts)
        normalized.append(
            {
                "event_uid": uid,
                "fh_movimiento": ts,
                "fecha_operativa": fecha_operativa,
                "turno_key": turno_key,
                "turno_label": SHIFT_LABELS[turno_key],
                "almacen": str(row.get("ALMACEN") or "SIN MAPEAR").strip() or "SIN MAPEAR",
                "copecrea": copecrea,
                "operario": str(row.get("OPERARIO") or "").strip(),
                "operacion": operacion,
                "zona_origen": str(row.get("ZONA_ORIGEN") or "").strip(),
                "ubic_origen": str(row.get("UBIC_ORIGEN") or "").strip(),
                "nro_pallet": str(row.get("NRO_PALLET") or "").strip(),
                "pedido": str(row.get("PEDIDO") or "").strip(),
                "cantidad": _to_float(row.get("CANTIDAD")),
                "peso": _to_float(row.get("PESO")),
                "source_table": str(row.get("SOURCE_TABLE") or "").strip(),
            }
        )
    return sorted(normalized, key=lambda item: (item["copecrea"], item["fh_movimiento"], item["operacion"]))


def _minutes(a: datetime, b: datetime) -> float:
    return max(0.0, round((b - a).total_seconds() / 60, 2))


def _build_sessions(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sessions = []
    for copecrea, items_iter in _group_by(events, "copecrea").items():
        current: dict[str, Any] | None = None
        operario = ""
        for event in items_iter:
            operario = event.get("operario") or operario
            op = event["operacion"]
            if op == "ENTREGA DE EQUIPO":
                if current and current["pickings"]:
                    current["fin"] = current["pickings"][-1]["fh_movimiento"]
                    current["incompleta"] = 1
                    sessions.append(current)
                current = {"copecrea": copecrea, "operario": operario, "inicio": event["fh_movimiento"], "fin": None, "pickings": [], "incompleta": 0}
            elif op == "PICKING":
                if not current:
                    current = {"copecrea": copecrea, "operario": operario, "inicio": event["fh_movimiento"], "fin": None, "pickings": [], "incompleta": 1}
                current["pickings"].append(event)
            elif op == "DEVOLUCION DE EQUIPO" and current:
                current["fin"] = event["fh_movimiento"]
                sessions.append(current)
                current = None
        if current and current["pickings"]:
            current["fin"] = current["pickings"][-1]["fh_movimiento"]
            current["incompleta"] = 1
            sessions.append(current)
    return sessions


def _group_by(rows: list[dict[str, Any]], key: str) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[str(row.get(key) or "")].append(row)
    return grouped


def _build_segments(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    segments = []
    for session_index, session in enumerate(_build_sessions(events)):
        pickings = session["pickings"]
        if not pickings:
            continue
        start = session["inicio"]
        end = session["fin"] or pickings[-1]["fh_movimiento"]
        for turno_key, fecha_operativa, shift_start, shift_end in _shift_windows_between(start, end):
            shift_pickings = [p for p in pickings if shift_start <= p["fh_movimiento"] < shift_end]
            if not shift_pickings:
                continue
            for almacen, almacen_pickings in _group_by(shift_pickings, "almacen").items():
                almacen_pickings.sort(key=lambda item: item["fh_movimiento"])
                first_pick = almacen_pickings[0]["fh_movimiento"]
                last_pick = almacen_pickings[-1]["fh_movimiento"]
                first_session_pick = pickings[0]["fh_movimiento"]
                if first_pick == first_session_pick and start <= first_pick and start < shift_start:
                    segment_start = start
                else:
                    segment_start = max(start, shift_start)
                segment_end = min(end, shift_end)
                gaps = [
                    _minutes(almacen_pickings[i - 1]["fh_movimiento"], almacen_pickings[i]["fh_movimiento"])
                    for i in range(1, len(almacen_pickings))
                ]
                bultos = sum(item["cantidad"] for item in almacen_pickings)
                traspasos = len(almacen_pickings)
                raw_uid = f'{session["copecrea"]}|{session_index}|{fecha_operativa}|{turno_key}|{almacen}'
                segments.append(
                    {
                        "segment_uid": hashlib.sha1(raw_uid.encode("utf-8")).hexdigest(),
                        "fecha_operativa": fecha_operativa,
                        "turno_key": turno_key,
                        "turno_label": SHIFT_LABELS[turno_key],
                        "almacen": almacen or "SIN MAPEAR",
                        "copecrea": session["copecrea"],
                        "operario": session.get("operario") or "",
                        "inicio_productivo": segment_start,
                        "fin_productivo": segment_end,
                        "primer_traspaso": first_pick,
                        "ultimo_traspaso": last_pick,
                        "minutos_productivos": _minutes(segment_start, segment_end),
                        "minutos_entrega_primer": _minutes(segment_start, first_pick),
                        "minutos_entre_traspasos": round(sum(gaps) / len(gaps), 2) if gaps else 0,
                        "minutos_ultimo_devol": _minutes(last_pick, segment_end),
                        "traspasos": traspasos,
                        "bultos": round(bultos, 2),
                        "peso": round(sum(item["peso"] for item in almacen_pickings), 2),
                        "promedio_bultos_traspaso": round(bultos / traspasos, 2) if traspasos else 0,
                        "sesion_incompleta": int(session.get("incompleta") or 0),
                        "excede_turno": int(start < shift_start or end > shift_end),
                    }
                )
    return segments


def _weighted_avg(rows: list[dict[str, Any]], field: str, weight: str = "traspasos") -> float:
    total_weight = sum(float(row.get(weight) or 0) for row in rows)
    if total_weight <= 0:
        values = [float(row.get(field) or 0) for row in rows]
        return round(sum(values) / len(values), 2) if values else 0
    return round(sum(float(row.get(field) or 0) * float(row.get(weight) or 0) for row in rows) / total_weight, 2)


def _build_productividad_payload(
    run_id: int,
    fecha_desde: str,
    fecha_hasta: str,
    events: list[dict[str, Any]],
    segments: list[dict[str, Any]],
    source_rows_count: int,
    source_name: str,
) -> dict[str, Any]:
    by_almacen = []
    for almacen, rows in sorted(_group_by(segments, "almacen").items()):
        bultos = sum(row["bultos"] for row in rows)
        traspasos = sum(row["traspasos"] for row in rows)
        operarios = len({row["copecrea"] for row in rows})
        by_almacen.append(
            {
                "almacen": almacen,
                "bultos": round(bultos, 2),
                "traspasos": traspasos,
                "operarios": operarios,
                "operarios_fuera_sector": len({row["copecrea"] for row in rows if row.get("dotacion_fuera_almacen")}),
                "segmentos_fuera_sector": sum(int(row.get("dotacion_fuera_almacen") or 0) for row in rows),
                "bultos_fuera_sector": round(sum(float(row.get("bultos") or 0) for row in rows if row.get("dotacion_fuera_almacen")), 2),
                "minutos_productivos": round(sum(row["minutos_productivos"] for row in rows), 2),
                "minutos_productivos_netos": round(sum(_net_minutes(row) for row in rows), 2),
                "productividad": round(bultos / (sum(row["minutos_productivos"] for row in rows) / 60), 2)
                if sum(row["minutos_productivos"] for row in rows) > 0
                else 0,
                "productividad_neta": round(bultos / (sum(_net_minutes(row) for row in rows) / 60), 2)
                if sum(_net_minutes(row) for row in rows) > 0
                else 0,
                "tiempo_entrega_primer": _weighted_avg(rows, "minutos_entrega_primer"),
                "tiempo_entre_traspasos": _weighted_avg(rows, "minutos_entre_traspasos"),
                "tiempo_ultimo_devolucion": _weighted_avg(rows, "minutos_ultimo_devol"),
                "promedio_bultos_traspaso": round(bultos / traspasos, 2) if traspasos else 0,
                "sesiones_incompletas": sum(row["sesion_incompleta"] for row in rows),
                "segmentos_extra_turno": sum(row["excede_turno"] for row in rows),
            }
        )
    summary_rows = segments
    total_bultos = sum(row["bultos"] for row in summary_rows)
    total_traspasos = sum(row["traspasos"] for row in summary_rows)
    total_minutos = sum(row["minutos_productivos"] for row in summary_rows)
    total_minutos_netos = sum(_net_minutes(row) for row in summary_rows)
    summary = {
        "bultos": round(total_bultos, 2),
        "traspasos": total_traspasos,
        "operarios": len({row["copecrea"] for row in summary_rows}),
        "almacenes": len(by_almacen),
        "operarios_fuera_sector": len({row["copecrea"] for row in summary_rows if row.get("dotacion_fuera_almacen")}),
        "segmentos_fuera_sector": sum(int(row.get("dotacion_fuera_almacen") or 0) for row in summary_rows),
        "bultos_fuera_sector": round(sum(float(row.get("bultos") or 0) for row in summary_rows if row.get("dotacion_fuera_almacen")), 2),
        "productividad": round(total_bultos / (total_minutos / 60), 2)
        if total_minutos > 0
        else 0,
        "productividad_neta": round(total_bultos / (total_minutos_netos / 60), 2)
        if total_minutos_netos > 0
        else 0,
        "minutos_productivos_netos": round(total_minutos_netos, 2),
        "tiempo_entrega_primer": _weighted_avg(summary_rows, "minutos_entrega_primer"),
        "tiempo_entre_traspasos": _weighted_avg(summary_rows, "minutos_entre_traspasos"),
        "tiempo_ultimo_devolucion": _weighted_avg(summary_rows, "minutos_ultimo_devol"),
        "promedio_bultos_traspaso": round(total_bultos / total_traspasos, 2) if total_traspasos else 0,
        "sesiones_incompletas": sum(row["sesion_incompleta"] for row in summary_rows),
        "segmentos_extra_turno": sum(row["excede_turno"] for row in summary_rows),
    }
    tendencias = _build_productividad_trends(segments)
    grid = sorted(
        [
            {
                **row,
                "inicio_productivo": _fmt_dt(row["inicio_productivo"]),
                "fin_productivo": _fmt_dt(row["fin_productivo"]),
                "primer_traspaso": _fmt_dt(row["primer_traspaso"]),
                "ultimo_traspaso": _fmt_dt(row["ultimo_traspaso"]),
            }
            for row in segments
        ],
        key=lambda item: (item["fecha_operativa"], item["turno_key"], item["almacen"], item["copecrea"]),
    )
    return {
        "run_id": run_id,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "source_name": source_name,
        "source_rows_count": source_rows_count,
        "event_rows_count": len(events),
        "segment_rows_count": len(segments),
        "summary": summary,
        "almacenes": by_almacen,
        "tendencias": tendencias,
        "grid": grid,
    }


def _build_productividad_trends(segments: list[dict[str, Any]]) -> dict[str, Any]:
    metrics = {
        "productividad": {"label": "Productividad bruta", "unit": " bultos/h", "description": "Bultos por hora desde entrega de equipo hasta devolucion o corte de turno."},
        "productividad_neta": {"label": "Productividad neta", "unit": " bultos/h", "description": "Bultos por hora descontando entrega a primer picking y ultimo picking a devolucion."},
        "bultos": {"label": "Bultos preparados", "unit": "", "description": "Volumen total preparado en picking durante el periodo seleccionado."},
        "operarios": {"label": "Cantidad de operarios", "unit": "", "description": "Personas distintas con actividad de picking en cada dia y almacen."},
        "operarios_fuera_sector": {
            "label": "Otros sectores en el almacen",
            "unit": "",
            "description": "Personas de otros sectores que hicieron picking en el almacen graficado. No mide gente de ese almacen trabajando afuera.",
        },
        "tiempo_entrega_primer": {"label": "Entrega a primer picking", "unit": " min", "description": "Minutos entre entrega de equipo y primer movimiento de picking."},
        "tiempo_entre_traspasos": {"label": "Tiempo medio entre pickings", "unit": " min", "description": "Pausa promedio entre movimientos de picking; ayuda a detectar esperas o recorridos."},
        "tiempo_ultimo_devolucion": {"label": "Ultimo picking a devolucion", "unit": " min", "description": "Minutos entre el ultimo picking y la devolucion del equipo."},
        "promedio_bultos_traspaso": {"label": "Promedio bultos por linea", "unit": "", "description": "Cantidad promedio de bultos procesados por cada linea de picking."},
    }
    dates = sorted({str(row.get("fecha_operativa") or "") for row in segments if row.get("fecha_operativa")})
    almacenes = sorted({str(row.get("almacen") or "SIN MAPEAR") for row in segments})
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in segments:
        grouped[(str(row.get("fecha_operativa") or ""), str(row.get("almacen") or "SIN MAPEAR"))].append(row)

    def aggregate(rows: list[dict[str, Any]]) -> dict[str, float]:
        bultos = sum(float(row.get("bultos") or 0) for row in rows)
        traspasos = sum(int(row.get("traspasos") or 0) for row in rows)
        minutos = sum(float(row.get("minutos_productivos") or 0) for row in rows)
        minutos_netos = sum(_net_minutes(row) for row in rows)
        return {
            "productividad": round(bultos / (minutos / 60), 2) if minutos > 0 else 0,
            "productividad_neta": round(bultos / (minutos_netos / 60), 2) if minutos_netos > 0 else 0,
            "bultos": round(bultos, 2),
            "operarios": float(len({row.get("copecrea") for row in rows if row.get("copecrea")})),
            "operarios_fuera_sector": float(
                len({row.get("copecrea") for row in rows if row.get("copecrea") and row.get("dotacion_fuera_almacen")})
            ),
            "tiempo_entrega_primer": _weighted_avg(rows, "minutos_entrega_primer"),
            "tiempo_entre_traspasos": _weighted_avg(rows, "minutos_entre_traspasos"),
            "tiempo_ultimo_devolucion": _weighted_avg(rows, "minutos_ultimo_devol"),
            "promedio_bultos_traspaso": round(bultos / traspasos, 2) if traspasos else 0,
        }

    charts = []
    for metric_key, meta in metrics.items():
        total_series = []
        for date in dates:
            rows = [row for almacen in almacenes for row in grouped.get((date, almacen), [])]
            total_series.append({"fecha": date, "valor": aggregate(rows)[metric_key] if rows else 0})
        almacen_series = []
        for almacen in almacenes:
            points = []
            for date in dates:
                rows = grouped.get((date, almacen), [])
                points.append({"fecha": date, "valor": aggregate(rows)[metric_key] if rows else 0})
            almacen_series.append({"almacen": almacen, "points": points})
        charts.append(
            {
                "metric": metric_key,
                "label": meta["label"],
                "unit": meta["unit"],
                "description": meta["description"],
                "total": total_series,
                "almacenes": almacen_series,
            }
        )
    return {"fechas": dates, "almacenes": almacenes, "charts": charts}


def _compact_productividad_for_ai(payload: dict[str, Any]) -> dict[str, Any]:
    charts = payload.get("tendencias", {}).get("charts") or []
    compact_charts = []
    for chart in charts:
        compact_charts.append(
            {
                "metric": chart.get("metric"),
                "label": chart.get("label"),
                "unit": chart.get("unit"),
                "total": chart.get("total", []),
                "almacenes": [
                    {
                        "almacen": serie.get("almacen"),
                        "points": serie.get("points", []),
                    }
                    for serie in chart.get("almacenes", [])[:8]
                ],
            }
        )
    grid = payload.get("grid") or []
    worst_rows = sorted(
        grid,
        key=lambda row: (
            -float(row.get("minutos_entrega_primer") or 0),
            -float(row.get("minutos_ultimo_devol") or 0),
            -float(row.get("minutos_entre_traspasos") or 0),
        ),
    )[:25]
    return {
        "fecha_desde": payload.get("fecha_desde"),
        "fecha_hasta": payload.get("fecha_hasta"),
        "source_name": payload.get("source_name"),
        "summary": payload.get("summary", {}),
        "almacenes": payload.get("almacenes", []),
        "tendencias": compact_charts,
        "casos_lentos_muestra": worst_rows,
    }


def _build_productividad_ai_context(payload: dict[str, Any]) -> str:
    compact = _compact_productividad_for_ai(payload)
    return (
        "Analiza productividad de picking en un centro de distribucion.\n"
        "Objetivo: dar una lectura gerencial muy breve de tendencias y estado actual vs pasado.\n"
        "La metrica fundamental es productividad bruta = bultos / hora desde entrega a devolucion o corte de turno.\n"
        "Compara tambien productividad neta = bultos / hora descontando arranque y cierre alrededor del picking.\n"
        "Inclui dotacion cruzada si aparece: operarios cuyo sector legajero no coincide con el almacen de picking.\n"
        "Usa numeros concretos y separa total vs almacenes cuando agregue valor.\n"
        "No se envian nombres completos ni datos personales sensibles; el contexto contiene metricas agregadas y una muestra operativa compacta.\n"
        "Datos compactos JSON:\n"
        f"{json.dumps(compact, ensure_ascii=False, separators=(',', ':'))}"
    )


SYSTEM_GESTION_PRODUCTIVIDAD_IA = (
    "Sos un analista operativo senior de un centro de distribucion.\n"
    "Tu audiencia es jefe/gerente: necesitan una lectura corta, directa y accionable.\n"
    "Tu tarea es detectar por que baja o sube la productividad de picking mirando tendencias dia a dia.\n"
    "Compara el estado actual contra el pasado del rango: toma los ultimos puntos disponibles como actual "
    "y los primeros puntos como referencia pasada. No inventes causas; inferi solo desde los indicadores.\n"
    "Se extremadamente conciso: resumen maximo 180 caracteres; comparacion maximo 160 caracteres; "
    "cada titulo maximo 4 palabras; cada detalle maximo 120 caracteres.\n"
    "Devuelve maximo 3 tendencias, 3 causas, 3 acciones y 3 almacenes.\n"
    "Responde SOLO JSON valido, sin markdown:\n"
    '{"resumen":"1 frase ejecutiva",'
    '"estado_actual":"mejor|estable|peor|mixto",'
    '"comparacion_vs_pasado":"1 frase con numeros",'
    '"tendencias":[{"titulo":"max 4 palabras","detalle":"max 120 chars","impacto":"alto|medio|bajo"}],'
    '"causas_probables":[{"titulo":"max 4 palabras","detalle":"max 120 chars","almacen":"..."}],'
    '"acciones":[{"prioridad":"alta|media|baja","titulo":"max 4 palabras","detalle":"max 120 chars"}],'
    '"almacenes_a_revisar":[{"titulo":"almacen","detalle":"max 120 chars"}]}'
)


async def _store_productividad_run(
    fecha_desde: str,
    fecha_hasta: str,
    events: list[dict[str, Any]],
    segments: list[dict[str, Any]],
    source_rows_count: int,
) -> int:
    rango_key = f"{fecha_desde}|{fecha_hasta}"
    resumen = {
        "event_rows_count": len(events),
        "segment_rows_count": len(segments),
    }
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("PRAGMA busy_timeout = 10000")
        await db.execute(
            """
            INSERT INTO gestion_productividad_picking_runs
                (rango_key, fecha_desde, fecha_hasta, source_rows_count, event_rows_count, segment_rows_count, resumen_json, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(rango_key) DO UPDATE SET
                source_rows_count=excluded.source_rows_count,
                event_rows_count=excluded.event_rows_count,
                segment_rows_count=excluded.segment_rows_count,
                resumen_json=excluded.resumen_json,
                updated_at=CURRENT_TIMESTAMP
            """,
            (rango_key, fecha_desde, fecha_hasta, source_rows_count, len(events), len(segments), json.dumps(resumen)),
        )
        async with db.execute("SELECT run_id FROM gestion_productividad_picking_runs WHERE rango_key = ?", (rango_key,)) as cur:
            run_id = int((await cur.fetchone())[0])
        await db.execute("DELETE FROM gestion_productividad_picking_events WHERE run_id = ?", (run_id,))
        await db.execute("DELETE FROM gestion_productividad_picking_segments WHERE run_id = ?", (run_id,))
        await db.executemany(
            """
            INSERT OR REPLACE INTO gestion_productividad_picking_events
                (event_uid, run_id, fh_movimiento, fecha_operativa, turno_key, turno_label, almacen, copecrea,
                 operario, operacion, zona_origen, ubic_origen, nro_pallet, pedido, cantidad, peso, source_table)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    item["event_uid"], run_id, _fmt_dt(item["fh_movimiento"]), item["fecha_operativa"], item["turno_key"],
                    item["turno_label"], item["almacen"], item["copecrea"], item["operario"], item["operacion"],
                    item["zona_origen"], item["ubic_origen"], item["nro_pallet"], item["pedido"], item["cantidad"],
                    item["peso"], item["source_table"],
                )
                for item in events
            ],
        )
        await db.executemany(
            """
            INSERT OR REPLACE INTO gestion_productividad_picking_segments
                (segment_uid, run_id, fecha_operativa, turno_key, turno_label, almacen, copecrea, operario,
                 inicio_productivo, fin_productivo, primer_traspaso, ultimo_traspaso, minutos_productivos,
                 minutos_entrega_primer, minutos_entre_traspasos, minutos_ultimo_devol, traspasos, bultos,
                 peso, promedio_bultos_traspaso, sesion_incompleta, excede_turno)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    item["segment_uid"], run_id, item["fecha_operativa"], item["turno_key"], item["turno_label"],
                    item["almacen"], item["copecrea"], item["operario"], _fmt_dt(item["inicio_productivo"]),
                    _fmt_dt(item["fin_productivo"]), _fmt_dt(item["primer_traspaso"]), _fmt_dt(item["ultimo_traspaso"]),
                    item["minutos_productivos"], item["minutos_entrega_primer"], item["minutos_entre_traspasos"],
                    item["minutos_ultimo_devol"], item["traspasos"], item["bultos"], item["peso"],
                    item["promedio_bultos_traspaso"], item["sesion_incompleta"], item["excede_turno"],
                )
                for item in segments
            ],
        )
        await db.commit()
    return run_id


async def _load_productividad_run(fecha_desde: str, fecha_hasta: str) -> dict[str, Any] | None:
    rango_key = f"{fecha_desde}|{fecha_hasta}"
    desde_dt = _parse_dt(fecha_desde)
    hasta_dt = _parse_dt(fecha_hasta)
    run_ids: list[int] = []
    cache_mode = "exact"
    cache_source_ranges: list[dict[str, Any]] = []
    source_rows_count = 0
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("PRAGMA busy_timeout = 10000")
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM gestion_productividad_picking_runs WHERE rango_key = ?", (rango_key,)) as cur:
            run = await cur.fetchone()
        if not run:
            async with db.execute(
                """
                SELECT *
                FROM gestion_productividad_picking_runs
                WHERE fecha_desde <= ?
                  AND fecha_hasta >= ?
                ORDER BY updated_at DESC
                LIMIT 1
                """,
                (fecha_desde, fecha_hasta),
            ) as cur:
                run = await cur.fetchone()
            cache_mode = "contained"
        if run:
            run_ids = [int(run["run_id"])]
            source_rows_count = int(run["source_rows_count"] or 0)
            cache_source_ranges = [{"fecha_desde": run["fecha_desde"], "fecha_hasta": run["fecha_hasta"]}]
        else:
            async with db.execute(
                """
                SELECT *
                FROM gestion_productividad_picking_runs
                WHERE fecha_hasta >= ?
                  AND fecha_desde <= ?
                ORDER BY fecha_desde ASC, fecha_hasta ASC
                """,
                (fecha_desde, fecha_hasta),
            ) as cur:
                candidates = await cur.fetchall()
            covered_until = desde_dt
            selected = []
            for candidate in candidates:
                candidate_start = _parse_dt(candidate["fecha_desde"])
                candidate_end = _parse_dt(candidate["fecha_hasta"])
                if not candidate_start or not candidate_end or not covered_until:
                    continue
                if candidate_end <= covered_until:
                    continue
                if candidate_start > covered_until:
                    selected = []
                    break
                selected.append(candidate)
                covered_until = max(covered_until, candidate_end)
                if hasta_dt and covered_until >= hasta_dt:
                    break
            if not selected or not hasta_dt or not covered_until or covered_until < hasta_dt:
                return None
            cache_mode = "combined"
            run_ids = [int(item["run_id"]) for item in selected]
            source_rows_count = sum(int(item["source_rows_count"] or 0) for item in selected)
            cache_source_ranges = [{"fecha_desde": item["fecha_desde"], "fecha_hasta": item["fecha_hasta"]} for item in selected]
        if not run_ids:
            return None
        placeholders = ",".join("?" for _ in run_ids)
        event_count = source_rows_count
        if cache_mode != "combined":
            async with db.execute(
                f"""
                SELECT COUNT(DISTINCT event_uid)
                FROM gestion_productividad_picking_events
                WHERE run_id IN ({placeholders})
                  AND fh_movimiento >= ?
                  AND fh_movimiento <= ?
                """,
                (*run_ids, fecha_desde, fecha_hasta),
            ) as cur:
                event_count = int((await cur.fetchone())[0] or 0)
            if not event_count and cache_mode == "exact":
                event_count = source_rows_count
        async with db.execute(
            f"""
            SELECT *
            FROM gestion_productividad_picking_segments
            WHERE run_id IN ({placeholders})
              AND primer_traspaso >= ?
              AND primer_traspaso <= ?
            """,
            (*run_ids, fecha_desde, fecha_hasta),
        ) as cur:
            segment_rows = [dict(row) for row in await cur.fetchall()]
    unique_segment_rows = {}
    for row in segment_rows:
        key = (
            row.get("fecha_operativa"),
            row.get("turno_key"),
            row.get("almacen"),
            row.get("copecrea"),
            row.get("inicio_productivo"),
            row.get("primer_traspaso"),
            row.get("ultimo_traspaso"),
        )
        unique_segment_rows[key] = row
    segments = [
        {
            **row,
            "inicio_productivo": _parse_dt(row["inicio_productivo"]),
            "fin_productivo": _parse_dt(row["fin_productivo"]),
            "primer_traspaso": _parse_dt(row["primer_traspaso"]),
            "ultimo_traspaso": _parse_dt(row["ultimo_traspaso"]),
            "traspasos": int(row["traspasos"] or 0),
            "bultos": _to_float(row["bultos"]),
            "peso": _to_float(row["peso"]),
            "sesion_incompleta": int(row["sesion_incompleta"] or 0),
            "excede_turno": int(row["excede_turno"] or 0),
        }
        for row in unique_segment_rows.values()
    ]
    segments = _enrich_segments_with_legajero(segments, await _load_latest_legajero_profiles())
    payload = _build_productividad_payload(
        run_ids[0] if len(run_ids) == 1 else 0,
        fecha_desde,
        fecha_hasta,
        [],
        segments,
        source_rows_count or event_count,
        "sqlite_cache",
    )
    payload["event_rows_count"] = event_count
    payload["cache_mode"] = cache_mode
    payload["cache_source_range"] = cache_source_ranges[0] if len(cache_source_ranges) == 1 else None
    payload["cache_source_ranges"] = cache_source_ranges
    return payload


async def _require_request_auth(request: Request) -> dict[str, Any]:
    auth = await current_auth(request)
    if not auth or auth.get("device_status") != "approved":
        raise HTTPException(status_code=401, detail="No autenticado.")
    return auth


def _daily_productividad_raw_end(daily: dict[str, Any]) -> datetime:
    fecha_fin = daily.get("fecha_fin")
    if isinstance(fecha_fin, datetime):
        end = fecha_fin
    else:
        end = datetime.fromisoformat(str(fecha_fin))
    return end + timedelta(minutes=90)


async def run_daily_auto_productividad_raw_precache(
    *,
    force: bool = False,
    trigger: str = "scheduler",
    usuario: str = "",
    daily_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    daily = daily_override or scheduled_daily_window()
    if not daily.get("can_run"):
        return {"process": "PRODUCTIVIDAD_RAW", "status": "skipped", "reason": daily.get("reason") or "Daily no habilitada.", "daily": daily}
    latest_run = await get_latest_run(daily["daily_key"], "PRODUCTIVIDAD_RAW")
    if not force and _is_current_process_run(latest_run, DAILY_AUTO_PRODUCTIVIDAD_RAW_QUERY_VERSION):
        return {"process": "PRODUCTIVIDAD_RAW", "status": "skipped", "reason": "Cache crudo de productividad ya disponible.", "daily": daily}

    fecha_desde = _fmt_daily_oracle_dt(daily.get("fecha_inicio"))
    fecha_hasta_raw = _daily_productividad_raw_end(daily)
    fecha_hasta = _fmt_daily_oracle_dt(fecha_hasta_raw)
    started_at = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
    started = time.perf_counter()
    logger.info("[daily-auto] Iniciando precarga cruda productividad %s %s..%s", daily["daily_key"], fecha_desde, fecha_hasta)
    try:
        rows = await asyncio.to_thread(query_productive_db_daily_productividad_raw, fecha_desde, fecha_hasta)
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        saved = await save_productividad_raw_cache(
            daily,
            rows,
            fecha_fin_raw=fecha_hasta_raw.isoformat(timespec="seconds"),
            started_at=started_at,
            duration_ms=elapsed_ms,
            timings={"productividad_raw_oracle_ms": elapsed_ms, "query_version": DAILY_AUTO_PRODUCTIVIDAD_RAW_QUERY_VERSION},
            trigger=trigger,
            usuario=usuario,
            retention_days=30 if str(trigger).startswith("manual_comparison") else 5,
        )
        logger.info("[daily-auto] Precarga cruda productividad OK: %s filas en %sms", len(rows), elapsed_ms)
        return {"process": "PRODUCTIVIDAD_RAW", "status": "success", "daily": daily, "rows": len(rows), "duration_ms": elapsed_ms, "saved": saved}
    except Exception as exc:
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        await mark_run_error(daily, "PRODUCTIVIDAD_RAW", str(exc), started_at, trigger=trigger, usuario=usuario)
        logger.exception("[daily-auto] Precarga cruda productividad fallo tras %sms", elapsed_ms)
        return {"process": "PRODUCTIVIDAD_RAW", "status": "error", "daily": daily, "duration_ms": elapsed_ms, "error": str(exc)}


def _build_picking_rows_from_productividad_raw(
    daily: dict[str, Any],
    raw_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    start = _parse_dt(daily.get("fecha_inicio"))
    end = _parse_dt(daily.get("fecha_fin"))
    if not start or not end:
        raise ValueError("Ventana Daily invalida para calcular Picking desde cache crudo.")

    scoped_rows = []
    for row in raw_rows:
        ts = _parse_dt(row.get("FCREAREG") or row.get("fecha"))
        legajo = _norm_legajo(row.get("COPECREA") or row.get("LEGAJO") or row.get("legajo"))
        if not ts or not legajo:
            continue
        if ts < start or ts >= end:
            continue
        scoped_rows.append({**row, "_ts": ts, "_legajo": legajo})

    scoped_rows.sort(key=lambda item: (item["_legajo"], item["_ts"], int(item.get("_row_index") or 0)))

    previous_by_legajo: dict[str, datetime] = {}
    picking_rows: list[dict[str, Any]] = []
    for row in scoped_rows:
        legajo = row["_legajo"]
        ts = row["_ts"]
        prev = previous_by_legajo.get(legajo)
        previous_by_legajo[legajo] = ts
        operacion = str(row.get("CDESCRIP") or row.get("OPERACION") or "").strip().upper()
        if operacion != "PICKING":
            continue
        seconds = max((ts - prev).total_seconds(), 0.0) if prev else 0.0
        hs_picking = seconds / 3600
        antiguedad = _picking_legajo_antiguedad(row, start)
        picking_rows.append(
            {
                "ALMACEN": _daily_productividad_raw_division(row),
                "COPECREA": legajo,
                "BULTOS_PICKING": _to_float(row.get("QCANTIDA") if row.get("QCANTIDA") is not None else row.get("cantidad")),
                "HS_PICKING": hs_picking if antiguedad["cuenta_legajo_bool"] else 0.0,
                "HS_PICKING_TOTAL": hs_picking,
                "FEC_ING": antiguedad["fec_ing"],
                "ANTIGUEDAD_DIAS": antiguedad["antiguedad_dias"],
                "CUENTA_LEGAJO": antiguedad["cuenta_legajo"],
                "MOTIVO_NO_CUENTA": antiguedad["motivo_no_cuenta"],
                "FCREAREG": ts.isoformat(sep=" "),
            }
        )
    return picking_rows


def _build_recepcion_rows_from_productividad_raw(
    daily: dict[str, Any],
    raw_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    start = _parse_dt(daily.get("fecha_inicio"))
    end = _parse_dt(daily.get("fecha_fin"))
    if not start or not end:
        raise ValueError("Ventana Daily invalida para calcular Recepcion desde cache crudo.")

    grouped: dict[str, dict[str, set[str]]] = {}
    for row in raw_rows:
        ts = _parse_dt(row.get("FCREAREG") or row.get("fecha"))
        if not ts or ts < start or ts >= end:
            continue
        operacion = str(row.get("CDESCRIP") or row.get("OPERACION") or "").strip().upper()
        if operacion != "REVISION PALETS ENTRADA":
            continue
        if _to_float(row.get("QCANTIDA") if row.get("QCANTIDA") is not None else row.get("cantidad")) <= 0:
            continue
        almacen = str(row.get("ALMACEN") or "").strip().upper()
        if almacen not in {"NOA", "SECOS", "REFRIGERADOS"}:
            continue
        bucket = grouped.setdefault(almacen, {"pallets": set(), "legajos": set()})
        pallet = str(row.get("CNUPALET") or "").strip()
        legajo = _norm_legajo(row.get("COPECREA") or row.get("LEGAJO") or row.get("legajo"))
        if pallet:
            bucket["pallets"].add(pallet)
        if legajo:
            bucket["legajos"].add(legajo)

    return [
        {
            "ALMACEN": almacen,
            "PALLETS": float(len(values["pallets"])),
            "LEGAJOS": float(len(values["legajos"])),
            "PRODUCCION": round(len(values["pallets"]) / len(values["legajos"]), 2) if values["legajos"] else 0.0,
        }
        for almacen, values in grouped.items()
    ]


def _build_clark_rows_from_productividad_raw(
    daily: dict[str, Any],
    raw_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    start = _parse_dt(daily.get("fecha_inicio"))
    end = _parse_dt(daily.get("fecha_fin"))
    if not start or not end:
        raise ValueError("Ventana Daily invalida para calcular Clark desde cache crudo.")

    clark_ops = {
        "GUARADO PALETS ENTRADA",
        "EXTRACCION DE REAPROS",
        "EXTRACCION TRASPASOS",
        "SURTIDO P.COMPLETOS",
    }
    scoped_rows = []
    for row in raw_rows:
        ts = _parse_dt(row.get("FCREAREG") or row.get("fecha"))
        legajo = _norm_legajo(row.get("COPECREA") or row.get("LEGAJO") or row.get("legajo"))
        if not ts or not legajo:
            continue
        if ts < start or ts >= end:
            continue
        scoped_rows.append({**row, "_ts": ts, "_legajo": legajo})

    scoped_rows.sort(key=lambda item: (item["_legajo"], item["_ts"], int(item.get("_row_index") or 0)))
    grouped: dict[str, dict[str, Any]] = {}
    previous_by_legajo: dict[str, datetime] = {}
    for row in scoped_rows:
        legajo = row["_legajo"]
        ts = row["_ts"]
        prev = previous_by_legajo.get(legajo)
        previous_by_legajo[legajo] = ts
        operacion = str(row.get("CDESCRIP") or row.get("OPERACION") or "").strip().upper()
        if operacion not in clark_ops:
            continue
        almacen = _daily_productividad_raw_division(row)
        if almacen not in {"NOA", "SECOS", "OC", "CONGELADOS"}:
            continue
        pallet = str(row.get("CNUPALET") or "").strip()
        if not _valid_clark_pallet(row):
            continue
        bucket = grouped.setdefault(
            almacen,
            {
                "pallets": set(),
                "hs_clark": 0.0,
                "legajos_clark": set(),
                "pallets_spc": set(),
                "legajos_spc": set(),
            },
        )
        seconds = max((ts - prev).total_seconds(), 0.0) if prev else 0.0
        bucket["hs_clark"] += seconds / 3600
        bucket["legajos_clark"].add(legajo)
        bucket["pallets"].add(pallet)
        if operacion == "SURTIDO P.COMPLETOS":
            bucket["pallets_spc"].add(pallet)
        if operacion == "SURTIDO P.COMPLETOS":
            bucket["legajos_spc"].add(legajo)

    return [
        {
            "ALMACEN": almacen,
            "PALLETS_DISTINTOS": float(len(values["pallets"])),
            "HS_CLARK_TOTAL": float(values["hs_clark"]),
            "LEGAJOS_CLARK": float(len(values["legajos_clark"])),
            "PALLETS_SPC_DISTINTOS": float(len(values["pallets_spc"])),
            "LEGAJOS_SPC": float(len(values["legajos_spc"])),
            "LEGAJOS_SPC_IDS": sorted(values["legajos_spc"]),
        }
        for almacen, values in grouped.items()
    ]


async def run_daily_auto_despacho_raw_precache(
    *,
    force: bool = False,
    trigger: str = "scheduler",
    usuario: str = "",
    daily_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    daily = daily_override or scheduled_daily_window()
    if not daily.get("can_run"):
        return {"process": "DESPACHO_RAW", "status": "skipped", "reason": daily.get("reason") or "Daily no habilitada.", "daily": daily}
    latest_run = await get_latest_run(daily["daily_key"], "DESPACHO_RAW")
    if not force and _is_current_process_run(latest_run, DAILY_AUTO_DESPACHO_RAW_QUERY_VERSION):
        return {"process": "DESPACHO_RAW", "status": "skipped", "reason": "Cache crudo Despacho ya disponible.", "daily": daily}

    fecha_desde = _fmt_daily_oracle_dt(daily.get("fecha_inicio"))
    fin_dt = _parse_dt(daily.get("fecha_fin"))
    fecha_hasta = _fmt_daily_oracle_dt(fin_dt + timedelta(minutes=90) if fin_dt else daily.get("fecha_fin"))
    started_at = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
    started = time.perf_counter()
    logger.info("[daily-auto] Iniciando precarga cruda Despacho %s %s..%s", daily["daily_key"], fecha_desde, fecha_hasta)
    try:
        rows = await asyncio.to_thread(query_productive_db_daily_despacho_raw, fecha_desde, fecha_hasta)
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        saved = await save_despacho_raw_cache(
            daily,
            rows,
            started_at=started_at,
            duration_ms=elapsed_ms,
            timings={"despacho_raw_oracle_ms": elapsed_ms, "query_version": DAILY_AUTO_DESPACHO_RAW_QUERY_VERSION},
            trigger=trigger,
            usuario=usuario,
            retention_days=30 if str(trigger).startswith("manual_comparison") else 5,
        )
        logger.info("[daily-auto] Precarga cruda Despacho OK: %s filas en %sms", len(rows), elapsed_ms)
        return {"process": "DESPACHO_RAW", "status": "success", "daily": daily, "rows": len(rows), "duration_ms": elapsed_ms, "saved": saved}
    except Exception as exc:
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        await mark_run_error(daily, "DESPACHO_RAW", str(exc), started_at, trigger=trigger, usuario=usuario)
        logger.exception("[daily-auto] Precarga cruda Despacho fallo tras %sms", elapsed_ms)
        return {"process": "DESPACHO_RAW", "status": "error", "daily": daily, "duration_ms": elapsed_ms, "error": str(exc)}


def _build_despacho_rows_from_raw(daily: dict[str, Any], raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    start = _parse_dt(daily.get("fecha_inicio"))
    end = _parse_dt(daily.get("fecha_fin"))
    if not start or not end:
        raise ValueError("Ventana Daily invalida para calcular Despacho desde cache crudo.")
    grouped: dict[str, dict[str, set[str]]] = {}
    for row in raw_rows:
        ts = _parse_dt(row.get("FECIERRE") or row.get("FECHA_CIERRE"))
        if not ts or ts < start or ts >= end:
            continue
        almacen = str(row.get("ALMACEN") or "").strip().upper()
        if almacen not in {"NOA", "SECOS", "REFRIGERADOS"}:
            continue
        bucket = grouped.setdefault(almacen, {"hojas_ruta": set(), "cargadores": set()})
        hoja_ruta = str(row.get("HOJARUTA") or row.get("CNUVIAJE") or row.get("VIAJE") or "").strip()
        cargador = _norm_legajo(row.get("CARGADOR"))
        if hoja_ruta:
            bucket["hojas_ruta"].add(hoja_ruta)
        if cargador and _despacho_cargador_elegible(row):
            bucket["cargadores"].add(cargador)
    return [
        {
            "ALMACEN": almacen,
            "VIAJES": float(len(values["hojas_ruta"])),
            "CARGADORES": float(len(values["cargadores"])),
            "PRODUCCION": round(len(values["hojas_ruta"]) / len(values["cargadores"]), 2) if values["cargadores"] else 0.0,
        }
        for almacen, values in grouped.items()
    ]


def _avance_window(daily: dict[str, Any]) -> tuple[datetime, datetime]:
    start = _parse_dt(daily.get("fecha_fin"))
    if not start:
        raise ValueError("Ventana Daily invalida para calcular Avance.")
    return start, start + timedelta(minutes=90)


def _append_avance(rows: list[dict[str, Any]], grouped: dict[str, dict[str, Any]], *, proceso: str, tipo: str, value_key: str) -> None:
    for almacen, values in grouped.items():
        rows.append(
            {
                "ALMACEN": almacen,
                "PROCESO": proceso,
                "TIPO": tipo,
                "VALOR": float(values.get(value_key, 0)),
                "LEGAJOS": float(len(values.get("legajos", set()))),
            }
        )


def _build_avance_real_rows_from_raw(
    daily: dict[str, Any],
    productividad_rows: list[dict[str, Any]],
    despacho_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    start, end = _avance_window(daily)
    picking: dict[str, dict[str, Any]] = {}
    spc: dict[str, dict[str, Any]] = {}
    recepcion: dict[str, dict[str, Any]] = {}
    for row in productividad_rows:
        ts = _parse_dt(row.get("FCREAREG") or row.get("fecha"))
        if not ts or ts < start or ts >= end:
            continue
        operacion = str(row.get("CDESCRIP") or row.get("OPERACION") or "").strip().upper()
        legajo = _norm_legajo(row.get("COPECREA") or row.get("LEGAJO") or row.get("legajo"))
        cantidad = _to_float(row.get("QCANTIDA") if row.get("QCANTIDA") is not None else row.get("cantidad"))
        if operacion == "PICKING":
            almacen = _daily_picking_raw_almacen(row)
            bucket = picking.setdefault(almacen, {"valor": 0.0, "legajos": set()})
            bucket["valor"] += cantidad
            if legajo:
                bucket["legajos"].add(legajo)
        elif operacion == "SURTIDO P.COMPLETOS":
            almacen = str(row.get("ALMACEN") or "").strip().upper()
            if almacen in {"NOA", "SECOS", "REFRIGERADOS"}:
                bucket = spc.setdefault(almacen, {"pallets": set(), "legajos": set()})
                pallet = str(row.get("CNUPALET") or "").strip()
                if pallet and pallet != "0":
                    bucket["pallets"].add(pallet)
                if legajo:
                    bucket["legajos"].add(legajo)
        elif operacion == "REVISION PALETS ENTRADA" and cantidad > 0:
            almacen = str(row.get("ALMACEN") or "").strip().upper()
            if almacen in {"NOA", "SECOS", "REFRIGERADOS"}:
                bucket = recepcion.setdefault(almacen, {"pallets": set(), "legajos": set()})
                pallet = str(row.get("CNUPALET") or "").strip()
                if pallet and pallet != "0":
                    bucket["pallets"].add(pallet)
                if legajo:
                    bucket["legajos"].add(legajo)

    despacho: dict[str, dict[str, Any]] = {}
    for row in despacho_rows:
        ts = _parse_dt(row.get("FECIERRE") or row.get("FECHA_CIERRE"))
        if not ts or ts < start or ts >= end:
            continue
        almacen = str(row.get("ALMACEN") or "").strip().upper()
        if almacen not in {"NOA", "SECOS", "REFRIGERADOS"}:
            continue
        bucket = despacho.setdefault(almacen, {"hojas_ruta": set(), "legajos": set()})
        hoja_ruta = str(row.get("HOJARUTA") or row.get("CNUVIAJE") or row.get("VIAJE") or "").strip()
        cargador = _norm_legajo(row.get("CARGADOR"))
        if hoja_ruta:
            bucket["hojas_ruta"].add(hoja_ruta)
        if cargador and _despacho_cargador_elegible(row):
            bucket["legajos"].add(cargador)

    rows: list[dict[str, Any]] = []
    _append_avance(rows, picking, proceso="PICKING", tipo="REAL", value_key="valor")
    _append_avance(
        rows,
        {almacen: {"valor": len(values["pallets"]), "legajos": values["legajos"]} for almacen, values in spc.items()},
        proceso="SPC",
        tipo="REAL",
        value_key="valor",
    )
    _append_avance(
        rows,
        {almacen: {"valor": len(values["pallets"]), "legajos": values["legajos"]} for almacen, values in recepcion.items()},
        proceso="RECEPCION",
        tipo="REAL",
        value_key="valor",
    )
    _append_avance(
        rows,
        {almacen: {"valor": len(values["hojas_ruta"]), "legajos": values["legajos"]} for almacen, values in despacho.items()},
        proceso="DESPACHO",
        tipo="REAL",
        value_key="valor",
    )
    return rows


def _build_avance_plan_rows(plan_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in plan_rows:
        almacen = str(row.get("ALMACEN") or "").strip().upper()
        if almacen not in {"NOA", "SECOS", "REFRIGERADOS"}:
            continue
        rows.extend(
            [
                {"ALMACEN": almacen, "PROCESO": "PICKING", "TIPO": "PLAN", "VALOR": _to_float(row.get("BULTOS_PICKING_PLANIFICADOS")), "LEGAJOS": 0},
                {"ALMACEN": almacen, "PROCESO": "SPC", "TIPO": "PLAN", "VALOR": _to_float(row.get("PALLETS_SPC_PLANIFICADOS")), "LEGAJOS": 0},
                {"ALMACEN": almacen, "PROCESO": "DESPACHO", "TIPO": "PLAN", "VALOR": _to_float(row.get("VIAJES_PLANIFICADOS")), "LEGAJOS": 0},
            ]
        )
    return rows


SHIFT_TURN_WINDOWS = {
    "TM": ((5, 55), (14, 5), 0),
    "TT": ((13, 55), (22, 5), 0),
    "TN": ((21, 55), (6, 5), 1),
}


def _shift_turn_window(turno: str, now: datetime | None = None) -> dict[str, Any]:
    current = (now or datetime.now(LOCAL_TZ)).astimezone(LOCAL_TZ)
    turn_key = str(turno or "TM").strip().upper()
    if turn_key not in SHIFT_TURN_WINDOWS:
        turn_key = "TM"
    (start_h, start_m), (end_h, end_m), end_offset = SHIFT_TURN_WINDOWS[turn_key]
    base_date = current.date()
    if turn_key == "TN" and current.time() <= dt_time(end_h, end_m):
        base_date = base_date - timedelta(days=1)
    start = datetime.combine(base_date, dt_time(start_h, start_m), tzinfo=LOCAL_TZ)
    end = datetime.combine(base_date + timedelta(days=end_offset), dt_time(end_h, end_m), tzinfo=LOCAL_TZ)
    return {
        "turno": turn_key,
        "fecha": current.date().isoformat(),
        "fecha_inicio": start.isoformat(timespec="seconds"),
        "fecha_fin": end.isoformat(timespec="seconds"),
        "label": f"{turn_key} {start.strftime('%H:%M')} - {end.strftime('%H:%M')}",
    }


def _shift_metric(real: float, plan: float) -> dict[str, Any]:
    pendiente = max(plan - real, 0.0)
    avance = (real / plan * 100) if plan else None
    return {"real": round(real, 2), "plan": round(plan, 2), "pendiente": round(pendiente, 2), "avance_pct": round(avance, 2) if avance is not None else None}


def _shift_build_summary(
    *,
    division: str,
    start: datetime,
    end: datetime,
    productividad_rows: list[dict[str, Any]],
    despacho_rows: list[dict[str, Any]],
    plan_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    almacen_target = DAILY_RAW_SECTOR_TO_ALMACEN.get(normalize_sector(division), "SECOS")
    real = {
        "picking_bultos": 0.0,
        "spc_pallets": set(),
        "hdr": set(),
        "picking_legajos": set(),
        "spc_legajos": set(),
        "despacho_legajos": set(),
    }
    scoped_rows: list[dict[str, Any]] = []
    for row in productividad_rows:
        ts = _parse_dt(row.get("FCREAREG") or row.get("fecha"))
        legajo = _norm_legajo(row.get("COPECREA") or row.get("LEGAJO") or row.get("legajo"))
        if not ts or not legajo or ts < start or ts >= end:
            continue
        scoped_rows.append({**row, "_ts": ts, "_legajo": legajo})
        operacion = str(row.get("CDESCRIP") or row.get("OPERACION") or "").strip().upper()
        cantidad = _to_float(row.get("QCANTIDA") if row.get("QCANTIDA") is not None else row.get("cantidad"))
        if operacion == "PICKING":
            almacen = _daily_productividad_raw_division(row)
            if almacen == almacen_target:
                real["picking_bultos"] += cantidad
                real["picking_legajos"].add(legajo)
        elif operacion == "SURTIDO P.COMPLETOS":
            almacen = str(row.get("ALMACEN") or "").strip().upper()
            if almacen == almacen_target:
                pallet = str(row.get("CNUPALET") or "").strip()
                if pallet and pallet != "0":
                    real["spc_pallets"].add(pallet)
                real["spc_legajos"].add(legajo)

    for row in despacho_rows:
        ts = _parse_dt(row.get("FECIERRE") or row.get("FECHA_CIERRE"))
        almacen = str(row.get("ALMACEN") or "").strip().upper()
        if not ts or ts < start or ts >= end or almacen != almacen_target:
            continue
        hoja_ruta = str(row.get("HOJARUTA") or row.get("CNUVIAJE") or row.get("VIAJE") or "").strip()
        cargador = _norm_legajo(row.get("CARGADOR"))
        if hoja_ruta:
            real["hdr"].add(hoja_ruta)
        if cargador and _despacho_cargador_elegible(row):
            real["despacho_legajos"].add(cargador)

    plan = {"picking_bultos": 0.0, "spc_pallets": 0.0, "hdr": 0.0}
    for row in plan_rows:
        almacen = str(row.get("ALMACEN") or "").strip().upper()
        if almacen != almacen_target:
            continue
        plan["picking_bultos"] += _to_float(row.get("BULTOS_PICKING_PLANIFICADOS"))
        plan["spc_pallets"] += _to_float(row.get("PALLETS_SPC_PLANIFICADOS"))
        plan["hdr"] += _to_float(row.get("VIAJES_PLANIFICADOS"))

    scoped_rows.sort(key=lambda item: (item["_legajo"], item["_ts"], int(item.get("_row_index") or 0)))
    previous_by_legajo: dict[str, datetime] = {}
    ranking: dict[str, dict[str, Any]] = {}
    for row in scoped_rows:
        legajo = row["_legajo"]
        ts = row["_ts"]
        prev = previous_by_legajo.get(legajo, start)
        previous_by_legajo[legajo] = ts
        operacion = str(row.get("CDESCRIP") or row.get("OPERACION") or "").strip().upper()
        if operacion != "PICKING" or _daily_productividad_raw_division(row) != almacen_target:
            continue
        seconds = max((ts - prev).total_seconds(), 0.0)
        bucket = ranking.setdefault(legajo, {"legajo": legajo, "bultos": 0.0, "horas": 0.0, "productividad": 0.0})
        bucket["bultos"] += _to_float(row.get("QCANTIDA") if row.get("QCANTIDA") is not None else row.get("cantidad"))
        bucket["horas"] += seconds / 3600
    ranking_rows = []
    for item in ranking.values():
        item["productividad"] = round(item["bultos"] / item["horas"], 2) if item["horas"] else 0.0
        item["bultos"] = round(item["bultos"], 2)
        item["horas"] = round(item["horas"], 4)
        ranking_rows.append(item)
    ranking_rows.sort(key=lambda item: item["productividad"], reverse=True)

    picking_horas = sum(item["horas"] for item in ranking_rows)
    picking_legajos = len(real["picking_legajos"])
    return {
        "division": normalize_sector(division),
        "almacen": almacen_target,
        "indicadores": [
            {"key": "picking", "label": "Bultos reales vs planificados", **_shift_metric(real["picking_bultos"], plan["picking_bultos"]), "unidad": "bultos"},
            {"key": "spc", "label": "SPC reales vs planificados", **_shift_metric(float(len(real["spc_pallets"])), plan["spc_pallets"]), "unidad": "pallets"},
            {"key": "hdr", "label": "HDR reales vs planificados", **_shift_metric(float(len(real["hdr"])), plan["hdr"]), "unidad": "hojas"},
        ],
        "kpis": {
            "legajos_picking": picking_legajos,
            "bultos_por_persona": round(real["picking_bultos"] / picking_legajos, 2) if picking_legajos else 0.0,
            "bultos_por_hora": round(real["picking_bultos"] / picking_horas, 2) if picking_horas else 0.0,
            "horas_picking": round(picking_horas, 2),
            "spc_legajos": len(real["spc_legajos"]),
            "despacho_legajos": len(real["despacho_legajos"]),
        },
        "ranking": ranking_rows[:25],
    }


def _olas_picking_plan_from_cache(cached_rows: list[dict[str, Any]]) -> dict[str, dict[str, float]]:
    plan = {
        "Secos": {"bultos": 0.0, "pallets": 0.0},
        "Refrigerados": {"bultos": 0.0, "pallets": 0.0},
        "Noa": {"bultos": 0.0, "pallets": 0.0},
    }
    for row in cached_rows:
        if str(row.get("process") or "").strip().upper() != "PLANIFICACION":
            continue
        if str(row.get("id_parametro") or "").strip().upper() != "OP_CUMP_PICKING_PLAN_6A6":
            continue
        sector = normalize_sector(row.get("sector") or "")
        if sector not in plan:
            continue
        plan[sector]["bultos"] += _to_float(row.get("cantidad") if row.get("cantidad") is not None else row.get("valor"))
        plan[sector]["pallets"] += _to_float(row.get("details_count"))
    return plan


def _olas_plan_from_cache(cached_rows: list[dict[str, Any]], operation: str) -> dict[str, dict[str, float]]:
    op = str(operation or "PICKING").strip().upper()
    plan = {
        "Secos": {"cantidad": 0.0, "pallets": 0.0},
        "Refrigerados": {"cantidad": 0.0, "pallets": 0.0},
        "OC": {"cantidad": 0.0, "pallets": 0.0},
        "Congelados": {"cantidad": 0.0, "pallets": 0.0},
        "Noa": {"cantidad": 0.0, "pallets": 0.0},
    }
    param_by_op = {
        "PICKING": "OP_CUMP_PICKING_PLAN_6A6",
        "DESPACHO": "OP_CUMP_DESPACHO_PLAN_6A6",
    }
    param_id = param_by_op.get(op)
    if not param_id:
        return plan
    for row in cached_rows:
        if str(row.get("process") or "").strip().upper() != "PLANIFICACION":
            continue
        if str(row.get("id_parametro") or "").strip().upper() != param_id:
            continue
        sector = normalize_sector(row.get("sector") or "")
        if sector not in plan:
            continue
        plan[sector]["cantidad"] += _to_float(row.get("cantidad") if row.get("cantidad") is not None else row.get("valor"))
        plan[sector]["pallets"] += _to_float(row.get("details_count"))
    return plan


def _olas_operation_config(operation: str) -> dict[str, Any]:
    op = str(operation or "PICKING").strip().upper()
    configs = {
        "PICKING": {
            "operation": "PICKING",
            "label": "Picking",
            "quantity_label": "Bultos",
            "unit": "bultos",
            "operation_filter": {"PICKING"},
            "sector_source": "picking",
            "quantity_mode": "sum_cantidad",
            "positive_quantity": False,
            "time_based": True,
        },
        "CLARK": {
            "operation": "CLARK",
            "label": "Clark",
            "quantity_label": "Pallets",
            "unit": "pallets",
            "operation_filter": {"GUARADO PALETS ENTRADA", "EXTRACCION DE REAPROS", "EXTRACCION TRASPASOS", "SURTIDO P.COMPLETOS"},
            "sector_source": "productividad_division",
            "quantity_mode": "distinct_pallets",
            "positive_quantity": False,
            "time_based": True,
        },
        "RECEPCION": {
            "operation": "RECEPCION",
            "label": "Recepcion",
            "quantity_label": "Pallets",
            "unit": "pallets",
            "operation_filter": {"REVISION PALETS ENTRADA"},
            "sector_source": "raw",
            "quantity_mode": "distinct_pallets",
            "positive_quantity": True,
            "time_based": True,
        },
        "DESPACHO": {
            "operation": "DESPACHO",
            "label": "Despacho",
            "quantity_label": "Hojas de ruta",
            "unit": "hojas",
            "time_based": False,
        },
    }
    return configs.get(op, configs["PICKING"])


def _olas_shift_label(turno: str, fecha: str) -> str:
    labels = {"manana": "TM", "tarde": "TT", "noche": "TN"}
    return f"{fecha} {labels.get(turno, turno.upper())}"


def _build_olas_raw_operation_summary(
    *,
    division: str,
    start: datetime,
    end: datetime,
    raw_rows: list[dict[str, Any]],
    plan: dict[str, dict[str, float]],
    config: dict[str, Any],
) -> dict[str, Any]:
    sector = normalize_sector(division)
    almacen_target = DAILY_RAW_SECTOR_TO_ALMACEN.get(sector, "SECOS")
    windows = _shift_windows_between(start, end)
    scoped_rows: list[dict[str, Any]] = []
    for row in raw_rows:
        ts = _parse_dt(row.get("FCREAREG") or row.get("fecha"))
        legajo = _norm_legajo(row.get("COPECREA") or row.get("LEGAJO") or row.get("legajo"))
        if not ts or not legajo or ts < start or ts >= end:
            continue
        scoped_rows.append({**row, "_ts": ts, "_legajo": legajo})
    scoped_rows.sort(key=lambda item: (item["_legajo"], item["_ts"], int(item.get("_row_index") or 0)))

    totals = {"cantidad": 0.0, "pallets": set(), "legajos": set(), "horas": 0.0, "lineas": 0}
    ranking: dict[str, dict[str, Any]] = {}
    turnos: dict[str, dict[str, Any]] = {}
    for turno, fecha, w_start, _w_end in windows:
        key = _olas_shift_label(turno, fecha)
        turnos[key] = {"turno": key, "fecha": fecha, "cantidad": 0.0, "pallets": set(), "legajos": set(), "horas": 0.0, "_start": w_start}

    previous_by_legajo: dict[str, datetime] = {}
    operations = config["operation_filter"]
    for row in scoped_rows:
        legajo = row["_legajo"]
        ts = row["_ts"]
        prev = previous_by_legajo.get(legajo, start)
        previous_by_legajo[legajo] = ts
        operacion = str(row.get("CDESCRIP") or row.get("OPERACION") or "").strip().upper()
        if operacion not in operations:
            continue
        cantidad = _to_float(row.get("QCANTIDA") if row.get("QCANTIDA") is not None else row.get("cantidad"))
        if config.get("positive_quantity") and cantidad <= 0:
            continue
        almacen = _daily_productividad_raw_division(row) if config.get("sector_source") in {"picking", "productividad_division"} else str(row.get("ALMACEN") or "").strip().upper()
        if almacen != almacen_target:
            continue
        seconds = max((ts - prev).total_seconds(), 0.0)
        horas = seconds / 3600
        pallet = str(row.get("CNUPALET") or "").strip()
        funcion = str(row.get("DESC_FUNCION") or row.get("desc_funcion") or "").strip()
        quantity_value = cantidad if config["quantity_mode"] == "sum_cantidad" else (1.0 if pallet and pallet != "0" else 0.0)

        totals["cantidad"] += quantity_value
        totals["horas"] += horas
        totals["lineas"] += 1
        totals["legajos"].add(legajo)
        if pallet and pallet != "0":
            totals["pallets"].add(pallet)

        bucket = ranking.setdefault(legajo, {"legajo": legajo, "funcion": funcion, "cantidad": 0.0, "pallets": set(), "horas": 0.0, "lineas": 0, "primer_mov": None, "ultimo_mov": None, "productividad": 0.0})
        if funcion and not bucket.get("funcion"):
            bucket["funcion"] = funcion
        bucket["cantidad"] += quantity_value
        bucket["horas"] += horas
        bucket["lineas"] += 1
        if pallet and pallet != "0":
            bucket["pallets"].add(pallet)
        if bucket["primer_mov"] is None or ts < bucket["primer_mov"]:
            bucket["primer_mov"] = ts
        if bucket["ultimo_mov"] is None or ts > bucket["ultimo_mov"]:
            bucket["ultimo_mov"] = ts

        for item in turnos.values():
            if item["_start"] <= ts < item["_start"] + timedelta(hours=8):
                item["cantidad"] += quantity_value
                item["horas"] += horas
                item["legajos"].add(legajo)
                if pallet and pallet != "0":
                    item["pallets"].add(pallet)
                break

    ranking_rows = []
    jornada = _daily_productividad_jornada(sector)
    for item in ranking.values():
        horas = float(item["horas"] or 0)
        item["productividad"] = round(item["cantidad"] / horas, 2) if horas else 0.0
        item["productividad_dia"] = round(item["productividad"] * jornada, 2)
        item["pallets"] = len(item["pallets"])
        item["cantidad"] = round(item["cantidad"], 2)
        item["horas"] = round(horas, 4)
        item["primer_mov"] = item["primer_mov"].strftime("%Y-%m-%d %H:%M:%S") if item["primer_mov"] else ""
        item["ultimo_mov"] = item["ultimo_mov"].strftime("%Y-%m-%d %H:%M:%S") if item["ultimo_mov"] else ""
        ranking_rows.append(item)
    ranking_rows.sort(key=lambda item: item["productividad"], reverse=True)

    turno_rows = []
    for item in sorted(turnos.values(), key=lambda row: row["_start"]):
        horas = float(item["horas"] or 0)
        prod_hora = round(item["cantidad"] / horas, 2) if horas else 0.0
        turno_rows.append({"turno": item["turno"], "fecha": item["fecha"], "cantidad": round(item["cantidad"], 2), "pallets": len(item["pallets"]), "legajos": len(item["legajos"]), "horas": round(horas, 4), "productividad_hora": prod_hora, "productividad_dia": round(prod_hora * jornada, 2)})

    plan_sector = plan.get(sector, {"cantidad": 0.0, "pallets": 0.0})
    legajos = len(totals["legajos"])
    horas_total = float(totals["horas"] or 0)
    cantidad_total = float(totals["cantidad"] or 0)
    productividad_hora = round(cantidad_total / horas_total, 2) if horas_total else 0.0
    return {
        "division": sector,
        "almacen": almacen_target,
        "operation": config["operation"],
        "operation_label": config["label"],
        "quantity_label": config["quantity_label"],
        "unit": config["unit"],
        "indicadores": [{"key": "cantidad", "label": f"{config['quantity_label']} acumulados vs plan", **_shift_metric(cantidad_total, plan_sector["cantidad"]), "unidad": config["unit"]}],
        "kpis": {
            "cantidad": round(cantidad_total, 2),
            "pallets": len(totals["pallets"]),
            "legajos": legajos,
            "horas": round(horas_total, 2),
            "ocupacion_pct": round((horas_total / (legajos * jornada) * 100), 2) if legajos else 0.0,
            "productividad_hora": productividad_hora,
            "cantidad_por_pallet": round(cantidad_total / len(totals["pallets"]), 2) if totals["pallets"] else 0.0,
            "productividad_dia": round(productividad_hora * jornada, 2),
            "lineas": totals["lineas"],
        },
        "turnos": turno_rows,
        "ranking": ranking_rows[:80],
    }


def _build_olas_despacho_summary(
    *,
    division: str,
    start: datetime,
    end: datetime,
    despacho_rows: list[dict[str, Any]],
    plan: dict[str, dict[str, float]],
) -> dict[str, Any]:
    sector = normalize_sector(division)
    almacen_target = DAILY_RAW_SECTOR_TO_ALMACEN.get(sector, "SECOS")
    windows = _shift_windows_between(start, end)
    totals = {"cantidad": 0.0, "legajos": set(), "hojas": set(), "lineas": 0}
    ranking: dict[str, dict[str, Any]] = {}
    turnos: dict[str, dict[str, Any]] = {}
    for turno, fecha, w_start, _w_end in windows:
        key = _olas_shift_label(turno, fecha)
        turnos[key] = {"turno": key, "fecha": fecha, "cantidad": 0.0, "legajos": set(), "_start": w_start}

    for row in despacho_rows:
        ts = _parse_dt(row.get("FECIERRE") or row.get("FECHA_CIERRE"))
        almacen = str(row.get("ALMACEN") or "").strip().upper()
        if not ts or ts < start or ts >= end or almacen != almacen_target:
            continue
        hoja = str(row.get("HOJARUTA") or row.get("CNUVIAJE") or row.get("VIAJE") or "").strip()
        cargador = _norm_legajo(row.get("CARGADOR"))
        funcion = str(row.get("DESC_FUNCION") or "").strip()
        if not hoja:
            continue
        totals["hojas"].add(hoja)
        totals["lineas"] += 1
        if cargador and _despacho_cargador_elegible(row):
            totals["legajos"].add(cargador)
            bucket = ranking.setdefault(cargador, {"legajo": cargador, "funcion": funcion, "cantidad": 0.0, "pallets": 0, "horas": 0.0, "lineas": 0, "primer_mov": None, "ultimo_mov": None, "productividad": 0.0, "productividad_dia": 0.0})
            bucket["cantidad"] += 1
            bucket["lineas"] += 1
            if bucket["primer_mov"] is None or ts < bucket["primer_mov"]:
                bucket["primer_mov"] = ts
            if bucket["ultimo_mov"] is None or ts > bucket["ultimo_mov"]:
                bucket["ultimo_mov"] = ts
        for item in turnos.values():
            if item["_start"] <= ts < item["_start"] + timedelta(hours=8):
                item["cantidad"] += 1
                if cargador and _despacho_cargador_elegible(row):
                    item["legajos"].add(cargador)
                break

    cantidad_total = float(len(totals["hojas"]))
    legajos = len(totals["legajos"])
    for item in ranking.values():
        item["productividad"] = round(item["cantidad"], 2)
        item["productividad_dia"] = round(item["cantidad"], 2)
        item["primer_mov"] = item["primer_mov"].strftime("%Y-%m-%d %H:%M:%S") if item["primer_mov"] else ""
        item["ultimo_mov"] = item["ultimo_mov"].strftime("%Y-%m-%d %H:%M:%S") if item["ultimo_mov"] else ""
    ranking_rows = sorted(ranking.values(), key=lambda row: row["cantidad"], reverse=True)[:80]
    turno_rows = [{"turno": item["turno"], "fecha": item["fecha"], "cantidad": round(item["cantidad"], 2), "pallets": 0, "legajos": len(item["legajos"]), "horas": 0.0, "productividad_hora": round(item["cantidad"], 2), "productividad_dia": round(item["cantidad"], 2)} for item in sorted(turnos.values(), key=lambda row: row["_start"])]
    plan_sector = plan.get(sector, {"cantidad": 0.0, "pallets": 0.0})
    return {
        "division": sector,
        "almacen": almacen_target,
        "operation": "DESPACHO",
        "operation_label": "Despacho",
        "quantity_label": "Hojas de ruta",
        "unit": "hojas",
        "indicadores": [{"key": "cantidad", "label": "Hojas de ruta acumuladas vs plan", **_shift_metric(cantidad_total, plan_sector["cantidad"]), "unidad": "hojas"}],
        "kpis": {"cantidad": cantidad_total, "pallets": 0, "legajos": legajos, "horas": 0.0, "ocupacion_pct": 0.0, "productividad_hora": round(cantidad_total / legajos, 2) if legajos else 0.0, "cantidad_por_pallet": 0.0, "productividad_dia": round(cantidad_total / legajos, 2) if legajos else 0.0, "lineas": totals["lineas"]},
        "turnos": turno_rows,
        "ranking": ranking_rows,
    }


def _build_olas_picking_summary(
    *,
    division: str,
    start: datetime,
    end: datetime,
    raw_rows: list[dict[str, Any]],
    plan: dict[str, dict[str, float]],
) -> dict[str, Any]:
    sector = normalize_sector(division)
    almacen_target = DAILY_RAW_SECTOR_TO_ALMACEN.get(sector, "SECOS")
    windows = _shift_windows_between(start, end)
    scoped_rows: list[dict[str, Any]] = []
    for row in raw_rows:
        ts = _parse_dt(row.get("FCREAREG") or row.get("fecha"))
        legajo = _norm_legajo(row.get("COPECREA") or row.get("LEGAJO") or row.get("legajo"))
        if not ts or not legajo or ts < start or ts >= end:
            continue
        scoped_rows.append({**row, "_ts": ts, "_legajo": legajo})
    scoped_rows.sort(key=lambda item: (item["_legajo"], item["_ts"], int(item.get("_row_index") or 0)))

    totals = {"bultos": 0.0, "pallets": set(), "legajos": set(), "horas": 0.0, "lineas": 0}
    ranking: dict[str, dict[str, Any]] = {}
    turnos: dict[str, dict[str, Any]] = {}
    for turno, fecha, w_start, _w_end in windows:
        key = _olas_shift_label(turno, fecha)
        turnos[key] = {
            "turno": key,
            "fecha": fecha,
            "bultos": 0.0,
            "pallets": set(),
            "legajos": set(),
            "horas": 0.0,
            "productividad_hora": 0.0,
            "productividad_dia": 0.0,
            "_start": w_start,
        }

    previous_by_legajo: dict[str, datetime] = {}
    for row in scoped_rows:
        legajo = row["_legajo"]
        ts = row["_ts"]
        prev = previous_by_legajo.get(legajo, start)
        previous_by_legajo[legajo] = ts
        operacion = str(row.get("CDESCRIP") or row.get("OPERACION") or "").strip().upper()
        if operacion != "PICKING" or _daily_productividad_raw_division(row) != almacen_target:
            continue
        seconds = max((ts - prev).total_seconds(), 0.0)
        horas = seconds / 3600
        cantidad = _to_float(row.get("QCANTIDA") if row.get("QCANTIDA") is not None else row.get("cantidad"))
        pallet = str(row.get("CNUPALET") or "").strip()
        funcion = str(row.get("DESC_FUNCION") or row.get("desc_funcion") or "").strip()

        totals["bultos"] += cantidad
        totals["horas"] += horas
        totals["lineas"] += 1
        totals["legajos"].add(legajo)
        if pallet and pallet != "0":
            totals["pallets"].add(pallet)

        bucket = ranking.setdefault(
            legajo,
            {
                "legajo": legajo,
                "funcion": funcion,
                "bultos": 0.0,
                "pallets": set(),
                "horas": 0.0,
                "lineas": 0,
                "primer_picking": None,
                "ultimo_picking": None,
                "productividad": 0.0,
            },
        )
        if funcion and not bucket.get("funcion"):
            bucket["funcion"] = funcion
        bucket["bultos"] += cantidad
        bucket["horas"] += horas
        bucket["lineas"] += 1
        if pallet and pallet != "0":
            bucket["pallets"].add(pallet)
        if bucket["primer_picking"] is None or ts < bucket["primer_picking"]:
            bucket["primer_picking"] = ts
        if bucket["ultimo_picking"] is None or ts > bucket["ultimo_picking"]:
            bucket["ultimo_picking"] = ts

        for turno_key, item in turnos.items():
            if item["_start"] <= ts < item["_start"] + timedelta(hours=8):
                item["bultos"] += cantidad
                item["horas"] += horas
                item["legajos"].add(legajo)
                if pallet and pallet != "0":
                    item["pallets"].add(pallet)
                break

    ranking_rows = []
    jornada = _daily_productividad_jornada(sector)
    for item in ranking.values():
        horas = float(item["horas"] or 0)
        item["productividad"] = round(item["bultos"] / horas, 2) if horas else 0.0
        item["productividad_dia"] = round(item["productividad"] * jornada, 2)
        item["pallets"] = len(item["pallets"])
        item["bultos"] = round(item["bultos"], 2)
        item["horas"] = round(horas, 4)
        item["primer_picking"] = item["primer_picking"].strftime("%Y-%m-%d %H:%M:%S") if item["primer_picking"] else ""
        item["ultimo_picking"] = item["ultimo_picking"].strftime("%Y-%m-%d %H:%M:%S") if item["ultimo_picking"] else ""
        ranking_rows.append(item)
    ranking_rows.sort(key=lambda item: item["productividad"], reverse=True)

    turno_rows = []
    for item in sorted(turnos.values(), key=lambda row: row["_start"]):
        horas = float(item["horas"] or 0)
        prod_hora = round(item["bultos"] / horas, 2) if horas else 0.0
        turno_rows.append(
            {
                "turno": item["turno"],
                "fecha": item["fecha"],
                "bultos": round(item["bultos"], 2),
                "pallets": len(item["pallets"]),
                "legajos": len(item["legajos"]),
                "horas": round(horas, 4),
                "productividad_hora": prod_hora,
                "productividad_dia": round(prod_hora * jornada, 2),
            }
        )

    plan_sector = plan.get(sector, {"bultos": 0.0, "pallets": 0.0})
    legajos = len(totals["legajos"])
    horas_total = float(totals["horas"] or 0)
    bultos_total = float(totals["bultos"] or 0)
    pallets_total = len(totals["pallets"])
    productividad_hora = round(bultos_total / horas_total, 2) if horas_total else 0.0
    ocupacion_pct = round((horas_total / (legajos * jornada) * 100), 2) if legajos else 0.0
    return {
        "division": sector,
        "almacen": almacen_target,
        "indicadores": [
            {"key": "bultos", "label": "Bultos acumulados vs plan", **_shift_metric(bultos_total, plan_sector["bultos"]), "unidad": "bultos"},
            {"key": "pallets", "label": "Pallets acumulados vs plan", **_shift_metric(float(pallets_total), plan_sector["pallets"]), "unidad": "pallets"},
        ],
        "kpis": {
            "bultos": round(bultos_total, 2),
            "pallets": pallets_total,
            "legajos": legajos,
            "horas": round(horas_total, 2),
            "ocupacion_pct": ocupacion_pct,
            "bultos_por_hora": productividad_hora,
            "bultos_por_pallet": round(bultos_total / pallets_total, 2) if pallets_total else 0.0,
            "productividad_dia": round(productividad_hora * jornada, 2),
            "lineas": totals["lineas"],
        },
        "turnos": turno_rows,
        "ranking": ranking_rows[:80],
    }


@router.get("/cambio-turno/resumen")
async def cambio_turno_resumen(
    request: Request,
    division: str = Query("Secos"),
    turno: str = Query("TM", pattern="^(TM|TT|TN|tm|tt|tn)$"),
):
    await _require_request_auth(request)
    window = _shift_turn_window(turno)
    start = _parse_dt(window["fecha_inicio"])
    end = _parse_dt(window["fecha_fin"])
    if not start or not end:
        raise HTTPException(status_code=400, detail="Ventana de turno invalida.")
    fecha_desde = _fmt_daily_oracle_dt(start)
    fecha_hasta = _fmt_daily_oracle_dt(end)
    try:
        productividad_rows, despacho_rows, plan_rows = await asyncio.gather(
            asyncio.to_thread(query_productive_db_daily_productividad_raw, fecha_desde, fecha_hasta),
            asyncio.to_thread(query_productive_db_daily_despacho_raw, fecha_desde, fecha_hasta),
            asyncio.to_thread(query_productive_db_daily_planificacion, fecha_desde, fecha_hasta),
        )
    except Exception as exc:
        logger.exception("[cambio-turno] No se pudo calcular resumen")
        raise HTTPException(status_code=500, detail=f"No se pudo consultar Oracle para Cambio Turno: {exc}") from exc
    division_names = ["Secos", "Refrigerados", "OC", "Congelados", "Noa"]
    summaries = {
        item: _shift_build_summary(
            division=item,
            start=start,
            end=end,
            productividad_rows=productividad_rows,
            despacho_rows=despacho_rows,
            plan_rows=plan_rows,
        )
        for item in division_names
    }
    selected = summaries.get(normalize_sector(division), summaries["Secos"])
    return {
        "window": window,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "source": "oracle_reuse_daily_queries",
        "divisiones": summaries,
        **selected,
    }


@router.get("/olas/resumen")
@router.get("/olas/picking/resumen")
async def olas_picking_resumen(
    request: Request,
    fecha: str | None = Query(None, description="Fecha de carga Daily YYYY-MM-DD"),
    operation: str = Query("PICKING", pattern="^(PICKING|CLARK|RECEPCION|DESPACHO|picking|clark|recepcion|despacho)$"),
):
    await _require_request_auth(request)
    if fecha:
        try:
            daily = _daily_window_for_fecha_carga(datetime.strptime(fecha, "%Y-%m-%d"))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Fecha invalida. Usa YYYY-MM-DD.") from exc
    else:
        daily = scheduled_daily_window()
    if not daily.get("can_run", True):
        raise HTTPException(status_code=400, detail=daily.get("reason") or "Daily no disponible.")
    start = _parse_dt(daily.get("fecha_inicio"))
    end = _parse_dt(daily.get("fecha_fin"))
    if not start or not end:
        raise HTTPException(status_code=400, detail="Ventana Daily invalida.")

    config = _olas_operation_config(operation)
    op = config["operation"]
    raw_process = "DESPACHO_RAW" if op == "DESPACHO" else "PRODUCTIVIDAD_RAW"
    raw_version = DAILY_AUTO_DESPACHO_RAW_QUERY_VERSION if op == "DESPACHO" else DAILY_AUTO_PRODUCTIVIDAD_RAW_QUERY_VERSION
    raw_run = await get_latest_run(daily["daily_key"], raw_process)
    if not _is_current_process_run(raw_run, raw_version):
        return {
            "daily": daily,
            "source": "cache_local",
            "status": "missing_cache",
            "operation": op,
            "message": f"No hay cache cruda vigente de {config['label']} para esta Daily.",
            "divisiones": {},
        }

    cached_rows = await get_cached_results(daily["daily_key"])
    plan = _olas_plan_from_cache(cached_rows, op)
    division_names = ["Secos", "Refrigerados", "OC", "Congelados", "Noa"]
    if op == "DESPACHO":
        despacho_rows = await get_despacho_raw_cache_rows(daily["daily_key"])
        summaries = {
            item: _build_olas_despacho_summary(
                division=item,
                start=start,
                end=end,
                despacho_rows=despacho_rows,
                plan=plan,
            )
            for item in division_names
        }
    else:
        raw_rows = await get_productividad_raw_cache_rows(daily["daily_key"])
        summaries = {
            item: _build_olas_raw_operation_summary(
                division=item,
                start=start,
                end=end,
                raw_rows=raw_rows,
                plan=plan,
                config=config,
            )
            for item in division_names
        }
    return {
        "daily": daily,
        "fecha_desde": _fmt_daily_oracle_dt(start),
        "fecha_hasta": _fmt_daily_oracle_dt(end),
        "source": "cache_local_despacho_raw" if op == "DESPACHO" else "cache_local_productividad_raw",
        "status": "success",
        "operation": op,
        "operation_label": config["label"],
        "run": raw_run,
        "divisiones": summaries,
    }


async def run_daily_auto_avance_precache(
    *,
    force: bool = False,
    trigger: str = "scheduler",
    usuario: str = "",
    daily_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    daily = daily_override or scheduled_daily_window()
    if not daily.get("can_run"):
        return {"process": "AVANCE", "status": "skipped", "reason": daily.get("reason") or "Daily no habilitada.", "daily": daily}
    latest_run = await get_latest_run(daily["daily_key"], "AVANCE")
    if not force and _is_current_process_run(latest_run, DAILY_AUTO_AVANCE_QUERY_VERSION):
        return {"process": "AVANCE", "status": "skipped", "reason": "Cache Avance ya disponible.", "daily": daily}

    started_at = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
    started = time.perf_counter()
    logger.info("[daily-auto] Iniciando calculo Avance 06:00-07:30 desde cache %s", daily["daily_key"])
    try:
        raw_run = await get_latest_run(daily["daily_key"], "PRODUCTIVIDAD_RAW")
        if not _is_current_process_run(raw_run, DAILY_AUTO_PRODUCTIVIDAD_RAW_QUERY_VERSION):
            raw_result = await run_daily_auto_productividad_raw_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily)
            if raw_result.get("status") not in {"success", "skipped"}:
                raise RuntimeError(raw_result.get("error") or raw_result.get("reason") or "No se pudo generar cache crudo de productividad.")
        despacho_raw_run = await get_latest_run(daily["daily_key"], "DESPACHO_RAW")
        if not _is_current_process_run(despacho_raw_run, DAILY_AUTO_DESPACHO_RAW_QUERY_VERSION):
            despacho_raw_result = await run_daily_auto_despacho_raw_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily)
            if despacho_raw_result.get("status") not in {"success", "skipped"}:
                raise RuntimeError(despacho_raw_result.get("error") or despacho_raw_result.get("reason") or "No se pudo generar cache crudo de Despacho.")

        avance_inicio, avance_fin = _avance_window(daily)
        fecha_desde = _fmt_daily_oracle_dt(avance_inicio)
        fecha_hasta = _fmt_daily_oracle_dt(avance_fin)
        productividad_rows = await get_productividad_raw_cache_rows(daily["daily_key"])
        despacho_rows = await get_despacho_raw_cache_rows(daily["daily_key"])
        real_rows = _build_avance_real_rows_from_raw(daily, productividad_rows, despacho_rows)
        plan_rows_oracle = await asyncio.to_thread(query_productive_db_daily_planificacion, fecha_desde, fecha_hasta)
        plan_rows = _build_avance_plan_rows(plan_rows_oracle)
        rows = real_rows + plan_rows
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        saved = await save_avance_summary_cache(
            daily,
            rows,
            avance_inicio=avance_inicio.isoformat(sep=" "),
            avance_fin=avance_fin.isoformat(sep=" "),
            started_at=started_at,
            duration_ms=elapsed_ms,
            timings={
                "avance_local_ms": elapsed_ms,
                "query_version": DAILY_AUTO_AVANCE_QUERY_VERSION,
                "productividad_raw_rows": len(productividad_rows),
                "despacho_raw_rows": len(despacho_rows),
                "plan_rows": len(plan_rows_oracle),
            },
            trigger=trigger,
            usuario=usuario,
        )
        logger.info("[daily-auto] Calculo Avance OK: %s filas en %sms", len(rows), elapsed_ms)
        return {"process": "AVANCE", "status": "success", "daily": daily, "rows": len(rows), "duration_ms": elapsed_ms, "saved": saved}
    except Exception as exc:
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        await mark_run_error(daily, "AVANCE", str(exc), started_at, trigger=trigger, usuario=usuario)
        logger.exception("[daily-auto] Calculo Avance fallo tras %sms", elapsed_ms)
        return {"process": "AVANCE", "status": "error", "daily": daily, "duration_ms": elapsed_ms, "error": str(exc)}


async def run_daily_auto_clark_precache(
    *,
    force: bool = False,
    trigger: str = "scheduler",
    usuario: str = "",
    daily_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    daily = daily_override or scheduled_daily_window()
    if not daily.get("can_run"):
        return {"process": "CLARK", "status": "skipped", "reason": daily.get("reason") or "Daily no habilitada.", "daily": daily}
    latest_run = await get_latest_run(daily["daily_key"], "CLARK")
    if not force and _is_current_process_run(latest_run, DAILY_AUTO_CLARK_QUERY_VERSION):
        return {"process": "CLARK", "status": "skipped", "reason": "Cache Clark ya disponible.", "daily": daily}

    started_at = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
    started = time.perf_counter()
    logger.info("[daily-auto] Iniciando calculo Clark desde cache crudo %s", daily["daily_key"])
    try:
        raw_run = await get_latest_run(daily["daily_key"], "PRODUCTIVIDAD_RAW")
        raw_result = None
        if not _is_current_process_run(raw_run, DAILY_AUTO_PRODUCTIVIDAD_RAW_QUERY_VERSION):
            raw_result = await run_daily_auto_productividad_raw_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily)
            if raw_result.get("status") not in {"success", "skipped"}:
                raise RuntimeError(raw_result.get("error") or raw_result.get("reason") or "No se pudo generar cache crudo de productividad.")
        raw_rows = await get_productividad_raw_cache_rows(daily["daily_key"])
        rows = _build_clark_rows_from_productividad_raw(daily, raw_rows)
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        saved = await save_clark_raw_summary_cache(
            daily,
            rows,
            started_at=started_at,
            duration_ms=elapsed_ms,
            timings={
                "clark_raw_local_ms": elapsed_ms,
                "query_version": DAILY_AUTO_CLARK_QUERY_VERSION,
                "source_process": "PRODUCTIVIDAD_RAW",
                "raw_rows": len(raw_rows),
                "raw_precache_status": raw_result.get("status") if raw_result else "hit",
            },
            trigger=trigger,
            usuario=usuario,
        )
        logger.info("[daily-auto] Calculo Clark OK desde crudo: %s almacenes en %sms", len(rows), elapsed_ms)
        return {"process": "CLARK", "status": "success", "daily": daily, "rows": len(rows), "raw_rows": len(raw_rows), "duration_ms": elapsed_ms, "saved": saved}
    except Exception as exc:
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        await mark_run_error(daily, "CLARK", str(exc), started_at, trigger=trigger, usuario=usuario)
        logger.exception("[daily-auto] Calculo Clark desde crudo fallo tras %sms", elapsed_ms)
        return {"process": "CLARK", "status": "error", "daily": daily, "duration_ms": elapsed_ms, "error": str(exc)}


async def run_daily_auto_picking_precache(
    *,
    force: bool = False,
    trigger: str = "scheduler",
    usuario: str = "",
    daily_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    daily = daily_override or scheduled_daily_window()
    if not daily.get("can_run"):
        return {"process": "PICKING", "status": "skipped", "reason": daily.get("reason") or "Daily no habilitada.", "daily": daily}
    latest_run = await get_latest_run(daily["daily_key"], "PICKING")
    if not force and _is_current_process_run(latest_run, DAILY_AUTO_PICKING_QUERY_VERSION):
        return {"process": "PICKING", "status": "skipped", "reason": "Cache Picking ya disponible.", "daily": daily}

    started_at = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
    started = time.perf_counter()
    logger.info("[daily-auto] Iniciando calculo Picking desde cache crudo %s", daily["daily_key"])
    try:
        raw_run = await get_latest_run(daily["daily_key"], "PRODUCTIVIDAD_RAW")
        raw_result = None
        if not _is_current_process_run(raw_run, DAILY_AUTO_PRODUCTIVIDAD_RAW_QUERY_VERSION):
            raw_result = await run_daily_auto_productividad_raw_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily)
            if raw_result.get("status") not in {"success", "skipped"}:
                raise RuntimeError(raw_result.get("error") or raw_result.get("reason") or "No se pudo generar cache crudo de productividad.")
        raw_rows = await get_productividad_raw_cache_rows(daily["daily_key"])
        rows = _build_picking_rows_from_productividad_raw(daily, raw_rows)
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        saved = await save_picking_summary_cache(
            daily,
            rows,
            started_at=started_at,
            duration_ms=elapsed_ms,
            timings={
                "picking_raw_local_ms": elapsed_ms,
                "query_version": DAILY_AUTO_PICKING_QUERY_VERSION,
                "source_process": "PRODUCTIVIDAD_RAW",
                "raw_rows": len(raw_rows),
                "raw_precache_status": raw_result.get("status") if raw_result else "hit",
            },
            trigger=trigger,
            usuario=usuario,
        )
        logger.info("[daily-auto] Calculo Picking OK desde crudo: %s movimientos picking en %sms", len(rows), elapsed_ms)
        return {"process": "PICKING", "status": "success", "daily": daily, "rows": len(rows), "raw_rows": len(raw_rows), "duration_ms": elapsed_ms, "saved": saved}
    except Exception as exc:
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        await mark_run_error(daily, "PICKING", str(exc), started_at, trigger=trigger, usuario=usuario)
        logger.exception("[daily-auto] Calculo Picking desde crudo fallo tras %sms", elapsed_ms)
        return {"process": "PICKING", "status": "error", "daily": daily, "duration_ms": elapsed_ms, "error": str(exc)}


async def run_daily_auto_recepcion_precache(
    *,
    force: bool = False,
    trigger: str = "scheduler",
    usuario: str = "",
    daily_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    daily = daily_override or scheduled_daily_window()
    if not daily.get("can_run"):
        return {"process": "RECEPCION", "status": "skipped", "reason": daily.get("reason") or "Daily no habilitada.", "daily": daily}
    latest_run = await get_latest_run(daily["daily_key"], "RECEPCION")
    if not force and _is_current_process_run(latest_run, DAILY_AUTO_RECEPCION_QUERY_VERSION):
        return {"process": "RECEPCION", "status": "skipped", "reason": "Cache Recepcion ya disponible.", "daily": daily}

    started_at = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
    started = time.perf_counter()
    logger.info("[daily-auto] Iniciando calculo Recepcion desde cache crudo %s", daily["daily_key"])
    try:
        raw_run = await get_latest_run(daily["daily_key"], "PRODUCTIVIDAD_RAW")
        raw_result = None
        if not _is_current_process_run(raw_run, DAILY_AUTO_PRODUCTIVIDAD_RAW_QUERY_VERSION):
            raw_result = await run_daily_auto_productividad_raw_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily)
            if raw_result.get("status") not in {"success", "skipped"}:
                raise RuntimeError(raw_result.get("error") or raw_result.get("reason") or "No se pudo generar cache crudo de productividad.")
        raw_rows = await get_productividad_raw_cache_rows(daily["daily_key"])
        rows = _build_recepcion_rows_from_productividad_raw(daily, raw_rows)
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        saved = await save_recepcion_summary_cache(
            daily,
            rows,
            started_at=started_at,
            duration_ms=elapsed_ms,
            timings={
                "recepcion_raw_local_ms": elapsed_ms,
                "query_version": DAILY_AUTO_RECEPCION_QUERY_VERSION,
                "source_process": "PRODUCTIVIDAD_RAW",
                "raw_rows": len(raw_rows),
                "raw_precache_status": raw_result.get("status") if raw_result else "hit",
            },
            trigger=trigger,
            usuario=usuario,
        )
        logger.info("[daily-auto] Calculo Recepcion OK desde crudo: %s almacenes en %sms", len(rows), elapsed_ms)
        return {"process": "RECEPCION", "status": "success", "daily": daily, "rows": len(rows), "raw_rows": len(raw_rows), "duration_ms": elapsed_ms, "saved": saved}
    except Exception as exc:
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        await mark_run_error(daily, "RECEPCION", str(exc), started_at, trigger=trigger, usuario=usuario)
        logger.exception("[daily-auto] Calculo Recepcion desde crudo fallo tras %sms", elapsed_ms)
        return {"process": "RECEPCION", "status": "error", "daily": daily, "duration_ms": elapsed_ms, "error": str(exc)}


async def run_daily_auto_despacho_precache(
    *,
    force: bool = False,
    trigger: str = "scheduler",
    usuario: str = "",
    daily_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    daily = daily_override or scheduled_daily_window()
    if not daily.get("can_run"):
        return {"process": "DESPACHO", "status": "skipped", "reason": daily.get("reason") or "Daily no habilitada.", "daily": daily}
    latest_run = await get_latest_run(daily["daily_key"], "DESPACHO")
    if not force and _is_current_process_run(latest_run, DAILY_AUTO_DESPACHO_QUERY_VERSION):
        return {"process": "DESPACHO", "status": "skipped", "reason": "Cache Despacho ya disponible.", "daily": daily}

    started_at = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
    started = time.perf_counter()
    logger.info("[daily-auto] Iniciando calculo Despacho desde cache crudo %s", daily["daily_key"])
    try:
        raw_run = await get_latest_run(daily["daily_key"], "DESPACHO_RAW")
        raw_result = None
        if not _is_current_process_run(raw_run, DAILY_AUTO_DESPACHO_RAW_QUERY_VERSION):
            raw_result = await run_daily_auto_despacho_raw_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily)
            if raw_result.get("status") not in {"success", "skipped"}:
                raise RuntimeError(raw_result.get("error") or raw_result.get("reason") or "No se pudo generar cache crudo de Despacho.")
        raw_rows = await get_despacho_raw_cache_rows(daily["daily_key"])
        rows = _build_despacho_rows_from_raw(daily, raw_rows)
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        saved = await save_despacho_summary_cache(
            daily,
            rows,
            started_at=started_at,
            duration_ms=elapsed_ms,
            timings={
                "despacho_raw_local_ms": elapsed_ms,
                "query_version": DAILY_AUTO_DESPACHO_QUERY_VERSION,
                "source_process": "DESPACHO_RAW",
                "raw_rows": len(raw_rows),
                "raw_precache_status": raw_result.get("status") if raw_result else "hit",
            },
            trigger=trigger,
            usuario=usuario,
        )
        logger.info("[daily-auto] Calculo Despacho OK desde crudo: %s almacenes en %sms", len(rows), elapsed_ms)
        return {"process": "DESPACHO", "status": "success", "daily": daily, "rows": len(rows), "raw_rows": len(raw_rows), "duration_ms": elapsed_ms, "saved": saved}
    except Exception as exc:
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        await mark_run_error(daily, "DESPACHO", str(exc), started_at, trigger=trigger, usuario=usuario)
        logger.exception("[daily-auto] Calculo Despacho desde crudo fallo tras %sms", elapsed_ms)
        return {"process": "DESPACHO", "status": "error", "daily": daily, "duration_ms": elapsed_ms, "error": str(exc)}


async def run_daily_auto_planificacion_precache(
    *,
    force: bool = False,
    trigger: str = "scheduler",
    usuario: str = "",
    daily_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    daily = daily_override or scheduled_daily_window()
    if not daily.get("can_run"):
        return {"process": "PLANIFICACION", "status": "skipped", "reason": daily.get("reason") or "Daily no habilitada.", "daily": daily}
    latest_run = await get_latest_run(daily["daily_key"], "PLANIFICACION")
    if not force and _is_current_process_run(latest_run, DAILY_AUTO_PLANIFICACION_QUERY_VERSION):
        return {"process": "PLANIFICACION", "status": "skipped", "reason": "Cache Planificacion ya disponible.", "daily": daily}

    fecha_desde = _fmt_daily_oracle_dt(daily.get("fecha_inicio"))
    fecha_hasta = _fmt_daily_oracle_dt(daily.get("fecha_fin"))
    started_at = datetime.now(LOCAL_TZ).isoformat(timespec="seconds")
    started = time.perf_counter()
    logger.info("[daily-auto] Iniciando precarga Planificacion %s %s..%s", daily["daily_key"], fecha_desde, fecha_hasta)
    try:
        rows = await asyncio.to_thread(query_productive_db_daily_planificacion, fecha_desde, fecha_hasta)
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        saved = await save_planificacion_summary_cache(
            daily,
            rows,
            started_at=started_at,
            duration_ms=elapsed_ms,
            timings={"planificacion_summary_oracle_ms": elapsed_ms, "query_version": DAILY_AUTO_PLANIFICACION_QUERY_VERSION},
            trigger=trigger,
            usuario=usuario,
        )
        logger.info("[daily-auto] Precarga Planificacion OK: %s filas en %sms", len(rows), elapsed_ms)
        return {"process": "PLANIFICACION", "status": "success", "daily": daily, "rows": len(rows), "duration_ms": elapsed_ms, "saved": saved}
    except Exception as exc:
        elapsed_ms = round((time.perf_counter() - started) * 1000, 1)
        await mark_run_error(daily, "PLANIFICACION", str(exc), started_at, trigger=trigger, usuario=usuario)
        logger.exception("[daily-auto] Precarga Planificacion fallo tras %sms", elapsed_ms)
        return {"process": "PLANIFICACION", "status": "error", "daily": daily, "duration_ms": elapsed_ms, "error": str(exc)}


async def run_daily_auto_precache_pending(
    *,
    force: bool = False,
    trigger: str = "scheduler",
    usuario: str = "",
    daily_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    daily = daily_override or scheduled_daily_window()
    blocked_processes: list[str] = []
    if not daily.get("can_run"):
        return {
            "status": "skipped",
            "daily": daily,
            "trigger": trigger,
            "processes": [],
            "blocked_processes": blocked_processes,
            "reason": daily.get("reason") or "Daily no habilitada.",
        }

    process_results = [
        await run_daily_auto_productividad_raw_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily),
        await run_daily_auto_clark_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily),
        await run_daily_auto_picking_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily),
        await run_daily_auto_recepcion_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily),
        await run_daily_auto_despacho_raw_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily),
        await run_daily_auto_despacho_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily),
        await run_daily_auto_planificacion_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily),
        await run_daily_auto_avance_precache(force=force, trigger=trigger, usuario=usuario, daily_override=daily),
    ]
    statuses = {item.get("status") for item in process_results}
    if "error" in statuses and "success" in statuses:
        status = "partial"
    elif "error" in statuses:
        status = "error"
    elif "success" in statuses:
        status = "success"
    else:
        status = "skipped"
    return {
        "status": status,
        "daily": daily,
        "trigger": trigger,
        "processes": process_results,
        "blocked_processes": blocked_processes,
    }


async def _daily_auto_cached_payload(daily: dict[str, Any]) -> dict[str, Any]:
    cached_rows = await get_cached_results(daily["daily_key"])
    latest_run = await get_latest_run(daily["daily_key"], "CLARK")
    latest_picking_run = await get_latest_run(daily["daily_key"], "PICKING")
    latest_recepcion_run = await get_latest_run(daily["daily_key"], "RECEPCION")
    latest_despacho_run = await get_latest_run(daily["daily_key"], "DESPACHO")
    latest_planificacion_run = await get_latest_run(daily["daily_key"], "PLANIFICACION")
    latest_avance_run = await get_latest_run(daily["daily_key"], "AVANCE")
    if latest_run and latest_run.get("status") == "success" and not _is_current_process_run(latest_run, DAILY_AUTO_CLARK_QUERY_VERSION):
        raise HTTPException(
            status_code=409,
            detail="El cache Clark disponible fue generado con una version anterior de la consulta. Requiere precarga de excepcion.",
        )
    if latest_picking_run and latest_picking_run.get("status") == "success" and not _is_current_process_run(latest_picking_run, DAILY_AUTO_PICKING_QUERY_VERSION):
        raise HTTPException(
            status_code=409,
            detail="El cache Picking disponible fue generado con una version anterior de la consulta. Requiere precarga de excepcion.",
        )
    if latest_recepcion_run and latest_recepcion_run.get("status") == "success" and not _is_current_process_run(latest_recepcion_run, DAILY_AUTO_RECEPCION_QUERY_VERSION):
        raise HTTPException(
            status_code=409,
            detail="El cache Recepcion disponible fue generado con una version anterior de la consulta. Requiere precarga de excepcion.",
        )
    if latest_despacho_run and latest_despacho_run.get("status") == "success" and not _is_current_process_run(latest_despacho_run, DAILY_AUTO_DESPACHO_QUERY_VERSION):
        raise HTTPException(
            status_code=409,
            detail="El cache Despacho disponible fue generado con una version anterior de la consulta. Requiere precarga de excepcion.",
        )
    if latest_planificacion_run and latest_planificacion_run.get("status") == "success" and not _is_current_process_run(latest_planificacion_run, DAILY_AUTO_PLANIFICACION_QUERY_VERSION):
        raise HTTPException(
            status_code=409,
            detail="El cache Planificacion disponible fue generado con una version anterior de la consulta. Requiere precarga de excepcion.",
        )
    if latest_avance_run and latest_avance_run.get("status") == "success" and not _is_current_process_run(latest_avance_run, DAILY_AUTO_AVANCE_QUERY_VERSION):
        raise HTTPException(
            status_code=409,
            detail="El cache Avance disponible fue generado con una version anterior de la consulta. Requiere precarga de excepcion.",
        )
    if not _is_current_process_run(latest_run, DAILY_AUTO_CLARK_QUERY_VERSION):
        if latest_run and latest_run.get("status") == "error":
            raise HTTPException(
                status_code=409,
                detail=f"La precarga automatica de Clark fallo: {latest_run.get('error') or 'sin detalle'}. No se consulta Oracle desde la pantalla.",
            )
        raise HTTPException(
            status_code=409,
            detail=f"Todavia no hay cache automatica vigente de Clark para esta Daily. La pantalla no consulta Oracle; espera la precarga de las {_daily_auto_schedule_label()} o ejecuta la precarga de excepcion.",
        )
    if not _is_current_process_run(latest_picking_run, DAILY_AUTO_PICKING_QUERY_VERSION):
        if latest_picking_run and latest_picking_run.get("status") == "error":
            raise HTTPException(
                status_code=409,
                detail=f"La precarga automatica de Picking fallo: {latest_picking_run.get('error') or 'sin detalle'}. No se consulta Oracle desde la pantalla.",
            )
        raise HTTPException(
            status_code=409,
            detail=f"Todavia no hay cache automatica vigente de Picking para esta Daily. La pantalla no consulta Oracle; espera la precarga de las {_daily_auto_schedule_label()} o ejecuta la precarga de excepcion.",
        )
    if not _is_current_process_run(latest_recepcion_run, DAILY_AUTO_RECEPCION_QUERY_VERSION):
        if latest_recepcion_run and latest_recepcion_run.get("status") == "error":
            raise HTTPException(
                status_code=409,
                detail=f"La precarga automatica de Recepcion fallo: {latest_recepcion_run.get('error') or 'sin detalle'}. No se consulta Oracle desde la pantalla.",
            )
        raise HTTPException(
            status_code=409,
            detail=f"Todavia no hay cache automatica vigente de Recepcion para esta Daily. La pantalla no consulta Oracle; espera la precarga de las {_daily_auto_schedule_label()} o ejecuta la precarga de excepcion.",
        )
    if not _is_current_process_run(latest_despacho_run, DAILY_AUTO_DESPACHO_QUERY_VERSION):
        if latest_despacho_run and latest_despacho_run.get("status") == "error":
            raise HTTPException(
                status_code=409,
                detail=f"La precarga automatica de Despacho fallo: {latest_despacho_run.get('error') or 'sin detalle'}. No se consulta Oracle desde la pantalla.",
            )
        raise HTTPException(
            status_code=409,
            detail=f"Todavia no hay cache automatica vigente de Despacho para esta Daily. La pantalla no consulta Oracle; espera la precarga de las {_daily_auto_schedule_label()} o ejecuta la precarga de excepcion.",
        )
    if not _is_current_process_run(latest_planificacion_run, DAILY_AUTO_PLANIFICACION_QUERY_VERSION):
        if latest_planificacion_run and latest_planificacion_run.get("status") == "error":
            raise HTTPException(
                status_code=409,
                detail=f"La precarga automatica de Planificacion fallo: {latest_planificacion_run.get('error') or 'sin detalle'}. No se consulta Oracle desde la pantalla.",
            )
        raise HTTPException(
            status_code=409,
            detail=f"Todavia no hay cache automatica vigente de Planificacion para esta Daily. La pantalla no consulta Oracle; espera la precarga de las {_daily_auto_schedule_label()} o ejecuta la precarga de excepcion.",
        )
    if not _is_current_process_run(latest_avance_run, DAILY_AUTO_AVANCE_QUERY_VERSION):
        if latest_avance_run and latest_avance_run.get("status") == "error":
            raise HTTPException(
                status_code=409,
                detail=f"La precarga automatica de Avance fallo: {latest_avance_run.get('error') or 'sin detalle'}. No se consulta Oracle desde la pantalla.",
            )
        raise HTTPException(
            status_code=409,
            detail=f"Todavia no hay cache automatica vigente de Avance para esta Daily. La pantalla no consulta Oracle; espera la precarga de las {_daily_auto_schedule_label()} o ejecuta la precarga de excepcion.",
        )
    if not cached_rows:
        raise HTTPException(
            status_code=409,
            detail=f"Todavia no hay cache automatica vigente para esta Daily. La pantalla no consulta Oracle; espera la precarga de las {_daily_auto_schedule_label()} o ejecuta la precarga de excepcion.",
        )

    results: list[dict[str, Any]] = []
    by_sector: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in cached_rows:
        row_process = str(row.get("process") or "")
        param_id = str(row.get("id_parametro") or "")
        display_process = row_process
        if row_process == "AVANCE":
            if "_PICKING_" in param_id:
                display_process = "PICKING"
            elif "_SPC_" in param_id:
                display_process = "SPC"
            elif "_RECEPCION_" in param_id:
                display_process = "RECEPCION"
            elif "_DESPACHO_" in param_id:
                display_process = "DESPACHO"
        cantidad = _to_float(row.get("cantidad"))
        item = {
            "sector": row.get("sector"),
            "sector_oracle": row.get("sector_oracle"),
            "proceso": display_process,
            "cache_process": row_process,
            "id_parametro": param_id,
            "valor": _to_float(row.get("valor")),
            "produccion": _to_float(row.get("valor")),
            "cantidad": cantidad,
            "pallets": cantidad if display_process in {"CLARK", "RECEPCION", "SPC"} else 0,
            "bultos": cantidad if display_process == "PICKING" else 0,
            "viajes": cantidad if display_process == "DESPACHO" else 0,
            "planificado": cantidad if row_process == "PLANIFICACION" or "_PLAN_" in param_id else 0,
            "legajos": _to_float(row.get("legajos")),
            "segundos": 0,
            "details_count": int(row.get("details_count") or 0),
            "cache_finished_at": row.get("finished_at"),
            "cache_duration_ms": row.get("duration_ms"),
        }
        results.append(item)
        by_sector[str(item["sector"] or "")].append(item)

    return {
        "daily": daily,
        "fecha_desde": _fmt_daily_oracle_dt(daily.get("fecha_inicio")),
        "fecha_hasta": _fmt_daily_oracle_dt(daily.get("fecha_fin")),
        "source": "daily_auto_cache",
        "db_path": str(DAILY_AUTO_DB_PATH),
        "blocked_processes": [],
        "message": "Modo cache local: Recepcion, Clark, Picking, Despacho y Planificacion habilitados. No se consulta Oracle desde la pantalla.",
        "results": results,
        "by_sector": dict(by_sector),
    }


def _next_daily_auto_run_after(now: datetime) -> datetime:
    current = now.astimezone(LOCAL_TZ)
    candidate = datetime.combine(current.date(), DAILY_AUTO_SCHEDULE_TIME, tzinfo=LOCAL_TZ)
    if candidate <= current:
        candidate += timedelta(days=1)
    return candidate


def _is_current_process_run(run: dict[str, Any] | None, query_version: str) -> bool:
    if not run or run.get("status") != "success":
        return False
    try:
        timings = json.loads(run.get("timings_json") or "{}")
    except Exception:
        timings = {}
    return timings.get("query_version") == query_version


async def _daily_auto_scheduler_loop() -> None:
    assert _daily_auto_scheduler_stop is not None
    await init_daily_auto_db()
    logger.info("[daily-auto] Scheduler iniciado. Proxima ejecucion diaria a las %s.", _daily_auto_schedule_label())
    while not _daily_auto_scheduler_stop.is_set():
        now = datetime.now(LOCAL_TZ)
        window = scheduled_daily_window(now)
        today_run_time = datetime.combine(now.date(), DAILY_AUTO_SCHEDULE_TIME, tzinfo=LOCAL_TZ)
        run_limit = today_run_time + timedelta(minutes=DAILY_AUTO_SCHEDULE_GRACE_MINUTES)
        if (
            window.get("can_run")
            and today_run_time <= now < run_limit
            and (
                not _is_current_process_run(await get_latest_run(window["daily_key"], "CLARK"), DAILY_AUTO_CLARK_QUERY_VERSION)
                or not _is_current_process_run(await get_latest_run(window["daily_key"], "PRODUCTIVIDAD_RAW"), DAILY_AUTO_PRODUCTIVIDAD_RAW_QUERY_VERSION)
                or not _is_current_process_run(await get_latest_run(window["daily_key"], "PICKING"), DAILY_AUTO_PICKING_QUERY_VERSION)
                or not _is_current_process_run(await get_latest_run(window["daily_key"], "RECEPCION"), DAILY_AUTO_RECEPCION_QUERY_VERSION)
                or not _is_current_process_run(await get_latest_run(window["daily_key"], "DESPACHO"), DAILY_AUTO_DESPACHO_QUERY_VERSION)
                or not _is_current_process_run(await get_latest_run(window["daily_key"], "PLANIFICACION"), DAILY_AUTO_PLANIFICACION_QUERY_VERSION)
                or not _is_current_process_run(await get_latest_run(window["daily_key"], "AVANCE"), DAILY_AUTO_AVANCE_QUERY_VERSION)
            )
        ):
            await run_daily_auto_precache_pending(trigger="scheduler")
            continue
        next_run = _next_daily_auto_run_after(now)
        sleep_seconds = max(30.0, min((next_run - now).total_seconds(), 1800.0))
        with suppress(asyncio.TimeoutError):
            await asyncio.wait_for(_daily_auto_scheduler_stop.wait(), timeout=sleep_seconds)


def start_daily_auto_scheduler() -> None:
    global _daily_auto_scheduler_task, _daily_auto_scheduler_stop
    if _daily_auto_scheduler_task and not _daily_auto_scheduler_task.done():
        return
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        return
    _daily_auto_scheduler_stop = asyncio.Event()
    _daily_auto_scheduler_task = loop.create_task(_daily_auto_scheduler_loop())


async def stop_daily_auto_scheduler() -> None:
    global _daily_auto_scheduler_task, _daily_auto_scheduler_stop
    if _daily_auto_scheduler_stop:
        _daily_auto_scheduler_stop.set()
    if _daily_auto_scheduler_task:
        _daily_auto_scheduler_task.cancel()
        with suppress(asyncio.CancelledError):
            await _daily_auto_scheduler_task
    _daily_auto_scheduler_task = None
    _daily_auto_scheduler_stop = None


@router.get("/analisis-legajos/sectores")
async def analisis_legajos_sectores(request: Request):
    await _require_request_auth(request)
    return await _load_legajos_filter_options()


@router.get("/analisis-legajos")
async def analisis_legajos(
    request: Request,
    fecha_desde: str = Query(..., description="YYYY-MM-DD"),
    fecha_hasta: str = Query(..., description="YYYY-MM-DD"),
    sectores: str = Query("", description="Sectores separados por coma"),
    dotacion: str = Query("ALL"),
    proveedor: str = Query("ALL"),
    antiguedad: str = Query("ALL"),
):
    await _require_request_auth(request)
    try:
        desde = datetime.strptime(fecha_desde[:10], "%Y-%m-%d")
        hasta = datetime.strptime(fecha_hasta[:10], "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Fechas invalidas. Usa YYYY-MM-DD.") from exc
    if hasta < desde:
        raise HTTPException(status_code=400, detail="fecha_hasta debe ser mayor o igual a fecha_desde.")
    selected = [item.strip() for item in sectores.split(",") if item.strip()]
    try:
        return await _build_legajos_productividad_payload(
            fecha_desde[:10],
            fecha_hasta[:10],
            selected,
            dotacion=dotacion,
            proveedor=proveedor,
            antiguedad=antiguedad,
        )
    except Exception as exc:
        logger.exception("[gestion-operativa:analisis-legajos] No se pudo completar la consulta")
        detail = str(exc).splitlines()[0].strip() or "Error interno sin detalle."
        raise HTTPException(
            status_code=502,
            detail=f"No se pudo consultar el crudo de Rendimiento Online: {detail}",
        ) from exc


@router.get("/daily/config")
async def daily_config(request: Request):
    auth = await _require_request_auth(request)
    return {
        "daily": calculate_daily_window(),
        "db_path": str(DAILY_DB_PATH),
        "usuario_carga": auth.get("display_name") or auth.get("username"),
        "tipo_daily_options": ["Operacion", "Logistica Inversa", "Planeamiento"],
        "sector_options": {
            "Operacion": ["Noa", "Secos", "Refrigerados", "OC", "Congelados"],
            "Logistica Inversa": ["Logistica Inversa"],
            "Planeamiento": ["Planeamiento"],
        },
        "turno_options": ["Manana", "Tarde", "Noche", "Completo"],
    }


@router.get("/daily/parametros")
async def daily_parametros(
    request: Request,
    tipo_daily: str = Query("Operacion"),
    sector: str = Query("Noa"),
):
    await _require_request_auth(request)
    parametros = await get_parametros(normalize_tipo_daily(tipo_daily), normalize_sector(sector))
    return {"parametros": parametros, "count": len(parametros)}


@router.get("/daily/parametros/admin")
async def daily_parametros_admin(request: Request, clave: str = Query("")):
    await _require_request_auth(request)
    if clave != "ingenieria":
        raise HTTPException(status_code=403, detail="Clave incorrecta.")
    parametros = await get_all_parametros()
    return {"parametros": parametros, "count": len(parametros)}


@router.post("/daily/parametros/admin")
async def daily_parametros_admin_update(req: DailyParametrosUpdateRequest, request: Request):
    await _require_request_auth(request)
    if req.clave != "ingenieria":
        raise HTTPException(status_code=403, detail="Clave incorrecta.")
    updated = await update_parametros(req.parametros)
    return {"updated": updated}


@router.post("/daily/exportar-csv")
async def daily_exportar_csv(request: Request):
    await _require_request_auth(request)
    csv_path = await export_powerbi_csv()
    consolidado_csv_path = await export_consolidado_powerbi_csv()
    return {
        "csv_path": str(csv_path),
        "consolidado_csv_path": str(consolidado_csv_path),
    }


@router.post("/daily/calcular-automatico")
async def daily_calcular_automatico(request: Request):
    await _require_request_auth(request)
    daily = calculate_daily_window()
    if not daily.get("can_load"):
        raise HTTPException(status_code=400, detail=daily.get("reason") or "La Daily no esta habilitada.")
    return await _daily_auto_cached_payload(daily)


def _daily_comparison_summary(rows: list[dict[str, Any]], fecha_desde: str, fecha_hasta: str) -> dict[str, Any]:
    ranking: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    heatmap: dict[tuple[str, str], dict[str, Any]] = {}
    timeseries: dict[str, dict[str, Any]] = {}
    missing_auto = 0
    for row in rows:
        diff_abs = _to_float(row.get("diferencia_abs"))
        key = (row["sector"], row["operacion"], row["metrica"], row["concepto"])
        bucket = ranking.setdefault(
            key,
            {
                "sector": row["sector"],
                "operacion": row["operacion"],
                "metrica": row["metrica"],
                "concepto": row["concepto"],
                "diferencia_abs_total": 0.0,
                "diferencia_abs_promedio": 0.0,
                "diferencia_pct_promedio": 0.0,
                "comparaciones": 0,
                "sin_automatico": 0,
            },
        )
        bucket["comparaciones"] += 1
        if row.get("valor_automatico") is None:
            bucket["sin_automatico"] += 1
            missing_auto += 1
        bucket["diferencia_abs_total"] += diff_abs
        if row.get("diferencia_pct") is not None:
            bucket["diferencia_pct_promedio"] += abs(_to_float(row.get("diferencia_pct")))

        hkey = (row["sector"], row["operacion"])
        hbucket = heatmap.setdefault(hkey, {"sector": row["sector"], "operacion": row["operacion"], "diferencia_abs_total": 0.0, "comparaciones": 0})
        hbucket["diferencia_abs_total"] += diff_abs
        hbucket["comparaciones"] += 1

        day = str(row["fecha_daily"])
        tbucket = timeseries.setdefault(day, {"fecha_daily": day, "diferencia_abs_total": 0.0, "comparaciones": 0})
        tbucket["diferencia_abs_total"] += diff_abs
        tbucket["comparaciones"] += 1

    ranking_rows = []
    for item in ranking.values():
        count = max(int(item["comparaciones"]), 1)
        item["diferencia_abs_promedio"] = item["diferencia_abs_total"] / count
        item["diferencia_pct_promedio"] = item["diferencia_pct_promedio"] / count
        ranking_rows.append(item)
    ranking_rows.sort(key=lambda item: item["diferencia_abs_total"], reverse=True)

    heatmap_rows = []
    for item in heatmap.values():
        count = max(int(item["comparaciones"]), 1)
        item["diferencia_abs_promedio"] = item["diferencia_abs_total"] / count
        heatmap_rows.append(item)
    heatmap_rows.sort(key=lambda item: (item["sector"], item["operacion"]))

    time_rows = []
    for item in timeseries.values():
        count = max(int(item["comparaciones"]), 1)
        item["diferencia_abs_promedio"] = item["diferencia_abs_total"] / count
        time_rows.append(item)
    time_rows.sort(key=lambda item: item["fecha_daily"])
    return {
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "rows": rows,
        "count": len(rows),
        "missing_auto": missing_auto,
        "ranking": ranking_rows[:30],
        "heatmap": heatmap_rows,
        "timeseries": time_rows,
    }


@router.get("/daily/comparacion-manual-auto")
async def daily_comparacion_manual_auto(
    request: Request,
    days: int = Query(15, ge=1, le=60),
):
    await _require_request_auth(request)
    path = DAILY_MANUAL_XLSX_DEFAULT
    if path.exists():
        dates = _manual_compare_dates_from_excel(path, days)
        if dates:
            fecha_desde = dates[0].date().isoformat()
            fecha_hasta = dates[-1].date().isoformat()
        else:
            today = datetime.now(LOCAL_TZ).date()
            fecha_desde = (today - timedelta(days=days - 1)).isoformat()
            fecha_hasta = today.isoformat()
    else:
        today = datetime.now(LOCAL_TZ).date()
        fecha_desde = (today - timedelta(days=days - 1)).isoformat()
        fecha_hasta = today.isoformat()
    rows = await get_daily_auto_manual_comparacion(fecha_desde, fecha_hasta)
    return _daily_comparison_summary(rows, fecha_desde, fecha_hasta)


@router.get("/daily/comparacion-cargas-auto")
async def daily_comparacion_cargas_auto(
    request: Request,
    days: int = Query(15, ge=1, le=60),
):
    await _require_request_auth(request)
    rows, fecha_desde, fecha_hasta = await get_daily_carga_auto_manual_comparacion(days)
    summary = _daily_comparison_summary(rows, fecha_desde or "-", fecha_hasta or "-")
    summary["source"] = "daily_cargas"
    return summary


@router.post("/daily/comparacion-manual-auto/importar")
async def daily_comparacion_manual_auto_importar(req: DailyManualComparacionImportRequest, request: Request):
    auth = await _require_request_auth(request)
    if auth.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Requiere administrador.")
    source = Path(req.source_path).expanduser()
    if not source.exists():
        raise HTTPException(status_code=404, detail=f"No existe el archivo {source}.")
    dates = _manual_compare_dates_from_excel(source, max(int(req.days or 15), 1))
    if not dates:
        raise HTTPException(status_code=400, detail="No se encontraron fechas validas en el Excel.")
    fecha_desde = dates[0].date().isoformat()
    fecha_hasta = dates[-1].date().isoformat()
    auto_results = []
    for fecha in dates:
        daily = _daily_window_for_fecha_carga(fecha)
        result = await run_daily_auto_precache_pending(
            force=bool(req.force_auto),
            trigger="manual_comparison",
            usuario=str(auth.get("display_name") or auth.get("username") or ""),
            daily_override=daily,
        )
        auto_results.append({"fecha_daily": fecha.date().isoformat(), "daily_key": daily["daily_key"], "status": result.get("status"), "processes": result.get("processes")})
    manual_rows = _manual_rows_from_excel(source, dates)
    imported = await replace_daily_manual_comparacion_rows(
        manual_rows,
        source_file=str(source),
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    compared = await rebuild_daily_auto_manual_comparacion(fecha_desde, fecha_hasta)
    rows = await get_daily_auto_manual_comparacion(fecha_desde, fecha_hasta)
    summary = _daily_comparison_summary(rows, fecha_desde, fecha_hasta)
    summary["imported"] = imported
    summary["compared"] = compared
    summary["auto_results"] = auto_results
    return summary


@router.get("/daily/raw-detail")
async def daily_raw_detail(
    request: Request,
    sector: str = Query(...),
    param_id: str = Query(...),
    process: str = Query(""),
):
    await _require_request_auth(request)
    daily = calculate_daily_window()
    if not daily.get("can_load"):
        raise HTTPException(status_code=400, detail=daily.get("reason") or "La Daily no esta habilitada.")
    return await _build_daily_raw_detail_payload(
        daily,
        sector=sector,
        param_id=param_id,
        process=process,
    )


@router.post("/daily/auto/retry-pending")
async def daily_auto_retry_pending(req: DailyAutoRetryRequest, request: Request):
    auth = await _require_request_auth(request)
    username = str(auth.get("username") or "").strip().lower()
    is_force = bool(req.force)
    if is_force and username not in DAILY_AUTO_FORCE_USERS:
        raise HTTPException(status_code=403, detail="No tenes permiso para forzar el recalculo automatico.")
    if not is_force and auth.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Requiere administrador.")
    usuario = str(auth.get("display_name") or auth.get("username") or "").strip()
    retry = await run_daily_auto_precache_pending(
        force=is_force,
        trigger="manual_force" if is_force else "manual_exception",
        usuario=usuario,
    )
    daily = calculate_daily_window()
    if not daily.get("can_load"):
        raise HTTPException(status_code=400, detail=daily.get("reason") or "La Daily no esta habilitada.")
    try:
        payload = await _daily_auto_cached_payload(daily)
    except HTTPException as exc:
        if exc.status_code == 409:
            raise HTTPException(status_code=409, detail={"message": exc.detail, "retry": retry}) from exc
        raise
    payload["retry"] = retry
    payload["message"] = "Precarga de excepcion ejecutada. La pantalla muestra datos desde cache local."
    return payload


def _daily_clark_sql_preview(fecha_desde: str, fecha_hasta: str) -> str:
    return (
        QUERY_DAILY_CLARK_REAL
        .replace(":fecha_desde", f"'{fecha_desde}'")
        .replace(":fecha_hasta", f"'{fecha_hasta}'")
        .strip()
    )


def _daily_picking_sql_preview(fecha_desde: str, fecha_hasta: str) -> str:
    return (
        QUERY_DAILY_PICKING_REAL
        .replace(":fecha_desde", f"'{fecha_desde}'")
        .replace(":fecha_hasta", f"'{fecha_hasta}'")
        .strip()
    )


def _daily_recepcion_sql_preview(fecha_desde: str, fecha_hasta: str) -> str:
    return (
        QUERY_DAILY_RECEPCION_REAL
        .replace(":fecha_desde", f"'{fecha_desde}'")
        .replace(":fecha_hasta", f"'{fecha_hasta}'")
        .strip()
    )


def _daily_despacho_sql_preview(fecha_desde: str, fecha_hasta: str) -> str:
    return (
        QUERY_DAILY_DESPACHO_REAL
        .replace(":fecha_desde", f"'{fecha_desde}'")
        .replace(":fecha_hasta", f"'{fecha_hasta}'")
        .strip()
    )


def _daily_despacho_raw_sql_preview(fecha_desde: str, fecha_hasta: str) -> str:
    return (
        QUERY_DAILY_DESPACHO_RAW
        .replace(":fecha_desde", f"'{fecha_desde}'")
        .replace(":fecha_hasta", f"'{fecha_hasta}'")
        .strip()
    )


def _daily_productividad_raw_sql_preview(fecha_desde: str, fecha_hasta: str) -> str:
    return (
        QUERY_DAILY_PRODUCTIVIDAD_RAW
        .replace(":fecha_desde", f"'{fecha_desde}'")
        .replace(":fecha_hasta", f"'{fecha_hasta}'")
        .strip()
    )


def _daily_planificacion_sql_preview(fecha_desde: str, fecha_hasta: str) -> str:
    return (
        QUERY_DAILY_PLANIFICACION
        .replace(":fecha_desde", f"'{fecha_desde}'")
        .replace(":fecha_hasta", f"'{fecha_hasta}'")
        .strip()
    )


def _daily_picking_raw_almacen(row: dict[str, Any]) -> str:
    almacen = str(row.get("ALMACEN") or "").strip().upper()
    zona = str(row.get("CZONAORI") or "").strip().upper()
    if zona == "T06":
        return "REFRIGERADOS"
    if zona and "T" in zona:
        return "REFRIGERADOS"
    if almacen in {"NOA", "SECOS", "REFRIGERADOS"}:
        return almacen
    if zona in {"N01", "N02", "N04", "N05", "N07", "N09", "N10", "N15"}:
        return "NOA"
    return "SECOS"


def _daily_productividad_raw_division(row: dict[str, Any]) -> str:
    base = _daily_picking_raw_almacen(row)
    if base != "REFRIGERADOS":
        return base
    zona = str(row.get("CZONAORI") or "").strip().upper()
    return "CONGELADOS" if zona == "T06" else "OC"


def _daily_productividad_jornada(division: str) -> float:
    sector = normalize_sector(division)
    return PRODUCTIVIDAD_JORNADA_BY_DIVISION.get(sector, 6.5)


def _valid_clark_pallet(row: dict[str, Any]) -> bool:
    pallet = str(row.get("CNUPALET") or row.get("pallet") or "").strip()
    return bool(pallet and pallet != "0")


def _clark_cuenta_movimiento(row: dict[str, Any]) -> dict[str, str]:
    if _valid_clark_pallet(row):
        return {"cuenta": "SI", "motivo_no_cuenta": ""}
    return {"cuenta": "NO", "motivo_no_cuenta": "Pallet vacio o cero"}


def _parse_wf_fec_ing(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    if "." in text:
        text = text.split(".", 1)[0]
    if len(text) != 8 or not text.isdigit():
        return None
    try:
        return datetime.strptime(text, "%Y%m%d")
    except ValueError:
        return None


def _picking_legajo_antiguedad(row: dict[str, Any], daily_start: datetime) -> dict[str, Any]:
    raw_fec_ing = row.get("FEC_ING") if row.get("FEC_ING") is not None else row.get("fec_ing")
    fec_ing = _parse_wf_fec_ing(raw_fec_ing)
    if not str(raw_fec_ing or "").strip():
        return {
            "fec_ing": "",
            "antiguedad_dias": None,
            "cuenta_legajo_bool": True,
            "cuenta_legajo": "SI",
            "motivo_no_cuenta": "Sin fecha de ingreso",
        }
    if not fec_ing:
        return {
            "fec_ing": str(raw_fec_ing).strip(),
            "antiguedad_dias": None,
            "cuenta_legajo_bool": True,
            "cuenta_legajo": "SI",
            "motivo_no_cuenta": "Fecha de ingreso invalida",
        }
    antiguedad = (daily_start.date() - fec_ing.date()).days
    cuenta = antiguedad >= 15
    return {
        "fec_ing": fec_ing.strftime("%Y-%m-%d"),
        "antiguedad_dias": antiguedad,
        "cuenta_legajo_bool": cuenta,
        "cuenta_legajo": "SI" if cuenta else "NO",
        "motivo_no_cuenta": "" if cuenta else "Menor a 15 dias",
    }


def _daily_raw_detail_mode(param_id: str, process: str) -> dict[str, Any] | None:
    param = str(param_id or "").strip().upper()
    proc = str(process or "").strip().upper()
    if proc == "PICKING" or param.startswith("OP_PROD_PICKING_") or param in {"OP_CUMP_PICKING_REAL_6A6", "OP_DOT_PICKING_LEGAJOS_6A8"}:
        return {"process": "PICKING", "operations": {"PICKING"}, "sector_source": "picking", "positive_quantity": False}
    if proc == "RECEPCION" or param.startswith("OP_PROD_RECEPCION_") or param in {"OP_CUMP_RECEPCION_REAL_6A6", "OP_DOT_RECEPCION_LEGAJOS_6A8"}:
        return {"process": "RECEPCION", "operations": {"REVISION PALETS ENTRADA"}, "sector_source": "raw", "positive_quantity": True}
    if proc == "SPC" or param in {"OP_CUMP_SPC_REAL_6A6", "OP_DOT_SPC_LEGAJOS_6A8"} or param.startswith("OP_AVANCE_SPC_"):
        return {"process": "SPC", "operations": {"SURTIDO P.COMPLETOS"}, "sector_source": "raw", "positive_quantity": False}
    if proc == "CLARK" or param.startswith("OP_PROD_CLARK_"):
        if param in {"OP_CUMP_SPC_REAL_6A6", "OP_DOT_SPC_LEGAJOS_6A8"}:
            return {"process": "SPC", "operations": {"SURTIDO P.COMPLETOS"}, "sector_source": "raw", "positive_quantity": False}
        return {
            "process": "CLARK",
            "operations": {
                "GUARADO PALETS ENTRADA",
                "EXTRACCION DE REAPROS",
                "EXTRACCION TRASPASOS",
                "SURTIDO P.COMPLETOS",
            },
            "sector_source": "productividad_division",
            "positive_quantity": False,
        }
    if proc == "DESPACHO" or param.startswith("OP_PROD_DESPACHO_") or param in {"OP_CUMP_DESPACHO_REAL_6A6", "OP_DOT_DESPACHO_LEGAJOS_6A8"}:
        return {"process": "DESPACHO", "source": "despacho_raw"}
    return None


def _daily_raw_detail_row(
    row: dict[str, Any],
    *,
    ts: datetime,
    legajo: str,
    almacen_calc: str,
    prev_ts: datetime | None,
    daily_start: datetime | None = None,
    cuenta_movimiento: dict[str, str] | None = None,
) -> dict[str, Any]:
    seconds = max((ts - prev_ts).total_seconds(), 0.0) if prev_ts else 0.0
    cantidad = _to_float(row.get("QCANTIDA") if row.get("QCANTIDA") is not None else row.get("cantidad"))
    cuenta = (cuenta_movimiento or {}).get("cuenta", "SI")
    detail = {
        "fecha": ts.strftime("%Y-%m-%d %H:%M:%S"),
        "legajo": legajo,
        "desc_funcion": str(row.get("DESC_FUNCION") or row.get("desc_funcion") or "").strip(),
        "operacion": str(row.get("CDESCRIP") or row.get("OPERACION") or "").strip(),
        "almacen": str(row.get("ALMACEN") or "").strip(),
        "almacen_calculo": almacen_calc,
        "zona_origen": str(row.get("CZONAORI") or "").strip(),
        "ubicacion": str(row.get("CUBIORIG") or "").strip(),
        "pallet": str(row.get("CNUPALET") or "").strip(),
        "referencia": str(row.get("CREFEREN") or "").strip(),
        "cantidad": cantidad,
        "minutos_desde_anterior": round(seconds / 60, 4),
        "cuenta": cuenta,
        "motivo_no_cuenta": (cuenta_movimiento or {}).get("motivo_no_cuenta", ""),
        "minutos_cuenta": round(seconds / 60, 4) if cuenta == "SI" else 0,
        "row_index": row.get("_row_index"),
    }
    if daily_start:
        antiguedad = _picking_legajo_antiguedad(row, daily_start)
        detail.update(
            {
                "fec_ing": antiguedad["fec_ing"],
                "antiguedad_dias": antiguedad["antiguedad_dias"],
                "cuenta_legajo": antiguedad["cuenta_legajo"],
                "motivo_no_cuenta": antiguedad["motivo_no_cuenta"],
            }
        )
    return detail


async def _build_daily_raw_detail_payload(
    daily: dict[str, Any],
    *,
    sector: str,
    param_id: str,
    process: str,
) -> dict[str, Any]:
    mode = _daily_raw_detail_mode(param_id, process)
    if not mode:
        raise HTTPException(status_code=400, detail="El parametro no se calcula desde la cache cruda.")
    sector_norm = normalize_sector(sector)
    param_upper = str(param_id or "").strip().upper()
    almacen_target = DAILY_RAW_SECTOR_TO_ALMACEN.get(sector_norm)
    if not almacen_target:
        raise HTTPException(status_code=400, detail="Sector invalido para detalle crudo.")
    if str(param_id or "").upper().startswith("OP_AVANCE_"):
        start, end = _avance_window(daily)
    else:
        start = _parse_dt(daily.get("fecha_inicio"))
        end = _parse_dt(daily.get("fecha_fin"))
    if not start or not end:
        raise HTTPException(status_code=400, detail="Ventana Daily invalida.")
    detail_fecha_desde = _fmt_daily_oracle_dt(start)
    detail_fecha_hasta = _fmt_daily_oracle_dt(end)
    cache_start = _parse_dt(daily.get("fecha_inicio"))
    cache_end = _parse_dt(daily.get("fecha_fin"))
    cache_fecha_desde = _fmt_daily_oracle_dt(cache_start) if cache_start else detail_fecha_desde
    cache_fecha_hasta = _fmt_daily_oracle_dt(cache_end + timedelta(minutes=90)) if cache_end else detail_fecha_hasta

    if mode.get("source") == "despacho_raw":
        raw_run = await get_latest_run(daily["daily_key"], "DESPACHO_RAW")
        if not _is_current_process_run(raw_run, DAILY_AUTO_DESPACHO_RAW_QUERY_VERSION):
            raise HTTPException(status_code=409, detail="La cache cruda de Despacho no esta vigente.")
        raw_rows = await get_despacho_raw_cache_rows(daily["daily_key"])
        detail_rows = []
        for row in raw_rows:
            ts = _parse_dt(row.get("FECIERRE") or row.get("FECHA_CIERRE"))
            almacen = str(row.get("ALMACEN") or "").strip().upper()
            if not ts or ts < start or ts >= end or almacen != almacen_target:
                continue
            detail_rows.append(
                {
                    "fecha_cierre": ts.strftime("%Y-%m-%d %H:%M:%S"),
                    "almacen": almacen,
                    "hoja_ruta": str(row.get("HOJARUTA") or row.get("CNUVIAJE") or row.get("VIAJE") or "").strip(),
                    "cargador": _norm_legajo(row.get("CARGADOR")),
                    "desc_funcion": str(row.get("DESC_FUNCION") or "").strip(),
                    "cuenta_legajo": "SI" if _despacho_cargador_elegible(row) else "NO",
                    "division": str(row.get("CDIVISIO") or "").strip(),
                    "calmacen": str(row.get("CALMACEN") or "").strip(),
                    "row_index": row.get("_row_index"),
                }
            )
        columns = [
            {"key": "fecha_cierre", "label": "Fecha cierre"},
            {"key": "almacen", "label": "Almacen"},
            {"key": "hoja_ruta", "label": "Hoja ruta"},
            {"key": "cargador", "label": "Cargador"},
            {"key": "desc_funcion", "label": "Funcion"},
            {"key": "cuenta_legajo", "label": "Cuenta legajo"},
            {"key": "division", "label": "Division"},
            {"key": "calmacen", "label": "Calmacen"},
            {"key": "row_index", "label": "Fila cache"},
        ]
        return {
            "daily": daily,
            "query_key": "daily_despacho_raw_cache",
            "source_process": "DESPACHO_RAW",
            "detail_process": "DESPACHO",
            "sector": normalize_sector(sector),
            "almacen": almacen_target,
            "param_id": param_id,
            "fecha_desde": detail_fecha_desde,
            "fecha_hasta": detail_fecha_hasta,
            "oracle_fecha_desde": cache_fecha_desde,
            "oracle_fecha_hasta": cache_fecha_hasta,
            "sql_template": QUERY_DAILY_DESPACHO_RAW.strip(),
            "sql_preview": _daily_despacho_raw_sql_preview(cache_fecha_desde, cache_fecha_hasta),
            "columns": columns,
            "rows": detail_rows,
            "count": len(detail_rows),
            "message": "Detalle generado desde cache cruda local de Despacho. No ejecuta Oracle.",
        }

    raw_run = await get_latest_run(daily["daily_key"], "PRODUCTIVIDAD_RAW")
    if not _is_current_process_run(raw_run, DAILY_AUTO_PRODUCTIVIDAD_RAW_QUERY_VERSION):
        raise HTTPException(status_code=409, detail="La cache cruda de productividad no esta vigente.")

    raw_rows = await get_productividad_raw_cache_rows(daily["daily_key"])
    scoped_rows = []
    for row in raw_rows:
        ts = _parse_dt(row.get("FCREAREG") or row.get("fecha"))
        legajo = _norm_legajo(row.get("COPECREA") or row.get("LEGAJO") or row.get("legajo"))
        if not ts or not legajo or ts < start or ts >= end:
            continue
        scoped_rows.append({**row, "_ts": ts, "_legajo": legajo})
    scoped_rows.sort(key=lambda item: (item["_legajo"], item["_ts"], int(item.get("_row_index") or 0)))

    detail_rows: list[dict[str, Any]] = []
    previous_by_legajo: dict[str, datetime] = {}
    operations = mode["operations"]
    aggregate_refri_picking = (
        mode["process"] == "PICKING"
        and sector_norm == "Refrigerados"
        and param_upper in {"OP_CUMP_PICKING_REAL_6A6", "OP_DOT_PICKING_LEGAJOS_6A8"}
    )
    for row in scoped_rows:
        legajo = row["_legajo"]
        ts = row["_ts"]
        prev = previous_by_legajo.get(legajo)
        previous_by_legajo[legajo] = ts
        operacion = str(row.get("CDESCRIP") or row.get("OPERACION") or "").strip().upper()
        if operacion not in operations:
            continue
        cantidad = _to_float(row.get("QCANTIDA") if row.get("QCANTIDA") is not None else row.get("cantidad"))
        if mode.get("positive_quantity") and cantidad <= 0:
            continue
        if aggregate_refri_picking:
            if _daily_picking_raw_almacen(row) != "REFRIGERADOS":
                continue
            almacen_calc = _daily_productividad_raw_division(row)
        else:
            almacen_calc = (
                _daily_productividad_raw_division(row)
                if mode["sector_source"] in {"picking", "productividad_division"}
                else str(row.get("ALMACEN") or "").strip().upper()
            )
        if not aggregate_refri_picking and almacen_calc != almacen_target:
            continue
        detail_rows.append(
            _daily_raw_detail_row(
                row,
                ts=ts,
                legajo=legajo,
                almacen_calc=almacen_calc,
                prev_ts=prev,
                daily_start=(cache_start or start) if mode["process"] == "PICKING" else None,
                cuenta_movimiento=_clark_cuenta_movimiento(row) if mode["process"] in {"CLARK", "SPC"} else None,
            )
        )

    columns = [
        {"key": "fecha", "label": "Fecha"},
        {"key": "legajo", "label": "Legajo"},
        *([{"key": "desc_funcion", "label": "Funcion"}] if mode["process"] == "PICKING" else []),
        *(
            [
                {"key": "fec_ing", "label": "Fec ing"},
                {"key": "antiguedad_dias", "label": "Antig dias"},
                {"key": "cuenta_legajo", "label": "Cuenta legajo"},
                {"key": "motivo_no_cuenta", "label": "Motivo no cuenta"},
            ]
            if mode["process"] == "PICKING"
            else []
        ),
        {"key": "operacion", "label": "Operacion"},
        {"key": "almacen", "label": "Division cruda"},
        {"key": "almacen_calculo", "label": "Division calculo"},
        {"key": "zona_origen", "label": "Zona origen"},
        {"key": "ubicacion", "label": "Ubicacion"},
        {"key": "pallet", "label": "Pallet"},
        {"key": "referencia", "label": "Referencia"},
        {"key": "cantidad", "label": "Cantidad"},
        {"key": "minutos_desde_anterior", "label": "Min prev"},
        *(
            [
                {"key": "cuenta", "label": "Cuenta"},
                {"key": "motivo_no_cuenta", "label": "Motivo no cuenta"},
                {"key": "minutos_cuenta", "label": "Min cuenta"},
            ]
            if mode["process"] in {"CLARK", "SPC"}
            else []
        ),
        {"key": "row_index", "label": "Fila cache"},
    ]
    return {
        "daily": daily,
        "query_key": "daily_productividad_raw_cache",
        "source_process": "PRODUCTIVIDAD_RAW",
        "detail_process": mode["process"],
        "sector": sector_norm,
        "almacen": almacen_target,
        "param_id": param_id,
        "fecha_desde": detail_fecha_desde,
        "fecha_hasta": detail_fecha_hasta,
        "oracle_fecha_desde": cache_fecha_desde,
        "oracle_fecha_hasta": cache_fecha_hasta,
        "sql_template": QUERY_DAILY_PRODUCTIVIDAD_RAW.strip(),
        "sql_preview": _daily_productividad_raw_sql_preview(cache_fecha_desde, cache_fecha_hasta),
        "columns": columns,
        "rows": detail_rows,
        "count": len(detail_rows),
        "message": "Detalle generado desde cache cruda local. No ejecuta Oracle.",
    }


@router.get("/daily/clark-sql")
async def daily_clark_sql(request: Request):
    auth = await _require_request_auth(request)
    if auth.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Requiere administrador.")
    daily = calculate_daily_window()
    if not daily.get("can_load"):
        raise HTTPException(status_code=400, detail=daily.get("reason") or "La Daily no esta habilitada.")
    fecha_desde = _fmt_daily_oracle_dt(daily.get("fecha_inicio"))
    fecha_hasta = _fmt_daily_oracle_dt(daily.get("fecha_fin"))
    return {
        "daily": daily,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "query_key": "daily_clark_real",
        "sql_template": QUERY_DAILY_CLARK_REAL.strip(),
        "sql_preview": _daily_clark_sql_preview(fecha_desde, fecha_hasta),
        "message": "Consulta agregada de Clark. Este endpoint no ejecuta Oracle; solo expone el SQL para revision admin.",
    }


@router.get("/daily/picking-sql")
async def daily_picking_sql(request: Request):
    auth = await _require_request_auth(request)
    if auth.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Requiere administrador.")
    daily = calculate_daily_window()
    if not daily.get("can_load"):
        raise HTTPException(status_code=400, detail=daily.get("reason") or "La Daily no esta habilitada.")
    fecha_desde = _fmt_daily_oracle_dt(daily.get("fecha_inicio"))
    fecha_hasta = _fmt_daily_oracle_dt(daily.get("fecha_fin"))
    return {
        "daily": daily,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "query_key": "daily_picking_real",
        "sql_template": QUERY_DAILY_PICKING_REAL.strip(),
        "sql_preview": _daily_picking_sql_preview(fecha_desde, fecha_hasta),
        "message": "Consulta agregada de Picking. Este endpoint no ejecuta Oracle; solo expone el SQL para revision admin.",
    }


@router.get("/daily/recepcion-sql")
async def daily_recepcion_sql(request: Request):
    auth = await _require_request_auth(request)
    if auth.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Requiere administrador.")
    daily = calculate_daily_window()
    if not daily.get("can_load"):
        raise HTTPException(status_code=400, detail=daily.get("reason") or "La Daily no esta habilitada.")
    fecha_desde = _fmt_daily_oracle_dt(daily.get("fecha_inicio"))
    fecha_hasta = _fmt_daily_oracle_dt(daily.get("fecha_fin"))
    return {
        "daily": daily,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "query_key": "daily_recepcion_real",
        "sql_template": QUERY_DAILY_RECEPCION_REAL.strip(),
        "sql_preview": _daily_recepcion_sql_preview(fecha_desde, fecha_hasta),
        "message": "Consulta agregada de Recepcion. Este endpoint no ejecuta Oracle; solo expone el SQL para revision admin.",
    }


@router.get("/daily/despacho-sql")
async def daily_despacho_sql(request: Request):
    auth = await _require_request_auth(request)
    if auth.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Requiere administrador.")
    daily = calculate_daily_window()
    if not daily.get("can_load"):
        raise HTTPException(status_code=400, detail=daily.get("reason") or "La Daily no esta habilitada.")
    fecha_desde = _fmt_daily_oracle_dt(daily.get("fecha_inicio"))
    fecha_hasta = _fmt_daily_oracle_dt(daily.get("fecha_fin"))
    return {
        "daily": daily,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "query_key": "daily_despacho_real",
        "sql_template": QUERY_DAILY_DESPACHO_REAL.strip(),
        "sql_preview": _daily_despacho_sql_preview(fecha_desde, fecha_hasta),
        "message": "Consulta agregada de Despacho. Este endpoint no ejecuta Oracle; solo expone el SQL para revision admin.",
    }


@router.get("/daily/planificacion-sql")
async def daily_planificacion_sql(request: Request):
    auth = await _require_request_auth(request)
    if auth.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Requiere administrador.")
    daily = calculate_daily_window()
    if not daily.get("can_load"):
        raise HTTPException(status_code=400, detail=daily.get("reason") or "La Daily no esta habilitada.")
    fecha_desde = _fmt_daily_oracle_dt(daily.get("fecha_inicio"))
    fecha_hasta = _fmt_daily_oracle_dt(daily.get("fecha_fin"))
    return {
        "daily": daily,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "query_key": "daily_planificacion",
        "sql_template": QUERY_DAILY_PLANIFICACION.strip(),
        "sql_preview": _daily_planificacion_sql_preview(fecha_desde, fecha_hasta),
        "message": "Consulta agregada de Planificacion. Este endpoint no ejecuta Oracle; solo expone el SQL para revision admin.",
    }


@router.get("/daily/existente")
async def daily_existente(
    request: Request,
    tipo_daily: str = Query("Operacion"),
    sector: str = Query("Noa"),
):
    await _require_request_auth(request)
    daily = calculate_daily_window()
    if not daily.get("can_load"):
        return {"daily": daily, "existing": []}
    rows = await get_existing_cargas(daily["daily_key"], tipo_daily, sector)
    return {"daily": daily, "existing": rows}


@router.post("/daily/cargas")
async def daily_guardar(req: DailyCargaRequest, request: Request):
    auth = await _require_request_auth(request)
    username = str(auth.get("display_name") or auth.get("username") or "").strip()
    if not username:
        raise HTTPException(status_code=401, detail="No se pudo identificar el usuario de carga.")
    try:
        result = await save_daily_carga(
            usuario_carga=username,
            tipo_daily=req.tipo_daily,
            sector=req.sector,
            turno=req.turno,
            plan=req.plan,
            respuestas=req.respuestas,
            action="replace",
        )
        return result
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/cambios-almacen")
async def get_cambios_almacen(
    fecha: str = Query(..., description="YYYY-MM-DD"),
    turno: str = Query(..., description="Mañana, Tarde o Noche"),
) -> dict[str, Any]:
    turno_key, fecha_desde, fecha_hasta = _turn_range_for_date(fecha, turno)
    logger.info(
        "[gestion-operativa:cambios-almacen] Consultando Oracle turno=%s rango=%s..%s",
        _turn_label(turno_key),
        fecha_desde,
        fecha_hasta,
    )
    try:
        detail_rows = await asyncio.to_thread(
            query_productive_db_picking_tiempos_muertos,
            fecha_desde,
            fecha_hasta,
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    except Exception as exc:
        logger.exception("Error consultando cambios de almacen")
        raise HTTPException(status_code=500, detail=f"No se pudo consultar Oracle: {exc}")

    base = _build_picking_idle_analysis(
        detail_rows,
        [],
        fecha,
        turno_key,
        fecha_desde,
        fecha_hasta,
    )
    crossings = base.get("almacen_crossings", {})
    return {
        "fecha": fecha,
        "turno": _turn_label(turno_key),
        "turno_key": turno_key,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "summary": {
            "movimientos_total": base.get("summary", {}).get("movimientos_total", 0),
            "pickings_total": base.get("summary", {}).get("pickings_total", 0),
            **crossings.get("summary", {}),
        },
        "transferencias": crossings.get("transferencias", []),
        "cruces": crossings.get("cruces", []),
        "source_name": "oracle_productiva",
    }


@router.get("/productividad-picking")
async def get_productividad_picking(
    fecha_desde: str = Query(..., description="YYYY-MM-DD HH:MM:SS"),
    fecha_hasta: str = Query(..., description="YYYY-MM-DD HH:MM:SS"),
    refresh: bool = Query(False, description="Forzar consulta Oracle"),
) -> dict[str, Any]:
    desde_dt = _parse_dt(fecha_desde)
    hasta_dt = _parse_dt(fecha_hasta)
    if not desde_dt or not hasta_dt:
        raise HTTPException(status_code=400, detail="Fechas invalidas. Usá formato YYYY-MM-DD HH:MM.")
    if hasta_dt <= desde_dt:
        raise HTTPException(status_code=400, detail="La fecha de fin debe ser posterior a la fecha de inicio.")
    fecha_desde_fmt = _fmt_dt(desde_dt)
    fecha_hasta_fmt = _fmt_dt(hasta_dt)
    if not refresh:
        cached = await _load_productividad_run(fecha_desde_fmt, fecha_hasta_fmt)
        if cached:
            return cached

    if hasta_dt - desde_dt > timedelta(days=14):
        raise HTTPException(
            status_code=400,
            detail="El rango no esta completo en cache local y supera 14 dias; para cuidar Oracle, consultalo en tramos.",
        )

    logger.info(
        "[gestion-operativa:productividad-picking] Consultando Oracle rango=%s..%s",
        fecha_desde_fmt,
        fecha_hasta_fmt,
    )
    try:
        raw_rows = await asyncio.to_thread(
            query_productive_db_gestion_productividad_picking,
            fecha_desde_fmt,
            fecha_hasta_fmt,
        )
    except RuntimeError as exc:
        cached = await _load_productividad_run(fecha_desde_fmt, fecha_hasta_fmt)
        if cached:
            cached["source_name"] = "sqlite_cache_oracle_error"
            cached["warning"] = str(exc)
            return cached
        raise HTTPException(status_code=500, detail=str(exc))
    except Exception as exc:
        logger.exception("Error consultando productividad de picking")
        cached = await _load_productividad_run(fecha_desde_fmt, fecha_hasta_fmt)
        if cached:
            cached["source_name"] = "sqlite_cache_oracle_error"
            cached["warning"] = f"No se pudo consultar Oracle: {exc}"
            return cached
        raise HTTPException(status_code=500, detail=f"No se pudo consultar Oracle: {exc}")

    events = _normalize_productividad_rows(raw_rows)
    segments = _build_segments(events)
    run_id = await _store_productividad_run(fecha_desde_fmt, fecha_hasta_fmt, events, segments, len(raw_rows))
    segments = _enrich_segments_with_legajero(segments, await _load_latest_legajero_profiles())
    return _build_productividad_payload(
        run_id,
        fecha_desde_fmt,
        fecha_hasta_fmt,
        events,
        segments,
        len(raw_rows),
        "oracle_productiva",
    )


@router.post("/productividad-picking/ia")
async def post_productividad_picking_ia(req: ProductividadPickingIARequest) -> dict[str, Any]:
    provider = (req.provider or os.getenv("AI_PROVIDER", "claude")).lower()
    payload = req.analisis_base or {}
    if not payload.get("summary"):
        raise HTTPException(status_code=400, detail="No hay datos de productividad para analizar.")
    logger.info(
        "[gestion-operativa:productividad-picking:ia] Ejecutando consulta IA provider=%s rango=%s..%s almacen=%s segmentos=%s charts=%s",
        provider,
        payload.get("fecha_desde"),
        payload.get("fecha_hasta"),
        payload.get("selected_almacen", "ALL"),
        payload.get("segment_rows_count", 0),
        len((payload.get("tendencias") or {}).get("charts") or []),
    )
    context = _build_productividad_ai_context(payload)
    try:
        text, model_used = await _call_ai(
            provider,
            SYSTEM_GESTION_PRODUCTIVIDAD_IA,
            [{"role": "user", "content": context}],
        )
        logger.info(
            "[gestion-operativa:productividad-picking:ia] Respuesta IA recibida provider=%s model=%s chars=%s",
            provider,
            model_used,
            len(text or ""),
        )
        try:
            parsed = json.loads(_extract_json(text))
        except Exception:
            parsed = {
                "resumen": text.strip()[:2000],
                "estado_actual": "mixto",
                "comparacion_vs_pasado": "La IA devolvio texto no estructurado.",
                "tendencias": [],
                "causas_probables": [],
                "acciones": [],
                "almacenes_a_revisar": [],
            }
        parsed["provider"] = provider
        parsed["model_used"] = model_used
        parsed["prompt_version"] = GESTION_PRODUCTIVIDAD_IA_PROMPT_VERSION
        parsed["audit"] = {
            "system_prompt": SYSTEM_GESTION_PRODUCTIVIDAD_IA,
            "user_prompt_template": (
                "Analiza productividad de picking en un centro de distribucion. Objetivo: lectura gerencial breve. "
                "Se adjunta JSON compacto con metricas agregadas, tendencias por dia/almacen y muestra operativa acotada."
            ),
            "user_prompt_real": context,
            "data_policy": (
                "No se envian nombres completos ni datos personales sensibles. "
                "El payload se limita a metricas agregadas de productividad, almacenes, fechas y una muestra operativa compacta."
            ),
            "payload_keys": sorted(list(_compact_productividad_for_ai(payload).keys())),
        }
        return parsed
    except Exception as exc:
        logger.exception("Error generando analisis IA de productividad picking")
        raise HTTPException(status_code=500, detail=f"No se pudo generar analisis IA: {exc}")
