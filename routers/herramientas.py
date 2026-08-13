from __future__ import annotations

import math
import base64
import csv
import io
import json
import logging
import os
import re
import subprocess
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

import aiosqlite
from openpyxl import load_workbook
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from db.schema import DB_PATH


router = APIRouter(prefix="/api/herramientas", tags=["herramientas"])
logger = logging.getLogger("vigia.herramientas")

ROOT_DIR = Path(__file__).resolve().parent.parent
JAVA_HELPER_SRC = ROOT_DIR / "scripts" / "OracleProductividadQuery.java"
JAVA_BUILD_DIR = ROOT_DIR / ".codex_tmp" / "java_build"


DESTINOS_KM = [
    {"id": "cd_mar_de_ajo", "nombre": "CD a Mar de Ajo", "km": 720},
    {"id": "cd_mar_de_ajo_madariaga", "nombre": "CD a Mar de Ajo - Madariaga", "km": 800},
    {"id": "cd_mar_del_plata", "nombre": "CD a Mar de Plata", "km": 850},
    {"id": "cd_mar_del_tuyu", "nombre": "CD a Mar del Tuyu", "km": 720},
    {"id": "cd_mar_del_tuyu_mdq", "nombre": "CD a Mar del Tuyu - MDQ", "km": 967},
    {"id": "cd_parana", "nombre": "CD a Parana", "km": 1100},
    {"id": "cd_rosario", "nombre": "CD a Rosario", "km": 680},
    {"id": "cd_santa_fe", "nombre": "CD a Santa Fe", "km": 1060},
    {"id": "cd_madariaga", "nombre": "CD Madariaga", "km": 750},
    {"id": "famaba_madariaga", "nombre": "Famaba - Madariaga", "km": 780},
    {"id": "famaba_mar_de_ajo", "nombre": "Famaba - Mar de Ajo", "km": 750},
    {"id": "famaba_mar_de_ajo_madariaga", "nombre": "Famaba - Mar de Ajo - Madariaga", "km": 810},
    {"id": "famaba_mar_del_plata_madariaga", "nombre": "Famaba - Mar del Plata - Madariaga", "km": 920},
    {"id": "famaba_parana", "nombre": "Famaba - Parana", "km": 1100},
    {"id": "famaba_rosario", "nombre": "Famaba - Rosario", "km": 720},
    {"id": "famaba_santa_fe", "nombre": "Famaba - Santa Fe", "km": 1060},
    {"id": "famaba_santa_fe_parana", "nombre": "Famaba - Sante Fe - Parana", "km": 1150},
    {"id": "rosario_santa_fe_parana", "nombre": "Rosario - Santa Fe - Parana", "km": 1150},
    {"id": "santa_fe_parana", "nombre": "Sanfe - Parana", "km": 1150},
]


FERIADOS_BASE = [
    {"fecha": "2026-01-01", "dia": "Jueves", "motivo": "Ano Nuevo", "tipo": "Feriado inamovible"},
    {"fecha": "2026-02-16", "dia": "Lunes", "motivo": "Carnaval", "tipo": "Feriado inamovible"},
    {"fecha": "2026-02-17", "dia": "Martes", "motivo": "Carnaval", "tipo": "Feriado inamovible"},
    {"fecha": "2026-03-23", "dia": "Lunes", "motivo": "Dia no laborable con fines turisticos", "tipo": "Turistico"},
    {"fecha": "2026-03-24", "dia": "Martes", "motivo": "Dia Nacional de la Memoria por la Verdad y la Justicia", "tipo": "Feriado inamovible"},
    {"fecha": "2026-04-02", "dia": "Jueves", "motivo": "Dia del Veterano y de los Caidos en la Guerra de Malvinas", "tipo": "Feriado inamovible"},
    {"fecha": "2026-04-03", "dia": "Viernes", "motivo": "Viernes Santo", "tipo": "Feriado inamovible"},
    {"fecha": "2026-05-01", "dia": "Viernes", "motivo": "Dia del Trabajador", "tipo": "Feriado inamovible"},
    {"fecha": "2026-05-25", "dia": "Lunes", "motivo": "Dia de la Revolucion de Mayo", "tipo": "Feriado inamovible"},
    {"fecha": "2026-06-15", "dia": "Lunes", "motivo": "Paso a la Inmortalidad del Gral. Martin Miguel de Guemes", "tipo": "Feriado trasladable"},
    {"fecha": "2026-06-20", "dia": "Sabado", "motivo": "Paso a la Inmortalidad del Gral. Manuel Belgrano", "tipo": "Feriado inamovible"},
    {"fecha": "2026-07-09", "dia": "Jueves", "motivo": "Dia de la Independencia", "tipo": "Feriado inamovible"},
    {"fecha": "2026-07-10", "dia": "Viernes", "motivo": "Dia no laborable con fines turisticos", "tipo": "Turistico"},
    {"fecha": "2026-08-17", "dia": "Lunes", "motivo": "Paso a la Inmortalidad del Gral. Jose de San Martin", "tipo": "Feriado trasladable"},
    {"fecha": "2026-10-12", "dia": "Lunes", "motivo": "Dia del Respeto a la Diversidad Cultural", "tipo": "Feriado trasladable"},
    {"fecha": "2026-11-23", "dia": "Lunes", "motivo": "Dia de la Soberania Nacional", "tipo": "Feriado trasladable"},
    {"fecha": "2026-12-07", "dia": "Lunes", "motivo": "Dia no laborable con fines turisticos", "tipo": "Turistico"},
    {"fecha": "2026-12-08", "dia": "Martes", "motivo": "Inmaculada Concepcion de Maria", "tipo": "Feriado inamovible"},
    {"fecha": "2026-12-25", "dia": "Viernes", "motivo": "Navidad", "tipo": "Feriado inamovible"},
]


class ProrrateoKmRequest(BaseModel):
    destino_id: str = Field(..., min_length=1)
    fecha_inicio: datetime
    fecha_fin: datetime
    km_override: float | None = Field(default=None, ge=0)
    demora_sucursal_horas: float = Field(default=0, ge=0)
    feriados_extra: list[date] = Field(default_factory=list)


class HojasRutaBuscarRequest(BaseModel):
    hojas_ruta: list[str] = Field(default_factory=list)
    feriados_extra: list[date] = Field(default_factory=list)


class HojaRutaEditable(BaseModel):
    hojaruta: str
    destino: str = ""
    distanciadirecta: float | None = Field(default=None, ge=0)
    fesalida: datetime | None = None
    fevuelta: datetime | None = None
    legajo: str = ""
    nombre: str = ""
    nrodoc: str = ""
    demora_sucursal_horas: float = Field(default=0, ge=0)
    estado: str = ""


class FeriadoRequest(BaseModel):
    fecha: date
    motivo: str = Field(..., min_length=1, max_length=200)
    tipo: str = Field(default="Manual", min_length=1, max_length=80)
    activo: bool = True


class HojasRutaCalcularRequest(BaseModel):
    rows: list[HojaRutaEditable] = Field(default_factory=list)
    feriados_extra: list[date] = Field(default_factory=list)


class ExcelConsultaRequest(BaseModel):
    filename: str = "modelo.xlsx"
    content_b64: str = Field(..., min_length=1)


class ExcelExportRequest(ExcelConsultaRequest):
    rows: list[dict[str, Any]] = Field(default_factory=list)


class ExcelRecalcularRequest(BaseModel):
    rows: list[dict[str, Any]] = Field(default_factory=list)


def _destino(destino_id: str) -> dict[str, Any] | None:
    return next((item for item in DESTINOS_KM if item["id"] == destino_id), None)


def _holiday_lookup(extra: list[date], configured: list[dict[str, Any]] | None = None) -> dict[date, dict[str, str]]:
    source = configured if configured is not None else FERIADOS_BASE
    holidays = {date.fromisoformat(str(item["fecha"])): item for item in source}
    for value in extra:
        holidays.setdefault(
            value,
            {"fecha": value.isoformat(), "dia": "", "motivo": "Feriado agregado manualmente", "tipo": "Manual"},
        )
    return holidays


def _overlap_minutes(start: datetime, end: datetime, window_start: datetime, window_end: datetime) -> float:
    overlap_start = max(start, window_start)
    overlap_end = min(end, window_end)
    if overlap_end <= overlap_start:
        return 0.0
    return (overlap_end - overlap_start).total_seconds() / 60


def _merge_windows(windows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not windows:
        return []
    ordered = sorted(windows, key=lambda item: (item["desde"], item["hasta"]))
    merged = [ordered[0].copy()]
    for item in ordered[1:]:
        current = merged[-1]
        if item["desde"] <= current["hasta"]:
            if item["hasta"] > current["hasta"]:
                current["hasta"] = item["hasta"]
            if item["motivo"] not in current["motivo"]:
                current["motivo"] = f"{current['motivo']} + {item['motivo']}"
        else:
            merged.append(item.copy())
    return merged


def _premium_windows(start: datetime, end: datetime, holidays: dict[date, dict[str, str]]) -> list[dict[str, Any]]:
    windows: list[dict[str, Any]] = []
    first_day = start.date() - timedelta(days=2)
    last_day = end.date() + timedelta(days=1)
    day_count = (last_day - first_day).days + 1
    for offset in range(day_count):
        current = first_day + timedelta(days=offset)
        if current.weekday() == 5:
            windows.append(
                {
                    "desde": datetime.combine(current, time(13, 0)),
                    "hasta": datetime.combine(current + timedelta(days=2), time(0, 0)),
                    "motivo": "Fin de semana desde sabado 13:00",
                    "tipo": "fin_semana",
                }
            )
        holiday = holidays.get(current)
        if holiday:
            windows.append(
                {
                    "desde": datetime.combine(current, time.min),
                    "hasta": datetime.combine(current + timedelta(days=1), time.min),
                    "motivo": holiday["motivo"],
                    "tipo": "feriado",
                }
            )
    return _merge_windows(windows)


def _round_pay_percent(value: float) -> int:
    if value <= 0:
        return 0
    return min(100, int(math.ceil(value / 5) * 5))


def _calculate_prorrateo(
    km_tramo: float,
    fecha_inicio: datetime,
    fecha_fin: datetime,
    feriados_extra: list[date],
    demora_sucursal_horas: float = 0,
    configured_holidays: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if fecha_fin <= fecha_inicio:
        raise ValueError("El fin del viaje debe ser posterior al inicio.")
    duration_minutes = (fecha_fin - fecha_inicio).total_seconds() / 60
    if duration_minutes > 62 * 24 * 60:
        raise ValueError("El rango maximo inicial es de 62 dias.")
    holidays = _holiday_lookup(feriados_extra, configured_holidays)
    windows = _premium_windows(fecha_inicio, fecha_fin, holidays)
    velocidad_km_h = km_tramo / (duration_minutes / 60) if duration_minutes else 0
    detail = []
    premium_minutes = 0.0
    for window in windows:
        minutes = _overlap_minutes(fecha_inicio, fecha_fin, window["desde"], window["hasta"])
        if minutes <= 0:
            continue
        km = velocidad_km_h * (minutes / 60)
        premium_minutes += minutes
        detail.append(
            {
                "desde": max(fecha_inicio, window["desde"]).isoformat(timespec="minutes"),
                "hasta": min(fecha_fin, window["hasta"]).isoformat(timespec="minutes"),
                "tipo": window["tipo"],
                "motivo": window["motivo"],
                "minutos": minutes,
                "km": km,
                "extra_pct": 100,
            }
        )
    km_100_real = min(km_tramo, velocidad_km_h * (premium_minutes / 60))
    porcentaje_real = (km_100_real * 100 / km_tramo) if km_tramo else 0
    porcentaje_a_pagar = _round_pay_percent(porcentaje_real)
    km_a_pagar = km_tramo * porcentaje_a_pagar / 100
    km_normal = max(0.0, km_tramo - km_100_real)
    conceptos: list[dict[str, Any]] = []
    if duration_minutes >= 24 * 60:
        conceptos.append({"concepto": "Pernocte", "motivo": "El viaje dura 24 horas o más."})
    if demora_sucursal_horas > 2:
        conceptos.append({"concepto": "Torito por demora", "motivo": "La demora en sucursal supera las 2 horas."})
    return {
        "total": km_tramo,
        "normal": km_normal,
        "extra_50": 0.0,
        "extra_100_real": km_100_real,
        "extra_100": km_a_pagar,
        "porcentaje_real_100": porcentaje_real,
        "porcentaje_a_pagar": porcentaje_a_pagar,
        "duracion_minutos": duration_minutes,
        "velocidad_km_h": velocidad_km_h,
        "detalle": detail,
        "conceptos": conceptos,
    }


def _productive_db_local_only_enabled() -> bool:
    return os.getenv("PRODUCTIVE_DB_LOCAL_ONLY", "0").strip().lower() in {"1", "true", "yes", "si"}


async def _load_holidays() -> list[dict[str, Any]]:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await db.execute(
            """
            CREATE TABLE IF NOT EXISTS herramientas_feriados (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT NOT NULL UNIQUE,
                motivo TEXT NOT NULL,
                tipo TEXT NOT NULL DEFAULT 'Manual',
                activo INTEGER NOT NULL DEFAULT 1,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        async with db.execute("SELECT COUNT(*) FROM herramientas_feriados") as cur:
            count = (await cur.fetchone())[0]
        if not count:
            await db.executemany(
                "INSERT OR IGNORE INTO herramientas_feriados (fecha, motivo, tipo) VALUES (?, ?, ?)",
                [(item["fecha"], item["motivo"], item["tipo"]) for item in FERIADOS_BASE],
            )
            await db.commit()
        async with db.execute(
            "SELECT id, fecha, motivo, tipo, activo FROM herramientas_feriados WHERE activo = 1 ORDER BY fecha"
        ) as cur:
            return [dict(row) for row in await cur.fetchall()]


async def _ensure_holiday_table() -> None:
    await _load_holidays()


def _ensure_java_helper_compiled() -> None:
    javac_bin = os.getenv(
        "PRODUCTIVE_DB_JAVAC_BIN",
        r"C:\Program Files\Android\openjdk\jdk-21.0.8\bin\javac.exe",
    ).strip()
    if not Path(javac_bin).exists():
        raise RuntimeError(f"No se encontro javac para consultar Oracle: {javac_bin}")
    if not JAVA_HELPER_SRC.exists():
        raise RuntimeError(f"No se encontro el helper JDBC de Oracle: {JAVA_HELPER_SRC}")

    JAVA_BUILD_DIR.mkdir(parents=True, exist_ok=True)
    class_file = JAVA_BUILD_DIR / "OracleProductividadQuery.class"
    if class_file.exists() and class_file.stat().st_mtime >= JAVA_HELPER_SRC.stat().st_mtime:
        return
    result = subprocess.run(
        [javac_bin, "-d", str(JAVA_BUILD_DIR), str(JAVA_HELPER_SRC)],
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"No se pudo compilar el helper JDBC de Oracle. STDERR: {result.stderr.strip() or result.stdout.strip()}")


def _query_oracle_raw_sql(sql: str) -> list[dict[str, Any]]:
    if _productive_db_local_only_enabled():
        raise RuntimeError("La BD productiva Oracle esta temporalmente bloqueada por configuracion.")

    user = os.getenv("PRODUCTIVE_DB_USER", "").strip()
    password = os.getenv("PRODUCTIVE_DB_PASSWORD", "").strip()
    host = os.getenv("PRODUCTIVE_DB_HOST", "").strip()
    port = os.getenv("PRODUCTIVE_DB_PORT", "1521").strip()
    service_name = os.getenv("PRODUCTIVE_DB_SERVICE_NAME", "").strip()
    java_bin = os.getenv(
        "PRODUCTIVE_DB_JAVA_BIN",
        r"C:\Users\207189\AppData\Local\DBeaver\jre\bin\java.exe",
    ).strip()
    ojdbc_jar = os.getenv(
        "PRODUCTIVE_DB_OJDBC_JAR",
        r"C:\Users\207189\AppData\Roaming\DBeaverData\drivers\maven\maven-central\com.oracle.database.jdbc\ojdbc11-23.2.0.0.jar",
    ).strip()

    if not all([user, password, host, service_name]):
        raise RuntimeError("Faltan variables JDBC PRODUCTIVE_DB_USER, PRODUCTIVE_DB_PASSWORD, PRODUCTIVE_DB_HOST o PRODUCTIVE_DB_SERVICE_NAME.")
    if not Path(java_bin).exists():
        raise RuntimeError(f"No se encontro Java para consultar Oracle: {java_bin}")
    if not Path(ojdbc_jar).exists():
        raise RuntimeError(f"No se encontro el driver JDBC Oracle: {ojdbc_jar}")

    _ensure_java_helper_compiled()
    jdbc_url = f"jdbc:oracle:thin:@//{host}:{port}/{service_name}"
    classpath = os.pathsep.join([str(JAVA_BUILD_DIR), ojdbc_jar])
    env = os.environ.copy()
    env["VIGIA_ORACLE_SQL_B64"] = base64.b64encode(sql.encode("utf-8")).decode("ascii")
    command = [
        java_bin,
        "-cp",
        classpath,
        "OracleProductividadQuery",
        jdbc_url,
        user,
        password,
        "2000-01-01 00:00:00",
        "2000-01-01 00:00:00",
        "raw_sql_env",
    ]
    result = subprocess.run(
        command,
        capture_output=True,
        text=True,
        check=False,
        env=env,
        timeout=int(os.getenv("PRODUCTIVE_DB_JDBC_TIMEOUT_SECONDS", "300")),
    )
    if result.returncode != 0:
        raise RuntimeError(f"No se pudo consultar Oracle via JDBC. STDERR: {result.stderr.strip() or result.stdout.strip()}")
    try:
        data = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"La respuesta JDBC no fue JSON valido. Salida: {result.stdout[:300]}") from exc
    return data if isinstance(data, list) else []


def _clean_hojas_ruta(values: list[str]) -> list[str]:
    found: list[str] = []
    for value in values:
        found.extend(re.findall(r"\d+", str(value or "")))
    unique = list(dict.fromkeys(item.lstrip("0") or "0" for item in found))
    if len(unique) > 200:
        raise HTTPException(status_code=400, detail="El maximo por consulta es 200 hojas de ruta.")
    return unique


def _build_hojas_ruta_sql(hojas_ruta: list[str]) -> str:
    in_list = ", ".join(hojas_ruta)
    return f"""
        WITH F810VIAJ_SOURCE AS (
            SELECT HOJARUTA, FESALIDA, FEVUELTA, DDNICOND
            FROM F810VIAJ
            UNION
            SELECT HOJARUTA, FESALIDA, FEVUELTA, DDNICOND
            FROM F810VIAJ_HIST
        )
        SELECT
            TO_CHAR(A.CODIGO) AS HOJARUTA,
            MAX(B.DISTANCIADIRECTA) AS DISTANCIADIRECTA,
            MIN(C.FESALIDA) AS FESALIDA,
            MAX(C.FEVUELTA) AS FEVUELTA,
            MAX(D.LEGAJO) AS LEGAJO,
            MAX(D.NOMBRE) AS NOMBRE,
            MAX(D.NRODOC) AS NRODOC,
            MAX(E.DESCRIPCION) AS DESTINO
        FROM TR_VIAJE A
        JOIN TR_TIEMPO_DE_VIAJE B
          ON (
            (A.CODIGODESITIODESTINO = B.CODIGODESITIO1 AND B.CODIGODESITIO2 = 93)
            OR
            (A.CODIGODESITIODESTINO = B.CODIGODESITIO2 AND B.CODIGODESITIO1 = 93)
          )
        JOIN F810VIAJ_SOURCE C
          ON C.HOJARUTA = TO_CHAR(A.CODIGO)
        LEFT JOIN ACTIVE_EMPLOYEE D
          ON D.NRODOC = C.DDNICOND
        JOIN TR_SITIO E
          ON E.CODIGO = A.CODIGODESITIODESTINO
        WHERE A.CODIGO IN ({in_list})
        GROUP BY TO_CHAR(A.CODIGO)
        ORDER BY TO_NUMBER(TO_CHAR(A.CODIGO))
    """


def _build_excel_hojas_ruta_sql(hojas_ruta: list[str]) -> str:
    in_list = ", ".join(f"'{item}'" for item in hojas_ruta)
    return f"""
        WITH F922_SOURCE AS (
            SELECT HOJARUTA, CLUGENTR, FESALIDA, FEVUELTA, DDNICOND, HORLLEGA, HORASALI
            FROM F922TRAF
            WHERE HOJARUTA IN ({in_list})
            UNION
            SELECT HOJARUTA, CLUGENTR, FESALIDA, FEVUELTA, DDNICOND, HORLLEGA, HORASALI
            FROM F922TRAF_HIST
            WHERE HOJARUTA IN ({in_list})
        )
        SELECT
            TO_CHAR(C.HOJARUTA) AS HOJARUTA,
            C.CLUGENTR AS SUCURSAL,
            C.FESALIDA AS FESALIDA,
            C.FEVUELTA AS FEVUELTA,
            NULL AS LEGAJO,
            D.NOMBRE AS NOMBRE,
            D.NRODOC AS NRODOC,
            C.HORLLEGA AS LLEGADA_SUCURSAL,
            C.HORASALI AS SALIDA_SUCURSAL
        FROM F922_SOURCE C
        LEFT JOIN ACTIVE_EMPLOYEE D ON D.NRODOC = C.DDNICOND
        ORDER BY TO_NUMBER(C.HOJARUTA), C.HORLLEGA
    """


def _excel_cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        if value.time() == time.min:
            return value.date().isoformat()
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    return value


def _read_excel_rows(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    sheet = workbook.worksheets[0]
    headers = [str(cell.value or "") for cell in sheet[1]]
    required = {"Fecha viaj", "Legajo", "Cantidad", "Num.Hoj.Ru"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"Faltan columnas obligatorias: {', '.join(missing)}")
    positions = {name: index for index, name in enumerate(headers)}
    rows = []
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in cells):
            continue
        route = str(cells[positions["Num.Hoj.Ru"]] or "").strip()
        if not route:
            continue
        rows.append({
            "excel_row": row_number,
            "fecha_viaje": _excel_cell_value(cells[positions["Fecha viaj"]]),
            "hora_viaje": _excel_cell_value(cells[positions.get("Hora viaje", 1)]),
            "excel_legajo": str(cells[positions["Legajo"]] or "").strip(),
            "cantidad": float(cells[positions["Cantidad"]] or 0),
            "hojaruta": route,
            "original": [_excel_cell_value(value) for value in cells],
        })
    return headers, rows


def _oracle_duration_minutes(start: Any, end: Any, reference: datetime | None = None) -> float:
    parsed_start = _parse_oracle_datetime(start)
    parsed_end = _parse_oracle_datetime(end)
    if parsed_start and parsed_end:
        if parsed_end < parsed_start:
            parsed_end += timedelta(days=1)
        return max(0.0, (parsed_end - parsed_start).total_seconds() / 60)
    def parse_clock(value: Any) -> time | None:
        if isinstance(value, time):
            return value
        text = str(value or "").strip()
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(text[:8], fmt).time()
            except ValueError:
                pass
        return None
    clock_start, clock_end = parse_clock(start), parse_clock(end)
    if not clock_start or not clock_end:
        return 0.0
    base = reference or datetime.combine(date.today(), time.min)
    left = datetime.combine(base.date(), clock_start)
    right = datetime.combine(base.date(), clock_end)
    if right < left:
        right += timedelta(days=1)
    return (right - left).total_seconds() / 60


def _parse_edit_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    return _parse_oracle_datetime(value)


async def _consolidate_excel_rows(excel_rows: list[dict[str, Any]], oracle_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_route: dict[str, list[dict[str, Any]]] = {}
    for row in oracle_rows:
        by_route.setdefault(str(_row_value(row, "HOJARUTA") or "").strip(), []).append(row)
    configured = await _load_holidays()
    result = []
    for route, source_rows in _group_excel_routes(excel_rows).items():
        source = source_rows[0]
        raw = by_route.get(route, [])
        normalized = [_normalize_oracle_excel_row(row) for row in raw]
        salida = min((row["fesalida"] for row in normalized if row["fesalida"]), default=None)
        vuelta = max((row["fevuelta"] for row in normalized if row["fevuelta"]), default=None)
        chosen = next((row for row in normalized if row["legajo"] or row["nombre"] or row["nrodoc"]), {})
        delay = max((row["demora_minutos"] for row in normalized), default=0.0)
        pernocte = bool(salida and vuelta and (vuelta - salida).total_seconds() >= 24 * 3600)
        km100 = 0.0
        if salida and vuelta:
            calc = _calculate_prorrateo(float(source["cantidad"]), salida, vuelta, [], 0, configured)
            km100 = float(calc["extra_100"])
        complete = bool(raw and salida and vuelta and chosen.get("nombre") and chosen.get("nrodoc"))
        result.append({
            **source,
            "fesalida": salida.isoformat(timespec="minutes") if salida else "",
            "fevuelta": vuelta.isoformat(timespec="minutes") if vuelta else "",
            "horas_totales_viaje": round((vuelta - salida).total_seconds() / 3600, 2) if salida and vuelta else 0,
            "legajo": chosen.get("legajo") or source.get("excel_legajo", ""),
            "nombre": chosen.get("nombre", ""),
            "nrodoc": chosen.get("nrodoc", ""),
            "torito": delay > 120,
            "demora_max_minutos": round(delay, 2),
            "demora_horas": round(delay / 60, 2),
            "llegada_local": "",
            "salida_local": "",
            "sucursales": normalized,
            "pernocte": pernocte,
            "km_al_100": round(km100, 2),
            "estado": "Validado" if complete else "Incompleto",
        })
    return result


def _group_excel_routes(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(row["hojaruta"], []).append(row)
    return grouped


def _normalize_oracle_excel_row(row: dict[str, Any]) -> dict[str, Any]:
    salida = _parse_oracle_datetime(_row_value(row, "FESALIDA"))
    vuelta = _parse_oracle_datetime(_row_value(row, "FEVUELTA"))
    reference = salida or datetime.now()
    return {
        "sucursal": str(_row_value(row, "SUCURSAL") or "").strip(),
        "fesalida": salida,
        "fevuelta": vuelta,
        "legajo": str(_row_value(row, "LEGAJO") or "").strip(),
        "nombre": str(_row_value(row, "NOMBRE") or "").strip(),
        "nrodoc": str(_row_value(row, "NRODOC") or "").strip(),
        "demora_minutos": _oracle_duration_minutes(_row_value(row, "LLEGADA_SUCURSAL"), _row_value(row, "SALIDA_SUCURSAL"), reference),
        "demora_horas": round(_oracle_duration_minutes(_row_value(row, "LLEGADA_SUCURSAL"), _row_value(row, "SALIDA_SUCURSAL"), reference) / 60, 2),
        "llegada_sucursal": str(_row_value(row, "LLEGADA_SUCURSAL") or "").strip(),
        "salida_sucursal": str(_row_value(row, "SALIDA_SUCURSAL") or "").strip(),
    }


def _row_value(row: dict[str, Any], key: str) -> Any:
    return row.get(key) if key in row else row.get(key.lower())


def _parse_oracle_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _normalize_hoja_row(row: dict[str, Any]) -> dict[str, Any]:
    salida = _parse_oracle_datetime(_row_value(row, "FESALIDA"))
    vuelta = _parse_oracle_datetime(_row_value(row, "FEVUELTA"))
    return {
        "hojaruta": str(_row_value(row, "HOJARUTA") or "").strip(),
        "destino": str(_row_value(row, "DESTINO") or "").strip(),
        "distanciadirecta": float(_row_value(row, "DISTANCIADIRECTA") or 0),
        "fesalida": salida.isoformat(timespec="minutes") if salida else "",
        "fevuelta": vuelta.isoformat(timespec="minutes") if vuelta else "",
        "legajo": str(_row_value(row, "LEGAJO") or "").strip(),
        "nombre": str(_row_value(row, "NOMBRE") or "").strip(),
        "nrodoc": str(_row_value(row, "NRODOC") or "").strip(),
        "demora_sucursal_horas": 0,
        "estado": "Datos encontrados.",
    }


async def _calculate_hoja_editable(row: HojaRutaEditable, feriados_extra: list[date]) -> dict[str, Any]:
    base = {
        "hojaruta": row.hojaruta,
        "destino": row.destino,
        "distanciadirecta": row.distanciadirecta,
        "fesalida": row.fesalida.isoformat(timespec="minutes") if row.fesalida else "",
        "fevuelta": row.fevuelta.isoformat(timespec="minutes") if row.fevuelta else "",
        "legajo": row.legajo,
        "nombre": row.nombre,
        "nrodoc": row.nrodoc,
        "demora_sucursal_horas": row.demora_sucursal_horas,
        "estado": row.estado or "",
        "totales": {},
        "detalle": [],
        "conceptos": [],
    }
    if not row.distanciadirecta or not row.fesalida or not row.fevuelta:
        base["estado"] = row.estado or "Faltan datos para calcular."
        return base
    try:
        result = _calculate_prorrateo(
            float(row.distanciadirecta),
            row.fesalida,
            row.fevuelta,
            feriados_extra,
            row.demora_sucursal_horas,
            await _load_holidays(),
        )
    except ValueError as exc:
        base["estado"] = str(exc)
        return base
    base["totales"] = {key: value for key, value in result.items() if key not in {"detalle", "conceptos"}}
    base["detalle"] = result["detalle"]
    base["conceptos"] = result["conceptos"]
    base["estado"] = "Calculado."
    return base


@router.get("/km/destinos")
async def list_km_destinos():
    return {"destinos": DESTINOS_KM, "feriados": await _load_holidays()}


@router.post("/excel/consultar")
async def consultar_excel(req: ExcelConsultaRequest):
    try:
        content = base64.b64decode(req.content_b64)
        _headers, excel_rows = _read_excel_rows(content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {exc}") from exc
    routes = sorted({row["hojaruta"] for row in excel_rows}, key=lambda value: int(value) if value.isdigit() else value)
    sql = _build_excel_hojas_ruta_sql(routes)
    try:
        oracle_rows = _query_oracle_raw_sql(sql)
        consolidated = await _consolidate_excel_rows(excel_rows, oracle_rows)
    except Exception as exc:
        logger.exception("No se pudo consultar el Excel contra Oracle.")
        raise HTTPException(status_code=502, detail=f"No se pudo consultar Oracle: {exc}") from exc
    return {"filename": req.filename, "rows": consolidated, "total": len(consolidated)}


@router.post("/excel/leer")
async def leer_excel(req: ExcelConsultaRequest):
    try:
        content = base64.b64decode(req.content_b64)
        headers, excel_rows = _read_excel_rows(content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {exc}") from exc
    rows = []
    for row in _group_excel_routes(excel_rows).values():
        item = dict(row[0])
        item.update({
            "fesalida": "",
            "fevuelta": "",
            "legajo": item.get("excel_legajo", ""),
            "nombre": "",
            "nrodoc": "",
            "torito": False,
            "llegada_local": "",
            "salida_local": "",
            "demora_horas": 0,
            "pernocte": False,
            "km_al_100": 0,
            "estado": "Pendiente de consulta Oracle.",
        })
        rows.append(item)
    return {"filename": req.filename, "headers": headers, "rows": rows, "total": len(rows)}


@router.post("/excel/recalcular")
async def recalcular_excel(req: ExcelRecalcularRequest):
    holidays = await _load_holidays()
    recalculated = []
    for row in req.rows:
        item = dict(row)
        start = _parse_oracle_datetime(item.get("fesalida"))
        end = _parse_oracle_datetime(item.get("fevuelta"))
        llegada_local = _parse_edit_datetime(item.get("llegada_local"))
        salida_local = _parse_edit_datetime(item.get("salida_local"))
        demora_minutos = _oracle_duration_minutes(llegada_local, salida_local)
        item["torito"] = demora_minutos > 120
        item["demora_horas"] = round(demora_minutos / 60, 2)
        if start and end and end > start:
            calc = _calculate_prorrateo(float(item.get("cantidad") or 0), start, end, [], 0, holidays)
            item["km_al_100"] = round(float(calc["extra_100"]), 2)
            item["pernocte"] = (end - start).total_seconds() >= 24 * 3600
            item["horas_totales_viaje"] = round((end - start).total_seconds() / 3600, 2)
            item["estado"] = "Validado manualmente." if item.get("nombre") and item.get("nrodoc") else "Datos incompletos; completar manualmente."
        else:
            item["km_al_100"] = 0
            item["pernocte"] = False
            item["horas_totales_viaje"] = 0
            item["estado"] = "Datos incompletos; completar manualmente."
        recalculated.append(item)
    return {"rows": recalculated, "total": len(recalculated)}


@router.post("/excel/exportar.xlsx")
async def exportar_excel(req: ExcelExportRequest):
    try:
        content = base64.b64decode(req.content_b64)
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook.worksheets[0]
        headers = [str(cell.value or "") for cell in sheet[1]]
        positions = {name: index + 1 for index, name in enumerate(headers)}
        route_col = positions.get("Num.Hoj.Ru")
        if not route_col:
            raise ValueError("No existe la columna Num.Hoj.Ru")
        extra_headers = ["KM al 100", "Pernocte", "Torito", "DNI", "Nombre del chofer"]
        start_col = sheet.max_column + 1
        for offset, header in enumerate(extra_headers):
            sheet.cell(row=1, column=start_col + offset, value=header)
        by_route = {str(row.get("hojaruta") or "").strip(): row for row in req.rows}
        for row_number in range(2, sheet.max_row + 1):
            route = str(sheet.cell(row=row_number, column=route_col).value or "").strip()
            if not route:
                continue
            row = by_route.get(route, {})
            sheet.cell(row=row_number, column=start_col, value=row.get("km_al_100", ""))
            sheet.cell(row=row_number, column=start_col + 1, value=1 if row.get("pernocte") else "")
            sheet.cell(row=row_number, column=start_col + 2, value=1 if row.get("torito") else "")
            sheet.cell(row=row_number, column=start_col + 3, value=row.get("nrodoc", ""))
            sheet.cell(row=row_number, column=start_col + 4, value=row.get("nombre", ""))
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo exportar el Excel: {exc}") from exc
    filename = Path(req.filename or "modelo.xlsx").stem + "_validado.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/km/feriados")
async def list_feriados():
    return {"feriados": await _load_holidays()}


@router.post("/km/feriados")
async def create_feriado(req: FeriadoRequest):
    await _ensure_holiday_table()
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute(
                "INSERT INTO herramientas_feriados (fecha, motivo, tipo, activo) VALUES (?, ?, ?, ?)",
                (req.fecha.isoformat(), req.motivo.strip(), req.tipo.strip(), int(req.activo)),
            )
            await db.commit()
    except aiosqlite.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="Ya existe un feriado para esa fecha.") from exc
    return {"feriados": await _load_holidays()}


@router.put("/km/feriados/{feriado_id}")
async def update_feriado(feriado_id: int, req: FeriadoRequest):
    await _ensure_holiday_table()
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "UPDATE herramientas_feriados SET fecha = ?, motivo = ?, tipo = ?, activo = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (req.fecha.isoformat(), req.motivo.strip(), req.tipo.strip(), int(req.activo), feriado_id),
        )
        await db.commit()
    if cur.rowcount == 0:
        raise HTTPException(status_code=404, detail="Feriado no encontrado.")
    return {"feriados": await _load_holidays()}


@router.delete("/km/feriados/{feriado_id}")
async def delete_feriado(feriado_id: int):
    await _ensure_holiday_table()
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("DELETE FROM herramientas_feriados WHERE id = ?", (feriado_id,))
        await db.commit()
    if cur.rowcount == 0:
        raise HTTPException(status_code=404, detail="Feriado no encontrado.")
    return {"feriados": await _load_holidays()}


@router.post("/km/prorratear")
async def prorratear_km(req: ProrrateoKmRequest):
    destino = _destino(req.destino_id)
    if not destino:
        raise HTTPException(status_code=400, detail="Destino no configurado.")
    km_tramo = float(req.km_override if req.km_override is not None else float(destino["km"]) / 2)
    try:
        calculated = _calculate_prorrateo(
            km_tramo,
            req.fecha_inicio,
            req.fecha_fin,
            req.feriados_extra,
            req.demora_sucursal_horas,
            await _load_holidays(),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    totals = {key: value for key, value in calculated.items() if key not in {"detalle", "conceptos"}}
    return {
        "destino": destino,
        "km_tramo": km_tramo,
        "fecha_inicio": req.fecha_inicio.isoformat(timespec="minutes"),
        "fecha_fin": req.fecha_fin.isoformat(timespec="minutes"),
        "dias": len({req.fecha_inicio.date() + timedelta(days=offset) for offset in range((req.fecha_fin.date() - req.fecha_inicio.date()).days + 1)}),
        "totales": totals,
        "detalle": calculated["detalle"],
        "criterio": "Version inicial segun Excel: se calcula la proporcion del tramo que cae en ventanas al 100%. Fin de semana desde sabado 13:00 hasta lunes 00:00 y feriados completos al 100%. El 50% queda fuera.",
        "conceptos": calculated["conceptos"],
    }


@router.post("/km/hojas-ruta/buscar")
async def buscar_hojas_ruta(req: HojasRutaBuscarRequest):
    hojas = _clean_hojas_ruta(req.hojas_ruta)
    if not hojas:
        raise HTTPException(status_code=400, detail="Pegue al menos una hoja de ruta.")
    sql = _build_hojas_ruta_sql(hojas)
    try:
        oracle_rows = _query_oracle_raw_sql(sql)
    except Exception as exc:
        logger.exception("No se pudieron consultar hojas de ruta en Oracle.")
        raise HTTPException(status_code=502, detail=f"No se pudo consultar Oracle: {exc}") from exc

    by_hoja = {_normalize_hoja_row(row)["hojaruta"]: _normalize_hoja_row(row) for row in oracle_rows}
    rows: list[dict[str, Any]] = []
    for hoja in hojas:
        item = by_hoja.get(hoja)
        if not item:
            rows.append(
                {
                    "hojaruta": hoja,
                    "distanciadirecta": None,
                    "fesalida": "",
                    "fevuelta": "",
                    "legajo": "",
                    "nombre": "",
                    "nrodoc": "",
                    "demora_sucursal_horas": 0,
                    "estado": "No encontrada en Oracle.",
                    "totales": {},
                    "detalle": [],
                    "conceptos": [],
                }
            )
            continue
        editable = HojaRutaEditable(**item)
        rows.append(await _calculate_hoja_editable(editable, req.feriados_extra))
    return {"rows": rows, "encontradas": sum(1 for row in rows if row["estado"] == "Calculado."), "total": len(rows)}


@router.post("/km/hojas-ruta/calcular")
async def calcular_hojas_ruta(req: HojasRutaCalcularRequest):
    rows = [await _calculate_hoja_editable(row, req.feriados_extra) for row in req.rows]
    return {"rows": rows, "total": len(rows)}


@router.post("/km/hojas-ruta/exportar.csv")
async def exportar_hojas_ruta(req: HojasRutaCalcularRequest):
    rows = [await _calculate_hoja_editable(row, req.feriados_extra) for row in req.rows]
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", lineterminator="\n")
    writer.writerow(
        [
            "Hoja de ruta",
            "Destino",
            "KM tramo",
            "Salida",
            "Vuelta",
            "Legajo",
            "Nombre",
            "Documento",
            "% real 100",
            "% pago 100",
            "KM paga 100",
            "KM normal",
            "Demora sucursal (h)",
            "Conceptos",
            "Estado",
        ]
    )
    for row in rows:
        totals = row.get("totales") or {}
        writer.writerow(
            [
                row.get("hojaruta", ""),
                row.get("destino", ""),
                row.get("distanciadirecta", ""),
                row.get("fesalida", ""),
                row.get("fevuelta", ""),
                row.get("legajo", ""),
                row.get("nombre", ""),
                row.get("nrodoc", ""),
                round(float(totals.get("porcentaje_real_100") or 0), 2),
                int(totals.get("porcentaje_a_pagar") or 0),
                round(float(totals.get("extra_100") or 0), 2),
                round(float(totals.get("normal") or 0), 2),
                row.get("demora_sucursal_horas", 0),
                ", ".join(item.get("concepto", "") for item in row.get("conceptos", [])),
                row.get("estado", ""),
            ]
        )
    filename = f"prorrateador_km_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter(["\ufeff" + output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
