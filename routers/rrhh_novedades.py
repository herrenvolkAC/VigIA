"""
VigIA - Novedades CD / RRHH.

Importa archivos fuente del proceso RRHH y expone reportes operativos.
RRHH solo deja los Excels; VigIA normaliza, guarda y filtra por permisos.
"""
from __future__ import annotations

import json
import hashlib
import re
import shutil
import sqlite3
import subprocess
import tempfile
import asyncio
import csv
import logging
import os
import time as time_module
from contextlib import suppress
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

import aiosqlite
from fastapi import APIRouter, HTTPException, Query, Request
from pydantic import BaseModel

from db.schema import DB_PATH
from routers.auth_local import current_auth

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

try:
    import xlrd
except Exception:  # pragma: no cover
    xlrd = None


router = APIRouter(prefix="/api/rrhh", tags=["rrhh-novedades"])
logger = logging.getLogger("vigia.rrhh_novedades")

SOURCE_ROOT = Path(__file__).parent.parent / "Docs" / "Panel_Choferes" / "PROCESADOS"
FULL_ACCESS_ROLES = {"admin", "rrhh"}
IMPORT_ROLES = {"admin", "rrhh"}
RRHH_SCOPE_MODULE = "novedades_cd"
_rrhh_import_lock: asyncio.Lock | None = None
_rrhh_monitor_task: asyncio.Task | None = None
_rrhh_monitor_stop: asyncio.Event | None = None
_rrhh_monitor_seen: dict[str, tuple[int, int, float]] = {}
REAL_SANCIONES = (
    "Amonestación",
    "Anotaciones Especiales",
    "Llamada de Atención Escrita",
    "Suspensión",
)


def _scheduled_day_sql(alias: str = "a") -> str:
    return (
        f"(COALESCE({alias}.ausentismo_contabiliza,0) = 1 "
        f"OR COALESCE({alias}.horas_teoricas,0) > 0 "
        f"OR TRIM(COALESCE({alias}.ingreso_teorico,'')) <> '')"
    )


class ImportFolderRequest(BaseModel):
    folder_path: str | None = None
    batch_key: str | None = None
    force: bool = False


def _env_bool(name: str, default: bool = False) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "yes", "si", "on"}


def _env_int(name: str, default: int, *, minimum: int | None = None, maximum: int | None = None) -> int:
    try:
        value = int(os.getenv(name, str(default)).strip())
    except (TypeError, ValueError):
        value = default
    if minimum is not None:
        value = max(minimum, value)
    if maximum is not None:
        value = min(maximum, value)
    return value


def _env_path(name: str, default: Path | None = None) -> Path | None:
    raw = os.getenv(name, "").strip()
    if not raw:
        return default
    return Path(raw).expanduser()


def _francos_fecha_inicial() -> str:
    raw = os.getenv("RRHH_FRANCOS_FECHA_INICIAL", "2026-05-26").strip()
    return _to_date(raw) or "2026-05-26"


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _norm(value: Any) -> str:
    return str(value or "").strip()


def _norm_legajo(value: Any) -> str:
    text = _norm(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+(?:\.0)?", text):
        text = text.split(".", 1)[0]
    stripped = text.lstrip("0")
    return stripped or "0"


def _norm_codigo(value: Any) -> str:
    text = _norm(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+(?:\.0)?", text):
        text = text.split(".", 1)[0]
    stripped = text.lstrip("0")
    return stripped or "0"


def _norm_key(value: Any) -> str:
    value = _norm(value).lower()
    value = (
        value.replace("á", "a").replace("é", "e").replace("í", "i")
        .replace("ó", "o").replace("ú", "u").replace("ñ", "n")
        .replace("ä", "a")
    )
    return re.sub(r"[^a-z0-9]+", "_", value).strip("_")


def _to_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = _norm(value).replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _to_hours(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, datetime):
        return value.hour + value.minute / 60 + value.second / 3600
    if isinstance(value, time):
        return value.hour + value.minute / 60 + value.second / 3600
    if isinstance(value, (int, float)):
        return float(value)
    text = _norm(value).replace(",", ".")
    if not text:
        return 0.0
    match = re.match(r"^(-?\d{1,4}):(\d{2})(?::(\d{2}))?$", text)
    if match:
        sign = -1 if match.group(1).startswith("-") else 1
        hours = abs(int(match.group(1)))
        minutes = int(match.group(2))
        seconds = int(match.group(3) or 0)
        return sign * (hours + minutes / 60 + seconds / 3600)
    return _to_float(text)


def _pick_value(data: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = data.get(key)
        if value is not None and value != "":
            return value
    return None


def _to_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _norm(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            pass
    return text


def _to_datetime(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return datetime.combine(value, time()).strftime("%Y-%m-%d %H:%M:%S")
    text = _norm(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    return text


def _to_time_text(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    text = _norm(value)
    return text or None


def _time_minus_minutes(value: Any, minutes: Any) -> str:
    text = _norm(value)
    if not text:
        return ""
    try:
        mins = int(round(float(minutes or 0)))
    except (TypeError, ValueError):
        mins = 0
    if mins <= 0:
        return text[:5] if re.match(r"^\d{1,2}:\d{2}", text) else text
    try:
        base = datetime.strptime(text[:8] if len(text) >= 8 else text[:5], "%H:%M:%S" if len(text) >= 8 else "%H:%M")
    except ValueError:
        return ""
    return (base - timedelta(minutes=mins)).strftime("%H:%M")


def _is_schedule_text(value: Any) -> bool:
    text = _norm(value)
    if not text:
        return False
    return bool(re.search(r"\bhs\b|\d+\s*\+\s*\d+|libre", text, re.IGNORECASE))


def _split_activity_schedule_code(data: dict[str, Any]) -> tuple[str, str, str]:
    first = _norm(data.get("horario"))
    second = _norm(data.get("pausa"))
    if _is_schedule_text(first) and not _is_schedule_text(second):
        return second, first, second
    if _is_schedule_text(second) and not _is_schedule_text(first):
        return first, second, first
    return first, first if _is_schedule_text(first) else second if _is_schedule_text(second) else "", second


def _parse_schedule(value: Any) -> dict[str, Any]:
    text = re.sub(r"\s+", " ", _norm(value)).strip()
    if not text or re.search(r"\blibre\b", text, re.IGNORECASE):
        return {"descripcion": text, "inicio": "", "fin": "", "horas": 0.0}
    time_match = re.search(r"(\d{1,2}):(\d{2})", text)
    hour_only_match = None if time_match else re.search(r"\b(\d{1,2})\s*hs\b", text, re.IGNORECASE)
    if not time_match and not hour_only_match:
        return {"descripcion": text, "inicio": "", "fin": "", "horas": 0.0}
    duration_part = text[:time_match.start()] if time_match else text[:hour_only_match.start()]
    numbers = [float(num.replace(",", ".")) for num in re.findall(r"\d+(?:[,.]\d+)?", duration_part)]
    hours = sum(numbers) if "+" in duration_part else (numbers[0] if numbers else 0.0)
    start_minutes = int((time_match or hour_only_match).group(1)) * 60 + (int(time_match.group(2)) if time_match else 0)
    if hours <= 0:
        return {
            "descripcion": text,
            "inicio": f"{start_minutes // 60:02d}:{start_minutes % 60:02d}",
            "fin": "",
            "horas": 0.0,
        }
    end_minutes = (start_minutes + int(round(hours * 60))) % (24 * 60)
    return {
        "descripcion": text,
        "inicio": f"{start_minutes // 60:02d}:{start_minutes % 60:02d}",
        "fin": f"{end_minutes // 60:02d}:{end_minutes % 60:02d}",
        "horas": hours,
    }


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return value


def _raw_json(headers: list[str], row: list[Any]) -> str:
    return json.dumps(
        {headers[i]: _json_value(row[i]) if i < len(row) else None for i in range(len(headers))},
        ensure_ascii=False,
    )


def _unique_headers(values: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    out = []
    for idx, value in enumerate(values):
        key = _norm_key(value) or f"col_{idx + 1}"
        count = seen.get(key, 0)
        seen[key] = count + 1
        out.append(key if count == 0 else f"{key}_{count + 1}")
    return out


def _is_gerencia(row: dict[str, Any]) -> int:
    haystack = " ".join(
        _norm(row.get(key))
        for key in (
            "desc_funcion",
            "desc_posicion",
            "desc_unidad_organizativa",
            "desc_sector_generico",
            "grupo_profesional",
        )
    ).upper()
    return 1 if any(token in haystack for token in ("GEREN", "JEF", "JEFE", "JEFAT")) else 0


PERSONA_MASTER_FIELDS = (
    "legajo", "nombre", "empresa", "division_personal", "sucursal",
    "unidad_organizativa", "desc_unidad_organizativa", "sector_generico",
    "desc_sector_generico", "clave_funcion", "desc_funcion", "posicion",
    "desc_posicion", "grupo_personal", "desc_grupo_personal", "area_personal",
    "desc_area_personal", "area_nomina", "clase_contrato", "desc_tipo_contrato",
    "regla_plan_horario", "jubilado", "centro_coste", "fecha_ingreso",
    "fecha_baja", "antiguedad_anios", "antiguedad_meses", "antiguedad_dias",
    "proveedor", "razon_social", "es_gerencia", "raw_json",
)


def _persona_hash(item: dict[str, Any]) -> str:
    data = {key: item.get(key) for key in PERSONA_MASTER_FIELDS if key != "raw_json"}
    encoded = json.dumps(data, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _read_xlsx(path: Path, sheet: str | None = None) -> list[list[Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl no esta disponible para leer archivos .xlsx.")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.worksheets[0]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _read_zipped_excel_with_legacy_extension(path: Path, sheet: str | None = None) -> list[list[Any]]:
    with tempfile.TemporaryDirectory() as tmp:
        copied = Path(tmp) / f"{path.stem}.xlsx"
        shutil.copy2(path, copied)
        return _read_xlsx(copied, sheet)


def _read_xls(path: Path, sheet: str | None = None) -> list[list[Any]]:
    with open(path, "rb") as fh:
        sample = fh.read(512)
        if sample[:2] == b"PK":
            return _read_zipped_excel_with_legacy_extension(path, sheet)

    if xlrd is not None:
        try:
            book = xlrd.open_workbook(str(path))
            ws = book.sheet_by_name(sheet) if sheet else book.sheet_by_index(0)
            rows: list[list[Any]] = []
            for r in range(ws.nrows):
                out = []
                for c in range(ws.ncols):
                    cell = ws.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        out.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                    else:
                        out.append(cell.value)
                rows.append(out)
            return rows
        except Exception:
            if b"\t" in sample or b";" in sample:
                return _read_delimited_legacy_xls(path, sample)
            raise

    converter = shutil.which("soffice") or shutil.which("libreoffice")
    if converter:
        with tempfile.TemporaryDirectory() as tmp:
            subprocess.run(
                [converter, "--headless", "--convert-to", "xlsx", "--outdir", tmp, str(path)],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
            )
            converted = Path(tmp) / f"{path.stem}.xlsx"
            if not converted.exists():
                raise RuntimeError(f"No se pudo convertir {path.name} a .xlsx.")
            return _read_xlsx(converted, sheet)

    raise RuntimeError(
        f"{path.name} es .xls legado. Instala xlrd>=2.0.1 o LibreOffice para importarlo."
    )


def _read_delimited_legacy_xls(path: Path, sample: bytes | None = None) -> list[list[Any]]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "cp1252", "latin1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("latin1", errors="replace")
    delimiter = "\t" if (sample or raw[:512]).count(b"\t") >= (sample or raw[:512]).count(b";") else ";"
    return [row for row in csv.reader(text.splitlines(), delimiter=delimiter)]


def _read_workbook_rows(path: Path, sheet: str | None = None) -> list[list[Any]]:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _read_xls(path, sheet)
    return _read_xlsx(path, sheet)


def _find_latest_folder() -> Path:
    if not SOURCE_ROOT.exists():
        raise RuntimeError(f"No existe la carpeta {SOURCE_ROOT}.")
    folders = [p for p in SOURCE_ROOT.iterdir() if p.is_dir()]
    if not folders:
        raise RuntimeError(f"No hay carpetas mensuales en {SOURCE_ROOT}.")
    return max(folders, key=lambda p: p.stat().st_mtime)


def _pick_files(folder: Path, patterns: list[str], exclude: list[str] | None = None) -> list[Path]:
    exclude = [item.lower() for item in (exclude or [])]
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(folder.glob(pattern))
    unique = {
        p.resolve(): p for p in candidates
        if p.is_file()
        and p.suffix.lower() in {".xlsx", ".xls"}
        and not p.name.startswith("~$")
        and not any(token in p.name.lower() for token in exclude)
    }
    return sorted(unique.values(), key=lambda p: (p.stat().st_mtime, p.name.lower()))


def _pick_file(folder: Path, patterns: list[str], exclude: list[str] | None = None) -> Path | None:
    candidates = _pick_files(folder, patterns, exclude)
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _excel_files(folder: Path) -> list[Path]:
    return sorted(
        [
            p for p in folder.iterdir()
            if p.is_file()
            and p.suffix.lower() in {".xlsx", ".xls"}
            and not p.name.startswith("~$")
        ],
        key=lambda p: (p.stat().st_mtime, p.name.lower()),
    )


def _detect_file_kind(path: Path) -> str | None:
    try:
        rows = _read_workbook_rows(path)
    except Exception:
        return None
    if not rows:
        return None
    header_idx = 1 if len(rows) > 1 and _norm_key(rows[0][0] if rows[0] else "") != "ubicacion" else 0
    headers = set(_unique_headers(rows[header_idx]))
    if {"legajo", "fecha"}.issubset(headers) and ("empleado" in headers or "sector" in headers):
        return "actividad"
    if {"legajo", "nombre_del_empleado_o_candidato"}.issubset(headers) or {"legajo", "desc_funcion", "desc_posicion"}.issubset(headers):
        return "legajero"
    if {"legajo_apellido_y_nombre", "fecha_fichada"}.issubset(headers):
        return "fichadas"
    if {"legajo", "apellido_y_nombre"}.issubset(headers) and ("causa_sancion" in headers or "descripcion_causa" in headers):
        return "sanciones"
    legajo_headers = {"legajo", "numero_de_personal", "nro_personal", "numero_personal"}
    saldo_headers = {"saldo", "saldo_francos", "francos", "dias", "dias_franco", "cuenta_corriente", "resto_global"}
    if headers.intersection(legajo_headers) and headers.intersection(saldo_headers):
        return "francos"
    return None


def _stringify_files(files: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in files.items():
        if isinstance(value, list):
            if value:
                out[key] = [str(item) for item in value]
        elif value is not None:
            out[key] = str(value)
    return out


def _as_paths(files: dict[str, Any], key: str) -> list[Path]:
    value = files.get(key)
    if value is None:
        return []
    if isinstance(value, list):
        return [Path(item) for item in value]
    return [Path(value)]


def _detect_files(folder: Path) -> dict[str, Any]:
    actividad_files = _pick_files(folder, ["*Actividad*.xlsx", "*ACTIVIDAD*.XLS", "*ACTIVIDAD*.xls"], ["full_base"])
    sanciones_files = _pick_files(folder, ["*SANCIONES*.xlsx", "*SANCIONES*.XLS", "*Sanciones*.xls"])
    files = {
        "actividad": actividad_files[-1] if actividad_files else None,
        "actividad_files": actividad_files,
        "fichadas": _pick_file(folder, ["*Fichadas*.xlsx", "*FICHADAS*.XLS", "*FICHADAS*.xls"]),
        "legajero": _pick_file(folder, ["*Legajero*.xlsx", "*LEGAJERO*.XLS", "*LEGAJERO*.xls"]),
        "sanciones": _pick_file(folder, ["*SANCIONES*.xlsx", "*SANCIONES*.XLS", "*Sanciones*.xls"]),
        "sanciones_files": sanciones_files,
        "codigos_ausentismo": _pick_file(folder, ["*Codigos_Ausentismo*.xlsx", "*Codigos_Ausentismo*.xls"]),
        "francos": _pick_file(folder, ["*Francos*.xlsx", "*FRANCOS*.XLS", "*Francos*.xls", "*francos*.xlsx", "*francos*.xls"]),
    }

    known_paths = {p.resolve() for value in files.values() for p in (value if isinstance(value, list) else [value]) if p}
    for path in _excel_files(folder):
        if path.resolve() in known_paths:
            continue
        kind = _detect_file_kind(path)
        if kind == "actividad":
            files["actividad_files"].append(path)
            files["actividad"] = path
        elif kind == "legajero" and not files.get("legajero"):
            files["legajero"] = path
        elif kind == "fichadas" and not files.get("fichadas"):
            files["fichadas"] = path
        elif kind == "sanciones":
            files["sanciones_files"].append(path)
            files["sanciones"] = path
        elif kind == "francos" and not files.get("francos"):
            files["francos"] = path

    if not any(files.get(key) for key in ("actividad", "legajero", "fichadas", "sanciones", "codigos_ausentismo", "francos")):
        raise RuntimeError(f"No se detectaron archivos RRHH compatibles en {folder}.")
    return _stringify_files(files)


def _row_dict(headers: list[str], row: list[Any]) -> dict[str, Any]:
    return {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}


def _sync_personas_master(cur: sqlite3.Cursor, batch_id: int, items: list[dict[str, Any]]) -> dict[str, int]:
    now = _now()
    stats = {"altas": 0, "modificaciones": 0, "bajas": 0, "reactivaciones": 0, "sin_cambios": 0}
    current_by_legajo = {item["legajo"]: item for item in items}
    seen_legajos = set(current_by_legajo)

    cur.execute("SELECT * FROM rrhh_personas")
    existing = {row["legajo"]: dict(row) for row in cur.fetchall()}

    for legajo, item in current_by_legajo.items():
        item_hash = _persona_hash(item)
        row = existing.get(legajo)
        if row is None:
            cur.execute(
                """
                INSERT INTO rrhh_personas (
                    legajo, nombre, empresa, division_personal, sucursal,
                    unidad_organizativa, desc_unidad_organizativa, sector_generico,
                    desc_sector_generico, clave_funcion, desc_funcion, posicion,
                    desc_posicion, grupo_personal, desc_grupo_personal, area_personal,
                    desc_area_personal, area_nomina, clase_contrato, desc_tipo_contrato,
                    regla_plan_horario, jubilado, centro_coste, fecha_ingreso, fecha_baja,
                    antiguedad_anios, antiguedad_meses, antiguedad_dias, proveedor,
                    razon_social, es_gerencia, active, first_seen_batch_id,
                    last_seen_batch_id, last_changed_batch_id, first_seen_at,
                    last_seen_at, data_hash, change_count, raw_json, updated_at
                ) VALUES (
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                )
                """,
                tuple(item.get(key) for key in PERSONA_MASTER_FIELDS[:-1])
                + (1, batch_id, batch_id, batch_id, now, now, item_hash, 1, item["raw_json"], now),
            )
            cur.execute(
                """
                INSERT INTO rrhh_personas_changes (batch_id, legajo, change_type, old_json, new_json)
                VALUES (?, ?, 'alta', NULL, ?)
                """,
                (batch_id, legajo, item["raw_json"]),
            )
            stats["altas"] += 1
            continue

        if row.get("data_hash") == item_hash and int(row.get("active") or 0) == 1:
            cur.execute(
                "UPDATE rrhh_personas SET last_seen_batch_id = ?, last_seen_at = ?, updated_at = ? WHERE legajo = ?",
                (batch_id, now, now, legajo),
            )
            stats["sin_cambios"] += 1
            continue

        change_type = "reactivacion" if int(row.get("active") or 0) == 0 else "modificacion"
        cur.execute(
            """
            UPDATE rrhh_personas
            SET nombre = ?, empresa = ?, division_personal = ?, sucursal = ?,
                unidad_organizativa = ?, desc_unidad_organizativa = ?, sector_generico = ?,
                desc_sector_generico = ?, clave_funcion = ?, desc_funcion = ?, posicion = ?,
                desc_posicion = ?, grupo_personal = ?, desc_grupo_personal = ?, area_personal = ?,
                desc_area_personal = ?, area_nomina = ?, clase_contrato = ?, desc_tipo_contrato = ?,
                regla_plan_horario = ?, jubilado = ?, centro_coste = ?, fecha_ingreso = ?,
                fecha_baja = ?, antiguedad_anios = ?, antiguedad_meses = ?, antiguedad_dias = ?,
                proveedor = ?, razon_social = ?, es_gerencia = ?, active = 1,
                last_seen_batch_id = ?, last_changed_batch_id = ?, last_seen_at = ?,
                deactivated_at = NULL, data_hash = ?, change_count = change_count + 1,
                raw_json = ?, updated_at = ?
            WHERE legajo = ?
            """,
            tuple(item.get(key) for key in PERSONA_MASTER_FIELDS[1:-1])
            + (batch_id, batch_id, now, item_hash, item["raw_json"], now, legajo),
        )
        cur.execute(
            """
            INSERT INTO rrhh_personas_changes (batch_id, legajo, change_type, old_json, new_json)
            VALUES (?, ?, ?, ?, ?)
            """,
            (batch_id, legajo, change_type, row.get("raw_json"), item["raw_json"]),
        )
        stats["reactivaciones" if change_type == "reactivacion" else "modificaciones"] += 1

    for legajo, row in existing.items():
        if legajo in seen_legajos or int(row.get("active") or 0) == 0:
            continue
        cur.execute(
            """
            UPDATE rrhh_personas
            SET active = 0, last_changed_batch_id = ?, deactivated_at = ?, updated_at = ?,
                change_count = change_count + 1
            WHERE legajo = ?
            """,
            (batch_id, now, now, legajo),
        )
        cur.execute(
            """
            INSERT INTO rrhh_personas_changes (batch_id, legajo, change_type, old_json, new_json)
            VALUES (?, ?, 'baja', ?, NULL)
            """,
            (batch_id, legajo, row.get("raw_json")),
        )
        stats["bajas"] += 1

    return stats


def _import_legajero(cur: sqlite3.Cursor, batch_id: int, path: Path) -> dict[str, int]:
    rows = _read_workbook_rows(path)
    headers = _unique_headers(rows[0])
    inserted = 0
    gerencia_by_legajo: dict[str, int] = {}
    payload = []
    master_items: list[dict[str, Any]] = []
    for row in rows[1:]:
        data = _row_dict(headers, row)
        legajo = _norm_legajo(data.get("legajo"))
        if not legajo:
            continue
        item = {
            "legajo": legajo,
            "nombre": _norm(data.get("nombre_del_empleado_o_candidato")),
            "empresa": _norm(data.get("empresa")),
            "division_personal": _norm(data.get("division_de_personal")),
            "sucursal": _norm(data.get("suc")),
            "unidad_organizativa": _norm(data.get("unidad_organizativa")),
            "desc_unidad_organizativa": _norm(data.get("desc_unid_organiz")),
            "sector_generico": _norm(data.get("sector_generico")),
            "desc_sector_generico": _norm(data.get("descrip_sector_generico")),
            "clave_funcion": _norm(data.get("clave_de_funcion")),
            "desc_funcion": _norm(data.get("desc_funcion")),
            "posicion": _norm(data.get("posicion")),
            "desc_posicion": _norm(data.get("desc_posicion")),
            "grupo_personal": _norm(data.get("grupo_de_personal")),
            "desc_grupo_personal": _norm(data.get("desc_grupo_de_personal")),
            "area_personal": _norm(data.get("area_de_personal")),
            "desc_area_personal": _norm(data.get("desc_area_de_personal")),
            "area_nomina": _norm(data.get("area_nomina")),
            "clase_contrato": _norm(data.get("clase_contrato")),
            "desc_tipo_contrato": _norm(data.get("desc_tipo_contrato")),
            "regla_plan_horario": _norm(data.get("regla_plan_hor_trab")),
            "jubilado": _norm(data.get("jubilado")),
            "centro_coste": _norm(data.get("centro_de_coste")),
            "fecha_ingreso": _to_date(data.get("fecha_de_ingreso")),
            "fecha_baja": _to_date(data.get("fecha_de_baja")),
            "antiguedad_anios": _to_float(data.get("antiguedad_anos")),
            "antiguedad_meses": _to_float(data.get("antiguedad_meses")),
            "antiguedad_dias": _to_float(data.get("antiguedad_dias")),
            "proveedor": _norm(data.get("proveedor")),
            "razon_social": _norm(data.get("razon_social")),
            "raw_json": _raw_json(headers, row),
        }
        es_gerencia = _is_gerencia(item)
        item["es_gerencia"] = es_gerencia
        gerencia_by_legajo[legajo] = es_gerencia
        master_items.append(item)
        payload.append((
            batch_id, item["legajo"], item["nombre"], item["empresa"], item["division_personal"],
            item["sucursal"], item["unidad_organizativa"], item["desc_unidad_organizativa"],
            item["sector_generico"], item["desc_sector_generico"], item["clave_funcion"],
            item["desc_funcion"], item["posicion"], item["desc_posicion"], item["grupo_personal"],
            item["desc_grupo_personal"], item["area_personal"], item["desc_area_personal"],
            item["fecha_ingreso"], item["fecha_baja"], item["proveedor"], item["razon_social"],
            item["raw_json"], es_gerencia,
        ))
        inserted += 1
    cur.executemany(
        """
        INSERT OR REPLACE INTO rrhh_legajero (
            batch_id, legajo, nombre, empresa, division_personal, sucursal,
            unidad_organizativa, desc_unidad_organizativa, sector_generico,
            desc_sector_generico, clave_funcion, desc_funcion, posicion,
            desc_posicion, grupo_personal, desc_grupo_personal, area_personal,
            desc_area_personal, fecha_ingreso, fecha_baja, proveedor, razon_social,
            raw_json, es_gerencia
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    master_stats = _sync_personas_master(cur, batch_id, master_items)
    return {"inserted": inserted, "gerencia_map": gerencia_by_legajo, **master_stats}


def _import_actividad(
    cur: sqlite3.Cursor,
    batch_id: int,
    path: Path,
    gerencia: dict[str, int],
    uo_sector_map: dict[str, str] | None = None,
    seen: set[tuple[str, str]] | None = None,
    codes: dict[str, dict[str, str]] | None = None,
    no_count_patterns: list[str] | None = None,
) -> int:
    rows = _read_workbook_rows(path)
    headers = _unique_headers(rows[0])
    payload = []
    seen = seen if seen is not None else set()
    uo_sector_map = uo_sector_map or {}
    codes = codes or {}
    no_count_patterns = no_count_patterns or []
    for row in rows[1:]:
        data = _row_dict(headers, row)
        legajo = _norm_legajo(data.get("legajo"))
        fecha = _to_date(data.get("fecha"))
        if not legajo or not fecha:
            continue
        dedupe_key = (legajo, fecha)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        sector = _norm(data.get("sector"))
        aus_pres = _norm(data.get("aus_pres"))
        motivo = _norm(data.get("motivo"))
        horario, horario_teorico_raw, pausa = _split_activity_schedule_code(data)
        horario_info = _parse_schedule(horario_teorico_raw)
        aus_info = _classify_ausentismo(aus_pres, motivo, horario_teorico_raw, codes, no_count_patterns)
        payload.append((
            batch_id, legajo, _norm(data.get("empleado")), fecha,
            _norm(data.get("division")), _norm(data.get("subdivision")), uo_sector_map.get(sector, sector),
            _norm(data.get("grupo_profesional")), _norm(data.get("area_de_personal")),
            _norm(data.get("dia")), pausa, horario,
            horario_info["descripcion"], horario_info["inicio"], horario_info["fin"], horario_info["horas"],
            aus_pres, aus_info["codigo_norm"], aus_info["tratamiento"], aus_info["tipo"],
            aus_info["clasificacion"], aus_info["contabiliza"], aus_info["regla"],
            motivo, _norm(data.get("comida")),
            _to_time_text(data.get("entrada")), _to_time_text(data.get("p1_ini")),
            _to_time_text(data.get("p1_fin")), _norm(data.get("mas")), _to_time_text(data.get("salida")),
            _to_hours(data.get("hs_trab")), _to_hours(_pick_value(data, "hs_ext_realiz", "hs_50")),
            _to_hours(_pick_value(data, "hs_50_autorizadas", "hs_extras_autoriz")),
            _to_hours(data.get("hs_100")),
            _to_hours(data.get("recargo_50")), _to_hours(data.get("recargo_100")),
            _to_hours(data.get("rec_noct")), _to_hours(data.get("hs_fer")),
            _to_hours(data.get("tarde")), _to_float(data.get("viajes_equiv")),
            gerencia.get(legajo, 0), _raw_json(headers, row),
        ))
    cur.executemany(
        """
        INSERT INTO rrhh_actividad_diaria (
            batch_id, legajo, empleado, fecha, division, subdivision, sector,
            grupo_profesional, area_personal, dia, pausa, horario,
            horario_teorico, ingreso_teorico, salida_teorica, horas_teoricas,
            aus_pres_codigo, aus_pres_codigo_norm, ausentismo_tratamiento, ausentismo_tipo,
            ausentismo_clasificacion, ausentismo_contabiliza, ausentismo_regla,
            motivo, comida, entrada, p1_ini, p1_fin, mas, salida, hs_trab,
            hs_ext_realiz, hs_50_autorizadas, hs_100, recargo_50, recargo_100,
            rec_noct, hs_fer, tarde, viajes_equiv, es_gerencia, raw_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def _unidad_sector_map(cur: sqlite3.Cursor, batch_id: int) -> dict[str, str]:
    cur.execute(
        """
        SELECT unidad_organizativa, desc_sector_generico, COUNT(*) c
        FROM rrhh_legajero
        WHERE batch_id = ?
          AND TRIM(COALESCE(unidad_organizativa,'')) <> ''
          AND TRIM(COALESCE(desc_sector_generico,'')) <> ''
        GROUP BY unidad_organizativa, desc_sector_generico
        ORDER BY unidad_organizativa, c DESC
        """,
        (batch_id,),
    )
    out: dict[str, str] = {}
    for row in cur.fetchall():
        out.setdefault(row["unidad_organizativa"], row["desc_sector_generico"])
    return out


def _load_ausentismo_classifier(cur: sqlite3.Cursor, batch_id: int) -> tuple[dict[str, dict[str, str]], list[str]]:
    cur.execute(
        """
        SELECT codigo_normalizado, descripcion, tratamiento, tipo_ausentismo
        FROM rrhh_codigos_ausentismo_maestro
        WHERE active = 1
        """
    )
    codes = {
        row["codigo_normalizado"]: {
            "descripcion": row["descripcion"] or "",
            "tratamiento": row["tratamiento"] or "",
            "tipo": row["tipo_ausentismo"] or "",
        }
        for row in cur.fetchall()
        if row["codigo_normalizado"]
    }
    if not codes:
        cur.execute(
            """
            SELECT codigo_normalizado, descripcion, tratamiento, tipo_ausentismo
            FROM rrhh_codigos_ausentismo
            WHERE batch_id = ?
            """,
            (batch_id,),
        )
        codes = {
            row["codigo_normalizado"]: {
                "descripcion": row["descripcion"] or "",
                "tratamiento": row["tratamiento"] or "",
                "tipo": row["tipo_ausentismo"] or "",
            }
            for row in cur.fetchall()
            if row["codigo_normalizado"]
        }
    cur.execute(
        """
        SELECT patron
        FROM rrhh_ausentismo_reglas
        WHERE (batch_id = ? OR batch_id IS NULL)
          AND active = 1
          AND clasificacion = 'NO CONSIDERAR'
        """,
        (batch_id,),
    )
    patterns = [row["patron"].upper() for row in cur.fetchall() if row["patron"]]
    for pattern in ("VACACION", "VACAC", "FRANCO", "DESCANSO", "LIBRE", "FERIADO NO", "NO CONVOCADO"):
        if pattern not in patterns:
            patterns.append(pattern)
    return codes, patterns


def _classify_ausentismo(codigo: Any, motivo: Any, horario: Any, codes: dict[str, dict[str, str]], no_count_patterns: list[str]) -> dict[str, Any]:
    codigo_text = _norm(codigo)
    codigo_norm = _norm_codigo(codigo_text)
    motivo_text = _norm(motivo)
    motivo_haystack = motivo_text.upper()
    horario_haystack = _norm(horario).upper()
    if not codigo_norm or codigo_norm == "0":
        return {
            "codigo_norm": codigo_norm,
            "tratamiento": "",
            "tipo": "",
            "clasificacion": "SIN NOVEDAD",
            "contabiliza": 0,
            "regla": "",
        }
    motivo_no_count = any(pattern and pattern in motivo_haystack for pattern in no_count_patterns)
    horario_no_count = any(pattern and pattern in horario_haystack for pattern in no_count_patterns)
    if codigo_norm == "666" or motivo_no_count or (horario_no_count and not motivo_text):
        return {
            "codigo_norm": codigo_norm,
            "tratamiento": "",
            "tipo": "",
            "clasificacion": "NO CONSIDERAR",
            "contabiliza": 0,
            "regla": "codigo_666" if codigo_norm == "666" else ("motivo_no_considerar" if motivo_no_count else "horario_no_considerar"),
        }
    info = codes.get(codigo_norm)
    if info:
        tipo = info["tipo"] or "SIN TIPO"
        return {
            "codigo_norm": codigo_norm,
            "tratamiento": info["tratamiento"],
            "tipo": tipo,
            "clasificacion": tipo,
            "contabiliza": 1 if tipo in {"CONTROLADO", "NO CONTROLADO"} else 0,
            "regla": "codigo_maestro",
        }
    return {
        "codigo_norm": codigo_norm,
        "tratamiento": "",
        "tipo": "",
        "clasificacion": "SIN CLASIFICAR",
        "contabiliza": 0,
        "regla": "codigo_no_maestro",
    }


def _import_actividad_files(cur: sqlite3.Cursor, batch_id: int, paths: list[Path], gerencia: dict[str, int]) -> int:
    seen: set[tuple[str, str]] = set()
    uo_sector_map = _unidad_sector_map(cur, batch_id)
    codes, no_count_patterns = _load_ausentismo_classifier(cur, batch_id)
    total = 0
    for path in paths:
        total += _import_actividad(cur, batch_id, path, gerencia, uo_sector_map, seen, codes, no_count_patterns)
    return total


def _parse_legajo_nombre(value: Any) -> tuple[str, str]:
    text = _norm(value)
    if " - " in text:
        legajo, nombre = text.split(" - ", 1)
        return _norm_legajo(legajo), _norm(nombre)
    return _norm_legajo(text), ""


def _import_fichadas(cur: sqlite3.Cursor, batch_id: int, path: Path, gerencia: dict[str, int]) -> int:
    rows = _read_workbook_rows(path)
    header_idx = 1 if rows and _norm_key(rows[0][0] if rows[0] else "") != "ubicacion" else 0
    headers = _unique_headers(rows[header_idx])
    payload = []
    for row in rows[header_idx + 1:]:
        data = _row_dict(headers, row)
        legajo, nombre = _parse_legajo_nombre(data.get("legajo_apellido_y_nombre"))
        fh = _to_datetime(data.get("fecha_fichada"))
        fecha = _to_date(data.get("fecha_fichada"))
        if not legajo or not fh or not fecha:
            continue
        payload.append((
            batch_id, legajo, nombre, fh, fecha, fh[11:19],
            _norm(data.get("sentido")), _norm(data.get("ubicacion")),
            _norm(data.get("origen")), _norm(data.get("destino")),
            _norm(data.get("tipo_lectura")), gerencia.get(legajo, 0), _raw_json(headers, row),
        ))
    cur.executemany(
        """
        INSERT INTO rrhh_fichadas (
            batch_id, legajo, empleado, fecha_fichada, fecha, hora, sentido,
            ubicacion, origen, destino, tipo_lectura, es_gerencia, raw_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def _merge_sancion_detalle(left: Any, right: Any) -> str:
    first = _norm(left)
    second = _norm(right)
    if not first:
        merged = second
    elif not second:
        merged = first
    elif re.search(r"\d$", first) and re.match(r"^\d{1,2}:", second):
        merged = first + second
    elif re.search(r"[A-ZÁÉÍÓÚÑ]{1,4}$", first) and re.match(r"^[A-ZÁÉÍÓÚÑ]{2,}", second):
        merged = first + second
    elif re.search(r"[-/([{]$", first):
        merged = first + second
    else:
        merged = f"{first} {second}"
    merged = re.sub(r"(\d{1,2}:\d{2})([A-ZÁÉÍÓÚÑ])", r"\1 \2", merged)
    merged = re.sub(r"\bCO\s+MUNICO\b", "COMUNICO", merged, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", merged).strip()


def _import_sanciones(
    cur: sqlite3.Cursor,
    batch_id: int,
    path: Path,
    gerencia: dict[str, int],
    seen: set[tuple[str, str | None, str | None, str, str, str]] | None = None,
) -> int:
    rows = _read_workbook_rows(path)
    headers = _unique_headers(rows[0])
    payload = []
    seen = seen if seen is not None else set()
    for row in rows[1:]:
        data = _row_dict(headers, row)
        legajo = _norm_legajo(data.get("legajo"))
        if not legajo:
            continue
        detalle = _merge_sancion_detalle(data.get("detalle"), data.get("detalle_2"))
        inicio = _to_date(data.get("inicio"))
        fin = _to_date(data.get("fin"))
        cod = _norm(data.get("cod"))
        creacion = _to_date(data.get("creacion"))
        descripcion = _norm(data.get("descripcion"))
        dedupe_key = (legajo, inicio, fin, cod, creacion or "", descripcion)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        payload.append((
            batch_id, legajo, _norm(data.get("apellido_y_nombre")),
            inicio, fin, cod,
            creacion, descripcion, detalle,
            _norm(data.get("ausentismo")), _norm(data.get("desc_ausentismo")),
            _norm(data.get("causa_sancion")), _norm(data.get("descripcion_causa")),
            _norm(data.get("u_o_sancion")), _norm(data.get("descripcion_u_o_san")),
            gerencia.get(legajo, 0), _raw_json(headers, row),
        ))
    cur.executemany(
        """
        INSERT INTO rrhh_sanciones (
            batch_id, legajo, nombre, inicio, fin, cod, creacion, descripcion,
            detalle, ausentismo, desc_ausentismo, causa_sancion, descripcion_causa,
            unidad_organizativa, desc_unidad_organizativa, es_gerencia, raw_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        payload,
    )
    return len(payload)


def _import_sanciones_files(cur: sqlite3.Cursor, batch_id: int, paths: list[Path], gerencia: dict[str, int]) -> int:
    seen: set[tuple[str, str | None, str | None, str, str, str]] = set()
    total = 0
    for path in paths:
        total += _import_sanciones(cur, batch_id, path, gerencia, seen)
    return total


def _import_codigos(cur: sqlite3.Cursor, batch_id: int, path: Path) -> dict[str, int]:
    total = 0
    reglas = 0
    seen_codes: set[str] = set()
    for sheet in ("No Controlado", "Controlado"):
        rows = _read_workbook_rows(path, sheet)
        headers = _unique_headers(rows[0])
        payload = []
        master_payload = []
        for row in rows[1:]:
            data = _row_dict(headers, row)
            codigo = _norm(data.get("cod_ausentismo"))
            codigo_norm = _norm_codigo(codigo)
            if not codigo_norm:
                continue
            seen_codes.add(codigo_norm)
            descripcion = _norm(data.get("descripcion"))
            tratamiento = _norm(data.get("tratamiento"))
            tipo = _norm(data.get("tipo_ausentismo"))
            raw = _raw_json(headers, row)
            payload.append((
                batch_id, codigo, codigo_norm, descripcion, tratamiento, tipo, sheet, raw,
            ))
            master_payload.append((
                codigo_norm, codigo, descripcion, tratamiento, tipo, sheet,
                batch_id, batch_id, _now(), raw,
            ))
        cur.executemany(
            """
            INSERT OR REPLACE INTO rrhh_codigos_ausentismo (
                batch_id, codigo, codigo_normalizado, descripcion, tratamiento,
                tipo_ausentismo, source_sheet, raw_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            payload,
        )
        cur.executemany(
            """
            INSERT INTO rrhh_codigos_ausentismo_maestro (
                codigo_normalizado, codigo_original, descripcion, tratamiento,
                tipo_ausentismo, source_sheet, first_seen_batch_id,
                last_seen_batch_id, updated_at, raw_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(codigo_normalizado) DO UPDATE SET
                codigo_original = excluded.codigo_original,
                descripcion = excluded.descripcion,
                tratamiento = excluded.tratamiento,
                tipo_ausentismo = excluded.tipo_ausentismo,
                source_sheet = excluded.source_sheet,
                active = 1,
                last_seen_batch_id = excluded.last_seen_batch_id,
                updated_at = excluded.updated_at,
                raw_json = excluded.raw_json
            """,
            master_payload,
        )
        total += len(payload)
    if seen_codes:
        placeholders = ", ".join("?" for _ in seen_codes)
        cur.execute(
            f"UPDATE rrhh_codigos_ausentismo_maestro SET active = 0, updated_at = ? WHERE codigo_normalizado NOT IN ({placeholders})",
            (_now(), *seen_codes),
        )

    try:
        rows = _read_workbook_rows(path, "Notas")
    except Exception:
        rows = []
    if rows:
        headers = _unique_headers(rows[0])
        payload = []
        for row in rows[1:]:
            data = _row_dict(headers, row)
            raw_values = [_norm(value) for value in row if _norm(value)]
            if not raw_values:
                continue
            patron = _norm(data.get("tipo_ausentismo")) or (raw_values[1] if len(raw_values) > 1 else raw_values[0])
            if not patron:
                continue
            payload.append((
                batch_id, "MOTIVO_CONTIENE", patron.upper(), "NO CONSIDERAR",
                0, "Notas", _raw_json(headers, row),
            ))
        cur.executemany(
            """
            INSERT OR REPLACE INTO rrhh_ausentismo_reglas (
                batch_id, regla_tipo, patron, clasificacion, contabiliza, source_sheet, raw_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            payload,
        )
        reglas += len(payload)
    macro_payload = [
        (batch_id, "MOTIVO_CONTIENE", pattern, "NO CONSIDERAR", 0, "macro_default", "{}")
        for pattern in ("VACACION", "VACAC", "FRANCO", "DESCANSO", "LIBRE", "FERIADO NO", "NO CONVOCADO")
    ]
    macro_payload.append((batch_id, "CODIGO_IGUAL", "666", "NO CONSIDERAR", 0, "macro_default", "{}"))
    cur.executemany(
        """
        INSERT OR REPLACE INTO rrhh_ausentismo_reglas (
            batch_id, regla_tipo, patron, clasificacion, contabiliza, source_sheet, raw_json
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        macro_payload,
    )
    reglas += len(macro_payload)
    return {"codigos": total, "reglas": reglas}


def _first_present(data: dict[str, Any], keys: tuple[str, ...]) -> Any:
    for key in keys:
        value = data.get(key)
        if value is not None and value != "":
            return value
    return None


def _import_francos_inicial(cur: sqlite3.Cursor, path: Path, imported_by: str) -> int:
    rows = _read_workbook_rows(path)
    if not rows:
        return 0
    header_idx = 0
    for idx, row in enumerate(rows[:5]):
        keys = set(_unique_headers(row))
        legajo_headers = {"legajo", "numero_de_personal", "nro_personal", "numero_personal"}
        saldo_headers = {"saldo", "saldo_francos", "francos", "dias", "dias_franco", "cuenta_corriente", "resto_global"}
        if keys.intersection(legajo_headers) and keys.intersection(saldo_headers):
            header_idx = idx
            break
    headers = _unique_headers(rows[header_idx])
    payload = []
    now = _now()
    legajo_keys = ("legajo", "numero_de_personal", "nro_personal", "numero_personal")
    saldo_keys = ("saldo_francos", "saldo", "francos", "dias", "dias_franco", "cuenta_corriente", "saldo_actual", "resto_global")
    nombre_keys = ("nombre", "empleado", "apellido_y_nombre", "nombre_del_empleado_o_candidato")
    fecha_keys = ("fecha_corte", "fecha", "corte", "fecha_saldo")
    fecha_inicial = _francos_fecha_inicial()
    for row in rows[header_idx + 1:]:
        data = _row_dict(headers, row)
        legajo = _norm_legajo(_first_present(data, legajo_keys))
        if not legajo:
            continue
        payload.append((
            legajo,
            _norm(_first_present(data, nombre_keys)),
            _to_float(_first_present(data, saldo_keys)),
            _to_date(_first_present(data, fecha_keys)) or fecha_inicial,
            str(path),
            _raw_json(headers, row),
            imported_by,
            now,
        ))
    cur.executemany(
        """
        INSERT INTO rrhh_francos_inicial (
            legajo, nombre, saldo_inicial, fecha_corte, source_file,
            raw_json, imported_by, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(legajo) DO UPDATE SET
            nombre = excluded.nombre,
            saldo_inicial = excluded.saldo_inicial,
            fecha_corte = excluded.fecha_corte,
            source_file = excluded.source_file,
            raw_json = excluded.raw_json,
            imported_by = excluded.imported_by,
            updated_at = excluded.updated_at
        """,
        payload,
    )
    return len(payload)


def _cleanup_rrhh_orphans(cur: sqlite3.Cursor) -> None:
    for table in (
        "rrhh_legajero",
        "rrhh_actividad_diaria",
        "rrhh_fichadas",
        "rrhh_sanciones",
        "rrhh_codigos_ausentismo",
        "rrhh_ausentismo_reglas",
        "rrhh_personas_changes",
    ):
        cur.execute(
            f"DELETE FROM {table} WHERE batch_id NOT IN (SELECT batch_id FROM rrhh_import_batches)"
        )


def _latest_complete_batch_with_legajero(cur: sqlite3.Cursor, exclude_batch_id: int | None = None) -> int | None:
    params: list[Any] = []
    exclude_sql = ""
    if exclude_batch_id is not None:
        exclude_sql = "AND b.batch_id <> ?"
        params.append(exclude_batch_id)
    cur.execute(
        f"""
        SELECT b.batch_id
        FROM rrhh_import_batches b
        WHERE b.status = 'complete'
          {exclude_sql}
          AND EXISTS (SELECT 1 FROM rrhh_legajero l WHERE l.batch_id = b.batch_id)
        ORDER BY b.imported_at DESC, b.batch_id DESC
        LIMIT 1
        """,
        params,
    )
    row = cur.fetchone()
    return int(row["batch_id"]) if row else None


def _copy_legajero_from_batch(cur: sqlite3.Cursor, target_batch_id: int, source_batch_id: int) -> dict[str, Any]:
    cur.execute(
        """
        INSERT OR REPLACE INTO rrhh_legajero (
            batch_id, legajo, nombre, empresa, division_personal, sucursal,
            unidad_organizativa, desc_unidad_organizativa, sector_generico,
            desc_sector_generico, clave_funcion, desc_funcion, posicion,
            desc_posicion, grupo_personal, desc_grupo_personal, area_personal,
            desc_area_personal, fecha_ingreso, fecha_baja, proveedor, razon_social,
            raw_json, es_gerencia
        )
        SELECT ?, legajo, nombre, empresa, division_personal, sucursal,
            unidad_organizativa, desc_unidad_organizativa, sector_generico,
            desc_sector_generico, clave_funcion, desc_funcion, posicion,
            desc_posicion, grupo_personal, desc_grupo_personal, area_personal,
            desc_area_personal, fecha_ingreso, fecha_baja, proveedor, razon_social,
            raw_json, es_gerencia
        FROM rrhh_legajero
        WHERE batch_id = ?
        """,
        (target_batch_id, source_batch_id),
    )
    inserted = int(cur.rowcount or 0)
    cur.execute(
        "SELECT legajo, COALESCE(es_gerencia, 0) es_gerencia FROM rrhh_legajero WHERE batch_id = ?",
        (target_batch_id,),
    )
    gerencia_map = {row["legajo"]: int(row["es_gerencia"] or 0) for row in cur.fetchall()}
    return {
        "inserted": inserted,
        "gerencia_map": gerencia_map,
        "altas": 0,
        "modificaciones": 0,
        "bajas": 0,
        "reactivaciones": 0,
        "sin_cambios": inserted,
        "referencia_batch_id": source_batch_id,
    }


def _delete_rrhh_batch_children(cur: sqlite3.Cursor, batch_id: int) -> None:
    for column in ("first_seen_batch_id", "last_seen_batch_id", "last_changed_batch_id"):
        cur.execute(
            f"UPDATE rrhh_personas SET {column} = NULL WHERE {column} = ?",
            (batch_id,),
        )
    for column in ("first_seen_batch_id", "last_seen_batch_id"):
        cur.execute(
            f"UPDATE rrhh_codigos_ausentismo_maestro SET {column} = NULL WHERE {column} = ?",
            (batch_id,),
        )
    for table in (
        "rrhh_legajero",
        "rrhh_actividad_diaria",
        "rrhh_fichadas",
        "rrhh_sanciones",
        "rrhh_codigos_ausentismo",
        "rrhh_ausentismo_reglas",
        "rrhh_personas_changes",
    ):
        cur.execute(f"DELETE FROM {table} WHERE batch_id = ?", (batch_id,))


def _rrhh_quality_warnings(cur: sqlite3.Cursor, batch_id: int, limit: int = 25) -> dict[str, Any]:
    cur.execute(
        """
        SELECT legajo, COALESCE(NULLIF(TRIM(nombre), ''), 'Sin nombre') nombre,
               COALESCE(NULLIF(TRIM(desc_posicion), ''), NULLIF(TRIM(desc_funcion), ''), '') referencia
        FROM rrhh_legajero
        WHERE batch_id = ?
          AND (
            TRIM(COALESCE(desc_sector_generico, '')) = ''
            OR TRIM(COALESCE(desc_funcion, '')) = ''
          )
        ORDER BY legajo
        LIMIT ?
        """,
        (batch_id, limit),
    )
    legajero_rows = [dict(row) for row in cur.fetchall()]
    cur.execute(
        """
        SELECT COUNT(*) c
        FROM rrhh_legajero
        WHERE batch_id = ?
          AND (
            TRIM(COALESCE(desc_sector_generico, '')) = ''
            OR TRIM(COALESCE(desc_funcion, '')) = ''
          )
        """,
        (batch_id,),
    )
    legajero_count = int(cur.fetchone()["c"] or 0)
    cur.execute(
        """
        SELECT a.legajo,
               COALESCE(NULLIF(TRIM(l.nombre), ''), NULLIF(TRIM(a.empleado), ''), 'Sin nombre') nombre,
               COUNT(*) registros,
               SUM(CASE WHEN COALESCE(a.ausentismo_contabiliza, 0) = 1 THEN 1 ELSE 0 END) ausencias
        FROM rrhh_actividad_diaria a
        LEFT JOIN rrhh_legajero l ON l.batch_id = a.batch_id AND l.legajo = a.legajo
        WHERE a.batch_id = ?
          AND COALESCE(NULLIF(TRIM(l.desc_sector_generico), ''), NULLIF(TRIM(a.sector), '')) IS NULL
        GROUP BY a.legajo, nombre
        ORDER BY ausencias DESC, registros DESC, a.legajo
        LIMIT ?
        """,
        (batch_id, limit),
    )
    activity_rows = [dict(row) for row in cur.fetchall()]
    cur.execute(
        """
        SELECT COUNT(DISTINCT a.legajo) legajos, COUNT(*) registros
        FROM rrhh_actividad_diaria a
        LEFT JOIN rrhh_legajero l ON l.batch_id = a.batch_id AND l.legajo = a.legajo
        WHERE a.batch_id = ?
          AND COALESCE(NULLIF(TRIM(l.desc_sector_generico), ''), NULLIF(TRIM(a.sector), '')) IS NULL
        """,
        (batch_id,),
    )
    activity_count = dict(cur.fetchone())
    return {
        "legajero_sin_sector_o_cargo": legajero_count,
        "legajero_sin_sector_o_cargo_muestra": legajero_rows,
        "actividad_sin_sector_legajos": int(activity_count.get("legajos") or 0),
        "actividad_sin_sector_registros": int(activity_count.get("registros") or 0),
        "actividad_sin_sector_muestra": activity_rows,
    }


def _import_folder_sync(folder: Path, batch_key: str, imported_by: str, force: bool) -> dict[str, Any]:
    files = _detect_files(folder)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()
    try:
        _cleanup_rrhh_orphans(cur)
        cur.execute("SELECT batch_id FROM rrhh_import_batches WHERE batch_key = ?", (batch_key,))
        existing = cur.fetchone()
        if existing and not force:
            raise RuntimeError(f"El lote {batch_key} ya existe. Usar force=true para reimportar.")
        existing_batch_id = int(existing["batch_id"]) if existing else None
        reference_legajero_batch_id = None
        if not files.get("legajero"):
            reference_legajero_batch_id = _latest_complete_batch_with_legajero(cur, existing_batch_id)
        if existing:
            _delete_rrhh_batch_children(cur, existing_batch_id)
            cur.execute("DELETE FROM rrhh_import_batches WHERE batch_id = ?", (existing_batch_id,))
        cur.execute(
            """
            INSERT INTO rrhh_import_batches (batch_key, source_dir, imported_by, status, files_json)
            VALUES (?, ?, ?, 'running', ?)
            """,
            (batch_key, str(folder), imported_by, json.dumps(files, ensure_ascii=False)),
        )
        batch_id = int(cur.lastrowid)
        import_mode = "completo"
        if files.get("legajero"):
            legajero_info = _import_legajero(cur, batch_id, Path(files["legajero"]))
        elif reference_legajero_batch_id:
            legajero_info = _copy_legajero_from_batch(cur, batch_id, reference_legajero_batch_id)
            import_mode = "parcial_con_legajero_referencia"
        else:
            legajero_info = {
                "inserted": 0,
                "gerencia_map": {},
                "altas": 0,
                "modificaciones": 0,
                "bajas": 0,
                "reactivaciones": 0,
                "sin_cambios": 0,
                "referencia_batch_id": None,
            }
            import_mode = "parcial_sin_legajero"
        gerencia_map = legajero_info["gerencia_map"]
        codigos_info = _import_codigos(cur, batch_id, Path(files["codigos_ausentismo"])) if files.get("codigos_ausentismo") else {"codigos": 0, "reglas": 0}
        francos_inicial = _import_francos_inicial(cur, Path(files["francos"]), imported_by) if files.get("francos") else 0
        actividad_paths = _as_paths(files, "actividad_files")
        sanciones_paths = _as_paths(files, "sanciones_files")
        summary = {
            "modo_importacion": import_mode,
            "archivos_detectados": {
                "actividad": len(actividad_paths),
                "legajero": 1 if files.get("legajero") else 0,
                "legajero_referencia_batch_id": legajero_info.get("referencia_batch_id"),
                "fichadas": 1 if files.get("fichadas") else 0,
                "sanciones": len(sanciones_paths),
                "codigos_ausentismo": 1 if files.get("codigos_ausentismo") else 0,
                "francos": 1 if files.get("francos") else 0,
            },
            "legajero": legajero_info["inserted"],
            "legajero_altas": legajero_info["altas"],
            "legajero_modificaciones": legajero_info["modificaciones"],
            "legajero_bajas": legajero_info["bajas"],
            "legajero_reactivaciones": legajero_info["reactivaciones"],
            "legajero_sin_cambios": legajero_info["sin_cambios"],
            "codigos_ausentismo": codigos_info["codigos"],
            "reglas_ausentismo": codigos_info["reglas"],
            "francos_inicial": francos_inicial,
            "actividad": _import_actividad_files(cur, batch_id, actividad_paths, gerencia_map) if actividad_paths else 0,
            "actividad_archivos": len(actividad_paths),
            "fichadas": _import_fichadas(cur, batch_id, Path(files["fichadas"]), gerencia_map) if files.get("fichadas") else 0,
            "sanciones": _import_sanciones_files(cur, batch_id, sanciones_paths, gerencia_map) if sanciones_paths else 0,
            "sanciones_archivos": len(sanciones_paths),
            "gerencia_legajos": sum(1 for value in gerencia_map.values() if value),
        }
        warnings = _rrhh_quality_warnings(cur, batch_id)
        summary["advertencias_calidad"] = warnings
        cur.execute(
            """
            UPDATE rrhh_import_batches
            SET status = 'complete', summary_json = ?, error = NULL
            WHERE batch_id = ?
            """,
            (json.dumps(summary, ensure_ascii=False), batch_id),
        )
        conn.commit()
        return {"batch_id": batch_id, "batch_key": batch_key, "files": files, "summary": summary}
    except Exception as exc:
        conn.rollback()
        raise RuntimeError(str(exc)) from exc
    finally:
        conn.close()


def _get_import_lock() -> asyncio.Lock:
    global _rrhh_import_lock
    if _rrhh_import_lock is None:
        _rrhh_import_lock = asyncio.Lock()
    return _rrhh_import_lock


async def _import_folder_locked(folder: Path, batch_key: str, imported_by: str, force: bool) -> dict[str, Any]:
    async with _get_import_lock():
        return await asyncio.to_thread(_import_folder_sync, folder, batch_key, imported_by, force)


def _watch_inbox() -> Path | None:
    return _env_path("RRHH_WATCH_INBOX", _env_path("RRHH_WATCH_FOLDER", SOURCE_ROOT))


def _watch_imported_dir(inbox: Path) -> Path:
    return _env_path("RRHH_WATCH_IMPORTED", inbox / "IMPORTADOS") or inbox / "IMPORTADOS"


def _watch_error_dir(inbox: Path) -> Path:
    return _env_path("RRHH_WATCH_ERROR", inbox / "ERROR") or inbox / "ERROR"


def _file_signature(path: Path) -> tuple[int, int] | None:
    try:
        stat = path.stat()
    except OSError:
        return None
    return int(stat.st_size), int(stat.st_mtime_ns)


def _stable_excel_files(folder: Path, stability_seconds: int) -> list[Path]:
    global _rrhh_monitor_seen
    now = time_module.monotonic()
    files = _excel_files(folder)
    current = {str(path.resolve()): path for path in files}
    for known in list(_rrhh_monitor_seen):
        if known not in current:
            _rrhh_monitor_seen.pop(known, None)

    stable: list[Path] = []
    for key, path in current.items():
        signature = _file_signature(path)
        if signature is None:
            _rrhh_monitor_seen.pop(key, None)
            continue
        previous = _rrhh_monitor_seen.get(key)
        if previous is None or previous[:2] != signature:
            _rrhh_monitor_seen[key] = (signature[0], signature[1], now)
            continue
        if now - previous[2] >= stability_seconds:
            stable.append(path)

    return stable if len(stable) == len(files) else []


def _make_auto_batch_key(paths: list[Path]) -> str:
    digest = hashlib.sha1(
        "|".join(
            f"{path.name}:{path.stat().st_size}:{path.stat().st_mtime_ns}"
            for path in sorted(paths, key=lambda item: item.name.lower())
            if path.exists()
        ).encode("utf-8")
    ).hexdigest()[:16]
    return f"auto_{digest}"


def _flatten_imported_paths(files: dict[str, Any]) -> list[Path]:
    paths: list[Path] = []
    for value in files.values():
        if isinstance(value, list):
            paths.extend(Path(item) for item in value)
        elif value:
            paths.append(Path(value))
    unique: dict[str, Path] = {}
    for path in paths:
        unique[str(path.resolve())] = path
    return list(unique.values())


def _unique_destination(dest_dir: Path, name: str) -> Path:
    candidate = dest_dir / name
    if not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    for index in range(1, 1000):
        candidate = dest_dir / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"No se pudo generar destino unico para {name}.")


def _move_files(paths: list[Path], dest_root: Path, batch_key: str) -> list[str]:
    dest_dir = dest_root / datetime.now().strftime("%Y_%m_%d") / batch_key
    dest_dir.mkdir(parents=True, exist_ok=True)
    moved: list[str] = []
    for path in paths:
        if not path.exists() or not path.is_file():
            continue
        dest = _unique_destination(dest_dir, path.name)
        shutil.move(str(path), str(dest))
        moved.append(str(dest))
        _rrhh_monitor_seen.pop(str(path.resolve()), None)
    return moved


async def _rrhh_monitor_loop() -> None:
    logger.info("Monitor RRHH iniciado.")
    interval = _env_int("RRHH_WATCH_POLL_SECONDS", 60, minimum=10, maximum=3600)
    stability = _env_int("RRHH_WATCH_STABILITY_SECONDS", 30, minimum=5, maximum=3600)
    while _rrhh_monitor_stop is not None and not _rrhh_monitor_stop.is_set():
        try:
            inbox = _watch_inbox()
            if inbox is None or not inbox.exists() or not inbox.is_dir():
                logger.warning("Monitor RRHH: carpeta no disponible: %s", inbox)
            else:
                stable_files = await asyncio.to_thread(_stable_excel_files, inbox, stability)
                if stable_files:
                    batch_key = _make_auto_batch_key(stable_files)
                    try:
                        result = await _import_folder_locked(inbox, batch_key, "rrhh_monitor", False)
                        moved = await asyncio.to_thread(
                            _move_files,
                            _flatten_imported_paths(result.get("files") or {}),
                            _watch_imported_dir(inbox),
                            batch_key,
                        )
                        logger.info("Monitor RRHH importo %s y movio %s archivo(s).", batch_key, len(moved))
                    except Exception as exc:
                        logger.exception("Monitor RRHH fallo al importar %s: %s", batch_key, exc)
                        if "ya existe" in str(exc):
                            moved = await asyncio.to_thread(_move_files, stable_files, _watch_imported_dir(inbox), batch_key)
                            logger.warning("Monitor RRHH movio %s archivo(s) ya importado(s) a IMPORTADOS.", len(moved))
                        else:
                            moved = await asyncio.to_thread(_move_files, stable_files, _watch_error_dir(inbox), batch_key)
                            logger.warning("Monitor RRHH movio %s archivo(s) a ERROR para %s.", len(moved), batch_key)
            await asyncio.wait_for(_rrhh_monitor_stop.wait(), timeout=interval)
        except asyncio.TimeoutError:
            continue
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            logger.warning("Monitor RRHH continuo tras error: %s", exc)
            await asyncio.sleep(interval)
    logger.info("Monitor RRHH detenido.")


def start_rrhh_folder_monitor() -> None:
    global _rrhh_monitor_task, _rrhh_monitor_stop
    if not _env_bool("RRHH_WATCH_ENABLED", False):
        logger.info("Monitor RRHH deshabilitado.")
        return
    if _rrhh_monitor_task and not _rrhh_monitor_task.done():
        return
    _rrhh_monitor_stop = asyncio.Event()
    _rrhh_monitor_task = asyncio.create_task(_rrhh_monitor_loop())


async def stop_rrhh_folder_monitor() -> None:
    global _rrhh_monitor_task, _rrhh_monitor_stop
    if _rrhh_monitor_stop:
        _rrhh_monitor_stop.set()
    if _rrhh_monitor_task:
        _rrhh_monitor_task.cancel()
        with suppress(asyncio.CancelledError):
            await _rrhh_monitor_task
    _rrhh_monitor_task = None
    _rrhh_monitor_stop = None


async def _require_auth(request: Request) -> dict[str, Any]:
    auth = await current_auth(request)
    if not auth or auth.get("device_status") != "approved":
        raise HTTPException(status_code=401, detail="No autenticado.")
    await _attach_rrhh_scope(auth)
    if auth.get("rrhh_scope") == "sin_acceso":
        raise HTTPException(status_code=403, detail="El usuario no tiene acceso a Novedades CD.")
    return auth


async def _attach_rrhh_scope(auth: dict[str, Any]) -> None:
    if auth.get("role") == "admin":
        auth["rrhh_scope"] = "global"
        auth["rrhh_sectors"] = []
        return
    sectors: list[str] = []
    scope = "operativo"
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await _ensure_consolidated_temp(db, sanciones=False, fichadas=False)
        async with db.execute(
            """
            SELECT scope, sector
            FROM auth_user_module_scopes
            WHERE username = ? AND module = ? AND active = 1
            ORDER BY scope
            """,
            (auth.get("username"), RRHH_SCOPE_MODULE),
        ) as cur:
            rows = [dict(row) for row in await cur.fetchall()]
    if any(row.get("scope") == "sin_acceso" for row in rows):
        scope = "sin_acceso"
    elif auth.get("role") in FULL_ACCESS_ROLES:
        scope = "global"
    elif any(row.get("scope") == "global" for row in rows):
        scope = "global"
    elif any(row.get("scope") == "sector_completo" for row in rows):
        scope = "sector_completo"
        for row in rows:
            sector = _norm(row.get("sector"))
            if row.get("scope") == "sector_completo" and sector and sector not in sectors:
                sectors.append(sector)
    auth["rrhh_scope"] = scope
    auth["rrhh_sectors"] = sectors


def _can_see_all(auth: dict[str, Any]) -> bool:
    return auth.get("role") in FULL_ACCESS_ROLES or auth.get("rrhh_scope") == "global"


def _can_see_gerencia(auth: dict[str, Any]) -> bool:
    return _can_see_all(auth) or auth.get("rrhh_scope") == "sector_completo"


def _require_import_role(auth: dict[str, Any]) -> None:
    if auth.get("role") not in IMPORT_ROLES:
        raise HTTPException(status_code=403, detail="Requiere perfil admin o RRHH.")


async def _latest_batch_id() -> int | None:
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(
            """
            SELECT b.batch_id
            FROM rrhh_import_batches b
            LEFT JOIN rrhh_actividad_diaria a ON a.batch_id = b.batch_id
            WHERE b.status = 'complete'
            GROUP BY b.batch_id
            ORDER BY CASE WHEN MAX(a.fecha) IS NULL THEN 1 ELSE 0 END,
                     MAX(a.fecha) DESC,
                     b.imported_at DESC,
                     b.batch_id DESC
            LIMIT 1
            """
        ) as cur:
            row = await cur.fetchone()
    return int(row[0]) if row else None


async def _resolve_batch_id(batch_id: int | None) -> int:
    if batch_id:
        return batch_id
    latest = await _latest_batch_id()
    if not latest:
        raise HTTPException(status_code=404, detail="Todavia no hay lotes RRHH importados.")
    return latest


def _sql_literals(values: list[str]) -> str:
    return ", ".join("'" + value.replace("'", "''") + "'" for value in values)


def _visibility_sql(auth: dict[str, Any], alias: str = "", sector_expression: str | None = None) -> str:
    if _can_see_all(auth):
        return "1=1"
    if auth.get("rrhh_scope") == "sector_completo":
        sectors = [_norm(sector) for sector in auth.get("rrhh_sectors", []) if _norm(sector)]
        if not sectors or not sector_expression:
            return "0=1"
        return f"{sector_expression} IN ({_sql_literals(sectors)})"
    prefix = f"{alias}." if alias else ""
    return f"COALESCE({prefix}es_gerencia, 0) = 0"


def _append_persona_filter(
    where: list[str],
    params: list[Any],
    search: str,
    *,
    record_alias: str,
    legajero_alias: str = "l",
    name_columns: tuple[str, ...],
) -> None:
    text = _norm(search)
    if not text:
        return
    numeric = _norm_legajo(text)
    like = f"%{text.upper()}%"
    clauses = [
        f"{record_alias}.legajo = ?",
        f"LTRIM({record_alias}.legajo, '0') = ?",
        f"UPPER(COALESCE({legajero_alias}.nombre, '')) LIKE ?",
    ]
    params.extend([numeric, numeric, like])
    for column in name_columns:
        clauses.append(f"UPPER(COALESCE({column}, '')) LIKE ?")
        params.append(like)
    where.append(f"({' OR '.join(clauses)})")


def _multi_query_values(request: Request, key: str, fallback: str = "ALL") -> list[str]:
    values = request.query_params.getlist(key) or ([fallback] if fallback else [])
    out: list[str] = []
    for value in values:
        for part in str(value).split("|"):
            item = _norm(part)
            if item and item != "ALL" and item not in out:
                out.append(item)
    return out


def _append_multi_filter(where: list[str], params: list[Any], expression: str, values: list[str]) -> None:
    if not values:
        return
    placeholders = ", ".join("?" for _ in values)
    where.append(f"{expression} IN ({placeholders})")
    params.extend(values)


CONSOLIDATED_CTES = """
latest_legajero AS (
    SELECT *
    FROM (
        SELECT l.*,
               ROW_NUMBER() OVER (
                   PARTITION BY l.legajo
                   ORDER BY b.imported_at DESC, l.batch_id DESC, l.id DESC
               ) rn
        FROM rrhh_legajero l
        JOIN rrhh_import_batches b ON b.batch_id = l.batch_id
        WHERE b.status = 'complete'
    )
    WHERE rn = 1
),
actividad_consolidada AS (
    SELECT *
    FROM (
        SELECT a.*,
               ROW_NUMBER() OVER (
                   PARTITION BY a.legajo, a.fecha
                   ORDER BY b.imported_at DESC, a.batch_id DESC, a.id DESC
               ) rn
        FROM rrhh_actividad_diaria a
        JOIN rrhh_import_batches b ON b.batch_id = a.batch_id
        WHERE b.status = 'complete'
    )
    WHERE rn = 1
),
sanciones_consolidadas AS (
    SELECT *
    FROM (
        SELECT s.*,
               ROW_NUMBER() OVER (
                   PARTITION BY s.legajo, COALESCE(s.inicio,''), COALESCE(s.fin,''), COALESCE(s.cod,''), COALESCE(s.creacion,''), COALESCE(s.descripcion,''), COALESCE(s.causa_sancion,''), COALESCE(s.descripcion_causa,'')
                   ORDER BY b.imported_at DESC, s.batch_id DESC, s.id DESC
               ) rn
        FROM rrhh_sanciones s
        JOIN rrhh_import_batches b ON b.batch_id = s.batch_id
        WHERE b.status = 'complete'
    )
    WHERE rn = 1
),
fichadas_consolidadas AS (
    SELECT *
    FROM (
        SELECT f.*,
               ROW_NUMBER() OVER (
                   PARTITION BY f.legajo, f.fecha_fichada, COALESCE(f.sentido,''), COALESCE(f.ubicacion,''), COALESCE(f.tipo_lectura,'')
                   ORDER BY b.imported_at DESC, f.batch_id DESC, f.id DESC
               ) rn
        FROM rrhh_fichadas f
        JOIN rrhh_import_batches b ON b.batch_id = f.batch_id
        WHERE b.status = 'complete'
    )
    WHERE rn = 1
)
"""


def _with_consolidated(extra_ctes: str = "") -> str:
    return f"WITH {extra_ctes}" if extra_ctes.strip() else ""


async def _ensure_consolidated_temp(
    db: aiosqlite.Connection,
    *,
    actividad: bool = True,
    sanciones: bool = True,
    fichadas: bool = True,
) -> None:
    await db.executescript(
        """
        DROP TABLE IF EXISTS temp.latest_legajero;
        DROP TABLE IF EXISTS temp.actividad_consolidada;
        DROP TABLE IF EXISTS temp.sanciones_consolidadas;
        DROP TABLE IF EXISTS temp.fichadas_consolidadas;

        CREATE TEMP TABLE latest_legajero AS
        SELECT *
        FROM (
            SELECT l.*,
                   ROW_NUMBER() OVER (
                       PARTITION BY l.legajo
                       ORDER BY b.imported_at DESC, l.batch_id DESC, l.id DESC
                   ) rn
            FROM rrhh_legajero l
            JOIN rrhh_import_batches b ON b.batch_id = l.batch_id
            WHERE b.status = 'complete'
        )
        WHERE rn = 1;
        CREATE INDEX temp.idx_tmp_latest_legajero ON latest_legajero(legajo);
        """
    )
    if actividad:
        await db.executescript(
            """
        CREATE TEMP TABLE actividad_consolidada AS
        SELECT *
        FROM (
            SELECT a.*,
                   ROW_NUMBER() OVER (
                       PARTITION BY a.legajo, a.fecha
                       ORDER BY b.imported_at DESC, a.batch_id DESC, a.id DESC
                   ) rn
            FROM rrhh_actividad_diaria a
            JOIN rrhh_import_batches b ON b.batch_id = a.batch_id
            WHERE b.status = 'complete'
        )
        WHERE rn = 1;
        CREATE INDEX temp.idx_tmp_actividad_fecha ON actividad_consolidada(fecha);
        CREATE INDEX temp.idx_tmp_actividad_legajo_fecha ON actividad_consolidada(legajo, fecha);
            """
        )
    if sanciones:
        await db.executescript(
            """
        CREATE TEMP TABLE sanciones_consolidadas AS
        SELECT *
        FROM (
            SELECT s.*,
                   ROW_NUMBER() OVER (
                       PARTITION BY s.legajo, COALESCE(s.inicio,''), COALESCE(s.fin,''), COALESCE(s.cod,''), COALESCE(s.creacion,''), COALESCE(s.descripcion,''), COALESCE(s.causa_sancion,''), COALESCE(s.descripcion_causa,'')
                       ORDER BY b.imported_at DESC, s.batch_id DESC, s.id DESC
                   ) rn
            FROM rrhh_sanciones s
            JOIN rrhh_import_batches b ON b.batch_id = s.batch_id
            WHERE b.status = 'complete'
        )
        WHERE rn = 1;
        CREATE INDEX temp.idx_tmp_sanciones_legajo ON sanciones_consolidadas(legajo);
            """
        )
    if fichadas:
        await db.executescript(
            """
        CREATE TEMP TABLE fichadas_consolidadas AS
        SELECT *
        FROM (
            SELECT f.*,
                   ROW_NUMBER() OVER (
                       PARTITION BY f.legajo, f.fecha_fichada, COALESCE(f.sentido,''), COALESCE(f.ubicacion,''), COALESCE(f.tipo_lectura,'')
                       ORDER BY b.imported_at DESC, f.batch_id DESC, f.id DESC
                   ) rn
            FROM rrhh_fichadas f
            JOIN rrhh_import_batches b ON b.batch_id = f.batch_id
            WHERE b.status = 'complete'
        )
        WHERE rn = 1;
        CREATE INDEX temp.idx_tmp_fichadas_fecha ON fichadas_consolidadas(fecha);
        CREATE INDEX temp.idx_tmp_fichadas_legajo ON fichadas_consolidadas(legajo);
        """
        )


@router.get("/config")
async def get_config(request: Request):
    auth = await _require_auth(request)
    latest_folder = None
    watch_inbox = _watch_inbox()
    if SOURCE_ROOT.exists():
        try:
            latest_folder = str(_find_latest_folder())
        except Exception:
            latest_folder = None
    return {
        "role": auth.get("role"),
        "data_scope": auth.get("rrhh_scope", "operativo"),
        "data_sectors": auth.get("rrhh_sectors", []),
        "can_import": auth.get("role") in IMPORT_ROLES,
        "can_see_gerencia": _can_see_gerencia(auth),
        "source_root": str(SOURCE_ROOT),
        "latest_folder": latest_folder,
        "watch": {
            "enabled": _env_bool("RRHH_WATCH_ENABLED", False),
            "inbox": str(watch_inbox) if watch_inbox else None,
            "imported": str(_watch_imported_dir(watch_inbox)) if watch_inbox else None,
            "error": str(_watch_error_dir(watch_inbox)) if watch_inbox else None,
            "poll_seconds": _env_int("RRHH_WATCH_POLL_SECONDS", 60, minimum=10, maximum=3600),
            "stability_seconds": _env_int("RRHH_WATCH_STABILITY_SECONDS", 30, minimum=5, maximum=3600),
        },
    }


@router.get("/batches")
async def list_batches(request: Request):
    await _require_auth(request)
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            """
            SELECT batch_id, batch_key, source_dir, imported_by, imported_at, status,
                   files_json, summary_json, error,
                   fecha_min, fecha_max, actividad_registros
            FROM (
                SELECT b.batch_id, b.batch_key, b.source_dir, b.imported_by, b.imported_at, b.status,
                       b.files_json, b.summary_json, b.error,
                       MIN(a.fecha) fecha_min,
                       MAX(a.fecha) fecha_max,
                       COUNT(a.id) actividad_registros
                FROM rrhh_import_batches b
                LEFT JOIN rrhh_actividad_diaria a ON a.batch_id = b.batch_id
                GROUP BY b.batch_id
            )
            ORDER BY CASE WHEN fecha_max IS NULL THEN 1 ELSE 0 END,
                     fecha_max DESC,
                     imported_at DESC,
                     batch_id DESC
            LIMIT 20
            """
        ) as cur:
            rows = [dict(row) for row in await cur.fetchall()]
    for row in rows:
        row["files"] = json.loads(row.pop("files_json") or "{}")
        row["summary"] = json.loads(row.pop("summary_json") or "{}")
        if "advertencias_calidad" not in row["summary"]:
            async with aiosqlite.connect(DB_PATH) as db:
                db.row_factory = aiosqlite.Row
                async with db.execute(
                    """
                    SELECT legajo, COALESCE(NULLIF(TRIM(nombre), ''), 'Sin nombre') nombre,
                           COALESCE(NULLIF(TRIM(desc_posicion), ''), NULLIF(TRIM(desc_funcion), ''), '') referencia
                    FROM rrhh_legajero
                    WHERE batch_id = ?
                      AND (
                        TRIM(COALESCE(desc_sector_generico, '')) = ''
                        OR TRIM(COALESCE(desc_funcion, '')) = ''
                      )
                    ORDER BY legajo
                    LIMIT 25
                    """,
                    (row["batch_id"],),
                ) as cur:
                    missing_master = [dict(item) for item in await cur.fetchall()]
                async with db.execute(
                    """
                    SELECT COUNT(*) c
                    FROM rrhh_legajero
                    WHERE batch_id = ?
                      AND (
                        TRIM(COALESCE(desc_sector_generico, '')) = ''
                        OR TRIM(COALESCE(desc_funcion, '')) = ''
                      )
                    """,
                    (row["batch_id"],),
                ) as cur:
                    missing_master_count = (await cur.fetchone())["c"]
                async with db.execute(
                    """
                    SELECT a.legajo,
                           COALESCE(NULLIF(TRIM(l.nombre), ''), NULLIF(TRIM(a.empleado), ''), 'Sin nombre') nombre,
                           COUNT(*) registros,
                           SUM(CASE WHEN COALESCE(a.ausentismo_contabiliza, 0) = 1 THEN 1 ELSE 0 END) ausencias
                    FROM rrhh_actividad_diaria a
                    LEFT JOIN rrhh_legajero l ON l.batch_id = a.batch_id AND l.legajo = a.legajo
                    WHERE a.batch_id = ?
                      AND COALESCE(NULLIF(TRIM(l.desc_sector_generico), ''), NULLIF(TRIM(a.sector), '')) IS NULL
                    GROUP BY a.legajo, nombre
                    ORDER BY ausencias DESC, registros DESC, a.legajo
                    LIMIT 25
                    """,
                    (row["batch_id"],),
                ) as cur:
                    missing_activity = [dict(item) for item in await cur.fetchall()]
                async with db.execute(
                    """
                    SELECT COUNT(DISTINCT a.legajo) legajos, COUNT(*) registros
                    FROM rrhh_actividad_diaria a
                    LEFT JOIN rrhh_legajero l ON l.batch_id = a.batch_id AND l.legajo = a.legajo
                    WHERE a.batch_id = ?
                      AND COALESCE(NULLIF(TRIM(l.desc_sector_generico), ''), NULLIF(TRIM(a.sector), '')) IS NULL
                    """,
                    (row["batch_id"],),
                ) as cur:
                    missing_activity_count = dict(await cur.fetchone())
            row["summary"]["advertencias_calidad"] = {
                "legajero_sin_sector_o_cargo": int(missing_master_count or 0),
                "legajero_sin_sector_o_cargo_muestra": missing_master,
                "actividad_sin_sector_legajos": int(missing_activity_count.get("legajos") or 0),
                "actividad_sin_sector_registros": int(missing_activity_count.get("registros") or 0),
                "actividad_sin_sector_muestra": missing_activity,
            }
    return {"batches": rows}


@router.post("/import/latest")
async def import_latest(req: ImportFolderRequest, request: Request):
    auth = await _require_auth(request)
    _require_import_role(auth)
    folder = _find_latest_folder()
    batch_key = req.batch_key or folder.name
    try:
        result = await _import_folder_locked(folder, batch_key, auth.get("username", ""), req.force)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return result


@router.post("/import/folder")
async def import_folder(req: ImportFolderRequest, request: Request):
    auth = await _require_auth(request)
    _require_import_role(auth)
    folder = Path(req.folder_path or "").expanduser()
    if not folder.exists() or not folder.is_dir():
        raise HTTPException(status_code=400, detail="La carpeta indicada no existe.")
    batch_key = req.batch_key or folder.name
    try:
        result = await _import_folder_locked(folder, batch_key, auth.get("username", ""), req.force)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return result


@router.get("/resumen")
async def get_resumen(request: Request, batch_id: int | None = None):
    auth = await _require_auth(request)
    batch_id = await _resolve_batch_id(batch_id)
    vis_legajero = _visibility_sql(auth, sector_expression="desc_sector_generico")
    vis_actividad = _visibility_sql(auth, sector_expression="sector")
    vis_sanciones = _visibility_sql(auth, sector_expression="desc_unidad_organizativa")
    vis_fichadas = _visibility_sql(auth, "f", "COALESCE(l.desc_sector_generico, '')")
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        queries = {
            "empleados": f"SELECT COUNT(DISTINCT legajo) v FROM rrhh_legajero WHERE batch_id = ? AND {vis_legajero}",
            "actividad": f"SELECT COUNT(*) v FROM rrhh_actividad_diaria WHERE batch_id = ? AND {vis_actividad}",
            "ausencias": f"SELECT COUNT(*) v FROM rrhh_actividad_diaria WHERE batch_id = ? AND {vis_actividad} AND COALESCE(ausentismo_contabiliza,0) = 1",
            "horas_trabajadas": f"SELECT COALESCE(SUM(hs_trab),0) v FROM rrhh_actividad_diaria WHERE batch_id = ? AND {vis_actividad}",
            "horas_extra": f"SELECT COALESCE(SUM(hs_ext_realiz + hs_50_autorizadas + hs_100),0) v FROM rrhh_actividad_diaria WHERE batch_id = ? AND {vis_actividad}",
            "llegadas_tarde": f"SELECT COUNT(*) v FROM rrhh_actividad_diaria WHERE batch_id = ? AND {vis_actividad} AND tarde > 0",
            "fichadas": f"SELECT COUNT(*) v FROM rrhh_fichadas f LEFT JOIN rrhh_legajero l ON l.batch_id = f.batch_id AND l.legajo = f.legajo WHERE f.batch_id = ? AND {vis_fichadas}",
            "sanciones": f"SELECT COUNT(*) v FROM rrhh_sanciones WHERE batch_id = ? AND {vis_sanciones}",
        }
        metrics = {}
        for key, sql in queries.items():
            async with db.execute(sql, (batch_id,)) as cur:
                metrics[key] = (await cur.fetchone())["v"]
        async with db.execute(
            f"""
            SELECT COALESCE(l.desc_sector_generico, a.sector, 'Sin sector') sector,
                   COUNT(DISTINCT a.legajo) empleados,
                   SUM(CASE WHEN COALESCE(a.ausentismo_contabiliza,0) = 1 THEN 1 ELSE 0 END) ausencias,
                   ROUND(SUM(a.hs_trab), 1) horas_trabajadas,
                   ROUND(SUM(a.hs_ext_realiz + a.hs_50_autorizadas + a.hs_100), 1) horas_extra,
                   SUM(CASE WHEN a.tarde > 0 THEN 1 ELSE 0 END) llegadas_tarde
            FROM rrhh_actividad_diaria a
            LEFT JOIN rrhh_legajero l ON l.batch_id = a.batch_id AND l.legajo = a.legajo
            WHERE a.batch_id = ? AND {_visibility_sql(auth, 'a', 'COALESCE(l.desc_sector_generico, a.sector)')}
            GROUP BY sector
            ORDER BY ausencias DESC, empleados DESC
            LIMIT 20
            """,
            (batch_id,),
        ) as cur:
            sectores = [dict(row) for row in await cur.fetchall()]
        async with db.execute(
            "SELECT * FROM rrhh_import_batches WHERE batch_id = ?",
            (batch_id,),
        ) as cur:
            batch = dict(await cur.fetchone())
    batch["files"] = json.loads(batch.pop("files_json") or "{}")
    batch["summary"] = json.loads(batch.pop("summary_json") or "{}")
    return {"batch": batch, "metrics": metrics, "sectores": sectores, "restricted": not _can_see_all(auth)}


@router.get("/rango-sugerido")
async def get_rango_sugerido(request: Request, batch_id: int | None = None):
    auth = await _require_auth(request)
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await _ensure_consolidated_temp(db)
        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT MIN(a.fecha) fecha_min, MAX(a.fecha) fecha_max
            FROM actividad_consolidada a
            LEFT JOIN latest_legajero l ON l.legajo = a.legajo
            WHERE {_visibility_sql(auth, "a", "COALESCE(l.desc_sector_generico, a.sector)")}
            """,
            (),
        ) as cur:
            row = dict(await cur.fetchone())
    fecha_max = row.get("fecha_max")
    fecha_min = row.get("fecha_min")
    suggested_from = fecha_min
    if fecha_max:
        try:
            max_dt = datetime.strptime(fecha_max, "%Y-%m-%d").date()
            min_dt = datetime.strptime(fecha_min, "%Y-%m-%d").date() if fecha_min else max_dt
            from_dt = max(max_dt - timedelta(days=6), min_dt)
            suggested_from = from_dt.isoformat()
        except ValueError:
            suggested_from = fecha_min
    return {
        "batch_id": None,
        "data_range": {"fecha_desde": fecha_min, "fecha_hasta": fecha_max},
        "suggested_range": {"fecha_desde": suggested_from, "fecha_hasta": fecha_max},
    }


@router.get("/indicadores")
async def get_indicadores(
    request: Request,
    batch_id: int | None = None,
    fecha_desde: str = "",
    fecha_hasta: str = "",
    sector: str = "ALL",
    cargo: str = "ALL",
    grupo: str = "ALL",
    motivo: str = "ALL",
    persona: str = "",
):
    auth = await _require_auth(request)
    sectores_sel = _multi_query_values(request, "sector", sector)
    cargos_sel = _multi_query_values(request, "cargo", cargo)
    grupos_sel = _multi_query_values(request, "grupo", grupo)
    motivos_sel = _multi_query_values(request, "motivo", motivo)

    where = [_visibility_sql(auth, "a", "COALESCE(l.desc_sector_generico, a.sector)")]
    params: list[Any] = []
    if fecha_desde:
        where.append("a.fecha >= ?")
        params.append(fecha_desde)
    if fecha_hasta:
        where.append("a.fecha <= ?")
        params.append(fecha_hasta)
    _append_multi_filter(where, params, "COALESCE(l.desc_sector_generico, a.sector)", sectores_sel)
    _append_multi_filter(where, params, "COALESCE(l.desc_funcion, '')", cargos_sel)
    _append_multi_filter(where, params, "COALESCE(l.desc_grupo_personal, '')", grupos_sel)
    _append_multi_filter(where, params, "COALESCE(NULLIF(TRIM(a.motivo), ''), 'Sin motivo')", motivos_sel)
    _append_persona_filter(where, params, persona, record_alias="a", name_columns=("a.empleado",))
    where_sql = " AND ".join(where)
    scheduled_a = _scheduled_day_sql("a")
    scheduled_af = _scheduled_day_sql("af")

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await _ensure_consolidated_temp(db)

        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT
                COUNT(DISTINCT a.legajo) empleados,
                SUM(CASE WHEN {scheduled_a} THEN 1 ELSE 0 END) registros,
                SUM(CASE WHEN COALESCE(a.ausentismo_contabiliza,0) = 1 THEN 1 ELSE 0 END) ausencias,
                SUM(CASE WHEN a.ausentismo_clasificacion = 'CONTROLADO' THEN 1 ELSE 0 END) ausencias_controladas,
                SUM(CASE WHEN a.ausentismo_clasificacion = 'NO CONTROLADO' THEN 1 ELSE 0 END) ausencias_no_controladas,
                SUM(CASE WHEN a.ausentismo_clasificacion = 'NO CONSIDERAR' THEN 1 ELSE 0 END) ausencias_no_considerar,
                SUM(CASE WHEN a.ausentismo_clasificacion = 'SIN CLASIFICAR' THEN 1 ELSE 0 END) ausencias_sin_clasificar,
                SUM(CASE WHEN a.tarde > 0 THEN 1 ELSE 0 END) llegadas_tarde,
                ROUND(COALESCE(SUM(a.hs_trab),0), 1) horas_trabajadas,
                ROUND(COALESCE(SUM(a.hs_ext_realiz + a.hs_50_autorizadas + a.hs_100),0), 1) horas_extra,
                ROUND(
                    SUM(CASE WHEN COALESCE(a.ausentismo_contabiliza,0) = 1 THEN 1 ELSE 0 END) * 100.0 / NULLIF(SUM(CASE WHEN {scheduled_a} THEN 1 ELSE 0 END), 0),
                    1
                ) ausentismo_pct
            FROM actividad_consolidada a
            LEFT JOIN latest_legajero l ON l.legajo = a.legajo
            WHERE {where_sql}
            """,
            tuple(params),
        ) as cur:
            kpis = dict(await cur.fetchone())

        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT MIN(a.fecha) fecha_desde, MAX(a.fecha) fecha_hasta
            FROM actividad_consolidada a
            LEFT JOIN latest_legajero l ON l.legajo = a.legajo
            WHERE {_visibility_sql(auth, "a", "COALESCE(l.desc_sector_generico, a.sector)")}
            """,
            (),
        ) as cur:
            data_range = dict(await cur.fetchone())

        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT a.fecha label,
                   SUM(CASE WHEN {scheduled_a} THEN 1 ELSE 0 END) registros,
                   SUM(CASE WHEN COALESCE(a.ausentismo_contabiliza,0) = 1 THEN 1 ELSE 0 END) ausencias,
                   SUM(CASE WHEN a.ausentismo_clasificacion = 'CONTROLADO' THEN 1 ELSE 0 END) ausencias_controladas,
                   SUM(CASE WHEN a.ausentismo_clasificacion = 'NO CONTROLADO' THEN 1 ELSE 0 END) ausencias_no_controladas,
                   SUM(CASE WHEN a.tarde > 0 THEN 1 ELSE 0 END) llegadas_tarde,
                   ROUND(SUM(a.hs_ext_realiz + a.hs_50_autorizadas + a.hs_100), 1) horas_extra
            FROM actividad_consolidada a
            LEFT JOIN latest_legajero l ON l.legajo = a.legajo
            WHERE {where_sql}
            GROUP BY a.fecha
            ORDER BY a.fecha
            """,
            tuple(params),
        ) as cur:
            diario = [dict(row) for row in await cur.fetchall()]

        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT COALESCE(NULLIF(TRIM(l.desc_sector_generico), ''), NULLIF(TRIM(a.sector), ''), 'Sin sector') label,
                   SUM(CASE WHEN {scheduled_a} THEN 1 ELSE 0 END) registros,
                   COUNT(DISTINCT a.legajo) empleados,
                   SUM(CASE WHEN COALESCE(a.ausentismo_contabiliza,0) = 1 THEN 1 ELSE 0 END) ausencias,
                   ROUND(SUM(a.hs_trab), 1) horas_trabajadas,
                   ROUND(SUM(a.hs_ext_realiz + a.hs_50_autorizadas + a.hs_100), 1) horas_extra,
                   SUM(CASE WHEN a.tarde > 0 THEN 1 ELSE 0 END) llegadas_tarde
            FROM actividad_consolidada a
            LEFT JOIN latest_legajero l ON l.legajo = a.legajo
            WHERE {where_sql}
              AND COALESCE(NULLIF(TRIM(l.desc_sector_generico), ''), NULLIF(TRIM(a.sector), '')) IS NOT NULL
            GROUP BY label
            ORDER BY ausencias DESC, horas_extra DESC
            LIMIT 12
            """,
            tuple(params),
        ) as cur:
            por_sector = [dict(row) for row in await cur.fetchall()]

        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT COALESCE(NULLIF(TRIM(l.desc_funcion), ''), NULLIF(TRIM(l.desc_posicion), ''), 'Sin cargo') || ' / ' ||
                   COALESCE(NULLIF(TRIM(l.desc_sector_generico), ''), NULLIF(TRIM(a.sector), ''), 'Sin sector') label,
                   SUM(CASE WHEN {scheduled_a} THEN 1 ELSE 0 END) registros,
                   COUNT(DISTINCT a.legajo) empleados,
                   SUM(CASE WHEN COALESCE(a.ausentismo_contabiliza,0) = 1 THEN 1 ELSE 0 END) ausencias,
                   ROUND(SUM(a.hs_trab), 1) horas_trabajadas,
                   ROUND(SUM(a.hs_ext_realiz + a.hs_50_autorizadas + a.hs_100), 1) horas_extra,
                   SUM(CASE WHEN a.tarde > 0 THEN 1 ELSE 0 END) llegadas_tarde
            FROM actividad_consolidada a
            LEFT JOIN latest_legajero l ON l.legajo = a.legajo
            WHERE {where_sql}
              AND COALESCE(NULLIF(TRIM(l.desc_sector_generico), ''), NULLIF(TRIM(a.sector), '')) IS NOT NULL
            GROUP BY label
            ORDER BY ausencias DESC, horas_extra DESC
            LIMIT 18
            """,
            tuple(params),
        ) as cur:
            por_sector_cargo = [dict(row) for row in await cur.fetchall()]

        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT TRIM(a.motivo) label,
                   COUNT(*) eventos
            FROM actividad_consolidada a
            LEFT JOIN latest_legajero l ON l.legajo = a.legajo
            WHERE {where_sql}
            GROUP BY label
            HAVING TRIM(COALESCE(label, '')) <> '' AND eventos > 0
            ORDER BY eventos DESC
            LIMIT 10
            """,
            tuple(params),
        ) as cur:
            motivos = [dict(row) for row in await cur.fetchall()]

        ranking_sancion_filters = [_visibility_sql(auth, "s", "COALESCE(l.desc_sector_generico, s.desc_unidad_organizativa, '')")]
        ranking_sancion_params: list[Any] = []
        if fecha_desde:
            ranking_sancion_filters.append("COALESCE(s.creacion, s.inicio) >= ?")
            ranking_sancion_params.append(fecha_desde)
        if fecha_hasta:
            ranking_sancion_filters.append("COALESCE(s.creacion, s.inicio) <= ?")
            ranking_sancion_params.append(fecha_hasta)
        ranking_sancion_filters.append(f"TRIM(s.descripcion) IN ({', '.join('?' for _ in REAL_SANCIONES)})")
        ranking_sancion_params.extend(REAL_SANCIONES)
        ranking_cte = f"""
            actividad_filtrada AS (
                SELECT a.*, COALESCE(l.desc_sector_generico, a.sector, 'Sin sector') unidad,
                       COALESCE(NULLIF(TRIM(l.desc_funcion), ''), 'Sin cargo') funcion,
                       COALESCE(NULLIF(TRIM(l.desc_grupo_personal), ''), '') grupo_personal,
                       COALESCE(NULLIF(TRIM(l.desc_area_personal), ''), '') area_personal
                FROM actividad_consolidada a
                LEFT JOIN latest_legajero l ON l.legajo = a.legajo
                WHERE {where_sql}
            )
        """
        async with db.execute(
            f"""
            {_with_consolidated(ranking_cte)},
            motivos_aus AS (
                SELECT legajo, GROUP_CONCAT(label, ', ') motivos_aus
                FROM (
                    SELECT legajo, TRIM(motivo) || '(' || COUNT(*) || ')' label
                    FROM actividad_filtrada
                    WHERE COALESCE(ausentismo_contabiliza,0) = 1
                      AND TRIM(COALESCE(motivo,'')) <> ''
                    GROUP BY legajo, TRIM(motivo)
                    ORDER BY COUNT(*) DESC, TRIM(motivo)
                )
                GROUP BY legajo
            ),
            otros_motivos AS (
                SELECT legajo, GROUP_CONCAT(label, ', ') otros_motivos
                FROM (
                    SELECT legajo, TRIM(motivo) || '(' || COUNT(*) || ')' label
                    FROM actividad_filtrada
                    WHERE COALESCE(ausentismo_contabiliza,0) = 0
                      AND ausentismo_clasificacion <> 'SIN NOVEDAD'
                      AND TRIM(COALESCE(motivo,'')) <> ''
                    GROUP BY legajo, TRIM(motivo)
                    ORDER BY COUNT(*) DESC, TRIM(motivo)
                )
                GROUP BY legajo
            ),
            sanciones_legajo AS (
                SELECT legajo, SUM(c) c_sanc, GROUP_CONCAT(label, ', ') tipo_sancion
                FROM (
                    SELECT s.legajo, COALESCE(NULLIF(TRIM(s.descripcion), ''), 'Sin descripcion') label, COUNT(*) c
                    FROM sanciones_consolidadas s
                    LEFT JOIN latest_legajero l ON l.legajo = s.legajo
                    WHERE {" AND ".join(ranking_sancion_filters)}
                    GROUP BY s.legajo, label
                    ORDER BY COUNT(*) DESC, label
                )
                GROUP BY legajo
            )
            SELECT af.legajo,
                   COALESCE(MAX(NULLIF(TRIM(af.empleado), '')), '') empleado,
                   MAX(af.funcion) funcion,
                   MAX(af.unidad) unidad,
                   COUNT(DISTINCT af.fecha) dias_reg,
                   SUM(CASE WHEN {scheduled_af} THEN 1 ELSE 0 END) dias_programados,
                   SUM(CASE WHEN TRIM(COALESCE(af.motivo,'')) <> ''
                              OR COALESCE(NULLIF(TRIM(af.aus_pres_codigo), ''), '0') <> '0'
                            THEN 1 ELSE 0 END) dias_novedad,
                   SUM(CASE WHEN (TRIM(COALESCE(af.motivo,'')) <> ''
                                  OR COALESCE(NULLIF(TRIM(af.aus_pres_codigo), ''), '0') <> '0')
                              AND COALESCE(af.ausentismo_contabiliza,0) = 0
                            THEN 1 ELSE 0 END) dias_no_considerar,
                   SUM(CASE WHEN COALESCE(af.ausentismo_contabiliza,0) = 1 THEN 1 ELSE 0 END) dias_aus,
                   ROUND(SUM(CASE WHEN COALESCE(af.ausentismo_contabiliza,0) = 1 THEN 1 ELSE 0 END) * 100.0 / NULLIF(SUM(CASE WHEN {scheduled_af} THEN 1 ELSE 0 END), 0), 1) pct_aus,
                   SUM(CASE WHEN af.ausentismo_clasificacion = 'CONTROLADO' THEN 1 ELSE 0 END) ctrl,
                   SUM(CASE WHEN af.ausentismo_clasificacion = 'NO CONTROLADO' THEN 1 ELSE 0 END) no_ctrl,
                   COALESCE(MAX(ma.motivos_aus), '') motivos_aus,
                   COALESCE(MAX(om.otros_motivos), '') otros_motivos,
                   ROUND(SUM(af.hs_trab), 1) hs_trab,
                   SUM(CASE WHEN (af.hs_ext_realiz + af.hs_50_autorizadas + af.hs_100) > 0 THEN 1 ELSE 0 END) dias_che,
                   SUM(CASE WHEN af.hs_50_autorizadas > 0 THEN 1 ELSE 0 END) dias_he_50,
                   SUM(CASE WHEN af.hs_100 > 0 THEN 1 ELSE 0 END) dias_he_100,
                   ROUND(SUM(af.hs_50_autorizadas), 1) hs_50,
                   ROUND(SUM(af.hs_100), 1) hs_100,
                   ROUND(SUM(af.hs_ext_realiz + af.hs_50_autorizadas + af.hs_100), 1) hs_extra,
                   ROUND(SUM(af.hs_fer), 1) hs_fer,
                   ROUND(SUM(af.rec_noct), 1) rec_noct,
                   SUM(CASE WHEN af.tarde > 0 THEN 1 ELSE 0 END) c_tarde,
                   ROUND(SUM(af.tarde) * 60, 0) min_tarde,
                   0 paus_30,
                   COALESCE(MAX(sl.c_sanc), 0) c_sanc,
                   COALESCE(MAX(sl.tipo_sancion), '') tipo_sancion,
                   MAX(af.grupo_personal) grupo_personal,
                   MAX(af.area_personal) area_personal
            FROM actividad_filtrada af
            LEFT JOIN motivos_aus ma ON ma.legajo = af.legajo
            LEFT JOIN otros_motivos om ON om.legajo = af.legajo
            LEFT JOIN sanciones_legajo sl ON sl.legajo = af.legajo
            GROUP BY af.legajo
            ORDER BY pct_aus DESC, dias_aus DESC, c_tarde DESC, hs_extra DESC
            LIMIT 1000
            """,
            tuple(params + ranking_sancion_params),
        ) as cur:
            ranking = [dict(row) for row in await cur.fetchall()]

        s_where = [_visibility_sql(auth, "s", "COALESCE(l.desc_sector_generico, s.desc_unidad_organizativa, '')")]
        s_params: list[Any] = []
        if fecha_desde:
            s_where.append("COALESCE(s.creacion, s.inicio) >= ?")
            s_params.append(fecha_desde)
        if fecha_hasta:
            s_where.append("COALESCE(s.creacion, s.inicio) <= ?")
            s_params.append(fecha_hasta)
        _append_multi_filter(s_where, s_params, "COALESCE(l.desc_sector_generico, s.desc_unidad_organizativa, '')", sectores_sel)
        _append_multi_filter(s_where, s_params, "COALESCE(l.desc_funcion, '')", cargos_sel)
        _append_multi_filter(s_where, s_params, "COALESCE(l.desc_grupo_personal, '')", grupos_sel)
        _append_persona_filter(s_where, s_params, persona, record_alias="s", name_columns=("s.nombre",))
        s_where.append(f"TRIM(s.descripcion) IN ({', '.join('?' for _ in REAL_SANCIONES)})")
        s_params.extend(REAL_SANCIONES)
        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT COALESCE(NULLIF(TRIM(s.descripcion), ''), 'Sin descripcion') label,
                   COUNT(*) eventos
            FROM sanciones_consolidadas s
            LEFT JOIN latest_legajero l ON l.legajo = s.legajo
            WHERE {" AND ".join(s_where)}
            GROUP BY label
            ORDER BY eventos DESC
            LIMIT 10
            """,
            tuple(s_params),
        ) as cur:
            sanciones = [dict(row) for row in await cur.fetchall()]
        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT COUNT(*) sanciones
            FROM sanciones_consolidadas s
            LEFT JOIN latest_legajero l ON l.legajo = s.legajo
            WHERE {" AND ".join(s_where)}
            """,
            tuple(s_params),
        ) as cur:
            kpis["sanciones"] = (await cur.fetchone())["sanciones"]

        f_where = [_visibility_sql(auth, "f", "COALESCE(l.desc_sector_generico, '')")]
        f_params: list[Any] = []
        if fecha_desde:
            f_where.append("f.fecha >= ?")
            f_params.append(fecha_desde)
        if fecha_hasta:
            f_where.append("f.fecha <= ?")
            f_params.append(fecha_hasta)
        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT COUNT(*) fichadas, COUNT(DISTINCT f.legajo) legajos_con_fichada
            FROM fichadas_consolidadas f
            LEFT JOIN latest_legajero l ON l.legajo = f.legajo
            WHERE {' AND '.join(f_where)}
            """,
            tuple(f_params),
        ) as cur:
            kpis.update(dict(await cur.fetchone()))

        filter_vis = _visibility_sql(auth, "l", "l.desc_sector_generico")
        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT DISTINCT desc_sector_generico value
            FROM latest_legajero l
            WHERE {filter_vis} AND TRIM(COALESCE(desc_sector_generico,'')) <> ''
            ORDER BY value
            """,
            (),
        ) as cur:
            sectores = [row["value"] for row in await cur.fetchall()]
        cargo_where = [filter_vis, "TRIM(COALESCE(l.desc_funcion,'')) <> ''"]
        cargo_params: list[Any] = []
        _append_multi_filter(cargo_where, cargo_params, "l.desc_sector_generico", sectores_sel)
        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT DISTINCT desc_funcion value
            FROM latest_legajero l
            WHERE {" AND ".join(cargo_where)}
            ORDER BY value
            """,
            tuple(cargo_params),
        ) as cur:
            cargos = [row["value"] for row in await cur.fetchall()]
        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT DISTINCT desc_grupo_personal value
            FROM latest_legajero l
            WHERE {filter_vis} AND TRIM(COALESCE(desc_grupo_personal,'')) <> ''
            ORDER BY value
            """,
            (),
        ) as cur:
            grupos = [row["value"] for row in await cur.fetchall()]
        async with db.execute(
            f"""
            {_with_consolidated()}
            SELECT DISTINCT TRIM(a.motivo) value
            FROM actividad_consolidada a
            WHERE {_visibility_sql(auth, "a", "a.sector")}
              AND COALESCE(a.ausentismo_contabiliza,0) = 1
              AND TRIM(COALESCE(a.motivo,'')) <> ''
            ORDER BY value
            """,
            (),
        ) as cur:
            motivos_filter = [row["value"] for row in await cur.fetchall()]

    return {
        "restricted": not _can_see_all(auth),
        "kpis": kpis,
        "data_range": data_range,
        "filters": {
            "sectores": sectores,
            "cargos": cargos,
            "grupos": grupos,
            "motivos": motivos_filter,
        },
        "series": {
            "diario": diario,
            "por_sector": por_sector,
            "por_sector_cargo": por_sector_cargo,
            "motivos": motivos,
            "sanciones": sanciones,
            "ranking": ranking,
        },
    }


async def _query_rows(sql: str, params: tuple[Any, ...]) -> list[dict[str, Any]]:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await _ensure_consolidated_temp(
            db,
            actividad="actividad_consolidada" in sql,
            sanciones="sanciones_consolidadas" in sql,
            fichadas="fichadas_consolidadas" in sql,
        )
        async with db.execute(sql, params) as cur:
            return [dict(row) for row in await cur.fetchall()]


def _franco_delta(row: dict[str, Any], rules: list[dict[str, Any]]) -> tuple[float, str, str]:
    haystack = " ".join(
        _norm(row.get(key)).upper()
        for key in ("motivo", "horario", "horario_teorico", "aus_pres_codigo", "ausentismo_clasificacion")
    )
    if _to_float(row.get("hs_100")) > 0 and any(token in haystack for token in ("FRANCO", "DESCANSO", "LIBRE")):
        haystack = f"{haystack} HS 100"
    for rule in rules:
        pattern = _norm(rule.get("patron")).upper()
        if pattern and pattern in haystack:
            delta = _to_float(rule.get("delta"))
            return delta, pattern, "credito" if delta > 0 else "debito" if delta < 0 else "neutro"
    return 0.0, "", ""


@router.get("/francos")
async def get_francos(
    request: Request,
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    persona: str = "",
    sector: str = "ALL",
    cargo: str = "ALL",
    limit: int = Query(500, ge=1, le=2000),
):
    auth = await _require_auth(request)
    sectores_sel = _multi_query_values(request, "sector", sector)
    cargos_sel = _multi_query_values(request, "cargo", cargo)

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await _ensure_consolidated_temp(db, actividad=True, sanciones=False, fichadas=False)
        async with db.execute(
            """
            SELECT tipo, patron, delta, prioridad
            FROM rrhh_francos_reglas
            WHERE active = 1
            ORDER BY prioridad, regla_id
            """
        ) as cur:
            rules = [dict(row) for row in await cur.fetchall()]

        base_visibility = _visibility_sql(auth, "l", "l.desc_sector_generico")
        initial_where = [base_visibility]
        initial_params: list[Any] = []
        _append_persona_filter(initial_where, initial_params, persona, record_alias="fi", legajero_alias="l", name_columns=("fi.nombre",))
        _append_multi_filter(initial_where, initial_params, "l.desc_sector_generico", sectores_sel)
        _append_multi_filter(initial_where, initial_params, "l.desc_funcion", cargos_sel)
        async with db.execute(
            f"""
            SELECT fi.legajo,
                   COALESCE(NULLIF(TRIM(l.nombre), ''), NULLIF(TRIM(fi.nombre), ''), 'Sin nombre') nombre,
                   COALESCE(NULLIF(TRIM(l.desc_sector_generico), ''), '') sector,
                   COALESCE(NULLIF(TRIM(l.desc_funcion), ''), NULLIF(TRIM(l.desc_posicion), ''), '') funcion,
                   COALESCE(fi.saldo_inicial, 0) saldo_inicial,
                   fi.fecha_corte
            FROM rrhh_francos_inicial fi
            LEFT JOIN latest_legajero l ON l.legajo = fi.legajo
            WHERE {" AND ".join(initial_where)}
            """,
            tuple(initial_params),
        ) as cur:
            balances = {row["legajo"]: dict(row) for row in await cur.fetchall()}

        activity_where = [_visibility_sql(auth, "a", "COALESCE(l.desc_sector_generico, a.sector)")]
        activity_params: list[Any] = []
        if fecha_desde:
            activity_where.append("a.fecha >= ?")
            activity_params.append(fecha_desde)
        if fecha_hasta:
            activity_where.append("a.fecha <= ?")
            activity_params.append(fecha_hasta)
        _append_persona_filter(activity_where, activity_params, persona, record_alias="a", legajero_alias="l", name_columns=("a.empleado",))
        _append_multi_filter(activity_where, activity_params, "COALESCE(l.desc_sector_generico, a.sector)", sectores_sel)
        _append_multi_filter(activity_where, activity_params, "l.desc_funcion", cargos_sel)
        async with db.execute(
            f"""
            SELECT a.legajo, a.fecha, a.empleado, a.sector, a.motivo, a.horario,
                   a.horario_teorico, a.aus_pres_codigo, a.ausentismo_clasificacion,
                   a.hs_100, a.hs_trab,
                   COALESCE(NULLIF(TRIM(l.nombre), ''), NULLIF(TRIM(a.empleado), ''), 'Sin nombre') nombre,
                   COALESCE(NULLIF(TRIM(l.desc_sector_generico), ''), NULLIF(TRIM(a.sector), ''), '') sector_legajero,
                   COALESCE(NULLIF(TRIM(l.desc_funcion), ''), NULLIF(TRIM(l.desc_posicion), ''), '') funcion
            FROM actividad_consolidada a
            LEFT JOIN latest_legajero l ON l.legajo = a.legajo
            WHERE {" AND ".join(activity_where)}
            ORDER BY a.fecha, a.legajo
            """,
            tuple(activity_params),
        ) as cur:
            activity_rows = [dict(row) for row in await cur.fetchall()]

    movements: list[dict[str, Any]] = []
    for row in activity_rows:
        delta, regla, tipo = _franco_delta(row, rules)
        if not delta:
            continue
        legajo = row["legajo"]
        corte = balances.get(legajo, {}).get("fecha_corte") or _francos_fecha_inicial()
        if row.get("fecha") and corte and row["fecha"] <= corte:
            continue
        item = balances.setdefault(
            legajo,
            {
                "legajo": legajo,
                "nombre": row.get("nombre") or "Sin nombre",
                "sector": row.get("sector_legajero") or row.get("sector") or "",
                "funcion": row.get("funcion") or "",
                "saldo_inicial": 0,
                "fecha_corte": None,
            },
        )
        item["nombre"] = item.get("nombre") or row.get("nombre") or "Sin nombre"
        item["sector"] = item.get("sector") or row.get("sector_legajero") or row.get("sector") or ""
        item["funcion"] = item.get("funcion") or row.get("funcion") or ""
        movement = {
            "fecha": row.get("fecha"),
            "legajo": legajo,
            "empleado": item["nombre"],
            "delta": delta,
            "tipo": tipo,
            "regla": regla,
            "motivo": row.get("motivo") or row.get("horario_teorico") or row.get("horario") or "",
            "hs_100": row.get("hs_100") or 0,
        }
        movements.append(movement)
        item["creditos"] = _to_float(item.get("creditos")) + (delta if delta > 0 else 0)
        item["debitos"] = _to_float(item.get("debitos")) + ((-delta) if delta < 0 else 0)
        item["movimientos"] = int(item.get("movimientos") or 0) + 1

    rows: list[dict[str, Any]] = []
    for item in balances.values():
        saldo_inicial = _to_float(item.get("saldo_inicial"))
        creditos = _to_float(item.get("creditos"))
        debitos = _to_float(item.get("debitos"))
        rows.append({
            **item,
            "creditos": creditos,
            "debitos": debitos,
            "movimientos": int(item.get("movimientos") or 0),
            "saldo_actual": saldo_inicial + creditos - debitos,
        })
    rows.sort(key=lambda item: (item["saldo_actual"], item["legajo"]))

    kpis = {
        "legajos": len(rows),
        "saldo_total": round(sum(_to_float(row.get("saldo_actual")) for row in rows), 1),
        "creditos": round(sum(_to_float(row.get("creditos")) for row in rows), 1),
        "debitos": round(sum(_to_float(row.get("debitos")) for row in rows), 1),
        "negativos": sum(1 for row in rows if _to_float(row.get("saldo_actual")) < 0),
    }
    return {"kpis": kpis, "rows": rows[:limit], "movimientos": movements[-500:]}


@router.get("/francos/candidatos")
async def get_francos_candidatos(
    request: Request,
    fecha_desde: str | None = None,
    fecha_hasta: str | None = None,
    persona: str = "",
    sector: str = "ALL",
    cargo: str = "ALL",
    limit: int = Query(1000, ge=1, le=3000),
):
    auth = await _require_auth(request)
    sectores_sel = _multi_query_values(request, "sector", sector)
    cargos_sel = _multi_query_values(request, "cargo", cargo)
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await _ensure_consolidated_temp(db, actividad=True, sanciones=False, fichadas=False)
        async with db.execute(
            """
            SELECT tipo, patron, delta, prioridad
            FROM rrhh_francos_reglas
            WHERE active = 1
            ORDER BY prioridad, regla_id
            """
        ) as cur:
            rules = [dict(row) for row in await cur.fetchall()]

        where = [_visibility_sql(auth, "a", "COALESCE(l.desc_sector_generico, a.sector)")]
        params: list[Any] = []
        if fecha_desde:
            where.append("a.fecha >= ?")
            params.append(fecha_desde)
        if fecha_hasta:
            where.append("a.fecha <= ?")
            params.append(fecha_hasta)
        _append_persona_filter(where, params, persona, record_alias="a", legajero_alias="l", name_columns=("a.empleado",))
        _append_multi_filter(where, params, "COALESCE(l.desc_sector_generico, a.sector)", sectores_sel)
        _append_multi_filter(where, params, "l.desc_funcion", cargos_sel)
        async with db.execute(
            f"""
            SELECT a.legajo, a.fecha, a.empleado, a.sector, a.motivo, a.horario,
                   a.horario_teorico, a.aus_pres_codigo, a.ausentismo_clasificacion,
                   a.hs_100, a.hs_trab,
                   COALESCE(fi.fecha_corte, ?) fecha_corte,
                   COALESCE(fi.saldo_inicial, 0) saldo_inicial,
                   COALESCE(NULLIF(TRIM(l.nombre), ''), NULLIF(TRIM(a.empleado), ''), 'Sin nombre') nombre,
                   COALESCE(NULLIF(TRIM(l.desc_sector_generico), ''), NULLIF(TRIM(a.sector), ''), '') sector_legajero,
                   COALESCE(NULLIF(TRIM(l.desc_funcion), ''), NULLIF(TRIM(l.desc_posicion), ''), '') funcion
            FROM actividad_consolidada a
            LEFT JOIN latest_legajero l ON l.legajo = a.legajo
            LEFT JOIN rrhh_francos_inicial fi ON fi.legajo = a.legajo
            WHERE {" AND ".join(where)}
            ORDER BY a.fecha DESC, a.legajo
            LIMIT ?
            """,
            (_francos_fecha_inicial(), *params, limit * 5),
        ) as cur:
            activity_rows = [dict(row) for row in await cur.fetchall()]

    candidatos: list[dict[str, Any]] = []
    for row in activity_rows:
        if row.get("fecha") and row.get("fecha_corte") and row["fecha"] <= row["fecha_corte"]:
            continue
        delta, regla, tipo = _franco_delta(row, rules)
        if not delta:
            continue
        candidatos.append({
            "fecha": row.get("fecha"),
            "fecha_corte": row.get("fecha_corte"),
            "legajo": row.get("legajo"),
            "empleado": row.get("nombre"),
            "sector": row.get("sector_legajero") or row.get("sector") or "",
            "funcion": row.get("funcion") or "",
            "tipo": tipo,
            "delta": delta,
            "regla": regla,
            "motivo": row.get("motivo") or row.get("horario_teorico") or row.get("horario") or "",
            "horario": row.get("horario_teorico") or row.get("horario") or "",
            "aus_pres_codigo": row.get("aus_pres_codigo") or "",
            "hs_100": row.get("hs_100") or 0,
        })
        if len(candidatos) >= limit:
            break
    return {
        "kpis": {
            "total": len(candidatos),
            "creditos": sum(1 for item in candidatos if item["delta"] > 0),
            "debitos": sum(1 for item in candidatos if item["delta"] < 0),
            "dias_netos": round(sum(_to_float(item["delta"]) for item in candidatos), 1),
        },
        "rows": candidatos,
    }


@router.get("/actividad")
async def get_actividad(
    request: Request,
    batch_id: int | None = None,
    fecha_desde: str = "",
    fecha_hasta: str = "",
    sector: str = "ALL",
    cargo: str = "ALL",
    motivo: str = "ALL",
    persona: str = "",
    limit: int = Query(200, ge=1, le=1000),
):
    auth = await _require_auth(request)
    sectores_sel = _multi_query_values(request, "sector", sector)
    cargos_sel = _multi_query_values(request, "cargo", cargo)
    motivos_sel = _multi_query_values(request, "motivo", motivo)
    where = [_visibility_sql(auth, "a", "COALESCE(l.desc_sector_generico, a.sector)")]
    params: list[Any] = []
    if fecha_desde:
        where.append("a.fecha >= ?")
        params.append(fecha_desde)
    if fecha_hasta:
        where.append("a.fecha <= ?")
        params.append(fecha_hasta)
    _append_multi_filter(where, params, "COALESCE(l.desc_sector_generico, a.sector)", sectores_sel)
    _append_multi_filter(where, params, "COALESCE(l.desc_funcion, '')", cargos_sel)
    _append_multi_filter(where, params, "COALESCE(NULLIF(TRIM(a.motivo), ''), 'Sin motivo')", motivos_sel)
    _append_persona_filter(where, params, persona, record_alias="a", name_columns=("a.empleado",))
    params.append(limit)
    rows = await _query_rows(
        f"""
        {_with_consolidated()}
        SELECT a.fecha, a.legajo, a.empleado, COALESCE(l.desc_sector_generico, a.sector) sector,
               a.horario, a.horario_teorico, a.ingreso_teorico, a.salida_teorica, a.horas_teoricas,
               a.entrada, a.salida, a.motivo, a.aus_pres_codigo,
               a.ausentismo_clasificacion, a.ausentismo_tratamiento,
               a.hs_trab, a.hs_ext_realiz, a.hs_50_autorizadas, a.hs_100,
               ROUND(a.tarde * 60, 0) tarde
        FROM actividad_consolidada a
        LEFT JOIN latest_legajero l ON l.legajo = a.legajo
        WHERE {' AND '.join(where)}
        ORDER BY a.fecha DESC, a.legajo
        LIMIT ?
        """,
        tuple(params),
    )
    for row in rows:
        if not row.get("ingreso_teorico"):
            row["ingreso_teorico"] = _time_minus_minutes(row.get("entrada"), row.get("tarde"))
    return {"rows": rows, "restricted": not _can_see_all(auth)}


@router.get("/fichadas")
async def get_fichadas(
    request: Request,
    batch_id: int | None = None,
    fecha_desde: str = "",
    fecha_hasta: str = "",
    sector: str = "ALL",
    cargo: str = "ALL",
    legajo: str = "",
    persona: str = "",
    limit: int = Query(200, ge=1, le=1000),
):
    auth = await _require_auth(request)
    sectores_sel = _multi_query_values(request, "sector", sector)
    cargos_sel = _multi_query_values(request, "cargo", cargo)
    where = [_visibility_sql(auth, "f", "COALESCE(l.desc_sector_generico, '')")]
    params: list[Any] = []
    if fecha_desde:
        where.append("f.fecha >= ?")
        params.append(fecha_desde)
    if fecha_hasta:
        where.append("f.fecha <= ?")
        params.append(fecha_hasta)
    _append_multi_filter(where, params, "COALESCE(l.desc_sector_generico, '')", sectores_sel)
    _append_multi_filter(where, params, "COALESCE(l.desc_funcion, '')", cargos_sel)
    if legajo:
        where.append("f.legajo = ?")
        params.append(_norm_legajo(legajo))
    _append_persona_filter(where, params, persona, record_alias="f", name_columns=("f.empleado",))
    params.append(limit)
    rows = await _query_rows(
        f"""
        {_with_consolidated()}
        SELECT f.fecha_fichada, f.legajo, f.empleado, COALESCE(l.desc_sector_generico, '') sector,
               COALESCE(l.desc_funcion, '') cargo, f.sentido, f.ubicacion, f.origen, f.destino, f.tipo_lectura
        FROM fichadas_consolidadas f
        LEFT JOIN latest_legajero l ON l.legajo = f.legajo
        WHERE {' AND '.join(where)}
        ORDER BY f.fecha_fichada DESC
        LIMIT ?
        """,
        tuple(params),
    )
    return {"rows": rows, "restricted": not _can_see_all(auth)}


@router.get("/sanciones")
async def get_sanciones(
    request: Request,
    batch_id: int | None = None,
    fecha_desde: str = "",
    fecha_hasta: str = "",
    sector: str = "ALL",
    cargo: str = "ALL",
    motivo: str = "ALL",
    persona: str = "",
    limit: int = Query(200, ge=1, le=1000),
):
    auth = await _require_auth(request)
    sectores_sel = _multi_query_values(request, "sector", sector)
    cargos_sel = _multi_query_values(request, "cargo", cargo)
    motivos_sel = _multi_query_values(request, "motivo", motivo)
    where = [_visibility_sql(auth, "s", "COALESCE(l.desc_sector_generico, s.desc_unidad_organizativa, '')")]
    params: list[Any] = []
    if fecha_desde:
        where.append("COALESCE(s.creacion, s.inicio) >= ?")
        params.append(fecha_desde)
    if fecha_hasta:
        where.append("COALESCE(s.creacion, s.inicio) <= ?")
        params.append(fecha_hasta)
    where.append(f"TRIM(s.descripcion) IN ({', '.join('?' for _ in REAL_SANCIONES)})")
    params.extend(REAL_SANCIONES)
    _append_multi_filter(where, params, "COALESCE(l.desc_sector_generico, s.desc_unidad_organizativa, '')", sectores_sel)
    _append_multi_filter(where, params, "COALESCE(l.desc_funcion, '')", cargos_sel)
    if motivos_sel:
        placeholders = ", ".join("?" for _ in motivos_sel)
        where.append(f"(COALESCE(NULLIF(TRIM(s.descripcion), ''), 'Sin motivo') IN ({placeholders}) OR COALESCE(NULLIF(TRIM(s.descripcion_causa), ''), 'Sin motivo') IN ({placeholders}))")
        params.extend(motivos_sel)
        params.extend(motivos_sel)
    _append_persona_filter(where, params, persona, record_alias="s", name_columns=("s.nombre",))
    params.append(limit)
    rows = await _query_rows(
        f"""
        {_with_consolidated()}
        SELECT s.legajo, s.nombre, COALESCE(l.desc_sector_generico, s.desc_unidad_organizativa, '') sector,
               COALESCE(l.desc_funcion, '') cargo, s.inicio, s.creacion, s.cod, s.descripcion,
               s.causa_sancion, s.descripcion_causa
        FROM sanciones_consolidadas s
        LEFT JOIN latest_legajero l ON l.legajo = s.legajo
        WHERE {' AND '.join(where)}
        ORDER BY COALESCE(s.creacion, s.inicio) DESC, s.legajo
        LIMIT ?
        """,
        tuple(params),
    )
    return {"rows": rows, "restricted": not _can_see_all(auth)}
