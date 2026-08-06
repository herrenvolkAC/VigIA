from __future__ import annotations

import csv
import io
import asyncio
import logging
import os
import sqlite3
from datetime import datetime, time, timedelta
from typing import Any
import unicodedata

import aiosqlite
from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
import xlrd

from db.auth import auth_db
from db.panol_insumos import panol_db
from routers.auth_local import current_auth
from routers.productividad_analisis import _query_productive_db_sql


router = APIRouter(prefix="/panol-insumos/api", tags=["panol-insumos"])
logger = logging.getLogger("vigia.panol_insumos")

FULL_PROFILES = {"ADMIN", "SUPERVISOR", "TODO"}
LIMITED_PROFILES = {"OPERACION", "OPERADOR", "USUARIO", ""}
REQUEST_PROFILES = {"SOLICITANTE", "PEDIDOS"}
MOVEMENT_TYPES = {"ALTA", "BAJA", "AJUSTE_POSITIVO", "AJUSTE_NEGATIVO", "TRANSFERENCIA"}
INCOMING_TYPES = {"ALTA", "AJUSTE_POSITIVO"}
OUTGOING_TYPES = {"BAJA", "AJUSTE_NEGATIVO"}
ORACLE_STOCK_CD_SOURCE = "ORACLE_STOCK_CD"
MERMA_TYPES = {"INSUMO", "PRODUCCION"}
MERMA_REASONS = {
    "IMPRESION_DUPLICADA": "Impresion duplicada",
    "ERROR_IMPRESION": "Error de impresion",
    "MATERIAL_DANADO": "Material danado",
    "VENCIDO": "Vencido",
    "NO_UTILIZABLE": "No utilizable",
    "OTRO": "Otro",
}
_stock_cd_scheduler_task: asyncio.Task | None = None
_stock_cd_scheduler_stop: asyncio.Event | None = None
_stock_cd_scheduler_last_attempt: str | None = None

QUERY_STOCK_CD_ORACLE = """
SELECT
F002.CREFEREN REFERENCIA,
SUM(P505.QSTKFISI * F209.QCOECONV) AS UNIDADES
FROM F002ARTI F002 JOIN P505STPA P505 ON P505.CREFEREN = F002.CREFEREN
JOIN F209CONV F209 ON F209.CREFEREN = P505.CREFEREN AND F209.CVARLOGI = P505.CVARLOGI
AND F209.CEMPRESA = P505.CEMPRESA
AND F209.CCONSIGN = P505.CCONSIGN
AND F209.CPRESENT = P505.CPRESENT
WHERE F002.CEMPRESA = '1'
AND F002.DDEPARTA IN (84)
AND F209.CVARLODP = '0'
AND F209.CEMPRESA = P505.CEMPRESA
GROUP BY
F002.CREFEREN
"""


class ArticleRequest(BaseModel):
    codigo: str
    descripcion: str
    categoria: str = ""
    unidad: str = "UN"
    uso: str = ""
    stock_minimo: float = 0
    activo: bool = True
    costo_unitario: float | None = None
    moneda: str = "ARS"
    fecha_costo: str = ""


class MovementRequest(BaseModel):
    articulo_id: int
    tipo: str
    ubicacion_origen_id: int | None = None
    ubicacion_destino_id: int | None = None
    cantidad: float
    motivo: str = ""
    observacion: str = ""


class ProductionRequest(BaseModel):
    articulo_id: int
    cantidad: float
    observacion: str = ""


class ProductionDeliveryRequest(BaseModel):
    articulo_id: int
    ubicacion_destino_id: int
    cantidad: float
    observacion: str = ""


class SupplyOrderItemRequest(BaseModel):
    articulo_id: int
    cantidad_insumo: float = 0
    cantidad_produccion: float = 0


class SupplyOrderRequest(BaseModel):
    sector_id: int
    observacion: str = ""
    items: list[SupplyOrderItemRequest]


class SupplyOrderConfirmItemRequest(BaseModel):
    item_id: int
    cantidad_insumo_confirmada: float = 0
    cantidad_produccion_confirmada: float = 0
    ubicacion_origen_insumo_id: int | None = None
    uso_entrega: str = ""


class SupplyOrderConfirmRequest(BaseModel):
    observacion: str = ""
    items: list[SupplyOrderConfirmItemRequest]


class WasteRequest(BaseModel):
    sector_id: int
    pedido_id: int | None = None
    articulo_id: int
    tipo: str = "INSUMO"
    cantidad: float
    motivo: str
    observacion: str


class InventoryLine(BaseModel):
    articulo_id: int
    stock_fisico: float


class InventoryRequest(BaseModel):
    fecha: str
    turno: str
    ubicacion_id: int | None = None
    ubicacion_codigo: str = "OFICINA_ADO"
    observacion: str = ""
    items: list[InventoryLine]


class OperationalResetRequest(BaseModel):
    clave: str


class UserSectorAssignRequest(BaseModel):
    usernames: list[str]
    sector_id: int


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _today() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _turno_por_hora(fecha_hora: str | None = None) -> str:
    dt = datetime.strptime(fecha_hora or _now(), "%Y-%m-%d %H:%M:%S")
    hour = dt.hour
    if 6 <= hour < 14:
        return "MANANA"
    if 14 <= hour < 22:
        return "TARDE"
    return "NOCHE"


def _logistic_day_bounds() -> tuple[str, str]:
    now = datetime.now()
    start = (now.date() - timedelta(days=1)).strftime("%Y-%m-%d") + " 22:00:00"
    return start, now.strftime("%Y-%m-%d %H:%M:%S")


def _norm_code(value: Any) -> str:
    return " ".join(str(value or "").strip().split()).upper()


def _norm_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return "".join(ch for ch in text.lower() if ch.isalnum())


def _clean(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def _row_dict(row: aiosqlite.Row | None) -> dict[str, Any] | None:
    return dict(row) if row is not None else None


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    text = str(value).strip().replace(",", ".")
    if not text:
        return 0.0
    return float(text)


def _usage_options(value: Any) -> list[str]:
    text = str(value or "").replace("\r", "\n").replace(";", "\n").replace(",", "\n")
    return [_clean(part) for part in text.split("\n") if _clean(part)]


async def _fetch_rows(db: aiosqlite.Connection, sql: str, args: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    async with db.execute(sql, args) as cur:
        return [dict(row) for row in await cur.fetchall()]


async def _fetch_one(db: aiosqlite.Connection, sql: str, args: tuple[Any, ...] = ()) -> dict[str, Any] | None:
    async with db.execute(sql, args) as cur:
        return _row_dict(await cur.fetchone())


def _current_cost_join(alias: str = "a") -> str:
    return f"""
    LEFT JOIN articulos_costos ac ON ac.id = (
        SELECT ac2.id
        FROM articulos_costos ac2
        WHERE ac2.articulo_id = {alias}.id
          AND ac2.fecha_desde <= date('now')
          AND (ac2.fecha_hasta IS NULL OR ac2.fecha_hasta > date('now'))
        ORDER BY ac2.fecha_desde DESC, ac2.id DESC
        LIMIT 1
    )
    """


def _cost_at_sql(articulo_expr: str, fecha_expr: str) -> str:
    return f"""
    COALESCE((
        SELECT ac.costo_unitario
        FROM articulos_costos ac
        WHERE ac.articulo_id = {articulo_expr}
          AND ac.fecha_desde <= date({fecha_expr})
          AND (ac.fecha_hasta IS NULL OR ac.fecha_hasta > date({fecha_expr}))
        ORDER BY ac.fecha_desde DESC, ac.id DESC
        LIMIT 1
    ), 0)
    """


async def _set_article_cost(
    db: aiosqlite.Connection,
    articulo_id: int,
    costo_unitario: float | None,
    moneda: str,
    fecha_desde: str,
    usuario: str,
) -> None:
    if costo_unitario is None:
        return
    costo = float(costo_unitario)
    if costo < 0:
        raise HTTPException(status_code=400, detail="El costo no puede ser negativo.")
    fecha = _clean(fecha_desde) or _today()
    try:
        datetime.strptime(fecha, "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="La fecha de vigencia de costo debe tener formato AAAA-MM-DD.") from exc
    money = _norm_code(moneda or "ARS") or "ARS"
    next_row = await _fetch_one(
        db,
        """
        SELECT MIN(fecha_desde) AS fecha_hasta
        FROM articulos_costos
        WHERE articulo_id = ? AND fecha_desde > ?
        """,
        (articulo_id, fecha),
    )
    fecha_hasta = (next_row or {}).get("fecha_hasta")
    await db.execute(
        """
        UPDATE articulos_costos
        SET fecha_hasta = ?
        WHERE articulo_id = ?
          AND fecha_desde < ?
          AND (fecha_hasta IS NULL OR fecha_hasta > ?)
        """,
        (fecha, articulo_id, fecha, fecha),
    )
    existing = await _fetch_one(
        db,
        "SELECT id FROM articulos_costos WHERE articulo_id = ? AND fecha_desde = ?",
        (articulo_id, fecha),
    )
    if existing:
        await db.execute(
            """
            UPDATE articulos_costos
            SET costo_unitario = ?, moneda = ?, fecha_hasta = ?, fuente = 'MAESTRO_PLUS',
                usuario = ?, fecha_hora = ?
            WHERE id = ?
            """,
            (costo, money, fecha_hasta, usuario, _now(), int(existing["id"])),
        )
    else:
        await db.execute(
            """
            INSERT INTO articulos_costos
                (articulo_id, costo_unitario, moneda, fecha_desde, fecha_hasta, fuente, usuario, fecha_hora)
            VALUES (?, ?, ?, ?, ?, 'MAESTRO_PLUS', ?, ?)
            """,
            (articulo_id, costo, money, fecha, fecha_hasta, usuario, _now()),
        )


async def _require_panol_access(request: Request, *, full: bool = False) -> dict[str, Any]:
    auth = await current_auth(request)
    if not auth or auth.get("device_status") != "approved":
        raise HTTPException(status_code=401, detail="No autenticado.")
    if auth.get("role") == "admin":
        auth["panol_profile"] = "ADMIN"
        auth["panol_full"] = True
        return auth

    async with auth_db(attach_operational=False) as db:
        try:
            row = await _fetch_one(
                db,
                """
                SELECT enabled, profile
                FROM auth_user_app_access
                WHERE username = ? AND module = 'panol'
                """,
                (auth["username"],),
            )
        except sqlite3.OperationalError:
            row = None

    if not row or not row.get("enabled"):
        raise HTTPException(status_code=403, detail="Sin acceso al modulo Panol.")

    profile = _norm_code(row.get("profile") or "OPERACION")
    can_full = profile in FULL_PROFILES
    if profile not in FULL_PROFILES and profile not in LIMITED_PROFILES and profile not in REQUEST_PROFILES:
        can_full = False
    if full and not can_full:
        raise HTTPException(status_code=403, detail="Requiere perfil completo de Panol.")
    auth["panol_profile"] = profile or "OPERACION"
    auth["panol_full"] = can_full
    auth["panol_request_only"] = profile in REQUEST_PROFILES
    return auth


async def _require_panol_operator(request: Request, *, full: bool = False) -> dict[str, Any]:
    auth = await _require_panol_access(request, full=full)
    if auth.get("panol_request_only"):
        raise HTTPException(status_code=403, detail="El perfil solicitante solo puede gestionar pedidos de insumos.")
    return auth


async def _ubicacion_id(db: aiosqlite.Connection, codigo: str) -> int:
    row = await _fetch_one(db, "SELECT id FROM ubicaciones WHERE codigo = ? AND activo = 1", (_norm_code(codigo),))
    if not row:
        raise HTTPException(status_code=500, detail=f"Falta ubicacion {codigo}.")
    return int(row["id"])


async def _stock_ubicacion(db: aiosqlite.Connection, articulo_id: int, ubicacion_id: int, until: str | None = None) -> float:
    until_sql = "AND fecha_hora <= ?" if until else ""
    args: list[Any] = [articulo_id, ubicacion_id, articulo_id, ubicacion_id]
    if until:
        args.append(until)
    row = await _fetch_one(
        db,
        f"""
        SELECT
            COALESCE(SUM(CASE
                WHEN articulo_id = ? AND ubicacion_destino_id = ?
                     AND tipo IN ('ALTA','AJUSTE_POSITIVO','TRANSFERENCIA') THEN cantidad
                ELSE 0 END), 0)
          - COALESCE(SUM(CASE
                WHEN articulo_id = ? AND ubicacion_origen_id = ?
                     AND tipo IN ('BAJA','AJUSTE_NEGATIVO','TRANSFERENCIA') THEN cantidad
                ELSE 0 END), 0) AS stock
        FROM movimientos
        WHERE 1 = 1 {until_sql}
        """,
        tuple(args),
    )
    return float((row or {}).get("stock") or 0)


async def _latest_inventory(db: aiosqlite.Connection, articulo_id: int, ubicacion_id: int | None = None) -> dict[str, Any] | None:
    location_clause = "AND ubicacion_id = ?" if ubicacion_id else ""
    args: list[Any] = [articulo_id]
    if ubicacion_id:
        args.append(ubicacion_id)
    return await _fetch_one(
        db,
        f"""
        SELECT *
        FROM inventario_turno
        WHERE articulo_id = ?
        {location_clause}
        ORDER BY fecha_hora DESC, id DESC
        LIMIT 1
        """,
        tuple(args),
    )


async def _office_transfers_since(db: aiosqlite.Connection, articulo_id: int, since: str | None, until: str) -> tuple[float, float]:
    oficina_id = await _ubicacion_id(db, "OFICINA_ADO")
    where_since = "AND fecha_hora >= ?" if since else ""
    args: list[Any] = [oficina_id, oficina_id, articulo_id, until]
    if since:
        args.append(since)
    row = await _fetch_one(
        db,
        f"""
        SELECT
            COALESCE(SUM(CASE
                WHEN ubicacion_destino_id = ?
                     AND tipo IN ('ALTA','AJUSTE_POSITIVO','TRANSFERENCIA') THEN cantidad
                ELSE 0 END), 0) AS ingresos,
            COALESCE(SUM(CASE
                WHEN ubicacion_origen_id = ?
                     AND tipo IN ('BAJA','AJUSTE_NEGATIVO','TRANSFERENCIA') THEN cantidad
                ELSE 0 END), 0) AS egresos
        FROM movimientos
        WHERE articulo_id = ?
          AND fecha_hora <= ?
          {where_since}
        """,
        tuple(args),
    )
    return float((row or {}).get("ingresos") or 0), float((row or {}).get("egresos") or 0)


async def _office_income_since(db: aiosqlite.Connection, articulo_id: int, since: str | None, until: str) -> float:
    ingresos, _ = await _office_transfers_since(db, articulo_id, since, until)
    return ingresos


async def _stock_oficina(db: aiosqlite.Connection, articulo_id: int) -> float:
    oficina_id = await _ubicacion_id(db, "OFICINA_ADO")
    latest = await _latest_inventory(db, articulo_id, oficina_id)
    if not latest:
        return 0.0
    ingresos, egresos = await _office_transfers_since(db, articulo_id, latest["fecha_hora"], _now())
    return float(latest["stock_fisico"] or 0) + ingresos - egresos


async def _stock_for_origin(db: aiosqlite.Connection, articulo_id: int, ubicacion_id: int) -> float:
    location = await _fetch_one(db, "SELECT codigo FROM ubicaciones WHERE id = ?", (ubicacion_id,))
    if _norm_code((location or {}).get("codigo")) == "OFICINA_ADO":
        return await _stock_oficina(db, articulo_id)
    return await _stock_ubicacion(db, articulo_id, ubicacion_id)


async def _stock_cd(db: aiosqlite.Connection, articulo_id: int) -> float:
    row = await _fetch_one(
        db,
        """
        SELECT stock_cd
        FROM stock_cd_importado
        WHERE articulo_id = ?
        ORDER BY fecha_importacion DESC, id DESC
        LIMIT 1
        """,
        (articulo_id,),
    )
    return float((row or {}).get("stock_cd") or 0)


async def _stock_producido(db: aiosqlite.Connection, articulo_id: int) -> float:
    row = await _fetch_one(
        db,
        """
        SELECT
            COALESCE(SUM(CASE WHEN tipo = 'PRODUCCION' THEN cantidad ELSE 0 END), 0)
          - COALESCE(SUM(CASE WHEN tipo = 'ENTREGA' THEN cantidad ELSE 0 END), 0) AS stock
        FROM produccion_movimientos
        WHERE articulo_id = ?
        """,
        (articulo_id,),
    )
    return float((row or {}).get("stock") or 0)


async def _request_sectors(db: aiosqlite.Connection) -> list[dict[str, Any]]:
    rows = await _fetch_rows(db, "SELECT * FROM ubicaciones WHERE activo = 1 ORDER BY codigo")
    filtered = [row for row in rows if _norm_code(row.get("codigo")) not in {"JAULA", "OFICINA_ADO"}]
    return filtered or rows


async def _active_user_sector(db: aiosqlite.Connection, username: str | None) -> dict[str, Any] | None:
    clean_user = _clean(username)
    if not clean_user:
        return None
    return await _fetch_one(
        db,
        """
        SELECT us.username, us.sector_id, u.codigo AS sector_codigo, u.descripcion AS sector_descripcion
        FROM usuarios_sectores_panol us
        JOIN ubicaciones u ON u.id = us.sector_id AND u.activo = 1
        WHERE us.username = ? AND us.activo = 1
        ORDER BY us.id DESC
        LIMIT 1
        """,
        (clean_user,),
    )


async def _validate_request_sector(
    db: aiosqlite.Connection,
    auth: dict[str, Any],
    sector_id: int | None,
) -> int | None:
    if not auth.get("panol_request_only"):
        if sector_id is None:
            return None
        sectors = await _request_sectors(db)
        if not any(int(row["id"]) == int(sector_id) for row in sectors):
            raise HTTPException(status_code=400, detail="Sector solicitante invalido.")
        return int(sector_id)
    assigned = await _active_user_sector(db, str(auth.get("username") or ""))
    if not assigned:
        raise HTTPException(
            status_code=403,
            detail="Tu usuario no tiene sector asignado en Panol. Solicita la configuracion a un administrador.",
        )
    assigned_id = int(assigned["sector_id"])
    if sector_id is not None and int(sector_id) != assigned_id:
        raise HTTPException(status_code=403, detail="Tu usuario solo puede operar con su sector asignado en Panol.")
    return assigned_id


@router.get("/context")
async def context(request: Request):
    auth = await _require_panol_access(request)
    async with panol_db() as db:
        ubicaciones = await _fetch_rows(db, "SELECT * FROM ubicaciones WHERE activo = 1 ORDER BY id")
        sectores_solicitantes = await _request_sectors(db)
        sector_asignado = await _active_user_sector(db, str(auth.get("username") or ""))
        turnos = await _fetch_rows(db, "SELECT * FROM turnos WHERE activo = 1 ORDER BY id")
    return {
        "user": {
            "username": auth.get("username"),
            "display_name": auth.get("display_name"),
            "role": auth.get("role"),
            "panol_profile": auth.get("panol_profile"),
            "can_full": bool(auth.get("panol_full")),
            "can_request": True,
            "can_operate": not bool(auth.get("panol_request_only")),
            "sector_asignado": sector_asignado,
        },
        "ubicaciones": ubicaciones,
        "sectores_solicitantes": sectores_solicitantes,
        "turnos": turnos,
        "merma_motivos": [{"codigo": code, "descripcion": label} for code, label in MERMA_REASONS.items()],
    }


@router.get("/usuarios-solicitantes")
async def list_request_users(request: Request):
    await _require_panol_operator(request, full=True)
    async with auth_db(attach_operational=False) as adb:
        users = await _fetch_rows(
            adb,
            """
            SELECT u.username, u.display_name, u.active, a.profile
            FROM auth_users u
            JOIN auth_user_app_access a ON a.username = u.username
            WHERE a.module = 'panol'
              AND a.enabled = 1
              AND UPPER(COALESCE(a.profile, '')) IN ('SOLICITANTE', 'PEDIDOS')
            ORDER BY u.username
            """,
        )
    async with panol_db() as db:
        sectors = await _request_sectors(db)
        assignments = {
            str(row["username"]): row
            for row in await _fetch_rows(
                db,
                """
                SELECT us.username, us.sector_id, u.codigo AS sector_codigo, u.descripcion AS sector_descripcion,
                       us.actualizado_en, us.fecha_hora
                FROM usuarios_sectores_panol us
                JOIN ubicaciones u ON u.id = us.sector_id
                WHERE us.activo = 1
                """,
            )
        }
    return {
        "items": [
            {
                **user,
                "sector_id": assignments.get(str(user["username"]), {}).get("sector_id"),
                "sector_codigo": assignments.get(str(user["username"]), {}).get("sector_codigo"),
                "sector_descripcion": assignments.get(str(user["username"]), {}).get("sector_descripcion"),
                "sector_actualizado": assignments.get(str(user["username"]), {}).get("actualizado_en")
                or assignments.get(str(user["username"]), {}).get("fecha_hora"),
            }
            for user in users
        ],
        "sectores": sectors,
    }


@router.post("/usuarios-solicitantes/asignar-sector")
async def assign_request_user_sector(req: UserSectorAssignRequest, request: Request):
    auth = await _require_panol_operator(request, full=True)
    usernames = sorted({_clean(username) for username in req.usernames if _clean(username)})
    if not usernames:
        raise HTTPException(status_code=400, detail="Selecciona al menos un usuario.")
    if len(usernames) > 500:
        raise HTTPException(status_code=400, detail="Demasiados usuarios seleccionados.")
    async with auth_db(attach_operational=False) as adb:
        placeholders = ",".join("?" for _ in usernames)
        valid_rows = await _fetch_rows(
            adb,
            f"""
            SELECT u.username
            FROM auth_users u
            JOIN auth_user_app_access a ON a.username = u.username
            WHERE u.username IN ({placeholders})
              AND u.active = 1
              AND a.module = 'panol'
              AND a.enabled = 1
              AND UPPER(COALESCE(a.profile, '')) IN ('SOLICITANTE', 'PEDIDOS')
            """,
            tuple(usernames),
        )
    valid = {str(row["username"]) for row in valid_rows}
    missing = [username for username in usernames if username not in valid]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Hay usuarios sin perfil solicitante activo en Panol: {', '.join(missing[:8])}.",
        )
    now = _now()
    async with panol_db() as db:
        sectors = await _request_sectors(db)
        if not any(int(row["id"]) == int(req.sector_id) for row in sectors):
            raise HTTPException(status_code=400, detail="Sector solicitante invalido.")
        await db.executemany(
            """
            UPDATE usuarios_sectores_panol
            SET activo = 0, actualizado_por = ?, actualizado_en = ?
            WHERE username = ? AND activo = 1
            """,
            [(auth.get("username"), now, username) for username in usernames],
        )
        await db.executemany(
            """
            INSERT INTO usuarios_sectores_panol
                (username, sector_id, activo, creado_por, fecha_hora, actualizado_por, actualizado_en)
            VALUES (?, ?, 1, ?, ?, ?, ?)
            """,
            [(username, int(req.sector_id), auth.get("username"), now, auth.get("username"), now) for username in usernames],
        )
        await db.commit()
    return {"ok": True, "updated": len(usernames), "sector_id": int(req.sector_id)}


@router.get("/articulos")
async def list_articles(request: Request, include_inactive: bool = Query(False)):
    await _require_panol_operator(request)
    where = "" if include_inactive else "WHERE a.activo = 1"
    async with panol_db() as db:
        return {
            "items": await _fetch_rows(
                db,
                f"""
                SELECT a.*,
                       ac.costo_unitario,
                       ac.moneda AS costo_moneda,
                       ac.fecha_desde AS costo_fecha_desde
                FROM articulos a
                {_current_cost_join("a")}
                {where}
                ORDER BY a.activo DESC, a.codigo
                """,
            )
        }


@router.post("/articulos")
async def create_article(req: ArticleRequest, request: Request):
    auth = await _require_panol_operator(request, full=True)
    codigo = _norm_code(req.codigo)
    descripcion = _clean(req.descripcion)
    if not codigo or not descripcion:
        raise HTTPException(status_code=400, detail="Codigo y descripcion son obligatorios.")
    now = _now()
    try:
        async with panol_db() as db:
            cur = await db.execute(
                """
                INSERT INTO articulos
                    (codigo, descripcion, categoria, unidad, uso, stock_minimo, activo, creado_en, actualizado_en)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    codigo,
                    descripcion,
                    _clean(req.categoria),
                    _clean(req.unidad) or "UN",
                    _clean(req.uso),
                    max(float(req.stock_minimo or 0), 0),
                    1 if req.activo else 0,
                    now,
                    now,
                ),
            )
            await _set_article_cost(
                db,
                int(cur.lastrowid),
                req.costo_unitario,
                req.moneda,
                req.fecha_costo,
                str(auth.get("username") or ""),
            )
            await db.commit()
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="Ya existe un articulo con ese codigo.")
    return {"ok": True}


@router.put("/articulos/{articulo_id}")
async def update_article(articulo_id: int, req: ArticleRequest, request: Request):
    auth = await _require_panol_operator(request)
    codigo = _norm_code(req.codigo)
    descripcion = _clean(req.descripcion)
    if not codigo or not descripcion:
        raise HTTPException(status_code=400, detail="Codigo y descripcion son obligatorios.")
    try:
        async with panol_db() as db:
            cur = await db.execute(
                """
                UPDATE articulos
                SET codigo = ?, descripcion = ?, categoria = ?, unidad = ?, uso = ?, stock_minimo = ?,
                    activo = ?, actualizado_en = ?
                WHERE id = ?
                """,
                (
                    codigo,
                    descripcion,
                    _clean(req.categoria),
                    _clean(req.unidad) or "UN",
                    _clean(req.uso),
                    max(float(req.stock_minimo or 0), 0),
                    1 if req.activo else 0,
                    _now(),
                    articulo_id,
                ),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Articulo no encontrado.")
            await _set_article_cost(
                db,
                articulo_id,
                req.costo_unitario,
                req.moneda,
                req.fecha_costo,
                str(auth.get("username") or ""),
            )
            await db.commit()
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="Ya existe otro articulo con ese codigo.")
    return {"ok": True}


@router.post("/movimientos")
async def create_movement(req: MovementRequest, request: Request):
    auth = await _require_panol_operator(request)
    tipo = _norm_code(req.tipo)
    cantidad = float(req.cantidad or 0)
    if tipo not in MOVEMENT_TYPES:
        raise HTTPException(status_code=400, detail="Tipo de movimiento invalido.")
    if cantidad <= 0:
        raise HTTPException(status_code=400, detail="La cantidad debe ser mayor a cero.")

    async with panol_db() as db:
        article = await _fetch_one(db, "SELECT id FROM articulos WHERE id = ? AND activo = 1", (req.articulo_id,))
        if not article:
            raise HTTPException(status_code=404, detail="Articulo no encontrado o inactivo.")

        origin = req.ubicacion_origen_id
        dest = req.ubicacion_destino_id
        if tipo in INCOMING_TYPES and not dest:
            raise HTTPException(status_code=400, detail="El movimiento requiere ubicacion destino.")
        if tipo in OUTGOING_TYPES and not origin:
            raise HTTPException(status_code=400, detail="El movimiento requiere ubicacion origen.")
        if tipo == "TRANSFERENCIA":
            if not origin or not dest:
                raise HTTPException(status_code=400, detail="La transferencia requiere origen y destino.")
            if origin == dest:
                raise HTTPException(status_code=400, detail="Origen y destino deben ser distintos.")
        if origin:
            stock = await _stock_for_origin(db, req.articulo_id, origin)
            if stock + 0.000001 < cantidad:
                raise HTTPException(status_code=400, detail="El movimiento dejaria stock negativo.")

        await db.execute(
            """
            INSERT INTO movimientos
                (articulo_id, tipo, ubicacion_origen_id, ubicacion_destino_id, cantidad,
                 motivo, observacion, usuario, fecha_hora)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                req.articulo_id,
                tipo,
                origin,
                dest,
                cantidad,
                _clean(req.motivo),
                _clean(req.observacion),
                auth.get("username"),
                _now(),
            ),
        )
        await db.commit()
    return {"ok": True}


@router.get("/movimientos")
async def list_movements(
    request: Request,
    fecha_desde: str = Query(""),
    fecha_hasta: str = Query(""),
    articulo_id: int | None = Query(None),
    tipo: str = Query(""),
    usuario: str = Query(""),
):
    await _require_panol_operator(request)
    where = []
    args: list[Any] = []
    if fecha_desde:
        where.append("m.fecha_hora >= ?")
        args.append(f"{fecha_desde} 00:00:00")
    if fecha_hasta:
        where.append("m.fecha_hora <= ?")
        args.append(f"{fecha_hasta} 23:59:59")
    if articulo_id:
        where.append("m.articulo_id = ?")
        args.append(articulo_id)
    if tipo:
        where.append("m.tipo = ?")
        args.append(_norm_code(tipo))
    if usuario:
        where.append("m.usuario LIKE ?")
        args.append(f"%{_clean(usuario)}%")
    clause = "WHERE " + " AND ".join(where) if where else ""
    async with panol_db() as db:
        rows = await _fetch_rows(
            db,
            f"""
            SELECT m.*, a.codigo, a.descripcion,
                   uo.codigo AS origen_codigo, ud.codigo AS destino_codigo
            FROM movimientos m
            JOIN articulos a ON a.id = m.articulo_id
            LEFT JOIN ubicaciones uo ON uo.id = m.ubicacion_origen_id
            LEFT JOIN ubicaciones ud ON ud.id = m.ubicacion_destino_id
            {clause}
            ORDER BY m.fecha_hora DESC, m.id DESC
            LIMIT 500
            """,
            tuple(args),
        )
    return {"items": rows}


async def _stock_insumo_total(db: aiosqlite.Connection, articulo_id: int) -> tuple[float, float, float]:
    jaula_id = await _ubicacion_id(db, "JAULA")
    stock_jaula = await _stock_ubicacion(db, articulo_id, jaula_id)
    stock_oficina = await _stock_oficina(db, articulo_id)
    return stock_jaula + stock_oficina, stock_jaula, stock_oficina


async def _pedido_rows(db: aiosqlite.Connection, where: str, args: tuple[Any, ...], limit: int = 100) -> list[dict[str, Any]]:
    headers = await _fetch_rows(
        db,
        f"""
        SELECT p.*, u.codigo AS sector_codigo
        FROM pedidos_insumos p
        JOIN ubicaciones u ON u.id = p.sector_id
        {where}
        ORDER BY p.fecha_solicitud DESC, p.id DESC
        LIMIT {int(limit)}
        """,
        args,
    )
    for pedido in headers:
        pedido["items"] = await _fetch_rows(
            db,
            """
            SELECT i.*, a.codigo, a.descripcion, a.unidad, a.uso, u.codigo AS origen_insumo_codigo
            FROM pedidos_insumos_items i
            JOIN articulos a ON a.id = i.articulo_id
            LEFT JOIN ubicaciones u ON u.id = i.ubicacion_origen_insumo_id
            WHERE i.pedido_id = ?
            ORDER BY a.codigo
            """,
            (pedido["id"],),
        )
    return headers


async def _received_for_waste(
    db: aiosqlite.Connection,
    *,
    sector_id: int,
    articulo_id: int,
    tipo: str,
    pedido_id: int | None = None,
) -> float:
    amount_column = "i.cantidad_insumo_confirmada" if tipo == "INSUMO" else "i.cantidad_produccion_confirmada"
    where = [
        "p.sector_id = ?",
        "i.articulo_id = ?",
        "p.estado IN ('CONFIRMADO', 'CONFIRMADO_PARCIAL')",
    ]
    args: list[Any] = [sector_id, articulo_id]
    if pedido_id:
        where.append("p.id = ?")
        args.append(pedido_id)
    row = await _fetch_one(
        db,
        f"""
        SELECT COALESCE(SUM({amount_column}), 0) AS cantidad
        FROM pedidos_insumos p
        JOIN pedidos_insumos_items i ON i.pedido_id = p.id
        WHERE {" AND ".join(where)}
        """,
        tuple(args),
    )
    return float((row or {}).get("cantidad") or 0)


async def _waste_registered(
    db: aiosqlite.Connection,
    *,
    sector_id: int,
    articulo_id: int,
    tipo: str,
    pedido_id: int | None = None,
) -> float:
    where = ["sector_id = ?", "articulo_id = ?", "tipo = ?"]
    args: list[Any] = [sector_id, articulo_id, tipo]
    if pedido_id:
        where.append("pedido_id = ?")
        args.append(pedido_id)
    row = await _fetch_one(
        db,
        f"""
        SELECT COALESCE(SUM(cantidad), 0) AS cantidad
        FROM mermas_insumos
        WHERE {" AND ".join(where)}
        """,
        tuple(args),
    )
    return float((row or {}).get("cantidad") or 0)


@router.get("/pedidos/catalogo")
async def supply_order_catalog(request: Request):
    await _require_panol_access(request)
    async with panol_db() as db:
        articles = await _fetch_rows(
            db,
            """
            SELECT id, codigo, descripcion, categoria, unidad, uso
            FROM articulos
            WHERE activo = 1
            ORDER BY codigo
            """,
        )
        items = []
        for art in articles:
            stock_insumo, stock_jaula, stock_oficina = await _stock_insumo_total(db, int(art["id"]))
            stock_produccion = await _stock_producido(db, int(art["id"]))
            if stock_insumo <= 0 and stock_produccion <= 0:
                continue
            items.append(
                {
                    **art,
                    "stock_insumo": stock_insumo,
                    "stock_jaula": stock_jaula,
                    "stock_oficina": stock_oficina,
                    "stock_produccion": stock_produccion,
                }
            )
    return {"items": items}


@router.post("/pedidos")
async def create_supply_order(req: SupplyOrderRequest, request: Request):
    auth = await _require_panol_access(request)
    if not req.items:
        raise HTTPException(status_code=400, detail="El pedido no tiene lineas.")
    now = _now()
    clean_obs = _clean(req.observacion)
    async with panol_db() as db:
        sector_id = await _validate_request_sector(db, auth, int(req.sector_id))
        sector = await _fetch_one(db, "SELECT id FROM ubicaciones WHERE id = ? AND activo = 1", (sector_id,))
        if not sector:
            raise HTTPException(status_code=400, detail="Sector invalido.")
        lines: list[tuple[int, float, float]] = []
        seen: set[int] = set()
        for item in req.items:
            articulo_id = int(item.articulo_id)
            if articulo_id in seen:
                raise HTTPException(status_code=400, detail="No repitas el mismo PLU en el pedido.")
            seen.add(articulo_id)
            cantidad_insumo = float(item.cantidad_insumo or 0)
            cantidad_produccion = float(item.cantidad_produccion or 0)
            if cantidad_insumo < 0 or cantidad_produccion < 0:
                raise HTTPException(status_code=400, detail="Las cantidades no pueden ser negativas.")
            if cantidad_insumo <= 0 and cantidad_produccion <= 0:
                continue
            article = await _fetch_one(db, "SELECT id FROM articulos WHERE id = ? AND activo = 1", (articulo_id,))
            if not article:
                raise HTTPException(status_code=404, detail=f"Articulo {articulo_id} no encontrado o inactivo.")
            stock_insumo, _, _ = await _stock_insumo_total(db, articulo_id)
            stock_produccion = await _stock_producido(db, articulo_id)
            if cantidad_insumo > stock_insumo + 0.000001:
                raise HTTPException(status_code=400, detail="La cantidad de insumo pedida supera el stock disponible.")
            if cantidad_produccion > stock_produccion + 0.000001:
                raise HTTPException(status_code=400, detail="La cantidad de produccion pedida supera el stock disponible.")
            lines.append((articulo_id, cantidad_insumo, cantidad_produccion))
        if not lines:
            raise HTTPException(status_code=400, detail="Carga al menos una cantidad a pedir.")
        cur = await db.execute(
            """
            INSERT INTO pedidos_insumos
                (sector_id, estado, usuario_solicita, fecha_solicitud, observacion_solicitud)
            VALUES (?, 'PENDIENTE', ?, ?, ?)
            """,
            (sector_id, auth.get("username"), now, clean_obs),
        )
        pedido_id = cur.lastrowid
        await db.executemany(
            """
            INSERT INTO pedidos_insumos_items
                (pedido_id, articulo_id, cantidad_insumo_solicitada, cantidad_produccion_solicitada)
            VALUES (?, ?, ?, ?)
            """,
            [(pedido_id, articulo_id, cantidad_insumo, cantidad_produccion) for articulo_id, cantidad_insumo, cantidad_produccion in lines],
        )
        await db.commit()
    return {"ok": True, "pedido_id": pedido_id, "estado": "PENDIENTE", "fecha_solicitud": now}


@router.get("/pedidos/mios")
async def my_supply_orders(request: Request, sector_id: int | None = Query(None)):
    auth = await _require_panol_access(request)
    async with panol_db() as db:
        checked_sector_id = await _validate_request_sector(db, auth, sector_id)
        if checked_sector_id:
            sector = await _fetch_one(db, "SELECT id FROM ubicaciones WHERE id = ? AND activo = 1", (checked_sector_id,))
            if not sector:
                raise HTTPException(status_code=400, detail="Sector invalido.")
            return {"items": await _pedido_rows(db, "WHERE p.sector_id = ?", (checked_sector_id,), 150)}
        return {"items": await _pedido_rows(db, "WHERE p.usuario_solicita = ?", (auth.get("username"),), 150)}


@router.get("/pedidos/recibido-dia")
async def supply_order_received_by_day(
    request: Request,
    sector_id: int,
    tipo: str = Query("insumo"),
    fecha_desde: str = Query(""),
    fecha_hasta: str = Query(""),
    articulo_id: int | None = Query(None),
):
    auth = await _require_panol_access(request)
    received_type = _norm_key(tipo)
    if received_type not in {"insumo", "produccion"}:
        raise HTTPException(status_code=400, detail="Tipo invalido.")
    amount_column = "i.cantidad_insumo_confirmada" if received_type == "insumo" else "i.cantidad_produccion_confirmada"
    where = ["p.sector_id = ?", "p.estado IN ('CONFIRMADO', 'CONFIRMADO_PARCIAL')"]
    args: list[Any] = [sector_id]
    if fecha_desde:
        where.append("p.fecha_confirmacion >= ?")
        args.append(f"{fecha_desde} 00:00:00")
    if fecha_hasta:
        where.append("p.fecha_confirmacion <= ?")
        args.append(f"{fecha_hasta} 23:59:59")
    if articulo_id:
        where.append("i.articulo_id = ?")
        args.append(articulo_id)
    clause = "WHERE " + " AND ".join(where)
    async with panol_db() as db:
        checked_sector_id = await _validate_request_sector(db, auth, int(sector_id))
        args[0] = checked_sector_id
        sector = await _fetch_one(db, "SELECT id, codigo FROM ubicaciones WHERE id = ? AND activo = 1", (checked_sector_id,))
        if not sector:
            raise HTTPException(status_code=400, detail="Sector invalido.")
        rows = await _fetch_rows(
            db,
            f"""
            SELECT date(p.fecha_confirmacion) AS fecha,
                   COALESCE(SUM({amount_column}), 0) AS cantidad,
                   COUNT(DISTINCT p.id) AS pedidos
            FROM pedidos_insumos p
            JOIN pedidos_insumos_items i ON i.pedido_id = p.id
            {clause}
            GROUP BY date(p.fecha_confirmacion)
            ORDER BY fecha
            """,
            tuple(args),
        )
    if fecha_desde and fecha_hasta:
        try:
            start = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
            end = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha invalido.")
        if end < start:
            raise HTTPException(status_code=400, detail="Fecha fin menor a fecha inicio.")
        if (end - start).days <= 370:
            by_day = {str(row["fecha"]): row for row in rows}
            rows = [
                {
                    "fecha": (start + timedelta(days=offset)).isoformat(),
                    "cantidad": float((by_day.get((start + timedelta(days=offset)).isoformat()) or {}).get("cantidad") or 0),
                    "pedidos": int((by_day.get((start + timedelta(days=offset)).isoformat()) or {}).get("pedidos") or 0),
                }
                for offset in range((end - start).days + 1)
            ]
    return {
        "sector_id": checked_sector_id,
        "sector": sector.get("codigo"),
        "tipo": received_type,
        "items": rows,
    }


@router.get("/pedidos/mermas")
async def list_wastes(
    request: Request,
    sector_id: int | None = Query(None),
    fecha_desde: str = Query(""),
    fecha_hasta: str = Query(""),
    articulo_id: int | None = Query(None),
    tipo: str = Query(""),
    pedido_id: int | None = Query(None),
):
    auth = await _require_panol_access(request)
    where = []
    args: list[Any] = []
    async with panol_db() as db:
        checked_sector_id = await _validate_request_sector(db, auth, sector_id)
    if checked_sector_id:
        where.append("m.sector_id = ?")
        args.append(checked_sector_id)
    if fecha_desde:
        where.append("m.fecha_hora >= ?")
        args.append(f"{fecha_desde} 00:00:00")
    if fecha_hasta:
        where.append("m.fecha_hora <= ?")
        args.append(f"{fecha_hasta} 23:59:59")
    if articulo_id:
        where.append("m.articulo_id = ?")
        args.append(articulo_id)
    clean_tipo = _norm_code(tipo)
    if clean_tipo:
        if clean_tipo not in MERMA_TYPES:
            raise HTTPException(status_code=400, detail="Tipo de merma invalido.")
        where.append("m.tipo = ?")
        args.append(clean_tipo)
    if pedido_id:
        where.append("m.pedido_id = ?")
        args.append(pedido_id)
    clause = "WHERE " + " AND ".join(where) if where else ""
    async with panol_db() as db:
        rows = await _fetch_rows(
            db,
            f"""
            SELECT m.*, s.codigo AS sector_codigo, a.codigo, a.descripcion, p.estado AS pedido_estado
            FROM mermas_insumos m
            JOIN ubicaciones s ON s.id = m.sector_id
            JOIN articulos a ON a.id = m.articulo_id
            LEFT JOIN pedidos_insumos p ON p.id = m.pedido_id
            {clause}
            ORDER BY m.fecha_hora DESC, m.id DESC
            LIMIT 200
            """,
            tuple(args),
        )
    return {
        "items": [
            {**row, "motivo_label": MERMA_REASONS.get(str(row.get("motivo") or ""), str(row.get("motivo") or ""))}
            for row in rows
        ],
        "motivos": [{"codigo": code, "descripcion": label} for code, label in MERMA_REASONS.items()],
    }


@router.post("/pedidos/mermas")
async def create_waste(req: WasteRequest, request: Request):
    auth = await _require_panol_access(request)
    sector_id = int(req.sector_id)
    articulo_id = int(req.articulo_id)
    tipo = _norm_code(req.tipo)
    motivo = _norm_code(req.motivo)
    cantidad = float(req.cantidad or 0)
    observacion = _clean(req.observacion)
    pedido_id = int(req.pedido_id) if req.pedido_id else None

    if tipo not in MERMA_TYPES:
        raise HTTPException(status_code=400, detail="Tipo de merma invalido.")
    if motivo not in MERMA_REASONS:
        raise HTTPException(status_code=400, detail="Motivo de merma invalido.")
    if cantidad <= 0:
        raise HTTPException(status_code=400, detail="La cantidad de merma debe ser mayor a cero.")
    if not observacion:
        raise HTTPException(status_code=400, detail="La observacion de merma es obligatoria.")

    async with panol_db() as db:
        sector_id = int(await _validate_request_sector(db, auth, sector_id) or sector_id)
        sector = await _fetch_one(db, "SELECT id FROM ubicaciones WHERE id = ? AND activo = 1", (sector_id,))
        if not sector:
            raise HTTPException(status_code=400, detail="Sector invalido.")
        article = await _fetch_one(db, "SELECT id FROM articulos WHERE id = ? AND activo = 1", (articulo_id,))
        if not article:
            raise HTTPException(status_code=404, detail="Articulo no encontrado o inactivo.")
        if pedido_id:
            pedido = await _fetch_one(
                db,
                """
                SELECT id, sector_id, estado
                FROM pedidos_insumos
                WHERE id = ?
                """,
                (pedido_id,),
            )
            if not pedido:
                raise HTTPException(status_code=404, detail="Pedido asociado no encontrado.")
            if int(pedido["sector_id"]) != sector_id:
                raise HTTPException(status_code=400, detail="El pedido asociado no corresponde al sector.")
            if pedido["estado"] not in {"CONFIRMADO", "CONFIRMADO_PARCIAL"}:
                raise HTTPException(status_code=400, detail="Solo se pueden asociar pedidos confirmados o parciales.")

        recibido = await _received_for_waste(
            db,
            sector_id=sector_id,
            articulo_id=articulo_id,
            tipo=tipo,
            pedido_id=pedido_id,
        )
        ya_mermado = await _waste_registered(
            db,
            sector_id=sector_id,
            articulo_id=articulo_id,
            tipo=tipo,
            pedido_id=pedido_id,
        )
        disponible_merma = max(recibido - ya_mermado, 0)
        if cantidad > disponible_merma + 0.000001:
            scope = "pedido asociado" if pedido_id else "sector y PLU seleccionados"
            raise HTTPException(
                status_code=400,
                detail=f"La merma supera lo recibido disponible para el {scope}. Disponible: {disponible_merma:.3f}.",
            )
        if pedido_id:
            recibido_total = await _received_for_waste(
                db,
                sector_id=sector_id,
                articulo_id=articulo_id,
                tipo=tipo,
                pedido_id=None,
            )
            mermado_total = await _waste_registered(
                db,
                sector_id=sector_id,
                articulo_id=articulo_id,
                tipo=tipo,
                pedido_id=None,
            )
            disponible_total = max(recibido_total - mermado_total, 0)
            if cantidad > disponible_total + 0.000001:
                raise HTTPException(
                    status_code=400,
                    detail=f"La merma supera lo recibido disponible para el sector y PLU seleccionados. Disponible: {disponible_total:.3f}.",
                )

        now = _now()
        cur = await db.execute(
            """
            INSERT INTO mermas_insumos
                (sector_id, pedido_id, articulo_id, tipo, cantidad, motivo, observacion, usuario, fecha_hora)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                sector_id,
                pedido_id,
                articulo_id,
                tipo,
                cantidad,
                motivo,
                observacion,
                auth.get("username"),
                now,
            ),
        )
        await db.commit()
    return {"ok": True, "merma_id": cur.lastrowid, "fecha_hora": now}


@router.get("/pedidos/pendientes")
async def pending_supply_orders(request: Request):
    await _require_panol_operator(request)
    async with panol_db() as db:
        items = await _pedido_rows(db, "WHERE p.estado = 'PENDIENTE'", (), 150)
    return {"items": items, "pending_count": len(items)}


@router.get("/pedidos/indicadores")
async def supply_order_indicators(
    request: Request,
    fecha_desde: str = Query(""),
    fecha_hasta: str = Query(""),
    articulo_id: int | None = Query(None),
):
    await _require_panol_operator(request)
    where = []
    args: list[Any] = []
    if fecha_desde:
        where.append("p.fecha_solicitud >= ?")
        args.append(f"{fecha_desde} 00:00:00")
    if fecha_hasta:
        where.append("p.fecha_solicitud <= ?")
        args.append(f"{fecha_hasta} 23:59:59")
    if articulo_id:
        where.append("i.articulo_id = ?")
        args.append(articulo_id)
    clause = "WHERE " + " AND ".join(where) if where else ""
    waste_where = []
    waste_args: list[Any] = []
    if fecha_desde:
        waste_where.append("m.fecha_hora >= ?")
        waste_args.append(f"{fecha_desde} 00:00:00")
    if fecha_hasta:
        waste_where.append("m.fecha_hora <= ?")
        waste_args.append(f"{fecha_hasta} 23:59:59")
    if articulo_id:
        waste_where.append("m.articulo_id = ?")
        waste_args.append(articulo_id)
    waste_clause = "WHERE " + " AND ".join(waste_where) if waste_where else ""
    order_cost = _cost_at_sql("i.articulo_id", "COALESCE(p.fecha_confirmacion, p.fecha_solicitud)")
    waste_cost = _cost_at_sql("m.articulo_id", "m.fecha_hora")
    async with panol_db() as db:
        metrics = await _fetch_one(
            db,
            f"""
            SELECT
                COUNT(DISTINCT p.id) AS total,
                COUNT(DISTINCT CASE WHEN p.estado = 'PENDIENTE' THEN p.id END) AS pendientes,
                COUNT(DISTINCT CASE WHEN p.estado = 'CONFIRMADO' THEN p.id END) AS confirmados,
                COUNT(DISTINCT CASE WHEN p.estado = 'CONFIRMADO_PARCIAL' THEN p.id END) AS parciales,
                COALESCE(SUM(i.cantidad_insumo_solicitada + i.cantidad_produccion_solicitada), 0) AS solicitado,
                COALESCE(SUM(i.cantidad_insumo_confirmada + i.cantidad_produccion_confirmada), 0) AS confirmado,
                COALESCE(SUM((i.cantidad_insumo_solicitada + i.cantidad_produccion_solicitada) * {order_cost}), 0) AS valor_solicitado,
                COALESCE(SUM((i.cantidad_insumo_confirmada + i.cantidad_produccion_confirmada) * {order_cost}), 0) AS valor_confirmado
            FROM pedidos_insumos p
            JOIN pedidos_insumos_items i ON i.pedido_id = p.id
            {clause}
            """,
            tuple(args),
        )
        waste_metrics = await _fetch_one(
            db,
            f"""
            SELECT COUNT(*) AS registros,
                   COALESCE(SUM(cantidad), 0) AS cantidad,
                   COALESCE(SUM(m.cantidad * {waste_cost}), 0) AS valor
            FROM mermas_insumos m
            {waste_clause}
            """,
            tuple(waste_args),
        )
        by_sector = await _fetch_rows(
            db,
            f"""
            SELECT u.codigo AS sector,
                   COUNT(DISTINCT p.id) AS pedidos,
                   COALESCE(SUM(i.cantidad_insumo_solicitada + i.cantidad_produccion_solicitada), 0) AS cantidad,
                   COALESCE(SUM((i.cantidad_insumo_solicitada + i.cantidad_produccion_solicitada) * {order_cost}), 0) AS valor
            FROM pedidos_insumos p
            JOIN ubicaciones u ON u.id = p.sector_id
            JOIN pedidos_insumos_items i ON i.pedido_id = p.id
            {clause}
            GROUP BY u.codigo
            ORDER BY pedidos DESC, cantidad DESC, u.codigo
            LIMIT 12
            """,
            tuple(args),
        )
        waste_by_sector = await _fetch_rows(
            db,
            f"""
            SELECT u.codigo AS sector,
                   COUNT(*) AS registros,
                   COALESCE(SUM(m.cantidad), 0) AS cantidad,
                   COALESCE(SUM(m.cantidad * {waste_cost}), 0) AS valor
            FROM mermas_insumos m
            JOIN ubicaciones u ON u.id = m.sector_id
            {waste_clause}
            GROUP BY u.codigo
            ORDER BY cantidad DESC, registros DESC, u.codigo
            LIMIT 12
            """,
            tuple(waste_args),
        )
        by_plu = await _fetch_rows(
            db,
            f"""
            SELECT a.codigo, a.descripcion,
                   COALESCE(SUM(i.cantidad_insumo_solicitada + i.cantidad_produccion_solicitada), 0) AS cantidad,
                   COALESCE(SUM((i.cantidad_insumo_solicitada + i.cantidad_produccion_solicitada) * {order_cost}), 0) AS valor
            FROM pedidos_insumos p
            JOIN pedidos_insumos_items i ON i.pedido_id = p.id
            JOIN articulos a ON a.id = i.articulo_id
            {clause}
            GROUP BY a.codigo, a.descripcion
            ORDER BY cantidad DESC, a.codigo
            LIMIT 12
            """,
            tuple(args),
        )
        waste_by_plu = await _fetch_rows(
            db,
            f"""
            SELECT a.codigo, a.descripcion,
                   COALESCE(SUM(m.cantidad), 0) AS cantidad,
                   COALESCE(SUM(m.cantidad * {waste_cost}), 0) AS valor,
                   COUNT(*) AS registros
            FROM mermas_insumos m
            JOIN articulos a ON a.id = m.articulo_id
            {waste_clause}
            GROUP BY a.codigo, a.descripcion
            ORDER BY cantidad DESC, a.codigo
            LIMIT 12
            """,
            tuple(waste_args),
        )
        waste_by_reason = await _fetch_rows(
            db,
            f"""
            SELECT m.motivo,
                   COALESCE(SUM(m.cantidad), 0) AS cantidad,
                   COALESCE(SUM(m.cantidad * {waste_cost}), 0) AS valor,
                   COUNT(*) AS registros
            FROM mermas_insumos m
            {waste_clause}
            GROUP BY m.motivo
            ORDER BY cantidad DESC, m.motivo
            LIMIT 12
            """,
            tuple(waste_args),
        )
        recent = await _fetch_rows(
            db,
            f"""
            SELECT p.id, p.estado, p.fecha_solicitud, p.usuario_solicita, u.codigo AS sector,
                   COUNT(i.id) AS lineas,
                   COALESCE(SUM(i.cantidad_insumo_solicitada + i.cantidad_produccion_solicitada), 0) AS cantidad,
                   COALESCE(SUM((i.cantidad_insumo_solicitada + i.cantidad_produccion_solicitada) * {order_cost}), 0) AS valor
            FROM pedidos_insumos p
            JOIN ubicaciones u ON u.id = p.sector_id
            JOIN pedidos_insumos_items i ON i.pedido_id = p.id
            {clause}
            GROUP BY p.id, p.estado, p.fecha_solicitud, p.usuario_solicita, u.codigo
            ORDER BY p.fecha_solicitud DESC, p.id DESC
            LIMIT 8
            """,
            tuple(args),
        )
    return {
        "metrics": {
            **(metrics or {}),
            "mermas_registros": int((waste_metrics or {}).get("registros") or 0),
            "mermas_cantidad": float((waste_metrics or {}).get("cantidad") or 0),
            "mermas_valor": float((waste_metrics or {}).get("valor") or 0),
        },
        "pedidos_por_sector": by_sector,
        "pedidos_por_plu": by_plu,
        "mermas_por_sector": waste_by_sector,
        "mermas_por_plu": waste_by_plu,
        "mermas_por_motivo": [
            {**row, "motivo_label": MERMA_REASONS.get(str(row.get("motivo") or ""), str(row.get("motivo") or ""))}
            for row in waste_by_reason
        ],
        "recientes": recent,
    }


@router.post("/pedidos/{pedido_id}/confirmar")
async def confirm_supply_order(pedido_id: int, req: SupplyOrderConfirmRequest, request: Request):
    auth = await _require_panol_operator(request)
    if not req.items:
        raise HTTPException(status_code=400, detail="No hay lineas para confirmar.")
    now = _now()
    turno = _turno_por_hora(now)
    clean_obs = _clean(req.observacion)
    async with panol_db() as db:
        pedido = await _fetch_one(
            db,
            """
            SELECT p.*, u.codigo AS sector_codigo
            FROM pedidos_insumos p
            JOIN ubicaciones u ON u.id = p.sector_id
            WHERE p.id = ?
            """,
            (pedido_id,),
        )
        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado.")
        if pedido["estado"] != "PENDIENTE":
            raise HTTPException(status_code=400, detail="El pedido ya fue procesado.")
        requested_items = {
            int(row["id"]): row
            for row in await _fetch_rows(
                db,
                """
                SELECT i.*, a.uso
                FROM pedidos_insumos_items i
                JOIN articulos a ON a.id = i.articulo_id
                WHERE i.pedido_id = ?
                """,
                (pedido_id,),
            )
        }
        confirm_by_id = {int(item.item_id): item for item in req.items}
        oficina_id = await _ubicacion_id(db, "OFICINA_ADO")
        any_partial = False
        any_confirmed = False
        for item_id, row in requested_items.items():
            confirm = confirm_by_id.get(item_id)
            qty_insumo = float((confirm.cantidad_insumo_confirmada if confirm else 0) or 0)
            qty_produccion = float((confirm.cantidad_produccion_confirmada if confirm else 0) or 0)
            selected_usage = _clean(confirm.uso_entrega if confirm else "")
            if qty_insumo < 0 or qty_produccion < 0:
                raise HTTPException(status_code=400, detail="Las cantidades confirmadas no pueden ser negativas.")
            req_insumo = float(row["cantidad_insumo_solicitada"] or 0)
            req_produccion = float(row["cantidad_produccion_solicitada"] or 0)
            if qty_insumo > req_insumo + 0.000001 or qty_produccion > req_produccion + 0.000001:
                raise HTTPException(status_code=400, detail="No se puede confirmar mas de lo solicitado.")
            if qty_insumo + 0.000001 < req_insumo or qty_produccion + 0.000001 < req_produccion:
                any_partial = True
            origin_id = int(confirm.ubicacion_origen_insumo_id) if confirm and confirm.ubicacion_origen_insumo_id else oficina_id
            if qty_insumo > 0:
                origin = await _fetch_one(db, "SELECT id FROM ubicaciones WHERE id = ? AND activo = 1", (origin_id,))
                if not origin:
                    raise HTTPException(status_code=400, detail="Origen de insumo invalido.")
                stock_origin = await _stock_for_origin(db, int(row["articulo_id"]), origin_id)
                if stock_origin + 0.000001 < qty_insumo:
                    raise HTTPException(status_code=400, detail="La confirmacion de insumo dejaria stock negativo.")
            if qty_produccion > 0:
                stock_produccion = await _stock_producido(db, int(row["articulo_id"]))
                if stock_produccion + 0.000001 < qty_produccion:
                    raise HTTPException(status_code=400, detail="La confirmacion de produccion dejaria stock negativo.")
            if qty_insumo > 0:
                await db.execute(
                    """
                    INSERT INTO movimientos
                        (articulo_id, tipo, ubicacion_origen_id, ubicacion_destino_id, cantidad,
                         motivo, observacion, usuario, fecha_hora)
                    VALUES (?, 'BAJA', ?, NULL, ?, 'PEDIDO_INSUMOS', ?, ?, ?)
                    """,
                    (
                        int(row["articulo_id"]),
                        origin_id,
                        qty_insumo,
                        clean_obs or f"Pedido #{pedido_id} - {pedido['sector_codigo']}",
                        auth.get("username"),
                        now,
                    ),
                )
            if qty_produccion > 0:
                await db.execute(
                    """
                    INSERT INTO produccion_movimientos
                        (articulo_id, tipo, ubicacion_destino_id, cantidad, turno, observacion, usuario, fecha_hora)
                    VALUES (?, 'ENTREGA', ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        int(row["articulo_id"]),
                        int(pedido["sector_id"]),
                        qty_produccion,
                        turno,
                        clean_obs or f"Pedido #{pedido_id}",
                        auth.get("username"),
                        now,
                    ),
                )
            if qty_insumo > 0 or qty_produccion > 0:
                article_usage = _clean(row.get("uso"))
                if article_usage and not selected_usage:
                    raise HTTPException(status_code=400, detail="Selecciona el uso de entrega para los PLUs que lo requieren.")
                if article_usage:
                    valid_usages = {usage.upper() for usage in _usage_options(article_usage)}
                    if selected_usage.upper() not in valid_usages:
                        raise HTTPException(status_code=400, detail="Uso de entrega invalido para el PLU.")
                any_confirmed = True
            await db.execute(
                """
                UPDATE pedidos_insumos_items
                SET cantidad_insumo_confirmada = ?,
                    cantidad_produccion_confirmada = ?,
                    ubicacion_origen_insumo_id = ?,
                    uso_entrega = ?
                WHERE id = ?
                """,
                (
                    qty_insumo,
                    qty_produccion,
                    origin_id if qty_insumo > 0 else None,
                    selected_usage if (qty_insumo > 0 or qty_produccion > 0) else "",
                    item_id,
                ),
            )
        estado = "CONFIRMADO_PARCIAL" if any_partial else "CONFIRMADO"
        if not any_confirmed:
            estado = "CONFIRMADO_PARCIAL"
        await db.execute(
            """
            UPDATE pedidos_insumos
            SET estado = ?, usuario_confirma = ?, fecha_confirmacion = ?, observacion_confirmacion = ?
            WHERE id = ?
            """,
            (estado, auth.get("username"), now, clean_obs, pedido_id),
        )
        await db.commit()
    return {"ok": True, "pedido_id": pedido_id, "estado": estado, "fecha_confirmacion": now}


@router.post("/pedidos/{pedido_id}/cancelar")
async def cancel_supply_order(pedido_id: int, req: SupplyOrderConfirmRequest, request: Request):
    auth = await _require_panol_access(request)
    async with panol_db() as db:
        pedido = await _fetch_one(
            db,
            """
            SELECT p.*, u.codigo AS sector_codigo
            FROM pedidos_insumos p
            JOIN ubicaciones u ON u.id = p.sector_id
            WHERE p.id = ?
            """,
            (pedido_id,),
        )
        if not pedido:
            raise HTTPException(status_code=404, detail="Pedido no encontrado.")
        if pedido["estado"] != "PENDIENTE":
            raise HTTPException(status_code=400, detail="El pedido ya fue procesado.")
        if auth.get("panol_request_only"):
            assigned = await _active_user_sector(db, str(auth.get("username") or ""))
            if not assigned:
                raise HTTPException(
                    status_code=403,
                    detail="Tu usuario no tiene sector asignado en Panol. Solicita la asignacion a ADO.",
                )
            if int(assigned["sector_id"]) != int(pedido["sector_id"]):
                raise HTTPException(
                    status_code=403,
                    detail="Solo el sector generador puede cancelar este pedido pendiente.",
                )
        cur = await db.execute(
            """
            UPDATE pedidos_insumos
            SET estado = 'CANCELADO', usuario_confirma = ?, fecha_confirmacion = ?, observacion_confirmacion = ?
            WHERE id = ? AND estado = 'PENDIENTE'
            """,
            (auth.get("username"), _now(), _clean(req.observacion), pedido_id),
        )
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Pedido pendiente no encontrado.")
        await db.commit()
    return {"ok": True, "pedido_id": pedido_id, "estado": "CANCELADO"}


@router.post("/admin/reset-operativo")
async def reset_operational_data(req: OperationalResetRequest, request: Request):
    await _require_panol_operator(request, full=True)
    if req.clave != "Ingenieria12345":
        raise HTTPException(status_code=403, detail="Clave de depuracion invalida.")
    tables = [
        "mermas_insumos",
        "pedidos_insumos_items",
        "pedidos_insumos",
        "movimientos",
        "stock_cd_importado",
        "consumos_calculados",
        "inventario_turno",
        "produccion_movimientos",
    ]
    deleted: dict[str, int] = {}
    async with panol_db() as db:
        await db.execute("BEGIN")
        try:
            for table in tables:
                row = await _fetch_one(db, f"SELECT COUNT(*) AS qty FROM {table}")
                deleted[table] = int((row or {}).get("qty") or 0)
                await db.execute(f"DELETE FROM {table}")
            placeholders = ",".join("?" for _ in tables)
            await db.execute(
                f"DELETE FROM sqlite_sequence WHERE name IN ({placeholders})",
                tuple(tables),
            )
            await db.commit()
        except Exception:
            await db.rollback()
            raise
    return {"ok": True, "deleted": deleted}


@router.post("/produccion")
async def create_production(req: ProductionRequest, request: Request):
    auth = await _require_panol_operator(request)
    cantidad = float(req.cantidad or 0)
    if cantidad <= 0:
        raise HTTPException(status_code=400, detail="La cantidad producida debe ser mayor a cero.")
    now = _now()
    turno = _turno_por_hora(now)
    async with panol_db() as db:
        article = await _fetch_one(db, "SELECT id FROM articulos WHERE id = ? AND activo = 1", (req.articulo_id,))
        if not article:
            raise HTTPException(status_code=404, detail="Articulo no encontrado o inactivo.")
        await db.execute(
            """
            INSERT INTO produccion_movimientos
                (articulo_id, tipo, ubicacion_destino_id, cantidad, turno, observacion, usuario, fecha_hora)
            VALUES (?, 'PRODUCCION', NULL, ?, ?, ?, ?, ?)
            """,
            (req.articulo_id, cantidad, turno, _clean(req.observacion), auth.get("username"), now),
        )
        await db.commit()
    return {"ok": True, "turno": turno, "fecha_hora": now}


@router.post("/produccion/entregas")
async def create_production_delivery(req: ProductionDeliveryRequest, request: Request):
    auth = await _require_panol_operator(request)
    cantidad = float(req.cantidad or 0)
    if cantidad <= 0:
        raise HTTPException(status_code=400, detail="La cantidad entregada debe ser mayor a cero.")
    now = _now()
    turno = _turno_por_hora(now)
    async with panol_db() as db:
        article = await _fetch_one(db, "SELECT id FROM articulos WHERE id = ? AND activo = 1", (req.articulo_id,))
        if not article:
            raise HTTPException(status_code=404, detail="Articulo no encontrado o inactivo.")
        destination = await _fetch_one(
            db,
            "SELECT id FROM ubicaciones WHERE id = ? AND activo = 1",
            (req.ubicacion_destino_id,),
        )
        if not destination:
            raise HTTPException(status_code=400, detail="Destino invalido.")
        stock_actual = await _stock_producido(db, req.articulo_id)
        if stock_actual + 0.000001 < cantidad:
            raise HTTPException(status_code=400, detail="La entrega dejaria stock producido negativo.")
        await db.execute(
            """
            INSERT INTO produccion_movimientos
                (articulo_id, tipo, ubicacion_destino_id, cantidad, turno, observacion, usuario, fecha_hora)
            VALUES (?, 'ENTREGA', ?, ?, ?, ?, ?, ?)
            """,
            (
                req.articulo_id,
                req.ubicacion_destino_id,
                cantidad,
                turno,
                _clean(req.observacion),
                auth.get("username"),
                now,
            ),
        )
        await db.commit()
    return {"ok": True, "turno": turno, "fecha_hora": now}


@router.get("/produccion/stock")
async def production_stock(request: Request, q: str = Query(""), articulo_id: int | None = Query(None)):
    await _require_panol_operator(request)
    async with panol_db() as db:
        rows = await _fetch_rows(
            db,
            f"""
            SELECT a.id AS articulo_id, a.codigo, a.descripcion, a.categoria, a.unidad,
                   ac.costo_unitario, ac.moneda AS costo_moneda, ac.fecha_desde AS costo_fecha_desde,
                   COALESCE(SUM(CASE WHEN pm.tipo = 'PRODUCCION' THEN pm.cantidad ELSE 0 END), 0) AS producido,
                   COALESCE(SUM(CASE WHEN pm.tipo = 'ENTREGA' THEN pm.cantidad ELSE 0 END), 0) AS entregado,
                   COALESCE(SUM(CASE WHEN pm.tipo = 'PRODUCCION' THEN pm.cantidad ELSE 0 END), 0)
                 - COALESCE(SUM(CASE WHEN pm.tipo = 'ENTREGA' THEN pm.cantidad ELSE 0 END), 0) AS stock_producido
            FROM articulos a
            {_current_cost_join("a")}
            LEFT JOIN produccion_movimientos pm ON pm.articulo_id = a.id
            WHERE a.activo = 1
              AND (? IS NULL OR a.id = ?)
            GROUP BY a.id, a.codigo, a.descripcion, a.categoria, a.unidad,
                     ac.costo_unitario, ac.moneda, ac.fecha_desde
            ORDER BY a.codigo
            """,
            (articulo_id, articulo_id),
        )
        today = await _fetch_one(
            db,
            """
            SELECT
                COALESCE(SUM(CASE WHEN tipo = 'PRODUCCION' THEN cantidad ELSE 0 END), 0) AS producido,
                COALESCE(SUM(CASE WHEN tipo = 'ENTREGA' THEN cantidad ELSE 0 END), 0) AS entregado
            FROM produccion_movimientos
            WHERE fecha_hora >= ?
              AND (? IS NULL OR articulo_id = ?)
            """,
            (f"{_today()} 00:00:00", articulo_id, articulo_id),
        )
    items = []
    for row in rows:
        if q and q.lower() not in f"{row['codigo']} {row['descripcion']}".lower():
            continue
        cost = float(row["costo_unitario"]) if row.get("costo_unitario") is not None else None
        items.append(
            {
                **row,
                "valor_stock_producido": None if cost is None else float(row.get("stock_producido") or 0) * cost,
            }
        )
    return {
        "items": items,
        "metrics": {
            "producido_hoy": float((today or {}).get("producido") or 0),
            "entregado_hoy": float((today or {}).get("entregado") or 0),
            "stock_total_producido": sum(float(row.get("stock_producido") or 0) for row in rows),
            "valor_stock_producido": sum(float(row.get("valor_stock_producido") or 0) for row in items),
            "articulos_sin_costo": sum(1 for row in items if row.get("costo_unitario") is None),
        },
    }


@router.get("/produccion/movimientos")
async def list_production_movements(
    request: Request,
    fecha_desde: str = Query(""),
    fecha_hasta: str = Query(""),
    articulo_id: int | None = Query(None),
    tipo: str = Query(""),
    destino_id: int | None = Query(None),
    turno: str = Query(""),
):
    await _require_panol_operator(request)
    where = []
    args: list[Any] = []
    if fecha_desde:
        where.append("pm.fecha_hora >= ?")
        args.append(f"{fecha_desde} 00:00:00")
    if fecha_hasta:
        where.append("pm.fecha_hora <= ?")
        args.append(f"{fecha_hasta} 23:59:59")
    if articulo_id:
        where.append("pm.articulo_id = ?")
        args.append(articulo_id)
    if tipo:
        where.append("pm.tipo = ?")
        args.append(_norm_code(tipo))
    if destino_id:
        where.append("pm.ubicacion_destino_id = ?")
        args.append(destino_id)
    if turno:
        where.append("pm.turno = ?")
        args.append(_norm_code(turno))
    clause = "WHERE " + " AND ".join(where) if where else ""
    async with panol_db() as db:
        rows = await _fetch_rows(
            db,
            f"""
            SELECT pm.*, a.codigo, a.descripcion, u.codigo AS destino_codigo
            FROM produccion_movimientos pm
            JOIN articulos a ON a.id = pm.articulo_id
            LEFT JOIN ubicaciones u ON u.id = pm.ubicacion_destino_id
            {clause}
            ORDER BY pm.fecha_hora DESC, pm.id DESC
            LIMIT 500
            """,
            tuple(args),
        )
    return {"items": rows}


@router.get("/produccion/dia-logistico")
async def production_logistic_day(request: Request):
    await _require_panol_operator(request)
    since, until = _logistic_day_bounds()
    async with panol_db() as db:
        produced = await _fetch_rows(
            db,
            """
            SELECT a.id AS articulo_id, a.codigo, a.descripcion,
                   COALESCE(SUM(pm.cantidad), 0) AS cantidad
            FROM produccion_movimientos pm
            JOIN articulos a ON a.id = pm.articulo_id
            WHERE pm.tipo = 'PRODUCCION'
              AND pm.fecha_hora >= ?
              AND pm.fecha_hora <= ?
            GROUP BY a.id, a.codigo, a.descripcion
            ORDER BY cantidad DESC, a.codigo
            LIMIT 20
            """,
            (since, until),
        )
        delivered = await _fetch_rows(
            db,
            """
            SELECT a.id AS articulo_id, a.codigo, a.descripcion,
                   u.codigo AS destino_codigo,
                   COALESCE(SUM(pm.cantidad), 0) AS cantidad
            FROM produccion_movimientos pm
            JOIN articulos a ON a.id = pm.articulo_id
            LEFT JOIN ubicaciones u ON u.id = pm.ubicacion_destino_id
            WHERE pm.tipo = 'ENTREGA'
              AND pm.fecha_hora >= ?
              AND pm.fecha_hora <= ?
            GROUP BY a.id, a.codigo, a.descripcion, u.codigo
            ORDER BY cantidad DESC, a.codigo, u.codigo
            LIMIT 20
            """,
            (since, until),
        )
    return {
        "desde": since,
        "hasta": until,
        "producido": produced,
        "entregado": delivered,
    }


@router.get("/produccion/indicadores")
async def production_indicators(
    request: Request,
    fecha_desde: str = Query(""),
    fecha_hasta: str = Query(""),
    articulo_id: int | None = Query(None),
):
    await _require_panol_operator(request)
    where = []
    args: list[Any] = []
    if fecha_desde:
        where.append("fecha_hora >= ?")
        args.append(f"{fecha_desde} 00:00:00")
    if fecha_hasta:
        where.append("fecha_hora <= ?")
        args.append(f"{fecha_hasta} 23:59:59")
    if articulo_id:
        where.append("articulo_id = ?")
        args.append(articulo_id)
    clause = "AND " + " AND ".join(where) if where else ""
    pm_clause = clause.replace("fecha_hora", "pm.fecha_hora")
    async with panol_db() as db:
        by_sector = await _fetch_rows(
            db,
            f"""
            SELECT u.id AS ubicacion_id, u.codigo AS sector, COALESCE(SUM(pm.cantidad), 0) AS cantidad
            FROM produccion_movimientos pm
            JOIN ubicaciones u ON u.id = pm.ubicacion_destino_id
            WHERE pm.tipo = 'ENTREGA'
              {pm_clause}
            GROUP BY u.id, u.codigo
            ORDER BY cantidad DESC
            """,
            tuple(args),
        )
        by_sector_plu = await _fetch_rows(
            db,
            f"""
            SELECT u.codigo AS sector, a.codigo, a.descripcion, COALESCE(SUM(pm.cantidad), 0) AS cantidad
            FROM produccion_movimientos pm
            JOIN ubicaciones u ON u.id = pm.ubicacion_destino_id
            JOIN articulos a ON a.id = pm.articulo_id
            WHERE pm.tipo = 'ENTREGA'
              {pm_clause}
            GROUP BY u.codigo, a.codigo, a.descripcion
            ORDER BY cantidad DESC, u.codigo, a.codigo
            LIMIT 200
            """,
            tuple(args),
        )
    return {
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "articulo_id": articulo_id,
        "entregas_por_sector": by_sector,
        "entregas_por_sector_plu": by_sector_plu,
    }


@router.get("/produccion/indicadores/exportar")
async def export_production_indicators(
    request: Request,
    fecha_desde: str = Query(""),
    fecha_hasta: str = Query(""),
    articulo_id: int | None = Query(None),
):
    await _require_panol_operator(request)
    where = []
    args: list[Any] = []
    if fecha_desde:
        where.append("pm.fecha_hora >= ?")
        args.append(f"{fecha_desde} 00:00:00")
    if fecha_hasta:
        where.append("pm.fecha_hora <= ?")
        args.append(f"{fecha_hasta} 23:59:59")
    if articulo_id:
        where.append("pm.articulo_id = ?")
        args.append(articulo_id)
    clause = "AND " + " AND ".join(where) if where else ""
    async with panol_db() as db:
        summary = await _fetch_rows(
            db,
            f"""
            SELECT u.codigo AS sector, a.codigo, a.descripcion, COALESCE(SUM(pm.cantidad), 0) AS cantidad
            FROM produccion_movimientos pm
            JOIN ubicaciones u ON u.id = pm.ubicacion_destino_id
            JOIN articulos a ON a.id = pm.articulo_id
            WHERE pm.tipo = 'ENTREGA'
              {clause}
            GROUP BY u.codigo, a.codigo, a.descripcion
            ORDER BY u.codigo, a.codigo
            """,
            tuple(args),
        )
        movements = await _fetch_rows(
            db,
            f"""
            SELECT pm.fecha_hora, u.codigo AS sector, a.codigo, a.descripcion,
                   pm.cantidad, pm.turno, pm.usuario, pm.observacion
            FROM produccion_movimientos pm
            JOIN ubicaciones u ON u.id = pm.ubicacion_destino_id
            JOIN articulos a ON a.id = pm.articulo_id
            WHERE pm.tipo = 'ENTREGA'
              {clause}
            ORDER BY pm.fecha_hora DESC, u.codigo, a.codigo
            """,
            tuple(args),
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Rango desde", fecha_desde or "Inicio", "Rango hasta", fecha_hasta or "Actual", "Articulo ID", articulo_id or "Todos"])
    ws.append([])
    ws.append(["Sector", "PLU", "Descripcion", "Cantidad entregada"])
    for row in summary:
        ws.append([row.get("sector"), row.get("codigo"), row.get("descripcion"), float(row.get("cantidad") or 0)])

    detail = wb.create_sheet("Movimientos")
    detail.append(["Fecha hora", "Sector", "PLU", "Descripcion", "Cantidad", "Turno", "Usuario", "Observacion"])
    for row in movements:
        detail.append(
            [
                row.get("fecha_hora"),
                row.get("sector"),
                row.get("codigo"),
                row.get("descripcion"),
                float(row.get("cantidad") or 0),
                row.get("turno"),
                row.get("usuario"),
                row.get("observacion"),
            ]
        )
    for sheet in (ws, detail):
        for column_cells in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 48)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    suffix = f"_{fecha_desde or 'inicio'}_{fecha_hasta or 'actual'}"
    if articulo_id:
        suffix += f"_plu_{articulo_id}"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=produccion_entregas_plu_sector{suffix}.xlsx"},
    )


@router.get("/stock")
async def stock(
    request: Request,
    q: str = Query(""),
    categoria: str = Query(""),
    bajo_minimo: bool = Query(False),
    articulo_id: int | None = Query(None),
):
    await _require_panol_operator(request)
    async with panol_db() as db:
        jaula_id = await _ubicacion_id(db, "JAULA")
        articles = await _fetch_rows(
            db,
            f"""
            SELECT a.*,
                   ac.costo_unitario,
                   ac.moneda AS costo_moneda,
                   ac.fecha_desde AS costo_fecha_desde
            FROM articulos a
            {_current_cost_join("a")}
            WHERE a.activo = 1
            ORDER BY a.codigo
            """,
        )
        items = []
        for art in articles:
            if articulo_id and int(art["id"]) != articulo_id:
                continue
            if q and q.lower() not in f"{art['codigo']} {art['descripcion']}".lower():
                continue
            if categoria and categoria.lower() not in str(art.get("categoria") or "").lower():
                continue
            stock_cd = await _stock_cd(db, int(art["id"]))
            stock_jaula = await _stock_ubicacion(db, int(art["id"]), jaula_id)
            stock_oficina = await _stock_oficina(db, int(art["id"]))
            total = stock_cd + stock_jaula + stock_oficina
            minimo = float(art.get("stock_minimo") or 0)
            is_low = minimo > 0 and total < minimo
            if bajo_minimo and not is_low:
                continue
            avg = await _average_daily_consumption(db, int(art["id"]), 30)
            cost = float(art["costo_unitario"]) if art.get("costo_unitario") is not None else None
            items.append(
                {
                    **art,
                    "stock_cd": stock_cd,
                    "stock_jaula": stock_jaula,
                    "stock_oficina": stock_oficina,
                    "stock_total": total,
                    "valor_stock_cd": None if cost is None else stock_cd * cost,
                    "valor_stock_jaula": None if cost is None else stock_jaula * cost,
                    "valor_stock_oficina": None if cost is None else stock_oficina * cost,
                    "valor_stock_total": None if cost is None else total * cost,
                    "estado": "BAJO_MINIMO" if is_low else "OK",
                    "consumo_promedio_diario": avg,
                    "cobertura_dias": None if avg <= 0 else total / avg,
                }
            )
        latest_cd = await _fetch_one(
            db,
            "SELECT fecha_importacion, archivo_origen FROM stock_cd_importado ORDER BY fecha_importacion DESC, id DESC LIMIT 1",
        )
        latest_inv = await _fetch_one(
            db,
            "SELECT fecha_hora, fecha, turno FROM inventario_turno ORDER BY fecha_hora DESC, id DESC LIMIT 1",
        )
        move_where = "fecha_hora >= ?"
        move_args: list[Any] = [f"{_today()} 00:00:00"]
        if articulo_id:
            move_where += " AND articulo_id = ?"
            move_args.append(articulo_id)
        today_moves = await _fetch_one(
            db,
            f"SELECT COUNT(*) AS qty FROM movimientos WHERE {move_where}",
            tuple(move_args),
        )
    return {
        "items": items,
        "metrics": {
            "total_articulos": len(items),
            "bajo_minimo": sum(1 for item in items if item["estado"] == "BAJO_MINIMO"),
            "movimientos_hoy": int((today_moves or {}).get("qty") or 0),
            "articulos_sin_costo": sum(1 for item in items if item.get("costo_unitario") is None),
            "valor_stock_total": sum(float(item.get("valor_stock_total") or 0) for item in items),
            "valor_stock_bajo_minimo": sum(
                float(item.get("valor_stock_total") or 0) for item in items if item["estado"] == "BAJO_MINIMO"
            ),
            "ultima_importacion_cd": latest_cd,
            "ultimo_inventario_oficina": latest_inv,
        },
    }


async def _average_daily_consumption(db: aiosqlite.Connection, articulo_id: int, days: int) -> float:
    row = await _fetch_one(
        db,
        """
        SELECT COALESCE(SUM(consumo_calculado), 0) AS consumo,
               COUNT(DISTINCT fecha) AS dias
        FROM consumos_calculados
        WHERE articulo_id = ?
          AND consumo_calculado > 0
          AND fecha >= date('now', ?)
        """,
        (articulo_id, f"-{int(days)} days"),
    )
    dias = int((row or {}).get("dias") or 0)
    if dias <= 0:
        return 0.0
    return float((row or {}).get("consumo") or 0) / dias


@router.post("/inventario-turno")
async def save_inventory(req: InventoryRequest, request: Request):
    auth = await _require_panol_operator(request)
    turno = _norm_code(req.turno)
    fecha = _clean(req.fecha)
    if not fecha or not turno:
        raise HTTPException(status_code=400, detail="Fecha y turno son obligatorios.")
    if not req.items:
        raise HTTPException(status_code=400, detail="No hay articulos para guardar.")
    now = _now()
    observacion = _clean(req.observacion)
    results = []
    async with panol_db() as db:
        turn = await _fetch_one(db, "SELECT 1 FROM turnos WHERE codigo = ? AND activo = 1", (turno,))
        if not turn:
            raise HTTPException(status_code=400, detail="Turno invalido.")
        if req.ubicacion_id:
            location = await _fetch_one(db, "SELECT id, codigo FROM ubicaciones WHERE id = ? AND activo = 1", (req.ubicacion_id,))
        else:
            location = await _fetch_one(db, "SELECT id, codigo FROM ubicaciones WHERE codigo = ? AND activo = 1", (_norm_code(req.ubicacion_codigo or "OFICINA_ADO"),))
        if not location:
            raise HTTPException(status_code=400, detail="Ubicacion invalida.")
        location_id = int(location["id"])
        location_code = _norm_code(location["codigo"])
        is_office = location_code == "OFICINA_ADO"
        is_jaula = location_code == "JAULA"
        if not (is_office or is_jaula):
            raise HTTPException(status_code=400, detail="Solo se permite inventario de Oficina ADO o Jaula.")
        for item in req.items:
            if item.stock_fisico < 0:
                raise HTTPException(status_code=400, detail="El stock fisico no puede ser negativo.")
            art = await _fetch_one(db, "SELECT id, codigo FROM articulos WHERE id = ? AND activo = 1", (item.articulo_id,))
            if not art:
                raise HTTPException(status_code=404, detail=f"Articulo {item.articulo_id} no encontrado.")
            consumo = None
            ingresos = 0.0
            if is_office:
                previous = await _latest_inventory(db, item.articulo_id, location_id)
                stock_initial = float((previous or {}).get("stock_fisico") or 0)
                since = (previous or {}).get("fecha_hora")
                ingresos, egresos = await _office_transfers_since(db, item.articulo_id, since, now)
                consumo = stock_initial + ingresos - egresos - float(item.stock_fisico)
                if consumo < -0.000001 and not observacion:
                    raise HTTPException(
                        status_code=400,
                        detail="Hay diferencias positivas de inventario. Carga una observacion o registra un ajuste.",
                    )
            else:
                stock_initial = await _stock_ubicacion(db, item.articulo_id, location_id)
                diff = float(item.stock_fisico) - stock_initial
                if abs(diff) > 0.000001:
                    await db.execute(
                        """
                        INSERT INTO movimientos
                            (articulo_id, tipo, ubicacion_origen_id, ubicacion_destino_id, cantidad,
                             motivo, observacion, usuario, fecha_hora)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            item.articulo_id,
                            "AJUSTE_POSITIVO" if diff > 0 else "AJUSTE_NEGATIVO",
                            None if diff > 0 else location_id,
                            location_id if diff > 0 else None,
                            abs(diff),
                            "INVENTARIO_JAULA",
                            observacion or "Ajuste automatico por inventario fisico de Jaula",
                            auth.get("username"),
                            now,
                        ),
                    )
            cur = await db.execute(
                """
                INSERT INTO inventario_turno
                    (articulo_id, ubicacion_id, turno, fecha, stock_fisico, usuario, fecha_hora, observacion)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (item.articulo_id, location_id, turno, fecha, float(item.stock_fisico), auth.get("username"), now, observacion),
            )
            inv_id = cur.lastrowid
            if is_office:
                await db.execute(
                    """
                    INSERT INTO consumos_calculados
                        (articulo_id, fecha, turno, stock_inicial, ingresos_turno, stock_final,
                         consumo_calculado, usuario, fecha_hora, inventario_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        item.articulo_id,
                        fecha,
                        turno,
                        stock_initial,
                        ingresos,
                        float(item.stock_fisico),
                        consumo,
                        auth.get("username"),
                        now,
                        inv_id,
                    ),
                )
            results.append(
                {
                    "articulo_id": item.articulo_id,
                    "codigo": art["codigo"],
                    "ubicacion": location_code,
                    "stock_inicial": stock_initial,
                    "ingresos_turno": ingresos,
                    "stock_final": float(item.stock_fisico),
                    "consumo_calculado": consumo,
                    "diferencia": float(item.stock_fisico) - stock_initial if is_jaula else None,
                }
            )
        await db.commit()
    return {"ok": True, "items": results}


@router.get("/inventario-turno")
async def list_inventory(
    request: Request,
    fecha: str = Query(""),
    turno: str = Query(""),
    articulo_id: int | None = Query(None),
    ubicacion_id: int | None = Query(None),
):
    await _require_panol_operator(request)
    where = []
    args: list[Any] = []
    if fecha:
        where.append("i.fecha = ?")
        args.append(fecha)
    if turno:
        where.append("i.turno = ?")
        args.append(_norm_code(turno))
    if articulo_id:
        where.append("i.articulo_id = ?")
        args.append(articulo_id)
    if ubicacion_id:
        where.append("i.ubicacion_id = ?")
        args.append(ubicacion_id)
    clause = "WHERE " + " AND ".join(where) if where else ""
    async with panol_db() as db:
        rows = await _fetch_rows(
            db,
            f"""
            SELECT i.*, a.codigo, a.descripcion, u.codigo AS ubicacion_codigo,
                   c.stock_inicial, c.ingresos_turno,
                   c.consumo_calculado
            FROM inventario_turno i
            JOIN articulos a ON a.id = i.articulo_id
            LEFT JOIN ubicaciones u ON u.id = i.ubicacion_id
            LEFT JOIN consumos_calculados c ON c.inventario_id = i.id
            {clause}
            ORDER BY i.fecha_hora DESC, i.id DESC
            LIMIT 500
            """,
            tuple(args),
        )
    return {"items": rows}


def _read_stock_file(file_name: str, content: bytes) -> list[dict[str, Any]]:
    suffix = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""
    rows: list[dict[str, Any]] = []
    if suffix == "csv":
        text = None
        for encoding in ("utf-8-sig", "latin-1", "cp1252"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise HTTPException(status_code=400, detail="No se pudo leer el CSV con una codificacion soportada.")
        sample = text[:2048]
        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t") if sample.strip() else csv.excel
        for row in csv.DictReader(io.StringIO(text), dialect=dialect):
            rows.append({_norm_key(key): value for key, value in row.items()})
        return rows
    if suffix in {"xlsx", "xlsm"}:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return []
        headers = [_norm_key(v) for v in values[0]]
        for raw in values[1:]:
            rows.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))})
        return rows
    if suffix == "xls":
        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        if sheet.nrows == 0:
            return []
        headers = [_norm_key(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
        for row_idx in range(1, sheet.nrows):
            rows.append({headers[col]: sheet.cell_value(row_idx, col) for col in range(sheet.ncols)})
        return rows
    raise HTTPException(status_code=400, detail="Formato no soportado. Usa CSV o Excel .xlsx.")


def _row_value(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(_norm_key(key))
        if value is not None and str(value).strip() != "":
            return value
    return None


async def _latest_oracle_stock_cd_today(db: aiosqlite.Connection) -> dict[str, Any] | None:
    return await _fetch_one(
        db,
        """
        SELECT fecha_reporte, fecha_importacion, usuario_importacion
        FROM stock_cd_importado
        WHERE archivo_origen = ?
          AND substr(fecha_reporte, 1, 10) = ?
        ORDER BY fecha_importacion DESC, id DESC
        LIMIT 1
        """,
        (ORACLE_STOCK_CD_SOURCE, _today()),
    )


async def _sync_cd_stock_from_oracle(usuario: str, *, force: bool = False) -> dict[str, Any]:
    today = _today()
    now = _now()
    async with panol_db() as db:
        latest = await _latest_oracle_stock_cd_today(db)
        if latest and not force:
            return {
                "ok": True,
                "skipped": True,
                "reason": "El stock CD de Oracle ya fue sincronizado hoy.",
                "fecha_importacion": latest.get("fecha_importacion"),
                "fecha_reporte": latest.get("fecha_reporte"),
                "matched_count": 0,
                "ignored": 0,
                "zeroed": 0,
                "matched": [],
            }

    rows = await asyncio.to_thread(
        _query_productive_db_sql,
        QUERY_STOCK_CD_ORACLE,
        fecha_desde=f"{today} 00:00:00",
        fecha_hasta=now,
    )

    stock_by_code: dict[str, float] = {}
    ignored = 0
    for row in rows:
        codigo = _norm_code(row.get("REFERENCIA") or row.get("referencia"))
        if not codigo:
            ignored += 1
            continue
        stock_by_code[codigo] = stock_by_code.get(codigo, 0.0) + _to_float(row.get("UNIDADES") or row.get("unidades"))

    async with panol_db() as db:
        articles = await _fetch_rows(db, "SELECT id, codigo FROM articulos WHERE activo = 1")
        inserts = []
        matched = []
        zeroed = 0
        for art in articles:
            codigo = _norm_code(art["codigo"])
            stock_cd = max(float(stock_by_code.get(codigo, 0.0)), 0.0)
            if codigo in stock_by_code:
                matched.append({"articulo_id": int(art["id"]), "codigo": codigo, "stock_cd": stock_cd})
            else:
                zeroed += 1
            inserts.append((int(art["id"]), stock_cd, today, ORACLE_STOCK_CD_SOURCE, usuario, now))
        await db.executemany(
            """
            INSERT INTO stock_cd_importado
                (articulo_id, stock_cd, fecha_reporte, archivo_origen, usuario_importacion, fecha_importacion)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            inserts,
        )
        await db.commit()

    return {
        "ok": True,
        "skipped": False,
        "fecha_importacion": now,
        "fecha_reporte": today,
        "oracle_rows": len(rows),
        "matched": matched[:100],
        "matched_count": len(matched),
        "ignored": ignored + max(len(stock_by_code) - len(matched), 0),
        "zeroed": zeroed,
    }


def _stock_cd_scheduler_enabled() -> bool:
    return os.getenv("PANOL_STOCK_CD_ORACLE_SYNC_ENABLED", "1").strip().lower() in {"1", "true", "yes", "si"}


def _stock_cd_scheduler_time() -> time:
    raw = os.getenv("PANOL_STOCK_CD_ORACLE_SYNC_TIME", "07:00").strip()
    try:
        hour, minute = raw.split(":", 1)
        return time(int(hour), int(minute))
    except Exception:
        logger.warning("[panol-stock-cd] Hora invalida %r; se usa 07:00.", raw)
        return time(7, 0)


async def _stock_cd_scheduler_loop() -> None:
    global _stock_cd_scheduler_last_attempt
    assert _stock_cd_scheduler_stop is not None
    logger.info("[panol-stock-cd] Scheduler iniciado. Hora diaria: %s.", _stock_cd_scheduler_time().strftime("%H:%M"))
    while not _stock_cd_scheduler_stop.is_set():
        try:
            if _stock_cd_scheduler_enabled():
                now = datetime.now()
                today = now.strftime("%Y-%m-%d")
                if now.time() >= _stock_cd_scheduler_time() and _stock_cd_scheduler_last_attempt != today:
                    _stock_cd_scheduler_last_attempt = today
                    result = await _sync_cd_stock_from_oracle("scheduler", force=False)
                    if result.get("skipped"):
                        logger.info("[panol-stock-cd] Stock CD ya sincronizado para %s.", today)
                    else:
                        logger.info(
                            "[panol-stock-cd] Stock CD sincronizado: %s PLUs propios, %s sin stock, %s filas Oracle.",
                            result.get("matched_count"),
                            result.get("zeroed"),
                            result.get("oracle_rows"),
                        )
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            logger.warning("[panol-stock-cd] No se pudo sincronizar stock CD desde Oracle: %s", exc)
        try:
            await asyncio.wait_for(_stock_cd_scheduler_stop.wait(), timeout=300)
        except asyncio.TimeoutError:
            pass
    logger.info("[panol-stock-cd] Scheduler detenido.")


def start_panol_stock_cd_scheduler() -> None:
    global _stock_cd_scheduler_task, _stock_cd_scheduler_stop
    if _stock_cd_scheduler_task and not _stock_cd_scheduler_task.done():
        return
    _stock_cd_scheduler_stop = asyncio.Event()
    _stock_cd_scheduler_task = asyncio.create_task(_stock_cd_scheduler_loop())


async def stop_panol_stock_cd_scheduler() -> None:
    global _stock_cd_scheduler_task, _stock_cd_scheduler_stop
    if _stock_cd_scheduler_stop:
        _stock_cd_scheduler_stop.set()
    if _stock_cd_scheduler_task:
        _stock_cd_scheduler_task.cancel()
        try:
            await _stock_cd_scheduler_task
        except asyncio.CancelledError:
            pass
    _stock_cd_scheduler_task = None
    _stock_cd_scheduler_stop = None


@router.post("/importar-stock-cd")
async def import_cd_stock(
    request: Request,
    filename: str = Query("stock_cd.csv"),
    preview: bool = Query(True),
    fecha_reporte: str = Query(""),
):
    auth = await _require_panol_operator(request, full=True)
    content = await request.body()
    raw_rows = _read_stock_file(filename or "stock_cd", content)
    now = _now()
    matched = []
    ignored = 0
    errors = []
    async with panol_db() as db:
        articles = await _fetch_rows(db, "SELECT id, codigo FROM articulos WHERE activo = 1")
        by_code = {_norm_code(row["codigo"]): int(row["id"]) for row in articles}
        for index, row in enumerate(raw_rows, start=2):
            codigo = _norm_code(_row_value(row, "codigo", "cod", "plu", "referencia", "referencia_plu", "articulo"))
            if not codigo or codigo not in by_code:
                ignored += 1
                continue
            try:
                stock_cd = _to_float(
                    _row_value(
                        row,
                        "stock_cd",
                        "stock",
                        "cantidad",
                        "unidades",
                        "unidad",
                        "cant_unidades",
                        "cant.unidades",
                        "cant unidades",
                    )
                )
            except ValueError:
                errors.append({"fila": index, "codigo": codigo, "error": "Stock CD invalido"})
                continue
            if stock_cd < 0:
                errors.append({"fila": index, "codigo": codigo, "error": "Stock CD negativo"})
                continue
            matched.append({"articulo_id": by_code[codigo], "codigo": codigo, "stock_cd": stock_cd})
        if errors:
            return {"ok": False, "preview": True, "matched": matched[:100], "matched_count": len(matched), "ignored": ignored, "errors": errors[:100]}
        if not preview:
            await db.executemany(
                """
                INSERT INTO stock_cd_importado
                    (articulo_id, stock_cd, fecha_reporte, archivo_origen, usuario_importacion, fecha_importacion)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        row["articulo_id"],
                        row["stock_cd"],
                        _clean(fecha_reporte) or now,
                        filename or "",
                        auth.get("username"),
                        now,
                    )
                    for row in matched
                ],
            )
            await db.commit()
    return {
        "ok": True,
        "preview": preview,
        "matched": matched[:100],
        "matched_count": len(matched),
        "ignored": ignored,
        "errors": [],
        "fecha_importacion": now if not preview else None,
    }


@router.post("/stock-cd/oracle/sincronizar")
async def sync_cd_stock_oracle(request: Request, force: bool = Query(False)):
    auth = await _require_panol_operator(request, full=True)
    try:
        return await _sync_cd_stock_from_oracle(str(auth.get("username") or ""), force=force)
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo sincronizar stock CD desde Oracle: {exc}") from exc


@router.get("/indicadores")
async def indicators(
    request: Request,
    fecha_desde: str = Query(""),
    fecha_hasta: str = Query(""),
    articulo_id: int | None = Query(None),
):
    await _require_panol_operator(request)
    where = ["c.consumo_calculado > 0"]
    args: list[Any] = []
    if fecha_desde:
        where.append("date(datetime(c.fecha_hora, '+2 hours')) >= ?")
        args.append(fecha_desde)
    if fecha_hasta:
        where.append("date(datetime(c.fecha_hora, '+2 hours')) <= ?")
        args.append(fecha_hasta)
    if articulo_id:
        where.append("c.articulo_id = ?")
        args.append(articulo_id)
    clause = " AND ".join(where)
    consumption_cost = _cost_at_sql("c.articulo_id", "c.fecha_hora")
    async with panol_db() as db:
        daily = await _fetch_rows(
            db,
            f"""
            SELECT c.articulo_id, a.codigo, a.descripcion,
                   date(datetime(c.fecha_hora, '+2 hours')) AS dia_logistico,
                   SUM(c.consumo_calculado) AS consumo,
                   SUM(c.consumo_calculado * {consumption_cost}) AS valor_consumo
            FROM consumos_calculados c
            JOIN articulos a ON a.id = c.articulo_id
            WHERE {clause}
            GROUP BY c.articulo_id, dia_logistico
            ORDER BY dia_logistico DESC, a.codigo
            LIMIT 200
            """,
            tuple(args),
        )
        by_logistic_day = await _fetch_rows(
            db,
            f"""
            SELECT date(datetime(c.fecha_hora, '+2 hours')) AS dia_logistico,
                   SUM(c.consumo_calculado) AS consumo,
                   SUM(c.consumo_calculado * {consumption_cost}) AS valor_consumo
            FROM consumos_calculados c
            JOIN articulos a ON a.id = c.articulo_id
            WHERE {clause}
            GROUP BY dia_logistico
            ORDER BY dia_logistico DESC
            """,
            tuple(args),
        )
    return {"consumo_diario": daily, "consumo_por_dia_logistico": by_logistic_day}


@router.get("/exportar-stock")
async def export_stock(request: Request):
    await _require_panol_operator(request)
    data = await stock(request)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "codigo",
        "descripcion",
        "categoria",
        "unidad",
        "stock_cd",
        "stock_jaula",
        "stock_oficina",
        "stock_total",
        "costo_unitario",
        "valor_stock_total",
        "stock_minimo",
        "estado",
        "cobertura_dias",
    ])
    for row in data["items"]:
        writer.writerow(
            [
                row.get("codigo"),
                row.get("descripcion"),
                row.get("categoria"),
                row.get("unidad"),
                row.get("stock_cd"),
                row.get("stock_jaula"),
                row.get("stock_oficina"),
                row.get("stock_total"),
                row.get("costo_unitario"),
                row.get("valor_stock_total"),
                row.get("stock_minimo"),
                row.get("estado"),
                row.get("cobertura_dias"),
            ]
        )
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=panol_stock.csv"},
    )
