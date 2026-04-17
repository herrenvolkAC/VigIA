"""
VigIA · cargar_historico.py
Extrae el historico del Excel de Dia Logistico y lo carga en vigia.db

Uso:
    python cargar_historico.py

Requiere:
    pip install openpyxl

Colocar en C:\\Ingenieria\\VigIA\\  junto con el archivo Excel.
"""

import sqlite3, re
from pathlib import Path
from datetime import date, timedelta
from openpyxl import load_workbook

EXCEL_PATH = Path(__file__).parent / "Copia_de_Día_Logístico_Olas__Template__v2_nuevo__1_.xlsx"
DB_PATH    = Path(__file__).parent / "vigia.db"

HOJAS_IGNORAR = {'52800','BASEE', 'Base', 'base1', 'Hoja1', 'Base S 18hs', 'Base S-18hs, D y L'}

# Programa 5193 arranca el 2023-08-19 (dia siguiente a la ultima hoja con fecha)
PROGRAMA_BASE  = 5193
FECHA_BASE     = date(2023, 8, 19)

# Filas de cada ola (basadas en el indice 0 de iter_rows)
# col 1=label, 2=progSecos, 3=progNOA, 4=progTotal,
# col 5=ejecSecos, 6=ejecNOA, 7=ejecTotal,
# col 10=dotProg, 12=dotEjec, 13=prodProg, 14=prodEjec
OLA_FILAS = {
    6:  {'num':1, 'label':'OLA 1', 'ini':'14:00','fin':'16:00','turno':'Tarde'},
    7:  {'num':2, 'label':'OLA 2', 'ini':'16:00','fin':'18:00','turno':'Tarde'},
    8:  {'num':3, 'label':'OLA 3', 'ini':'18:00','fin':'20:00','turno':'Tarde'},
    9:  {'num':4, 'label':'OLA 4', 'ini':'20:00','fin':'22:00','turno':'Tarde'},
    11: {'num':5, 'label':'OLA 5', 'ini':'22:00','fin':'00:00','turno':'Noche'},
    12: {'num':6, 'label':'OLA 6', 'ini':'00:00','fin':'02:00','turno':'Noche'},
    13: {'num':7, 'label':'OLA 7', 'ini':'02:00','fin':'04:00','turno':'Noche'},
    14: {'num':8, 'label':'OLA 8', 'ini':'04:00','fin':'06:00','turno':'Noche'},
    16: {'num':9, 'label':'OLA 9', 'ini':'06:00','fin':'08:00','turno':'Mañana'},
    17: {'num':10,'label':'OLA 10','ini':'08:00','fin':'10:00','turno':'Mañana'},
    18: {'num':11,'label':'OLA 11','ini':'10:00','fin':'12:00','turno':'Mañana'},
    19: {'num':12,'label':'OLA 12','ini':'12:00','fin':'14:00','turno':'Mañana'},
}

def sf(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def si(v):
    f = sf(v)
    return int(f) if 0 < f < 150 else 0  # dotacion validada

def parse_fecha(nombre):
    n = nombre.strip()
    m = re.match(r'^(\d{1,2})-(\d{2})-(\d{2})$', n)
    if m:
        d,mo,y=int(m.group(1)),int(m.group(2)),int(m.group(3))
        try: return date(2000+y,mo,d).isoformat(), None
        except: return None, None
    m = re.match(r'^(\d{1,2})-(\d{1,2})$', n)
    if m:
        d,mo=int(m.group(1)),int(m.group(2))
        anio=2022 if mo>=7 else 2023
        try: return date(anio,mo,d).isoformat(), None
        except: return None, None
    m = re.match(r'^(\d{4,5})$', n)
    if m:
        p=int(m.group(1))
        return (FECHA_BASE+timedelta(days=p-PROGRAMA_BASE)).isoformat(), p
    return None, None

def extraer(ws):
    rows = [list(r) for r in ws.iter_rows(max_row=35, values_only=True)]

    def r(i): return rows[i] if i < len(rows) else [None]*20

    # Numero de programa (fila 2, col B = indice 1)
    programa = None
    fila3 = r(2)
    if fila3 and isinstance(fila3[1], (int,float)) and fila3[1] > 1000:
        programa = int(fila3[1])

    # Olas (indice 0-based)
    olas = []
    for idx, meta in OLA_FILAS.items():
        row = r(idx)
        if not row or row[1] is None: continue
        if 'OLA' not in str(row[1]).upper(): continue

        pt = sf(row[4]); et = sf(row[7])
        if pt == 0 and et == 0: continue

        olas.append({**meta,
            'prog_secos': round(sf(row[2]),1),
            'prog_noa':   round(sf(row[3]),1),
            'prog_total': round(pt,1),
            'ejec_secos': round(sf(row[5]),1),
            'ejec_noa':   round(sf(row[6]),1),
            'ejec_total': round(et,1),
            'desvio':     round(et-pt,1),
            'dot_prog':   si(row[10]),
            'dot_ejec':   si(row[12]),
            'prod_prog':  round(sf(row[13]),1),
            'prod_ejec':  round(sf(row[14]),1),
        })

    # Totales por turno (filas 10,15,20 = indices 10,15,20)
    turnos = {}
    for idx_t, nt in [(10,'Tarde'),(15,'Noche'),(20,'Mañana')]:
        row = r(idx_t)
        if row and row[1] and 'Turno' in str(row[1]):
            turnos[nt] = {'prog_total': round(sf(row[4]),1),
                          'ejec_total': round(sf(row[7]),1)}

    # Ausentismo (filas 24,25,26 = indices 24,25,26)
    aus = {}
    for idx_a, na in [(24,'Tarde'),(25,'Noche'),(26,'Mañana')]:
        row = r(idx_a)
        if row and row[1]:
            aus[na] = {'secos': int(sf(row[2])), 'noa': int(sf(row[3]))}

    return {'programa': programa, 'olas': olas, 'turnos': turnos, 'aus': aus}

def init_db(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS historico_dias (
            dia_id       INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha        DATE    NOT NULL UNIQUE,
            hoja_nombre  TEXT,
            programa     INTEGER,
            prog_total   REAL DEFAULT 0,
            ejec_total   REAL DEFAULT 0,
            desvio_dia   REAL DEFAULT 0,
            aus_tarde_s  INTEGER DEFAULT 0,
            aus_tarde_n  INTEGER DEFAULT 0,
            aus_noche_s  INTEGER DEFAULT 0,
            aus_noche_n  INTEGER DEFAULT 0,
            aus_manana_s INTEGER DEFAULT 0,
            aus_manana_n INTEGER DEFAULT 0,
            dia_semana   INTEGER,
            created_at   DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS historico_olas (
            ola_id     INTEGER PRIMARY KEY AUTOINCREMENT,
            dia_id     INTEGER NOT NULL REFERENCES historico_dias(dia_id),
            fecha      DATE NOT NULL,
            ola_num    INTEGER NOT NULL,
            ola_label  TEXT,
            hora_ini   TEXT,
            hora_fin   TEXT,
            turno      TEXT,
            prog_secos REAL DEFAULT 0,
            prog_noa   REAL DEFAULT 0,
            prog_total REAL DEFAULT 0,
            ejec_secos REAL DEFAULT 0,
            ejec_noa   REAL DEFAULT 0,
            ejec_total REAL DEFAULT 0,
            desvio     REAL DEFAULT 0,
            dot_prog   INTEGER DEFAULT 0,
            dot_ejec   INTEGER DEFAULT 0,
            prod_prog  REAL DEFAULT 0,
            prod_ejec  REAL DEFAULT 0
        );
        CREATE INDEX IF NOT EXISTS idx_hdias_fecha  ON historico_dias(fecha);
        CREATE INDEX IF NOT EXISTS idx_hdias_semana ON historico_dias(dia_semana);
        CREATE INDEX IF NOT EXISTS idx_holas_fecha  ON historico_olas(fecha);
        CREATE INDEX IF NOT EXISTS idx_holas_turno  ON historico_olas(turno);
        CREATE INDEX IF NOT EXISTS idx_holas_num    ON historico_olas(ola_num);
    """)
    conn.commit()

def main():
    print("=" * 55)
    print("VigIA - Carga de Historico")
    print("=" * 55)

    if not EXCEL_PATH.exists():
        print(f"\nNo encontre el Excel: {EXCEL_PATH.name}")
        print("Copialo a la carpeta del proyecto y volvé a correr.")
        input("\nPresiona Enter para cerrar...")
        return

    if not DB_PATH.exists():
        print(f"\nNo encontre vigia.db")
        print("Corre 'python main.py' primero.")
        input("\nPresiona Enter para cerrar...")
        return

    print(f"\n  Excel : {EXCEL_PATH.name}")
    print(f"  BD    : {DB_PATH.name}")

    conn = sqlite3.connect(str(DB_PATH))
    init_db(conn)

    print("\nCargando Excel...")
    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    hojas = [h for h in wb.sheetnames if h.strip() not in HOJAS_IGNORAR]
    print(f"{len(hojas)} hojas a procesar\n")

    ok=0; sin_fecha=0; vacias=0; errores=0; duplicadas=0

    for i, h in enumerate(hojas, 1):
        if i % 100 == 0:
            print(f"  {i}/{len(hojas)} - {ok} cargadas OK")

        fecha, prog_num = parse_fecha(h)
        if fecha is None:
            sin_fecha += 1
            continue

        try:
            ws = wb[h]
            d = extraer(ws)

            if not d['olas']:
                vacias += 1
                continue

            pt = (d['turnos'].get('Tarde',{}).get('prog_total',0) +
                  d['turnos'].get('Noche',{}).get('prog_total',0) +
                  d['turnos'].get('Mañana',{}).get('prog_total',0))
            et = (d['turnos'].get('Tarde',{}).get('ejec_total',0) +
                  d['turnos'].get('Noche',{}).get('ejec_total',0) +
                  d['turnos'].get('Mañana',{}).get('ejec_total',0))
            if pt == 0: pt = sum(o['prog_total'] for o in d['olas'])
            if et == 0: et = sum(o['ejec_total'] for o in d['olas'])

            ds = date.fromisoformat(fecha).weekday()
            a  = d['aus']

            try:
                cur = conn.execute(
                    """INSERT INTO historico_dias
                       (fecha,hoja_nombre,programa,prog_total,ejec_total,desvio_dia,
                        aus_tarde_s,aus_tarde_n,aus_noche_s,aus_noche_n,
                        aus_manana_s,aus_manana_n,dia_semana)
                       VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (fecha, h.strip(), d['programa'] or prog_num,
                     round(pt,1), round(et,1), round(et-pt,1),
                     a.get('Tarde',{}).get('secos',0), a.get('Tarde',{}).get('noa',0),
                     a.get('Noche',{}).get('secos',0), a.get('Noche',{}).get('noa',0),
                     a.get('Mañana',{}).get('secos',0), a.get('Mañana',{}).get('noa',0),
                     ds))
                dia_id = cur.lastrowid
            except sqlite3.IntegrityError:
                duplicadas += 1
                continue

            conn.executemany(
                """INSERT INTO historico_olas
                   (dia_id,fecha,ola_num,ola_label,hora_ini,hora_fin,turno,
                    prog_secos,prog_noa,prog_total,ejec_secos,ejec_noa,ejec_total,
                    desvio,dot_prog,dot_ejec,prod_prog,prod_ejec)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                [(dia_id,fecha,o['num'],o['label'],o['ini'],o['fin'],o['turno'],
                  o['prog_secos'],o['prog_noa'],o['prog_total'],
                  o['ejec_secos'],o['ejec_noa'],o['ejec_total'],o['desvio'],
                  o['dot_prog'],o['dot_ejec'],o['prod_prog'],o['prod_ejec'])
                 for o in d['olas']])

            conn.commit()
            ok += 1

        except Exception as e:
            errores += 1
            print(f"  Error en '{h}': {e}")

    r1 = conn.execute("SELECT COUNT(*) FROM historico_dias").fetchone()[0]
    r2 = conn.execute("SELECT COUNT(*) FROM historico_olas").fetchone()[0]
    r3 = conn.execute("SELECT MIN(fecha),MAX(fecha) FROM historico_dias").fetchone()

    print(f"\n{'='*55}")
    print(f"COMPLETADO")
    print(f"{'='*55}")
    print(f"  Dias cargados   : {ok}")
    print(f"  Sin fecha       : {sin_fecha}")
    print(f"  Hojas vacias    : {vacias}")
    print(f"  Duplicadas      : {duplicadas}")
    print(f"  Errores         : {errores}")
    print(f"\n  En la base de datos:")
    print(f"  - Dias logisticos : {r1}")
    print(f"  - Registros olas  : {r2}")
    if r3[0]: print(f"  - Rango           : {r3[0]} a {r3[1]}")
    print(f"\n  Conectate en DBeaver a:")
    print(f"  {DB_PATH}")
    print(f"\n  Tablas:")
    print(f"  - historico_dias  (un registro por dia logistico)")
    print(f"  - historico_olas  (un registro por ola por dia)")

    conn.close()
    input("\nPresiona Enter para cerrar...")

if __name__ == "__main__":
    main()