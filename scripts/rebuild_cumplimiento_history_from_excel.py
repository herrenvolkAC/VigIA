"""
Recarga historico_dias e historico_olas desde el Excel de cumplimiento.

Reglas de negocio:
- Solo usa hojas numéricas.
- Hoja 5441 => ayer respecto de la fecha base configurada.
- 5440 => antes de ayer, y así sucesivamente.
- Ignora hojas posteriores al ancla (ej. 5442+).
- No inserta datos con fecha anterior a 1 año desde la fecha ancla.
- Si una hoja está incompleta o tiene formato inesperado, la saltea y sigue.
"""

from __future__ import annotations

import re
import sqlite3
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "vigia.db"
DEFAULT_WORKBOOK = Path(r"C:\Users\207189\Downloads\Copia de Día Logístico Olas (Template) v2 nuevo (2).xlsx")
ANCHOR_SHEET = 5441
ANCHOR_DATE = date(2026, 4, 21)  # ayer respecto de 2026-04-22
MAX_DAYS_BACK = 366

OLA_REGEX = re.compile(r"^\s*(\d{2}:\d{2})\s*a\s*(\d{2}:\d{2})\s+OLA\s+(\d+)\s*$", re.IGNORECASE)


@dataclass
class DiaRecord:
    fecha: str
    hoja_nombre: str
    programa: int
    prog_total: float
    ejec_total: float
    desvio_dia: float
    aus_tarde_s: int
    aus_tarde_n: int
    aus_noche_s: int
    aus_noche_n: int
    aus_manana_s: int
    aus_manana_n: int
    dia_semana: int


@dataclass
class OlaRecord:
    fecha: str
    ola_num: int
    ola_label: str
    hora_ini: str
    hora_fin: str
    turno: str
    prog_secos: float
    prog_noa: float
    prog_total: float
    ejec_secos: float
    ejec_noa: float
    ejec_total: float
    desvio: float
    dot_prog: int
    dot_ejec: int
    prod_prog: float
    prod_ejec: float


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch)).lower()


def to_float(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def to_int(value) -> int:
    return int(round(to_float(value)))


def infer_turno(ola_num: int) -> str:
    if 1 <= ola_num <= 4:
        return "Tarde"
    if 5 <= ola_num <= 8:
        return "Noche"
    if 9 <= ola_num <= 12:
        return "Manana"
    return "Otro"


def iter_sheet_rows(ws) -> Iterable[tuple]:
    for row in ws.iter_rows(values_only=True):
        yield row


def parse_sheet(ws, logical_date: date) -> tuple[DiaRecord, list[OlaRecord]]:
    programa = None
    prog_total = ejec_total = desvio_dia = 0.0
    aus = {
        "tarde_s": 0,
        "tarde_n": 0,
        "noche_s": 0,
        "noche_n": 0,
        "manana_s": 0,
        "manana_n": 0,
    }
    olas: list[OlaRecord] = []
    saw_ausentismo_header = False

    for row in iter_sheet_rows(ws):
        row = tuple(row) + (None,) * (20 - len(row))
        label = row[1]
        label_norm = normalize_text(label)

        if label_norm == "programa n°":
            continue

        if isinstance(row[1], (int, float)) and programa is None:
            programa = int(row[1])

        if label_norm == "dia logistico" and any(row[idx] is not None for idx in (4, 7, 8)):
            prog_total = to_float(row[4])
            ejec_total = to_float(row[7])
            desvio_dia = to_float(row[8])
            continue

        if label_norm == "ausentismo":
            saw_ausentismo_header = True
            continue

        if saw_ausentismo_header and label_norm.startswith("turno "):
            secos = to_int(row[2])
            noa = to_int(row[3])
            if "tarde" in label_norm:
                aus["tarde_s"], aus["tarde_n"] = secos, noa
            elif "noche" in label_norm:
                aus["noche_s"], aus["noche_n"] = secos, noa
            elif "manana" in label_norm:
                aus["manana_s"], aus["manana_n"] = secos, noa
            continue

        match = OLA_REGEX.match(str(label).strip()) if label else None
        if not match:
            continue

        hora_ini, hora_fin, ola_num_raw = match.groups()
        ola_num = int(ola_num_raw)
        olas.append(
            OlaRecord(
                fecha=logical_date.isoformat(),
                ola_num=ola_num,
                ola_label=str(label).strip(),
                hora_ini=hora_ini,
                hora_fin=hora_fin,
                turno=infer_turno(ola_num),
                prog_secos=to_float(row[2]),
                prog_noa=to_float(row[3]),
                prog_total=to_float(row[4]),
                ejec_secos=to_float(row[5]),
                ejec_noa=to_float(row[6]),
                ejec_total=to_float(row[7]),
                desvio=to_float(row[8]),
                dot_prog=to_int(row[10]),
                dot_ejec=to_int(row[12]),
                prod_prog=to_float(row[13]),
                prod_ejec=to_float(row[14]),
            )
        )

    if programa is None:
        raise ValueError("No se encontró el número de programa")
    if not olas:
        raise ValueError("No se encontraron filas de OLA válidas")

    dia = DiaRecord(
        fecha=logical_date.isoformat(),
        hoja_nombre=ws.title,
        programa=programa,
        prog_total=prog_total,
        ejec_total=ejec_total,
        desvio_dia=desvio_dia,
        aus_tarde_s=aus["tarde_s"],
        aus_tarde_n=aus["tarde_n"],
        aus_noche_s=aus["noche_s"],
        aus_noche_n=aus["noche_n"],
        aus_manana_s=aus["manana_s"],
        aus_manana_n=aus["manana_n"],
        dia_semana=logical_date.weekday(),
    )
    return dia, olas


def main() -> int:
    workbook_path = DEFAULT_WORKBOOK
    if len(sys.argv) > 1:
        workbook_path = Path(sys.argv[1])
    if not workbook_path.exists():
        raise FileNotFoundError(f"No existe el Excel: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    numeric_sheets = sorted((int(name), name) for name in wb.sheetnames if name.isdigit())
    eligible = [(num, name) for num, name in numeric_sheets if num <= ANCHOR_SHEET]

    min_allowed_date = ANCHOR_DATE - timedelta(days=MAX_DAYS_BACK - 1)
    parsed: list[tuple[DiaRecord, list[OlaRecord]]] = []
    warnings: list[str] = []

    for sheet_num, sheet_name in sorted(eligible, reverse=True):
        logical_date = ANCHOR_DATE - timedelta(days=(ANCHOR_SHEET - sheet_num))
        if logical_date < min_allowed_date:
            continue
        try:
            dia, olas = parse_sheet(wb[sheet_name], logical_date)
            parsed.append((dia, olas))
        except Exception as exc:
            warnings.append(f"{sheet_name}: {exc}")

    wb.close()

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("PRAGMA foreign_keys = ON")
    cur.execute("BEGIN")
    cur.execute("DELETE FROM historico_olas")
    cur.execute("DELETE FROM historico_dias")

    inserted_days = 0
    inserted_olas = 0
    for dia, olas in sorted(parsed, key=lambda item: item[0].fecha):
        cur.execute(
            """
            INSERT INTO historico_dias
            (fecha, hoja_nombre, programa, prog_total, ejec_total, desvio_dia,
             aus_tarde_s, aus_tarde_n, aus_noche_s, aus_noche_n, aus_manana_s, aus_manana_n, dia_semana)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                dia.fecha,
                dia.hoja_nombre,
                dia.programa,
                dia.prog_total,
                dia.ejec_total,
                dia.desvio_dia,
                dia.aus_tarde_s,
                dia.aus_tarde_n,
                dia.aus_noche_s,
                dia.aus_noche_n,
                dia.aus_manana_s,
                dia.aus_manana_n,
                dia.dia_semana,
            ),
        )
        dia_id = cur.lastrowid
        inserted_days += 1

        for ola in olas:
            cur.execute(
                """
                INSERT INTO historico_olas
                (dia_id, fecha, ola_num, ola_label, hora_ini, hora_fin, turno,
                 prog_secos, prog_noa, prog_total, ejec_secos, ejec_noa, ejec_total,
                 desvio, dot_prog, dot_ejec, prod_prog, prod_ejec)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    dia_id,
                    ola.fecha,
                    ola.ola_num,
                    ola.ola_label,
                    ola.hora_ini,
                    ola.hora_fin,
                    ola.turno,
                    ola.prog_secos,
                    ola.prog_noa,
                    ola.prog_total,
                    ola.ejec_secos,
                    ola.ejec_noa,
                    ola.ejec_total,
                    ola.desvio,
                    ola.dot_prog,
                    ola.dot_ejec,
                    ola.prod_prog,
                    ola.prod_ejec,
                ),
            )
            inserted_olas += 1

    conn.commit()

    min_fecha = cur.execute("SELECT MIN(fecha) FROM historico_dias").fetchone()[0]
    max_fecha = cur.execute("SELECT MAX(fecha) FROM historico_dias").fetchone()[0]
    conn.close()

    print(f"Workbook: {workbook_path}")
    print(f"Hojas elegibles: {len(eligible)}")
    print(f"Dias insertados: {inserted_days}")
    print(f"Olas insertadas: {inserted_olas}")
    print(f"Rango insertado: {min_fecha} -> {max_fecha}")
    print(f"Warnings: {len(warnings)}")
    for warning in warnings[:20]:
        print(f" - {warning}")
    if len(warnings) > 20:
        print(f" - ... {len(warnings) - 20} warnings más")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
